"""
Alpha Containers - MRP Updater v1
----------------------------------------------------------------------
PURPOSE:
  Reads pending.xlsx (latest sheet = most recent date) and updates
  the MRP order rows (rows 3-16) in the AlphaContainers file.

WHAT IT UPDATES (MRP sheet only):
  Col A  Dia            - from Product_Catalog via alias resolution
  Col B  Customer       - from Product_Catalog via alias resolution
  Col C  Product Name   - catalog name (resolved from pending alias)
  Col D  PID            - formula =INDEX(Product_Catalog!A:A,MATCH(...))
  Col E  Job Order #    - from pending file (col A = JOB NO)
  Col F  Required Qty   - hardcoded = produced_to_date + pending_balance
                          (recalculated each run; stays accurate between runs
                          because G formula auto-updates from Production_Log)
  Col G  Produced       - live SUMPRODUCT formula from Production_Log
                          (Printing-03 and Printing-04 machines only)
  Col H  Remaining Bal  - formula =F-G (unchanged)

DESIGN RULE — FIXED ZONE:
  Order rows are always written to rows 3-16 (14 slots max).
  Rows are never inserted or deleted. This preserves all formula
  references in the MRP material section ($D$3:$D$16, $H$3:$H$16,
  and the individual $D$3...$D$16 refs in cols I/J/K).
  Unused slots in rows 3-16 are left blank.

ALIASES:
  Maps (pending_product_name.upper().strip(), dia_float) to catalog name.
  Add entries here when Imran's pending file uses a different name
  than the Product_Catalog. Dia is used to disambiguate products with
  the same name at different diameters (e.g. SAMSOL 43 at 19mm vs 25mm).

FOLDER STRUCTURE:
  AlphaContainers_v9_xx.xlsx   (any version - latest is used)
  pending.xlsx                  (Imran's pending orders file)
  update_mrp.py                 (this script)
  Update MRP.bat                (double-click launcher)
"""

import os
import re
import glob
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment


# -----------------------------------------------------------------------
# ALIASES
# Maps (pending_name.strip().upper(), dia_as_float) -> catalog_product_name
# dia_as_float: use None if dia is ambiguous or not needed for this product
#
# TO ADD A NEW ENTRY: copy any line below.
# Key = (PENDING NAME UPPERCASE, dia as float or None)
# Val = exact Product_Catalog col D name
#
# If the same pending name appears at different diameters, add separate
# entries with specific dia values to disambiguate.
# -----------------------------------------------------------------------
ALIASES = {
    # Pending name               Dia     Catalog name
    ("ECZEMUS OINTMENT",         16.0):  "ECZEMUS OINTMENT 0.03% 10G",
    ("PHLOGIN GEL",              19.0):  "PHLOGIN GEL 20G",
    ("CONTRATUBEX GEL",          19.0):  "CONTRATUBEX GEL 20G",
    ("PYODINE GEL",              19.0):  "PYODINE GEL 20GM",
    ("SAMSOL 43",                19.0):  "S-43 DIA 19",
    ("S-43 DIA 19",              19.0):  "S-43 DIA 19",
    ("SAMSOL 43",                20.5):  "S-43 DIA 20.5",
    ("SAMSOL 45",                20.5):  "S-45 DIA 20.5",
    ("SAMSOL 43",                25.0):  "S 43 25MM",
    ("S 43 25MM",                25.0):  "S 43 25MM",
    ("TUBES MENS",               25.0):  "TUBES MEN BLUE",
    ("TUBE MENS",                25.0):  "TUBES MEN BLUE",
    ("TUBES MEN",                25.0):  "TUBES MEN BLUE",
    ("TUBES MEN BLUE",           25.0):  "TUBES MEN BLUE",
    ("TUBES COMMON RED",         25.0):  "TUBES COMMON RED",
    ("TUBES",                    25.0):  "TUBES",
    ("TUBE COMMON PURPLE",       25.0):  "TUBE COMMON PURPLE",
    ("GOLDEN PEARL",             30.0):  "HELLO HAIR COLOR",
    ("GOLDEN PEARL ",            30.0):  "HELLO HAIR COLOR",   # trailing space variant
    ("HELLO HAIR COLOR",         30.0):  "HELLO HAIR COLOR",
    ("DOWFEN GEL",               30.0):  "DOWFEN GEL 50G",
    ("DOWFEN GEL 50G",           30.0):  "DOWFEN GEL 50G",
    ("MEGA GREY",                32.0):  "M.G",
    ("M.G",                      32.0):  "M.G",
    ("HOLA HAIR",                32.0):  "H.H 100GM",
    ("H.H 100GM",                32.0):  "H.H 100GM",
    ("ANVIL 43",                 35.0):  "ANVIL 43",
    ("ANVIL 45",                 35.0):  "ANVIL 45",
    ("V- HC BROWN",              35.0):  "V- HC BROWN",
    ("V-HC BLACK",               35.0):  "V-HC BLACK",
    # Dia-agnostic fallbacks (used when dia cannot be determined from pending)
    ("ECZEMUS OINTMENT",         None):  "ECZEMUS OINTMENT 0.03% 10G",
    ("GOLDEN PEARL",             None):  "HELLO HAIR COLOR",
    ("GOLDEN PEARL ",            None):  "HELLO HAIR COLOR",
    ("DOWFEN GEL",               None):  "DOWFEN GEL 50G",
    ("MEGA GREY",                None):  "M.G",
    ("HOLA HAIR",                None):  "H.H 100GM",
}

# -----------------------------------------------------------------------
# Constants
# -----------------------------------------------------------------------
MRP_SHEET        = "MRP"
CATALOG_SHEET    = "Product_Catalog"
PROD_LOG_SHEET   = "Production_Log"

ORDER_ROW_START  = 3
ORDER_ROW_END    = 16   # Fixed zone — DO NOT CHANGE without rebuilding material formulas
ORDER_ROW_MAX    = ORDER_ROW_END - ORDER_ROW_START + 1  # = 14 slots

# MRP column indices (1-based)
COL_DIA      = 1   # A
COL_CUSTOMER = 2   # B
COL_PRODUCT  = 3   # C
COL_PID      = 4   # D
COL_JOB      = 5   # E
COL_REQ      = 6   # F
COL_PROD     = 7   # G
COL_REM      = 8   # H

# Product_Catalog column indices (1-based)
CAT_PID      = 1   # A
CAT_BOMID    = 2   # B
CAT_CUST     = 3   # C
CAT_NAME     = 4   # D
CAT_DIA      = 5   # E

# Production_Log column indices (1-based)
PLOG_MACHINE  = 2   # B
PLOG_PID      = 6   # F
PLOG_GOOD     = 8   # H

PRINTING_MACHINES = {"Printing-03", "Printing-04"}

# Production_Log formula range — generous to handle growth
PLOG_RANGE_END = 10000


# -----------------------------------------------------------------------
def find_files(folder):
    ac_files = glob.glob(os.path.join(folder, "AlphaContainers*.xlsx"))
    if not ac_files:
        print("  ERROR: No AlphaContainers*.xlsx found in: " + folder)
        return None, None
    ac = sorted(ac_files)[-1]

    pending = os.path.join(folder, "pending.xlsx")
    if not os.path.exists(pending):
        print("  ERROR: pending.xlsx not found in: " + folder)
        return None, None

    return ac, pending


def load_catalog(wb):
    """
    Returns {catalog_name_upper: {'pid': int, 'customer': str, 'dia': float/str}}
    """
    ws = wb[CATALOG_SHEET]
    catalog = {}
    for r in range(3, ws.max_row + 1):
        pid_v  = ws.cell(r, CAT_PID).value
        name_v = ws.cell(r, CAT_NAME).value
        cust_v = ws.cell(r, CAT_CUST).value
        dia_v  = ws.cell(r, CAT_DIA).value
        if pid_v is None or name_v is None:
            continue
        try:
            pid = int(float(str(pid_v)))
        except (ValueError, TypeError):
            continue
        name = str(name_v).strip()
        catalog[name.upper()] = {
            'pid':      pid,
            'customer': str(cust_v).strip() if cust_v else "",
            'dia':      dia_v,
            'name':     name,
        }
    return catalog


def load_production_totals(wb):
    """
    Sum Good Production from Production_Log for Printing-03/04 machines only.
    Returns {pid: total_good_qty}
    """
    ws   = wb[PROD_LOG_SHEET]
    totals = {}
    for r in range(3, ws.max_row + 1):
        machine = ws.cell(r, PLOG_MACHINE).value
        pid_v   = ws.cell(r, PLOG_PID).value
        good_v  = ws.cell(r, PLOG_GOOD).value
        if machine is None or pid_v is None:
            continue
        if str(machine).strip() not in PRINTING_MACHINES:
            continue
        try:
            pid = int(float(str(pid_v)))
            qty = float(good_v) if good_v else 0
        except (ValueError, TypeError):
            continue
        totals[pid] = totals.get(pid, 0) + qty
    return totals


def get_latest_sheet_name(wb):
    """
    Return the last sheet name in the pending workbook.
    Sheet names are dates in DD-MM-YYYY format — last = most recent.
    Handles trailing spaces in sheet names (e.g. '10-04-2026 ').
    """
    sheets = wb.sheetnames
    return sheets[-1]


def parse_pending(wb):
    """
    Parse the latest date sheet from pending workbook.
    Returns list of dicts: {name, dia, balance, job_no}

    Dia inheritance rules:
    - Group header row: col D is blank, col E has a number → sets current_dia
    - Data row: if col E has a numeric value, use it as the row's dia AND
      update current_dia for subsequent rows without explicit dia
    - Data row with no col E value: inherits current_dia

    Duplicate catalog products (same pending name + dia) are aggregated:
    their balances are summed, first job_no is kept.
    """
    sheet_name = get_latest_sheet_name(wb)
    ws = wb[sheet_name]
    print("  Pending sheet: '%s'" % sheet_name)

    SKIP_NAMES = {'total qty', 'total', 'grand total', 'job no', ''}
    raw_orders  = []
    current_dia = None

    for row in ws.iter_rows(values_only=True):
        col_a = row[0]   # JOB NO
        col_d = row[3]   # Product Name
        col_e = row[4]   # Dia (explicit per-row or group header)
        col_h = row[7]   # Balance

        # Detect dia group header: col E has a number, col D is blank
        if col_e is not None and col_d is None:
            try:
                current_dia = float(col_e)
            except (ValueError, TypeError):
                pass
            continue

        # Skip summary rows
        if col_d is None:
            continue
        name = str(col_d).strip()
        if name.upper() in SKIP_NAMES or name == '':
            continue

        # If this data row has an explicit dia in col E, use it and update current
        if col_e is not None:
            try:
                current_dia = float(col_e)
            except (ValueError, TypeError):
                pass

        # Balance
        try:
            balance = int(float(col_h)) if col_h is not None else 0
        except (ValueError, TypeError):
            balance = 0

        if balance <= 0:
            continue   # skip fully completed orders

        # Job number
        job_no = None
        if col_a is not None:
            try:
                job_no = int(float(str(col_a)))
            except (ValueError, TypeError):
                job_no = str(col_a).strip() if str(col_a).strip() else None

        raw_orders.append({
            'name':    name,
            'dia':     current_dia,
            'balance': balance,
            'job_no':  job_no,
        })

    # Aggregate duplicates: same (name, dia) → sum balance, keep first job_no
    seen    = {}   # (name, dia) -> index in aggregated list
    aggregated = []
    for o in raw_orders:
        key = (o['name'].upper().strip(), o['dia'])
        if key in seen:
            aggregated[seen[key]]['balance'] += o['balance']
            # Append additional job numbers if different
            existing_job = aggregated[seen[key]]['job_no']
            if o['job_no'] and o['job_no'] != existing_job:
                aggregated[seen[key]]['extra_jobs'].append(o['job_no'])
        else:
            o['extra_jobs'] = []
            seen[key] = len(aggregated)
            aggregated.append(o)

    return aggregated


def resolve_order(order, catalog):
    """
    Resolve a pending order dict to a catalog entry.
    Returns (catalog_entry_dict, was_aliased) or (None, False).
    catalog_entry_dict: {'pid', 'customer', 'dia', 'name'}
    """
    name = order['name'].strip()
    dia  = order['dia']

    # Try alias lookup: (name_upper, dia) then (name_upper, None) as fallback
    for key in [(name.upper(), dia), (name.upper(), None)]:
        if key in ALIASES:
            catalog_name = ALIASES[key]
            entry = catalog.get(catalog_name.upper())
            if entry:
                return entry, True

    # Try direct catalog match (no alias needed)
    entry = catalog.get(name.upper())
    if entry:
        return entry, False

    return None, False


def produced_formula(row):
    """
    SUMPRODUCT formula for col G — sums Good Production from Production_Log
    for Printing-03 and Printing-04 machines only, matching PID in col D.
    """
    end = PLOG_RANGE_END
    return (
        "=SUMPRODUCT("
        "((Production_Log!$B$3:$B${end}=\"Printing-03\")"
        "+(Production_Log!$B$3:$B${end}=\"Printing-04\"))"
        "*(Production_Log!$F$3:$F${end}=D{row})"
        "*(Production_Log!$H$3:$H${end})"
        ")"
    ).format(end=end, row=row)


def pid_formula(row):
    return "=INDEX(Product_Catalog!A:A,MATCH(MRP!C{r},Product_Catalog!D:D,0))".format(r=row)


def remaining_formula(row):
    return "=F{r}-G{r}".format(r=row)


def clear_order_rows(ws):
    """Clear all data in order zone (rows 3-16, cols A-H). Preserve row structure."""
    for r in range(ORDER_ROW_START, ORDER_ROW_END + 1):
        for c in range(1, 9):
            ws.cell(r, c).value = None


def write_order_rows(ws, resolved_orders, prod_totals):
    """
    Write resolved orders into rows 3-16.
    resolved_orders: list of (catalog_entry, job_no, balance) dicts
    """
    written = []

    for i, item in enumerate(resolved_orders):
        if i >= ORDER_ROW_MAX:
            break
        r = ORDER_ROW_START + i
        entry   = item['entry']
        job_no  = item['job_no']
        balance = item['balance']
        pid     = entry['pid']
        produced = int(prod_totals.get(pid, 0))
        req_qty  = produced + balance

        # Write dia — use float if whole number, else as-is
        dia_val = entry['dia']
        try:
            dia_f = float(str(dia_val))
            dia_write = int(dia_f) if dia_f == int(dia_f) else dia_f
        except (ValueError, TypeError):
            dia_write = dia_val

        # Build job number string — combine multiple job orders if aggregated
        extra = item.get('extra_jobs', [])
        if extra:
            all_jobs = ([str(job_no)] if job_no else []) + [str(j) for j in extra]
            job_write = " & ".join(all_jobs)
        else:
            job_write = job_no

        ws.cell(r, COL_DIA).value      = dia_write
        ws.cell(r, COL_CUSTOMER).value = entry['customer']
        ws.cell(r, COL_PRODUCT).value  = entry['name']
        ws.cell(r, COL_PID).value      = pid_formula(r)
        ws.cell(r, COL_JOB).value      = job_write
        ws.cell(r, COL_REQ).value      = req_qty
        ws.cell(r, COL_PROD).value     = produced_formula(r)
        ws.cell(r, COL_REM).value      = remaining_formula(r)

        written.append({
            'row':      r,
            'dia':      dia_write,
            'product':  entry['name'],
            'pid':      pid,
            'job_no':   job_no,
            'produced': produced,
            'balance':  balance,
            'req_qty':  req_qty,
        })

    return written


def update_mrp(ac_path, resolved_orders, prod_totals):
    wb = load_workbook(ac_path)
    ws = wb[MRP_SHEET]

    clear_order_rows(ws)
    written = write_order_rows(ws, resolved_orders, prod_totals)

    wb.save(ac_path)
    return written


def main():
    SEP = "=" * 62

    print("")
    print(SEP)
    print("  Alpha Containers - MRP Updater v1")
    print(SEP)
    print("")

    folder = os.path.dirname(os.path.abspath(__file__))

    print("[1/5] Finding files...")
    ac_path, pending_path = find_files(folder)
    if not ac_path:
        return
    print("  Alpha File: " + os.path.basename(ac_path))
    print("  Pending:    " + os.path.basename(pending_path))

    print("")
    print("[2/5] Loading Product_Catalog and Production_Log...")
    wb_ac = load_workbook(ac_path, read_only=True, data_only=True)
    catalog    = load_catalog(wb_ac)
    prod_totals = load_production_totals(wb_ac)
    wb_ac.close()
    print("  Catalog: %d products loaded" % len(catalog))
    print("  Production totals (Printing-03/04 only):")
    for pid in sorted(prod_totals):
        print("    PID %-8d  %9d pcs" % (pid, int(prod_totals[pid])))

    print("")
    print("[3/5] Parsing pending.xlsx (latest sheet)...")
    wb_p = load_workbook(pending_path, read_only=True, data_only=True)
    pending_orders = parse_pending(wb_p)
    wb_p.close()
    print("  %d pending orders with balance > 0" % len(pending_orders))

    print("")
    print("[4/5] Resolving aliases and building order list...")
    resolved   = []
    unmapped   = []

    for order in pending_orders:
        entry, was_aliased = resolve_order(order, catalog)
        if entry is None:
            unmapped.append(order)
        else:
            resolved.append({
                'entry':      entry,
                'job_no':     order['job_no'],
                'balance':    order['balance'],
                'was_aliased': was_aliased,
            })

    print("  %d orders resolved, %d unmapped" % (len(resolved), len(unmapped)))

    if len(resolved) > ORDER_ROW_MAX:
        print("")
        print("  !! WARNING: %d orders exceed the %d-slot MRP zone (rows %d-%d)." % (
            len(resolved), ORDER_ROW_MAX, ORDER_ROW_START, ORDER_ROW_END))
        print("     Only the first %d orders will be written." % ORDER_ROW_MAX)
        print("     To add more slots, the MRP material formulas must be manually")
        print("     extended to cover the new rows. Contact Claude.")

    if unmapped:
        print("")
        print("  !! UNMAPPED PRODUCTS (not written to MRP):")
        for o in unmapped:
            print("    Pending: '%-30s'  Dia=%-5s  Balance=%d" % (
                o['name'], str(o['dia']), o['balance']))
            print("    Fix: add to ALIASES dict in update_mrp.py")

    print("")
    print("[5/5] Writing MRP sheet rows %d-%d..." % (ORDER_ROW_START, ORDER_ROW_END))
    written = update_mrp(ac_path, resolved, prod_totals)

    print("")
    print("  Orders written:")
    print("  %-4s  %-5s  %-40s  %-8s  %-9s  %-9s  %-9s" % (
        "Row", "Dia", "Product", "PID", "Produced", "Pending", "F=Req"))
    print("  " + "-" * 85)
    for w in written:
        print("  %-4d  %-5s  %-40s  %-8d  %9d  %9d  %9d" % (
            w['row'], str(w['dia']), w['product'][:40], w['pid'],
            w['produced'], w['balance'], w['req_qty']))
    blank_slots = ORDER_ROW_MAX - len(written)
    if blank_slots > 0:
        print("  (%d blank slot(s) remaining in rows %d-%d)" % (
            blank_slots, ORDER_ROW_START + len(written), ORDER_ROW_END))

    print("")
    print("  Saved: " + os.path.basename(ac_path))
    print("")
    print("  REMINDER: Open Excel and press Ctrl+Shift+F9 to recalculate.")
    print("  Col G (Produced) is now a live formula — it will auto-update")
    print("  as Production_Log grows. Col F is refreshed each run of this script.")
    print(SEP)


if __name__ == "__main__":
    main()
