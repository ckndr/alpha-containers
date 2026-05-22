"""
Fix off-by-one formula errors in Tubex_v10_27.xlsx MRP sheet.

Three sections have the same bug pattern - each row's G/H/F/I/K formulas
reference the row above instead of the current row:

1. PET Orders (rows 105-110): G and H columns
2. PET MRP (rows 115-122): F, G, I, K columns
3. FG Stock Caps (rows 163-166): B, C, F columns
"""

import openpyxl
import shutil
import os

# Backup original
src = 'Tubex_v10_27.xlsx'
bak = 'Tubex_v10_27_BACKUP.xlsx'
shutil.copy2(src, bak)
print(f"Backup created: {bak}")

wb = openpyxl.load_workbook(src)
ws = wb['MRP']

fixes = []

# =============================================================================
# SECTION 1: PET Orders rows 105-110
# Bug: G[n] references D[n-1], H[n] references F[n-1]-G[n-1]
# Fix: G[n] should reference D[n], H[n] should reference F[n]-G[n]
# =============================================================================
for row in range(105, 111):  # rows 105,106,107,108,109,110
    prev = row - 1
    # Fix G column (col 7): references D{prev} -> should be D{row}
    g_cell = ws.cell(row=row, column=7)
    if g_cell.value and f'D{prev}' in str(g_cell.value):
        old_val = g_cell.value
        g_cell.value = g_cell.value.replace(f'D{prev}', f'D{row}')
        fixes.append(f"  G{row}: D{prev} -> D{row}")
        
    # Fix H column (col 8): references F{prev}-G{prev} -> should be F{row}-G{row}
    h_cell = ws.cell(row=row, column=8)
    if h_cell.value and f'F{prev}-G{prev}' in str(h_cell.value):
        old_val = h_cell.value
        h_cell.value = h_cell.value.replace(f'F{prev}-G{prev}', f'F{row}-G{row}')
        fixes.append(f"  H{row}: F{prev}-G{prev} -> F{row}-G{row}")

# =============================================================================
# SECTION 2: PET MRP rows 115-122
# Bug: F[n] has MATCH(A[n-1],...) -> should be MATCH(A[n],...)
#      G[n] is =F[n-1]-E[n-1]  -> should be =F[n]-E[n]
#      I[n] references E[n-1], G[n-1], F[n-1] -> should reference current row
#      K[n] references E[n-1], F[n-1] -> should reference current row
# =============================================================================
for row in range(115, 123):  # rows 115-122
    prev = row - 1

    # Fix F column (col 6): MATCH(A{prev}, -> MATCH(A{row},
    f_cell = ws.cell(row=row, column=6)
    if f_cell.value and f'A{prev},' in str(f_cell.value):
        f_cell.value = f_cell.value.replace(f'A{prev},', f'A{row},')
        fixes.append(f"  F{row}: MATCH(A{prev},... -> MATCH(A{row},...")
    # Also handle second MATCH in same formula
    if f_cell.value and f'A{prev},' in str(f_cell.value):
        f_cell.value = f_cell.value.replace(f'A{prev},', f'A{row},')

    # Fix G column (col 7): =F{prev}-E{prev} -> =F{row}-E{row}
    g_cell = ws.cell(row=row, column=7)
    if g_cell.value and f'F{prev}-E{prev}' in str(g_cell.value):
        g_cell.value = g_cell.value.replace(f'F{prev}-E{prev}', f'F{row}-E{row}')
        fixes.append(f"  G{row}: F{prev}-E{prev} -> F{row}-E{row}")

    # Fix I column (col 9): references E{prev}, G{prev}, F{prev} -> current row
    i_cell = ws.cell(row=row, column=9)
    if i_cell.value and str(prev) in str(i_cell.value):
        new_val = i_cell.value
        # Replace E{prev}, G{prev}, F{prev} patterns
        for col_letter in ['E', 'G', 'F']:
            if f'{col_letter}{prev}' in new_val:
                new_val = new_val.replace(f'{col_letter}{prev}', f'{col_letter}{row}')
        if new_val != i_cell.value:
            i_cell.value = new_val
            fixes.append(f"  I{row}: references -> row {row}")

    # Fix K column (col 11): references E{prev}, F{prev} -> current row
    k_cell = ws.cell(row=row, column=11)
    if k_cell.value and str(prev) in str(k_cell.value):
        new_val = k_cell.value
        for col_letter in ['E', 'F']:
            if f'{col_letter}{prev}' in new_val:
                new_val = new_val.replace(f'{col_letter}{prev}', f'{col_letter}{row}')
        if new_val != k_cell.value:
            k_cell.value = new_val
            fixes.append(f"  K{row}: references -> row {row}")

# =============================================================================
# SECTION 3: FG Stock Caps lookup rows 163-166
# Bug: B[n], C[n], F[n] reference A[n-1] instead of A[n]
# =============================================================================
for row in range(163, 167):  # rows 163,164,165,166
    prev = row - 1

    # Fix B column (col 2): VLOOKUP(A{prev},...) -> VLOOKUP(A{row},...)
    b_cell = ws.cell(row=row, column=2)
    if b_cell.value and f'A{prev},' in str(b_cell.value):
        b_cell.value = b_cell.value.replace(f'A{prev},', f'A{row},')
        fixes.append(f"  B{row}: VLOOKUP(A{prev},... -> VLOOKUP(A{row},...")

    # Fix C column (col 3): VLOOKUP(A{prev},...) -> VLOOKUP(A{row},...)
    c_cell = ws.cell(row=row, column=3)
    if c_cell.value and f'A{prev},' in str(c_cell.value):
        c_cell.value = c_cell.value.replace(f'A{prev},', f'A{row},')
        fixes.append(f"  C{row}: VLOOKUP(A{prev},... -> VLOOKUP(A{row},...")

    # Fix F column (col 6): $B$4:$B$100=A{prev} -> $B$4:$B$100=A{row}
    f_cell = ws.cell(row=row, column=6)
    if f_cell.value and f'=A{prev}' in str(f_cell.value):
        f_cell.value = f_cell.value.replace(f'=A{prev}', f'=A{row}')
        fixes.append(f"  F{row}: =A{prev} -> =A{row}")

# Save
wb.save(src)
print(f"\nFixed {len(fixes)} formulas:")
for f in fixes:
    print(f)
print(f"\nSaved to {src}")
