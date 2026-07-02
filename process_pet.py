import pandas as pd
import openpyxl
import re

excel_path = r"D:\Alpha\Tubex_v10_30.xlsx"
output_path = r"D:\Alpha\PET_SKUs.xlsx"

erp_products = [
    {"Customer": "SAMSOL INTERNATIONAL PRIVATE LIMITED", "Product Name": "PET BOTTLE SMALL (120 ML) YELLOW"},
    {"Customer": "SAMSOL INTERNATIONAL PRIVATE LIMITED", "Product Name": "PET BOTTLE LARGE (200 ML) YELLOW"},
    {"Customer": "ABID MASOOD KHAN", "Product Name": "BT-300 ML TRANSPARENT"},
    {"Customer": "SAMSOL INTERNATIONAL PRIVATE LIMITED", "Product Name": "PET BOTTLE LARGE (200ML) COMPACT BLACK"},
    {"Customer": "SAMSOL INTERNATIONAL PRIVATE LIMITED", "Product Name": "PET BOTTLE SMALL (120ML) COMPACT BLACK"},
    {"Customer": "SAMSOL INTERNATIONAL PRIVATE LIMITED", "Product Name": "PET BOTTLE LARGE 200ML WHITE"},
    {"Customer": "ALPHA LABS PVT LTD", "Product Name": "PET BOTTLE (150ML)TRANSPARENT BODY MIST"},
    {"Customer": "ABID MASOOD KHAN", "Product Name": "BT-120 ML YELLOW"},
    {"Customer": "ABID MASOOD KHAN", "Product Name": "BT-200 ML YELLOW"},
    {"Customer": "ABID MASOOD KHAN", "Product Name": "BT-130 ML TRANSPARENT"},
    {"Customer": "ALPHA LABS PVT LTD", "Product Name": "PET BOTTLE (500ML) GREEN"},
]

wb = openpyxl.load_workbook(excel_path)
ws_catalog = wb['Product_Catalog']
ws_bom = wb['BOM']

# Find header row in Product_Catalog
catalog_header_row = 1
for row in ws_catalog.iter_rows():
    if row[0].value == "Product ID":
        catalog_header_row = row[0].row
        break

# Find header row in BOM
bom_header_row = 1
for row in ws_bom.iter_rows():
    if row[0].value == "Product ID":
        bom_header_row = row[0].row
        break

# Read existing products
existing_products = []
max_8k_pid = 8000
for row in ws_catalog.iter_rows(min_row=catalog_header_row+1, values_only=True):
    pid = row[0]
    prod_name = row[3]
    if pd.isna(prod_name): continue
    if isinstance(prod_name, str):
        existing_products.append(prod_name.strip().upper())
    if isinstance(pid, (int, float)):
        if 8000 <= pid < 9000:
            max_8k_pid = max(max_8k_pid, int(pid))
    elif isinstance(pid, str) and pid.isdigit():
        pid_int = int(pid)
        if 8000 <= pid_int < 9000:
            max_8k_pid = max(max_8k_pid, pid_int)

# Check and add new products
new_pids_added = 0
for prod in erp_products:
    name_upper = prod["Product Name"].strip().upper()
    if name_upper not in existing_products:
        max_8k_pid += 1
        new_pid = max_8k_pid
        # Append to Product_Catalog
        # Columns: Product ID, BOM ID, Customer, Product Name, ...
        # Based on pandas output, Customer is index 2, Product Name is index 3
        new_row_catalog = [new_pid, None, prod["Customer"], prod["Product Name"]]
        ws_catalog.append(new_row_catalog)
        
        # Append to BOM
        # Columns: Product ID, BOM ID, Customer, Product Name, Dia, Material Category, Item ID, Item Name, UOM, Per 1000 Units, Material Group, Scrap %, Change Note
        new_row_bom = [new_pid, None, prod["Customer"], prod["Product Name"]]
        ws_bom.append(new_row_bom)
        
        existing_products.append(name_upper)
        new_pids_added += 1

if new_pids_added > 0:
    wb.save(excel_path)
    print(f"Added {new_pids_added} new PET products to Tubex_v10_30.xlsx")
else:
    print("All ERP products are already present.")

# Now extract all 8xxx products to make a complete PET list
all_pet_products = []
for row in ws_catalog.iter_rows(min_row=catalog_header_row+1, values_only=True):
    pid = row[0]
    prod_name = row[3]
    if pd.isna(prod_name) or prod_name is None: continue
    
    is_8k = False
    if isinstance(pid, (int, float)) and 8000 <= pid < 9000:
        is_8k = True
    elif isinstance(pid, str) and pid.isdigit() and 8000 <= int(pid) < 9000:
        is_8k = True
        
    if is_8k:
        # Parse product name
        name_upper = str(prod_name).upper()
        
        # Volume (ml)
        ml_match = re.search(r'(\d+)\s*ML', name_upper)
        volume = f"{ml_match.group(1)} ML" if ml_match else "Unknown"
        
        # Type
        if "JAR" in name_upper:
            type_str = "Jar"
        else:
            type_str = "Bottle"
            
        # Color
        color = "Unknown"
        if "TRANSPARENT" in name_upper:
            color = "Transparent"
        elif "COMPACT BLACK" in name_upper:
            color = "Colored (Compact Black)"
        elif "BLACK" in name_upper:
            color = "Colored (Black)"
        elif "YELLOW" in name_upper:
            color = "Colored (Yellow)"
        elif "WHITE" in name_upper:
            color = "Colored (White)"
        elif "GREEN" in name_upper:
            color = "Colored (Green)"
        elif "MUSTARD" in name_upper:
            color = "Colored (Mustard)"
            
        all_pet_products.append({
            "Original Product Name": prod_name,
            "Type": type_str,
            "Color": color,
            "Volume": volume
        })

df_pet = pd.DataFrame(all_pet_products)
df_pet.to_excel(output_path, index=False)
print(f"Exported complete PET SKUs list to {output_path}")

