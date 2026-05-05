"""
Alpha Containers - Production Log Auto-Updater v13
──────────────────────────────────────────────────
CHANGE LOG v13 (vs v12):
  BUGFIX: Replaced Unicode arrow (→) with ASCII (->) in 3 print statements.
  Caused UnicodeEncodeError on Windows cp1252 consoles when output was
  redirected to log files by Run_All_Updates.bat.
  ADDED: warnings.filterwarnings to suppress openpyxl Data Validation
  UserWarning from cluttering the issues summary.

CHANGE LOG v12 (vs v11):
  STRUCTURE: ALIASES dict flattened — all customer-name section headings removed.
  One flat list sorted by dia/volume. Add new entries at the bottom.
  ALIAS ADD: ("samsol men blue", "25") → TUBES MEN BLUE (PID 6506)
  ALIAS ADD: ("hola hair", "30") and ("hola hair (varnish)", "30") → H.H 100GM (6515)
    Imran enters dia=30 for Hola Hair; actual product is 32mm. Both "30" and "32"
    entries retained so either Imran input is handled correctly.
  ALIAS ADD: ("samsol 43", "19") and ("samsol hc 43", "19") → S-43 DIA 19 (PID 9003)
  ALIAS ADD: ("dowfen gel", "30") → DOWFEN GEL 50G (PID 6595) — Seatle (Pvt) Ltd

CHANGE LOG v11 (vs v10):
  BUGFIX: Date column candidate '`' (backtick) added. Imran's file has a
  backtick as the column A header (corrupted 'Date'). That column contains
  the real timestamps (2026-04-01 00:00:00). 'month' removed from candidates
  -- the 'Month' column contains text strings like "April-26" not actual
  dates, causing all 168 rows to fail the year sanity check.

CHANGE LOG v10 (vs v9):
  BUGFIX: Replaced hardcoded skiprows=2 with auto-detect header row.
  Script now scans rows 0-8 of the source file and picks whichever row
  has the most matching known column names. A single typo or extra row
  in Imran's file will no longer break the script.

  BUGFIX: Added parse_date() helper with year sanity check (2020-2035).
  Previously, corrupted date values (e.g. -693478) were passed directly
  to pandas.Timestamp(), producing dates around year 2 AD which openpyxl
  converted back to a deeply negative Excel serial (-693478 in col A).
  Now: any date outside 2020-2035 is rejected, the row is skipped, and
  a WARNING is printed with the raw value so Imran can fix it.

  BUGFIX: col A number_format = 'DD-MMM-YYYY' is now set explicitly on
  every cell write, not just inherited from template. Template was
  unreliable if first data row was empty after a failed previous run.

  BUGFIX: Template style grab now scans for the first non-empty data row
  instead of blindly using row 3 (which may be wiped/corrupt).

CHANGE LOG v9 (vs v8):
  BUGFIX: Added 'month' to Date column candidates. Imran's source file
  now uses "Month" as the column header instead of "Date". Script was
  aborting with "ERROR: Missing columns: ['Date']" and crashing with
  ValueError on unpack. Both fixed.
  BUGFIX: read_production_source() now returns ([], []) on column error
  instead of [] -- prevents ValueError: not enough values to unpack.

CHANGE LOG v8 (vs v7):
  ALIAS ADD: "Samsol 45/43" at Dia=25 -> TUBES (PID 3726)
  ALIAS ADD: "Samsol Common Red" at Dia=25 -> TUBES COMMON RED (PID 6470)
  PID_TO_CUSTOMER: PID 6470 updated to Samsol International Private Limited
  STRUCTURE: All Samsol 25mm ALIASES consolidated into one labelled block

CHANGE LOG v7 (vs v6):
  ALIAS FIX: Samsol 43 at 20.5mm was mapped to PID 5732 (SAMSOL HAIR COLOR 43
  DIA 20.5 MM) — WRONG. ERP April 2026 production report (POF# 6865, 6873)
  confirms the correct product is "S-43 DIA 20.5" = PID 5699 (BOM#21).
  Both ("samsol 43", "20.5") and ("samsol hc 43", "20.5") updated to PID 5699.

CHANGE LOG v6 (vs v5):
  Added backfill_missing_pids() function. DATA_START_ROW = 3.

CHANGE LOG v5 (vs v4):
  WIPE-AND-REWRITE replaces upsert logic.
  Every run: ALL Production_Log data rows are deleted, then rewritten fresh
  from the source file. This eliminates the duplicate row bug that occurred
  when product names changed between runs (upsert couldn't match old names
  to new corrected names, so both rows survived).

  IMPORTANT: Production_Log after each run will exactly mirror the source file.
  Make sure the source Production*.xlsx contains COMPLETE history before running.
  If source file only has April, the log will only have April.

KEY BUSINESS RULES:
  - MHK (Muhammad Hashim Khan) is a cash/no-GST placeholder account in ERP, NOT a real customer.
    Real brand customers use MHK for cash transactions. PID_TO_CUSTOMER maps each product
    to its actual brand customer. Customer resolved from PID, never from Imran's name.
  - Samsol International is a SEPARATE customer from MHK. PID disambiguates.
  - Varnish pass rule: "(Varnish)" in product name → PID 9002 (placeholder).
    Regular print pass → actual ERP PID (e.g. Vince Nurtural → 5814).
  - PET BOMs are placeholder only (PIDs 8001–8013). No ERP BOMs yet.
  - Alpha Lab = Alpha Containers' Karachi subsidiary. Maps to "Alpha Lab".
  - Mabley's 130ml bottle is written "150ml" by Imran. Script fixes to PID 8010.
  - Samsol 43 at 20.5mm = ERP "S-43 DIA 20.5" (PID 5699). Verified Apr-2026.
"""

import os
import glob
import warnings
warnings.filterwarnings("ignore", message=".*Data Validation.*")
warnings.filterwarnings("ignore", message=".*extension.*")
from datetime import datetime


# ─────────────────────────────────────────────────────────────────────────────
# ALIASES: (imran_product_name.lower().strip(), dia_string) → (catalog_name, PID)
# dia_string must match EXACTLY what Imran types.
# Tubes: "19","16","20.5","25","30","32","35"
# PET:   "120 ml","150 ml","200 ml","300 ml" (note space before ml)
# ─────────────────────────────────────────────────────────────────────────────
ALIASES = {
    # ── Dia 16 ──────────────────────────────────────────────────────────────
    ("eczumas ointment",           "16"): ("ECZEMUS OINTMENT 0.03% 10G",           6561),

    # ── Dia 19 ──────────────────────────────────────────────────────────────
    ("pyodine gel",                "19"): ("PYODINE GEL 20GM",                     1085),
    ("philogin gel",               "19"): ("PHLOGIN GEL 20G",                      6559),
    ("phlogin gel",                "19"): ("PHLOGIN GEL 20G",                      6559),
    ("contractubex gel",           "19"): ("CONTRATUBEX GEL 20G",                  6560),
    ("samsol 43",                  "19"): ("S-43 DIA 19",                          9003),  # Added v12
    ("samsol hc 43",               "19"): ("S-43 DIA 19",                          9003),  # Added v12

    # ── Dia 20.5 ────────────────────────────────────────────────────────────
    # VERIFIED Apr-2026: ERP (POF# 6865, 6873) confirms "S-43 DIA 20.5" = PID 5699. PID 5732 was wrong (fixed v7).
    ("samsol 43",                "20.5"): ("S-43 DIA 20.5",                        5699),
    ("samsol hc 43",             "20.5"): ("S-43 DIA 20.5",                        5699),
    ("samsol 45",                "20.5"): ("SAMSOL HAIR COLOR 45 DIA 20.5 MM",     5731),
    ("samsol hc 45",             "20.5"): ("SAMSOL HAIR COLOR 45 DIA 20.5 MM",     5731),
    ("samsol red",               "20.5"): ("S-43 DIA 20.5",                        5699),

    # ── Dia 25 ──────────────────────────────────────────────────────────────
    ("s 43 25mm",                  "25"): ("S 43 25MM",                            3447),
    ("samsol 43",                  "25"): ("S 43 25MM",                            3447),
    ("samsol 45",                  "25"): ("TUBES",                                3726),
    ("samsol 45/43",               "25"): ("TUBES",                                3726),
    ("samsol common red",          "25"): ("TUBES COMMON RED",                     6470),
    ("tubes men blue",             "25"): ("TUBES MEN BLUE",                       6506),
    ("samsol men blue",            "25"): ("TUBES MEN BLUE",                       6506),  # Added v12
    ("samsol common purple",       "25"): ("TUBE COMMON PURPLE",                   6532),
    ("tube common purple",         "25"): ("TUBE COMMON PURPLE",                   6532),
    ("samsol red",                 "25"): ("SAMSOL RED 25MM",                      9006),
    ("vince his",                  "25"): ("VINCE HIS BEARD & MUSTACHE COLOR",     6077),
    ("vincehis",                   "25"): ("VINCE HIS BEARD & MUSTACHE COLOR",     6077),
    ("vince his (varnish)",        "25"): ("VINCE HIS (Varnish)",                  6077),

    # ── Dia 30 ──────────────────────────────────────────────────────────────
    ("hello hair",                 "30"): ("HELLO HAIR COLOR",                     6206),
    ("eazi color red",             "30"): ("EAZICOLOR PREMIUM HAIR COLOUR 60ML",   4050),
    ("hiba h.c",                   "30"): ("HIBA S HAIR COLOR 60ML",               6312),
    ("hiba h.c (varnish)",         "30"): ("HIBA S HAIR COLOR 60ML (Varnish)",     6312),
    ("signature h.c",              "30"): ("SIGNATURE HAIR COLOR CREAM 60ML",      6416),
    ("active pro",                 "30"): ("ACTIVE PRO HAIR COLOR CREAM 60ML",     6337),
    ("hola hair",                  "30"): ("H.H 100GM",                            6515),  # Imran enters dia=30; actual product is 32mm — Added v12
    ("hola hair (varnish)",        "30"): ("H.H 100GM (Varnish)",                  6515),  # Imran enters dia=30; actual product is 32mm — Added v12
    ("vince nurtural",             "30"): ("VINCE NURTURAL",                       5814),  # print pass — ERP BOM#95
    ("vince nurtural (varnish)",   "30"): ("VINCE NURTURAL (Varnish)",             9002),  # varnish pass — placeholder PID
    ("dowfen gel",                 "30"): ("DOWFEN GEL 50G",                       6595),  # Added v12

    # ── Dia 32 ──────────────────────────────────────────────────────────────
    ("hola hair",                  "32"): ("H.H 100GM",                            6515),
    ("hola hair (varnish)",        "32"): ("H.H 100GM (Varnish)",                  6515),
    ("mega grey",                  "32"): ("M.G",                                  5782),

    # ── Dia 35 ──────────────────────────────────────────────────────────────
    ("anvil 43",                   "35"): ("ANVIL 43",                             6020),
    ("anvil 45",                   "35"): ("ANVIL 45",                             6021),

    # ── PET (PF Machine — all placeholder BOMs) ─────────────────────────────
    ("transparent bottle",        "150 ml"): ("TRANSPARENT BOTTLE 150ML",          8001),
    ("transparent bottle",        "300 ml"): ("TRANSPARENT BOTTLE 300ML",          8009),
    ("yellow small bottle",       "120 ml"): ("YELLOW SMALL BOTTLE 120ML",         8005),
    ("yellow large bottle",       "200 ml"): ("YELLOW LARGE BOTTLE 200ML",         8006),
    ("yellow bottle",             "200 ml"): ("YELLOW LARGE BOTTLE 200ML",         8006),
    ("white bottle",              "200 ml"): ("WHITE BOTTLE 200ML",                8007),
    ("black bottle",              "200 ml"): ("BLACK BOTTLE 200ML",                8008),
    ("samsol black bottle",       "120 ml"): ("BLACK SMALL BOTTLE 120ML",          8011),
}


# ─────────────────────────────────────────────────────────────────────────────
# CUSTOMER_MAP: Imran shorthand → ERP name (fallback; PID_TO_CUSTOMER overrides)
# ─────────────────────────────────────────────────────────────────────────────
CUSTOMER_MAP = {
    "brooks pharma":          "Brookes Pharma Private Limited",
    "samsol international":   "Samsol International Private Limited",
    "samsol":                 "Samsol International Private Limited",
    "golden pearl":           "Golden Pearl Cosmetics (PVT) LTD",
    "golden pearl cosmetics": "Golden Pearl Cosmetics (PVT) LTD",
    "adore":                  "Muhammad Hashim Khan",
    "pbs":                    "Professional Beauty Solution (PVT) LTD.Pakistan",
    "mabley beauty":          "Mablay Beauty PVT LTD.",
    "mega grey":              "Muhammad Hashim Khan",
    "al rehman group":        "Al-Rehman Group",
    "alpha lab":              "Alpha Lab",
    "alpha container":        "Alpha Lab",   # Alpha Containers' own Karachi subsidiary
    "hashim khan":            "Muhammad Hashim Khan",
    "m hashim khan":          "Muhammad Hashim Khan",
    "active pro":             "Al-Rehman Group",
    "seatle":                 "Seatle (Private) Limited",
    "dtm":                    "Muhammad Hashim Khan",
    "bahadur":                "Bahadur",
    "swaleheen akhtar":       "Muhammad Hashim Khan",
    "mr. swaleheen akhtar":   "Muhammad Hashim Khan",
}


# ─────────────────────────────────────────────────────────────────────────────
# PID_TO_CUSTOMER: ERP-authoritative customer. ALWAYS overrides CUSTOMER_MAP.
# Source: BOMs.xlsx row 1 (customer name) for each BOM sheet. Verified Apr-2026.
# This resolves the MHK/Samsol ambiguity — PID wins, Imran's customer loses.
# ─────────────────────────────────────────────────────────────────────────────
PID_TO_CUSTOMER = {
    # Brookes Pharma
    1085: "Brookes Pharma Private Limited",
    6559: "Brookes Pharma Private Limited",
    6560: "Brookes Pharma Private Limited",
    6561: "Brookes Pharma Private Limited",
    # Real customers — MHK is a cash/no-GST placeholder in ERP, not a real customer.
    # Confirmed by Sikander (plant visit 16-Apr-2026). Real brands mapped by product.
    3447: "Samsol International Private Limited",   # S 43 25MM — Samsol brand
    5699: "Samsol International Private Limited",   # S-43 DIA 20.5 — Samsol brand
    5782: "Mega Grey",                              # M.G — Mega Grey brand
    6020: "Adore",                                  # ANVIL 43 — Adore brand
    6021: "Adore",                                  # ANVIL 45 — Adore brand
    6338: "DTM",                                    # DTM DIA 25MM — DTM brand
    6470: "Samsol International Private Limited",   # TUBES COMMON RED — confirmed Sikander v8
    6515: "Samsol International Private Limited",   # H.H 100GM (Hola Hair) — Samsol brand
    6530: "Adore",                                  # V- HC BROWN — Adore brand
    6531: "Adore",                                  # V-HC BLACK — Adore brand
    # Samsol International (separate ERP entity from MHK)
    3726: "Samsol International Private Limited",   # TUBES
    5731: "Samsol International Private Limited",   # SAMSOL HC 45 20.5mm
    5732: "Samsol International Private Limited",   # SAMSOL HC 43 20.5mm
    6506: "Samsol International Private Limited",   # TUBES MEN BLUE
    6532: "Samsol International Private Limited",   # TUBE COMMON PURPLE
    6556: "Samsol International Private Limited",   # HC 39 SMALL PURPLE
    6557: "Samsol International Private Limited",   # HC 41 SMALL PURPLE
    # PBS
    4050: "Professional Beauty Solution (PVT) LTD.Pakistan",
    6312: "Professional Beauty Solution (PVT) LTD.Pakistan",
    # Golden Pearl
    6206: "Golden Pearl Cosmetics (PVT) LTD",
    # Mablay Beauty
    5814: "Mablay Beauty PVT LTD.",   # VINCE NURTURAL print pass (ERP BOM#95)
    6077: "Mablay Beauty PVT LTD.",   # VINCE HIS
    6228: "Mablay Beauty PVT LTD.",   # HIS ONLY HAIR COLOR
    # Al-Rehman
    6337: "Al-Rehman Group",
    6416: "Al-Rehman Group",
    # Seatle
    6595: "Seatle (Private) Limited",
    # Placeholders
    9002: "Mablay Beauty PVT LTD.",   # VINCE NURTURAL varnish pass
    9003: "Samsol International Private Limited",
    9004: "Bahadur",
    9006: "Samsol International Private Limited",
    # PET placeholders
    8001: "Alpha Lab",
    8005: "Samsol International Private Limited",
    8006: "Samsol International Private Limited",
    8007: "Samsol International Private Limited",
    8008: "Samsol International Private Limited",
    8009: "Mablay Beauty PVT LTD.",
    8010: "Mablay Beauty PVT LTD.",
    8011: "Samsol International Private Limited",
    8012: "Samsol International Private Limited",
    8013: "Mablay Beauty PVT LTD.",
}


# ─────────────────────────────────────────────────────────────────────────────
COLUMN_CANDIDATES = {
    'Date':                 ['date', '`'],  # '`' = Imran's corrupted "Date" header
    'Machines':             ['machines', 'machine'],
    'Customer':             ['customer'],
    'Product Name':         ['product name', 'product', 'productname'],
    'Dia(mm)/Volume':       ['dia(mm)/volume', 'dia', 'dia(mm)', 'volume', 'dia/volume'],
    'Good Production':      ['good production', 'good qty', 'good', 'goodproduction'],
    'Wastage':              ['wastage', 'reject', 'rejects', 'rejection'],
    'Total Production':     ['total production', 'total', 'totalproduction'],
    'Mechanical Downtime':  ['mechanical downtime', 'mech dt', 'mechanical'],
    'Electrical Downtime':  ['electrical downtime', 'elec dt', 'electrical'],
    'Material Shortage ':   ['material shortage', 'material shortage ', 'mat shortage'],
    'ChangeOver':           ['changeover', 'change over', 'co'],
    'Operations Downtime':  ['operations downtime', 'ops dt', 'operations'],
    'Power Shut Down':      ['power shut down', 'power shutdown', 'power'],
    'Gas Shut Down':        ['gas shut down', 'gas shutdown', 'gas'],
    'Workers Shortage':     ['workers shortage', 'worker shortage', 'workers'],
}


def build_col_map(df_columns):
    df_lower = {str(c).strip().lower(): c for c in df_columns}
    result = {}
    for key, candidates in COLUMN_CANDIDATES.items():
        for cand in candidates:
            if cand in df_lower:
                result[key] = df_lower[cand]
                break
    return result


def get_val(row, col_map, key, default=None):
    actual = col_map.get(key)
    if actual is None:
        return default
    import pandas as pd
    val = row.get(actual, default)
    try:
        if pd.isna(val):
            return default
    except Exception:
        pass
    return val


def find_files():
    folder = os.path.dirname(os.path.abspath(__file__))
    ac_files   = glob.glob(os.path.join(folder, "AlphaContainers*.xlsx"))
    prod_files = glob.glob(os.path.join(folder, "Production*.xlsx"))
    if not ac_files:
        print("  ERROR: No AlphaContainers*.xlsx found in: " + folder)
        return None, None
    if not prod_files:
        print("  ERROR: No Production*.xlsx found in: " + folder)
        return None, None
    ac   = sorted(ac_files)[-1]
    prod = sorted(prod_files)[-1]
    print("  Alpha File: " + os.path.basename(ac))
    print("  Production: " + os.path.basename(prod))
    return ac, prod


def detect_header_row(prod_path):
    """
    Scan rows 0-8 of the source file to find the real header row.
    Picks whichever row has the most matching known column names.
    Robustness: a single corrupted/renamed header or extra title row
    will not break detection.
    Returns the row index to use as skiprows.
    """
    import pandas as pd
    # All known column name fragments across any version of Imran's file
    KNOWN = {
        'date', 'month', 'machines', 'machine', 'customer',
        'product name', 'product', 'dia', 'dia(mm)', 'volume',
        'good production', 'good', 'wastage', 'reject', 'total production',
        'mechanical downtime', 'electrical downtime', 'changeover',
        'operations downtime', 'power shut down', 'gas shut down',
        'workers shortage', 'material shortage', 'actual time', 'downtime',
    }
    raw = pd.read_excel(
        prod_path, sheet_name='Production Day wise',
        header=None, nrows=10
    )
    best_row  = 2   # fallback: original assumption
    best_score = 0
    for i, row in raw.iterrows():
        vals = [str(v).strip().lower() for v in row if str(v).strip() not in ('', 'nan')]
        score = sum(1 for v in vals if any(v.startswith(k) or k in v for k in KNOWN))
        if score > best_score:
            best_score = score
            best_row   = i
    print("  Header row detected at row %d (score=%d)" % (best_row + 1, best_score))
    return best_row


def parse_date(date_raw):
    """
    Safely parse a date value from the source file.
    Returns a datetime.date object, or None if parsing fails or the
    year is outside the expected range (2020-2035).
    Prints a WARNING for bad values so Imran can fix the source file.
    """
    import pandas as pd
    try:
        if pd.isna(date_raw):
            return None
    except Exception:
        pass
    try:
        ts = pd.Timestamp(date_raw)
        d  = ts.date()
        if not (2020 <= d.year <= 2035):
            print("  WARNING: Skipping row -- date out of range (got '%s' -> %s). "
                  "Check source file for corrupted date values." % (date_raw, d))
            return None
        return d
    except Exception:
        print("  WARNING: Skipping row — cannot parse date value '%s'." % date_raw)
        return None


def read_production_source(prod_path):
    import pandas as pd

    header_row = detect_header_row(prod_path)
    df = pd.read_excel(
        prod_path, sheet_name='Production Day wise', skiprows=header_row
    )

    actual_cols = list(df.columns)
    print("  Columns found:")
    for c in actual_cols:
        print("    '" + str(c) + "'")

    col_map = build_col_map(actual_cols)
    missing = [k for k in ['Date', 'Machines', 'Product Name'] if k not in col_map]
    if missing:
        print("  ERROR: Missing columns: " + str(missing))
        print("  Tip: Check that Imran's file has 'Date'/'Month', 'Machines', "
              "and 'Product Name' columns on the detected header row.")
        return [], []

    rows      = []
    no_pid    = []
    bad_dates = 0

    for _, r in df.iterrows():
        date_raw    = get_val(r, col_map, 'Date')
        machine_raw = get_val(r, col_map, 'Machines')
        if date_raw is None or machine_raw is None:
            continue
        try:
            if pd.isna(machine_raw):
                continue
        except Exception:
            pass

        # Safe date parse — skips rows with corrupt/missing dates
        d = parse_date(date_raw)
        if d is None:
            bad_dates += 1
            continue

        name_raw = str(get_val(r, col_map, 'Product Name', '')).strip()
        dia_raw  = str(get_val(r, col_map, 'Dia(mm)/Volume', '')).strip()
        cust_raw = str(get_val(r, col_map, 'Customer', '')).strip()

        # Step 1: Alias lookup
        catalog_name, pid = ALIASES.get((name_raw.lower().strip(), dia_raw), (None, None))
        if catalog_name is None:
            catalog_name, pid = ALIASES.get((name_raw.lower().rstrip(), dia_raw), (None, None))
        if catalog_name is None:
            catalog_name = name_raw
            pid = None

        # Step 2: Customer from CUSTOMER_MAP
        cust_norm = CUSTOMER_MAP.get(cust_raw.lower().strip(), cust_raw)

        # Step 3: Mabley 150ml special case → PID 8010 + correct catalog name
        if 'transparent' in name_raw.lower() and cust_raw.lower() == 'mabley beauty':
            if '150' in dia_raw:
                pid = 8010
                catalog_name = 'TRANSPARENT BOTTLE 130ML'
            elif '300' in dia_raw:
                pid = 8009
                catalog_name = 'TRANSPARENT BOTTLE 300ML'

        # Step 4: PID_TO_CUSTOMER overrides — ERP authoritative customer
        if pid is not None and pid in PID_TO_CUSTOMER:
            cust_norm = PID_TO_CUSTOMER[pid]

        try:
            dia_val = float(dia_raw)
        except ValueError:
            dia_val = dia_raw

        def sf(val, default=None):
            if val is None: return default
            try:
                if pd.isna(val): return default
            except Exception:
                pass
            try:
                return float(val)
            except Exception:
                return default

        if pid is None:
            no_pid.append((catalog_name, dia_raw, cust_norm))

        rows.append({
            'date':                 d,
            'machine':              str(machine_raw).strip(),
            'customer':             cust_norm,
            'product_name':         catalog_name,
            'dia':                  dia_val,
            'pid':                  pid,
            'good_qty':             sf(get_val(r, col_map, 'Good Production'), 0),
            'reject_qty':           sf(get_val(r, col_map, 'Wastage'), 0),
            'total_production':     sf(get_val(r, col_map, 'Total Production'), 0),
            'mechanical_dt':        sf(get_val(r, col_map, 'Mechanical Downtime')),
            'electrical_dt':        sf(get_val(r, col_map, 'Electrical Downtime')),
            'material_shortage_dt': sf(get_val(r, col_map, 'Material Shortage ')),
            'changeover_dt':        sf(get_val(r, col_map, 'ChangeOver')),
            'operations_dt':        sf(get_val(r, col_map, 'Operations Downtime')),
            'power_shutdown_dt':    sf(get_val(r, col_map, 'Power Shut Down')),
            'gas_shutdown_dt':      sf(get_val(r, col_map, 'Gas Shut Down')),
            'workers_shortage_dt':  sf(get_val(r, col_map, 'Workers Shortage')),
            'original_name':        name_raw,
        })

    if bad_dates:
        print("  WARNING: %d row(s) skipped due to bad/corrupt date values "
              "(see warnings above). Fix in source file and re-run." % bad_dates)

    return rows, list(set(no_pid))


# Row 3 in Excel = first data row (row 2 = column headers). NEVER touch row 2.
DATA_START_ROW = 3


def write_production_log(ac_path, source_rows):
    """
    WIPE-AND-REWRITE: clears all existing data rows in Production_Log,
    then writes all source_rows fresh. Sorts by date ascending.
    Rows are written in source order, then sorted.
    """
    import openpyxl
    from copy import copy

    wb = openpyxl.load_workbook(ac_path)
    ws = wb['Production_Log']

    # ── Grab formatting from the first non-empty data row as template ────
    # Scan instead of blindly using DATA_START_ROW, which may be empty
    # after a failed previous run.
    template_row = DATA_START_ROW
    for probe in range(DATA_START_ROW, DATA_START_ROW + 15):
        if any(ws.cell(row=probe, column=c).value is not None for c in range(1, 20)):
            template_row = probe
            break

    template_styles = {}
    for c in range(1, 20):
        cell = ws.cell(row=template_row, column=c)
        template_styles[c] = {
            'font':          copy(cell.font),
            'fill':          copy(cell.fill),
            'border':        copy(cell.border),
            'alignment':     copy(cell.alignment),
            'number_format': cell.number_format,
        }
    # Always override number formats — never trust what's in the template row
    template_styles[1]['number_format']  = 'DD-MMM-YYYY'   # col A: date
    template_styles[7]['number_format']  = '#,##0'
    template_styles[8]['number_format']  = '#,##0'
    template_styles[9]['number_format']  = '#,##0'
    template_styles[10]['number_format'] = '0.0%'

    # ── WIPE all existing data rows ───────────────────────────────────────
    cleared = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, 20)):
            for c in range(1, 20):
                ws.cell(row=r, column=c).value = None
            cleared += 1
    print(f"  Cleared {cleared} existing rows")

    # ── Sort source rows by date, then machine ────────────────────────────
    source_rows.sort(key=lambda x: (x['date'] or datetime(2099,1,1).date(), x['machine']))

    # ── Write all rows fresh ──────────────────────────────────────────────
    for i, row in enumerate(source_rows):
        r = DATA_START_ROW + i
        if row['date'] is None:
            continue

        d = row['date']
        col_a = ws.cell(row=r, column=1)
        col_a.value         = datetime(d.year, d.month, d.day)
        col_a.number_format = 'DD-MMM-YYYY'   # set directly — never rely on template alone
        ws.cell(row=r, column=2).value  = row['machine']
        ws.cell(row=r, column=3).value  = row['customer']
        ws.cell(row=r, column=4).value  = row['product_name']
        ws.cell(row=r, column=5).value  = row['dia']
        ws.cell(row=r, column=6).value  = row['pid']
        ws.cell(row=r, column=7).value  = row['total_production']
        ws.cell(row=r, column=8).value  = row['good_qty']
        ws.cell(row=r, column=9).value  = row['reject_qty']
        ws.cell(row=r, column=10).value = '=IFERROR(I{0}/(H{0}+I{0}),\"\")'.format(r)
        ws.cell(row=r, column=11).value = row['mechanical_dt']
        ws.cell(row=r, column=12).value = row['electrical_dt']
        ws.cell(row=r, column=13).value = row['material_shortage_dt']
        ws.cell(row=r, column=14).value = row['changeover_dt']
        ws.cell(row=r, column=15).value = row['operations_dt']
        ws.cell(row=r, column=16).value = row['power_shutdown_dt']
        ws.cell(row=r, column=18).value = row['gas_shutdown_dt']
        ws.cell(row=r, column=19).value = row['workers_shortage_dt']

        # Apply formatting
        for c in range(1, 20):
            cell = ws.cell(row=r, column=c)
            s = template_styles.get(c, {})
            if s:
                cell.font          = copy(s['font'])
                cell.fill          = copy(s['fill'])
                cell.border        = copy(s['border'])
                cell.alignment     = copy(s['alignment'])
                cell.number_format = s['number_format']

    wb.save(ac_path)
    return len(source_rows)


def main():
    print("\n" + "="*55)
    print("  Alpha Containers -- Production Log Updater v13")
    print("="*55 + "\n")

    print("[1/3] Finding files...")
    ac_path, prod_path = find_files()
    if not ac_path:
        return

    print("\n[2/3] Reading production data...")
    try:
        result = read_production_source(prod_path)
    except ImportError:
        print("\n  ERROR: pandas not installed. Run: pip install pandas openpyxl")
        return
    except Exception as e:
        import traceback
        print(f"\n  ERROR reading file: {e}")
        traceback.print_exc()
        return

    source_rows, no_pid = result

    if not source_rows:
        print("  No rows parsed — check column diagnostics above.")
        return

    dates = {r['date'] for r in source_rows if r['date']}
    print(f"  Parsed {len(source_rows)} rows  |  {min(dates)} -> {max(dates)}")

    # Show name corrections
    corrections = [(r['original_name'], r['product_name'], r['pid'])
                   for r in source_rows if r['original_name'] != r['product_name']]
    if corrections:
        shown = set()
        print("\n  Name corrections applied:")
        for orig, corr, pid in sorted(set(corrections)):
            if orig not in shown:
                shown.add(orig)
                print(f"    '{orig}' -> '{corr}' (PID={pid})")

    print("\n[3/3] Wiping Production_Log and rewriting from scratch...")
    written = write_production_log(ac_path, source_rows)
    print(f"  Written: {written} rows")
    print("  Saved:   " + os.path.basename(ac_path))

    if no_pid:
        print(f"\n  WARNING — {len(set(no_pid))} products with NO PID (add to ALIASES):")
        for name, dia, cust in sorted(set(no_pid)):
            print(f"    {cust}: {name}  (Dia={dia})")

    print("\n  Press Ctrl+Shift+F9 in Excel to recalculate.")
    print("="*55)


if __name__ == "__main__":
    main()
