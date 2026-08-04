import openpyxl

wb = openpyxl.load_workbook("D:/Alpha/Tubex_July26.xlsx")

# Check tables
for sn in ["Product_Catalog", "BOM", "Tubex_Dashboard"]:
    ws = wb[sn]
    print(f"\nTables in {sn}:")
    for table in ws.tables.values():
        print(f"  Table: {table.name}, Ref: {table.ref}")

# Product Catalog - Last 15 rows
print("\n=== PRODUCT CATALOG - Last 15 rows ===")
ws = wb["Product_Catalog"]
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
for r in range(max(2, ws.max_row - 15), ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    if any(row_vals):
        print(f"  Row {r}: {row_vals}")

# Check for PET COCONUT in ALL sheets
print("\n=== Search for COCONUT in all sheets ===")
for sn in wb.sheetnames:
    ws = wb[sn]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and "COCONUT" in str(val).upper():
                print(f"  {sn} Row {r} Col {c}: {val}")

# BOM - Last 15 rows
print("\n=== BOM - Last 15 rows ===")
ws = wb["BOM"]
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
for r in range(max(2, ws.max_row - 15), ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    if any(row_vals):
        print(f"  Row {r}: {row_vals}")

# Dashboard - rows 55-70
print("\n=== DASHBOARD - rows 55-70 ===")
ws = wb["Tubex_Dashboard"]
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
for r in range(55, min(75, ws.max_row + 1)):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
    if any(row_vals):
        print(f"  Row {r}: {row_vals}")

# Dashboard row 23 - full details (MUSTARD OIL 200ML as template)
print("\n=== DASHBOARD Row 23 (MUSTARD OIL template) - All cols ===")
ws = wb["Tubex_Dashboard"]
for c in range(1, ws.max_column + 1):
    val = ws.cell(row=23, column=c).value
    print(f"  Col {c}: {val}")

# BOM rows for 8007 (PET BOTTLE LARGE 200ML WHITE) - the closest match
print("\n=== BOM rows for Product ID 8007 ===")
ws = wb["BOM"]
for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=1).value == 8007:
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"  Row {r}: {row_vals}")

# Catalog row 46 (PET BOTTLE LARGE 200ML WHITE) - all cols
print("\n=== CATALOG Row 46 (PET BOTTLE LARGE 200ML WHITE) - All cols ===")
ws = wb["Product_Catalog"]
for c in range(1, ws.max_column + 1):
    val = ws.cell(row=46, column=c).value
    print(f"  Col {c} ({ws.cell(row=1, column=c).value}): {val}")
