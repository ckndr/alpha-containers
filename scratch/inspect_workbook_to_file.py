import openpyxl

wb = openpyxl.load_workbook("d:\\Alpha\\Tubex_July26.xlsx")

with open("d:\\Alpha\\scratch\\inspected_data.txt", "w", encoding="utf-8") as f:
    f.write("=== SHEETS ===\n")
    f.write(", ".join(wb.sheetnames) + "\n\n")

    if "MRP" in wb.sheetnames:
        ws = wb["MRP"]
        f.write("=== MRP ROWS ===\n")
        for r in range(1, 200):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
            # We want to print non-empty rows, or just all rows to see their exact structure
            # Let's print rows that have any value
            if any(val is not None for val in row_vals):
                f.write(f"Row {r:03d}: {row_vals}\n")
        f.write("\n")

    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        f.write("=== DASHBOARD ROWS ===\n")
        for r in range(1, 200):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
            if any(val is not None for val in row_vals):
                f.write(f"Row {r:03d}: {row_vals}\n")
        f.write("\n")
