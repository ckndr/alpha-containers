import openpyxl

wb = openpyxl.load_workbook("D:/Alpha/Tubex_July26.xlsx")
print("Sheet names:", wb.sheetnames)

# 1. Inspect Product_Catalog for PET products
for sn in wb.sheetnames:
    if "catalog" in sn.lower() or "catalo" in sn.lower() or "product" in sn.lower():
        ws = wb[sn]
        print(f"\n=== {sn} (rows={ws.max_row}, cols={ws.max_column}) ===")
        # Print header
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"Header: {header}")
        # Find PET rows
        for r in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            row_str = str(row_vals).upper()
            if "PET" in row_str and "200" in row_str:
                print(f"  Row {r}: {row_vals}")

# 2. Inspect BOM for PET products
for sn in wb.sheetnames:
    if "bom" in sn.lower():
        ws = wb[sn]
        print(f"\n=== {sn} (rows={ws.max_row}, cols={ws.max_column}) ===")
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"Header: {header}")
        # Find PET rows
        for r in range(1, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            row_str = str(row_vals).upper()
            if "PET" in row_str and "200" in row_str:
                print(f"  Row {r}: {row_vals}")

# 3. Inspect Dashboard for PET products
for sn in wb.sheetnames:
    if "dashboard" in sn.lower():
        ws = wb[sn]
        print(f"\n=== {sn} (rows={ws.max_row}, cols={ws.max_column}) ===")
        # Print first 3 rows as headers
        for r in range(1, 4):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            print(f"  Row {r}: {row_vals}")
        # Find PET rows
        for r in range(4, ws.max_row + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            row_str = str(row_vals).upper()
            if "PET" in row_str and "200" in row_str:
                print(f"  Row {r}: {row_vals}")
