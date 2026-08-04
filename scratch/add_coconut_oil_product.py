"""
Add "PET BOTTLE COCONUT OIL (200ML) WHITE" to Tubex_July26.xlsx
- Product Catalog (after row 48 / near other PET 200ML products)
- BOM (after row 277, using same materials as 8007 PET BOTTLE LARGE 200ML WHITE)
- Dashboard (after row 23, near other active PET 200ML products)

New product ID: 8016, BOM ID: 816
Reference: Product 8007 (PET BOTTLE LARGE 200ML WHITE) - same size, white color
Dispatch this month: 9,000
"""
import openpyxl
from copy import copy

FILE = "D:/Alpha/Tubex_July26.xlsx"
PRODUCT_NAME = "PET BOTTLE COCONUT OIL (200ML) WHITE"
PRODUCT_ID = 8016
BOM_ID = 816
CUSTOMER = "Samsol International Private Limited"
DIA = "200 ml"
DISPATCH = 9000

# Reference product: 8007 PET BOTTLE LARGE 200ML WHITE
REF_PRODUCT_ID = 8007

print("Loading workbook...")
wb = openpyxl.load_workbook(FILE)

# Helper to copy cell style
def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)

# =====================================================================
# 1. ADD TO PRODUCT CATALOG
# =====================================================================
print("\n1. Adding to Product_Catalog...")
ws = wb["Product_Catalog"]

# Find row 48 (PET BOTTLE MUSTARD OIL 200ML) and insert after it
# Insert a new row after row 48
INSERT_CAT_ROW = 49  # After row 48 (MUSTARD OIL)
ws.insert_rows(INSERT_CAT_ROW)

# Copy styles from row 48 (now the row above our new row)
ref_row = 48
for c in range(1, ws.max_column + 1):
    copy_style(ws.cell(row=ref_row, column=c), ws.cell(row=INSERT_CAT_ROW, column=c))

# Set values
new_row = INSERT_CAT_ROW
ws.cell(row=new_row, column=1, value=PRODUCT_ID)   # Product ID
ws.cell(row=new_row, column=2, value=BOM_ID)        # BOM ID
ws.cell(row=new_row, column=3, value=CUSTOMER)       # Customer
ws.cell(row=new_row, column=4, value=PRODUCT_NAME)   # Product Name
ws.cell(row=new_row, column=5, value=DIA)             # Dia (mm) / Size
ws.cell(row=new_row, column=6, value=0)               # Length (mm)
ws.cell(row=new_row, column=7, value="No")            # Latex
ws.cell(row=new_row, column=8, value=0)               # Pcs/Carton

# Columns 9 onwards are formulas - copy same pattern as ref row but with correct row reference
# Col 9: # of Pieces (None/empty)
ws.cell(row=new_row, column=9, value=None)

# Cols 10-16: Material requirement formulas (same pattern, row reference updated automatically)
r = new_row
ws.cell(row=r, column=10, value=f'=IF(I{r}="","",IFERROR(SUMPRODUCT((TableBOM[Product ID]=A{r})*(TableBOM[Material Category]="SLUG")*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %]))*I{r}/1000,0))')
ws.cell(row=r, column=11, value=f'=IF(I{r}="","",IFERROR(SUMPRODUCT((TableBOM[Product ID]=A{r})*(TableBOM[Material Category]="BASE COAT")*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %]))*I{r}/1000,0))')
ws.cell(row=r, column=12, value=f'=IF(I{r}="","",IFERROR(SUMPRODUCT((TableBOM[Product ID]=A{r})*(TableBOM[Material Category]="LACQUER")*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %]))*I{r}/1000,0))')
ws.cell(row=r, column=13, value=f'=IF(I{r}="","",IFERROR(SUMPRODUCT((TableBOM[Product ID]=A{r})*(TableBOM[Material Category]="LATEX")*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %]))*I{r}/1000,0))')
ws.cell(row=r, column=14, value=f'=IF(I{r}="","",IFERROR(SUMPRODUCT((TableBOM[Product ID]=A{r})*(TableBOM[Material Category]="ZINC POWDER")*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %]))*I{r}/1000,0))')
ws.cell(row=r, column=15, value=f'=IF(I{r}="","",IFERROR(SUMPRODUCT((TableBOM[Product ID]=A{r})*(TableBOM[Material Category]="CAP")*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %]))*I{r}/1000,0))')
ws.cell(row=r, column=16, value=f'=IF(I{r}="","",IFERROR(SUMPRODUCT((TableBOM[Product ID]=A{r})*(TableBOM[Material Category]="CARTON")*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %]))*I{r}/1000,0))')

print(f"  Added at row {INSERT_CAT_ROW}: {PRODUCT_NAME} (ID: {PRODUCT_ID})")

# =====================================================================
# 2. ADD TO BOM 
# =====================================================================
print("\n2. Adding to BOM...")
ws = wb["BOM"]

# Find BOM entries for 8007 (rows 274-277 originally, but may have shifted)
# Since we inserted a row in Product_Catalog (different sheet), BOM rows haven't shifted
# BOM for 8007: rows 274-277
# PET RESIN, MASTERBATCH, POLY BAG, PLASTIC DOORI
# We'll add after row 277 (last 8007 entry)

# But first, let's find the exact rows for 8007 to be safe
ref_bom_rows = []
for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=1).value == REF_PRODUCT_ID:
        ref_bom_rows.append(r)
print(f"  Reference BOM rows for {REF_PRODUCT_ID}: {ref_bom_rows}")

# Insert 4 new rows after the last 8007 BOM entry
insert_after = max(ref_bom_rows)
INSERT_BOM_START = insert_after + 1

# BOM materials for 8007 (PET BOTTLE LARGE 200ML WHITE):
# Row 274: PET RESIN, 2680, PET RESIN A-84, kg, 23.75, PET Resin, 0.15
# Row 275: MASTERBATCH, 5816, MASTER BATCH WHITE PRC-MC ET02340000, kg, 1.25, Masterbatch, 0.15
# Row 276: POLY BAG, 3006, POLY BAG 24X36 (LDPE), PCS, 8.333, Poly Bag, 0.15
# Row 277: PLASTIC DOORI, 3043, PLASTIC DOORI, ROLL, 0.005, Packing, 0.15

bom_materials = [
    # (Material Category, Item ID, Item Name, UOM, Per 1000 Units, Material Group, Scrap %)
    ("PET RESIN", 2680, "PET RESIN A-84", "kg", 23.75, "PET Resin", 0.15),
    ("MASTERBATCH", 5816, "MASTER BATCH WHITE PRC-MC ET02340000", "kg", 1.25, "Masterbatch", 0.15),
    ("POLY BAG", 3006, "POLY BAG 24X36 (LDPE)", "PCS", 8.333, "Poly Bag", 0.15),
    ("PLASTIC DOORI", 3043, "PLASTIC DOORI", "ROLL", 0.005, "Packing", 0.15),
]

# Insert 4 rows
ws.insert_rows(INSERT_BOM_START, amount=4)

for i, (mat_cat, item_id, item_name, uom, per_1000, mat_group, scrap) in enumerate(bom_materials):
    r = INSERT_BOM_START + i
    # Copy styles from ref row
    ref_r = ref_bom_rows[min(i, len(ref_bom_rows) - 1)]
    for c in range(1, 13 + 1):
        copy_style(ws.cell(row=ref_r, column=c), ws.cell(row=r, column=c))
    
    ws.cell(row=r, column=1, value=PRODUCT_ID)       # Product ID
    ws.cell(row=r, column=2, value=BOM_ID)            # BOM ID
    ws.cell(row=r, column=3, value=CUSTOMER)           # Customer
    ws.cell(row=r, column=4, value=PRODUCT_NAME)       # Product Name
    ws.cell(row=r, column=5, value=DIA)                 # Dia
    ws.cell(row=r, column=6, value=mat_cat)             # Material Category
    ws.cell(row=r, column=7, value=item_id)             # Item ID
    ws.cell(row=r, column=8, value=item_name)           # Item Name
    ws.cell(row=r, column=9, value=uom)                 # UOM
    ws.cell(row=r, column=10, value=per_1000)           # Per 1000 Units
    ws.cell(row=r, column=11, value=mat_group)          # Material Group
    ws.cell(row=r, column=12, value=scrap)              # Scrap %
    ws.cell(row=r, column=13, value=None)               # Change Note

print(f"  Added 4 BOM rows at {INSERT_BOM_START}-{INSERT_BOM_START + 3}")

# Update the TableBOM reference to include new rows
for table in ws.tables.values():
    if table.name == "TableBOM":
        old_ref = table.ref
        # Parse the ref: A2:M359 -> expand to include 4 more rows
        parts = old_ref.split(":")
        start = parts[0]
        end_col = ''.join(c for c in parts[1] if c.isalpha())
        end_row = int(''.join(c for c in parts[1] if c.isdigit()))
        new_end_row = end_row + 4
        new_ref = f"{start}:{end_col}{new_end_row}"
        table.ref = new_ref
        print(f"  Updated TableBOM ref: {old_ref} -> {new_ref}")

# =====================================================================
# 3. ADD TO DASHBOARD
# =====================================================================
print("\n3. Adding to Dashboard...")
ws = wb["Tubex_Dashboard"]

# Insert after row 23 (MUSTARD OIL 200ML) to keep PET 200ML products together
INSERT_DASH_ROW = 24
ws.insert_rows(INSERT_DASH_ROW)

# Copy styles from row 23 (MUSTARD OIL - now the row above)
ref_dash_row = 23
for c in range(1, ws.max_column + 1):
    copy_style(ws.cell(row=ref_dash_row, column=c), ws.cell(row=INSERT_DASH_ROW, column=c))

r = INSERT_DASH_ROW
ws.cell(row=r, column=1, value=None)                # Col A - empty
ws.cell(row=r, column=2, value="PET")               # Col B - Type
ws.cell(row=r, column=3, value=CUSTOMER)              # Col C - Customer
ws.cell(row=r, column=4, value=PRODUCT_NAME)          # Col D - Product Name
ws.cell(row=r, column=5, value=DIA)                    # Col E - Dia/Size
ws.cell(row=r, column=6, value=PRODUCT_ID)             # Col F - Prod ID
ws.cell(row=r, column=7, value=9000)                   # Col G - Orders (same as dispatch for now)
# Col H - MTD Produced (formula matching PET pattern, no printing filter)
ws.cell(row=r, column=8, value=f'=SUMPRODUCT((Production_Log!$F$3:$F$8963=F{r})*Production_Log!$H$3:$H$8963)')
# Col I - Remaining
ws.cell(row=r, column=9, value=f'=G{r}-H{r}')
# Col J - Compliance
ws.cell(row=r, column=10, value=f'=IF(G{r}=0,"-",H{r}/G{r})')
# Col K - Dispatch
ws.cell(row=r, column=11, value=DISPATCH)               # 9,000 dispatch

print(f"  Added at row {INSERT_DASH_ROW}: {PRODUCT_NAME}")
print(f"  Dispatch amount: {DISPATCH}")
print(f"  Orders: 9,000")

# =====================================================================
# SAVE
# =====================================================================
print("\nSaving workbook...")
wb.save(FILE)
print("Done! File saved successfully.")

# =====================================================================
# VERIFY
# =====================================================================
print("\n" + "="*60)
print("VERIFICATION")
print("="*60)

wb2 = openpyxl.load_workbook(FILE)

# Verify Product Catalog
ws = wb2["Product_Catalog"]
print("\nProduct Catalog - New product:")
for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=1).value == PRODUCT_ID:
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
        print(f"  Row {r}: {row_vals}")
        break

# Verify BOM
ws = wb2["BOM"]
print(f"\nBOM - New entries:")
for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=1).value == PRODUCT_ID:
        mat_cat = ws.cell(row=r, column=6).value
        item_name = ws.cell(row=r, column=8).value
        per_1000 = ws.cell(row=r, column=10).value
        print(f"  Row {r}: {mat_cat} - {item_name} ({per_1000})")

for table in ws.tables.values():
    if table.name == "TableBOM":
        print(f"\nTableBOM ref: {table.ref}")

# Verify Dashboard
ws = wb2["Tubex_Dashboard"]
print(f"\nDashboard - New entry:")
for r in range(4, ws.max_row + 1):
    if ws.cell(row=r, column=6).value == PRODUCT_ID:
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 12)]
        print(f"  Row {r}: {row_vals}")
        break

print("\nAll done!")
