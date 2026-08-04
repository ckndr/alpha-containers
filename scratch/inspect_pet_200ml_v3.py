import openpyxl

wb = openpyxl.load_workbook("D:/Alpha/Tubex_July26.xlsx")

# 1. Product Catalog - full PET section
print("="*80)
print("PRODUCT CATALOG - All PET rows")
print("="*80)
ws = wb["Product_Catalog"]
header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print(f"Header: {header}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

for r in range(2, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    row_str = str(row_vals).upper()
    if "PET" in row_str:
        print(f"  Row {r}: {row_vals}")

# 2. BOM - full headers and structure
print("\n" + "="*80)
print("BOM - Headers and first data row")
print("="*80)
ws = wb["BOM"]
for r in range(1, 6):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"  Row {r}: {row_vals}")

# Check table names
print("\nDefined names in workbook:")
for dn in wb.defined_names.definedNameList:
    print(f"  {dn.name}: {dn.attr_text}")

# Check if BOM is a table
print("\nTables in BOM sheet:")
ws = wb["BOM"]
for table in ws.tables.values():
    print(f"  Table: {table.name}, Ref: {table.ref}")

print("\nTables in Product_Catalog sheet:")
ws = wb["Product_Catalog"]
for table in ws.tables.values():
    print(f"  Table: {table.name}, Ref: {table.ref}")

print("\nTables in Tubex_Dashboard sheet:")
ws = wb["Tubex_Dashboard"]
for table in ws.tables.values():
    print(f"  Table: {table.name}, Ref: {table.ref}")

# Product Catalog - Check the last few rows for the next available slot
print("\n" + "="*80)
print("PRODUCT CATALOG - Last 15 rows")
print("="*80)
ws = wb["Product_Catalog"]
for r in range(max(2, ws.max_row - 15), ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    if any(row_vals):
        print(f"  Row {r}: {row_vals}")

# BOM - check last rows  
print("\n" + "="*80)
print("BOM - Last 15 rows")
print("="*80)
ws = wb["BOM"]
for r in range(max(2, ws.max_row - 15), ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    if any(row_vals):
        print(f"  Row {r}: {row_vals}")

# Dashboard - check rows around PET 200ML area (rows 20-25) and the end area (rows 58-65)
print("\n" + "="*80)
print("DASHBOARD - Around PET 200ML and end of list")  
print("="*80)
ws = wb["Tubex_Dashboard"]
for r in range(17, 30):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(15, ws.max_column + 1))]
    print(f"  Row {r}: {row_vals}")
print("...")
for r in range(55, 70):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(15, ws.max_column + 1))]
    if any(row_vals):
        print(f"  Row {r}: {row_vals}")
