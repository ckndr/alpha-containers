import openpyxl

wb = openpyxl.load_workbook("d:\\Alpha\\Tubex_July26.xlsx")
if "BOM" in wb.sheetnames:
    ws = wb["BOM"]
    print("=== BOM SHEET ===")
    for r in range(1, 200):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        if any(row_vals):
            # Check if Product ID matches 8001, 8020, 8005, 8016
            pid = row_vals[3]
            try:
                pid_int = int(pid) if pid is not None else None
                if pid_int in (8001, 8005, 8016, 8020):
                    print(f"Row {r:03d}: {row_vals}")
            except (ValueError, TypeError):
                if r < 5:
                    print(f"Header Row {r:03d}: {row_vals}")
else:
    print("BOM sheet not found!")
