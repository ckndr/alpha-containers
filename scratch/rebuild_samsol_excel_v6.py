import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
from datetime import datetime
import os
import sys

# Append script path
sys.path.append("D:/Alpha/Scripts")
import update_production
import update_dispatch

catalog_path = "D:/Alpha/Tubex_July26.xlsx"
prod_path = "D:/Alpha/Tubex Records/production nov to jul.xls"
pet_prod_path = "D:/Alpha/Tubex Records/Production report Jan-2026 till Date.xlsx"
disp_path = "D:/Alpha/Tubex Records/dispatch nov to jul.xls"
pet_disp_path = "D:/Alpha/Tubex Records/dispatch pet nov to jul.xls"
template_path = "D:/Alpha/Tubex Records/Samsol_Production_and_Dispatch_Formatting.xlsx"
output_path = "D:/Alpha/Tubex Records/Samsol_Production_and_Dispatch.xlsx"

# Alignments
align_left = Alignment(horizontal="left", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

def get_float(val):
    if pd.isna(val): return 0.0
    try: return float(val)
    except: return 0.0

# 1. Load the template workbook to extract styles
print("Loading template workbook to extract styles...")
wb_tmpl = openpyxl.load_workbook(template_path)

# Extract styles from ws_sum ("Monthly Summary")
ws_style_sum = wb_tmpl["Monthly Summary"]

sum_title_font = copy(ws_style_sum["A1"].font)
sum_title_align = copy(ws_style_sum["A1"].alignment)

sum_header_styles = {}
for col in range(1, 5):
    cell = ws_style_sum.cell(row=2, column=col)
    sum_header_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment)
    }

sum_data_styles = {}
for col in range(1, 5):
    cell = ws_style_sum.cell(row=3, column=col)
    sum_data_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment),
        'number_format': cell.number_format
    }

sum_total_styles = {}
for col in range(1, 5):
    cell = ws_style_sum.cell(row=12, column=col)
    sum_total_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment),
        'number_format': cell.number_format
    }

# Extract styles from ws_dash ("Interactive Dashboard")
ws_style_dash = wb_tmpl["Interactive Dashboard"]

dash_title_font = copy(ws_style_dash["A1"].font)
dash_title_align = copy(ws_style_dash["A1"].alignment)

dash_header_styles = {}
for col in range(1, 8):
    cell = ws_style_dash.cell(row=2, column=col)
    dash_header_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment)
    }

dash_data_styles = {}
for col in range(1, 8):
    cell = ws_style_dash.cell(row=3, column=col)
    dash_data_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment),
        'number_format': cell.number_format
    }

dash_total_styles = {}
for col in range(1, 8):
    cell = ws_style_dash.cell(row=17, column=col)
    dash_total_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment),
        'number_format': cell.number_format
    }

panel_header_styles = {}
for col in [9, 10]:
    cell = ws_style_dash.cell(row=3, column=col)
    panel_header_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment)
    }

panel_data_styles = {}
for col in [9, 10]:
    cell = ws_style_dash.cell(row=4, column=col)
    panel_data_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment),
        'number_format': cell.number_format
    }

# Extract styles from ws_prod_log ("Samsol Production Logging")
ws_style_prod = wb_tmpl["Samsol Production Logging"]

prod_title_font = copy(ws_style_prod["A1"].font)
prod_title_align = copy(ws_style_prod["A1"].alignment)

prod_header_styles = {}
for col in range(1, 10):
    cell = ws_style_prod.cell(row=2, column=col)
    prod_header_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment)
    }

prod_data_styles = {}
for col in range(1, 10):
    cell = ws_style_prod.cell(row=3, column=col)
    prod_data_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment),
        'number_format': cell.number_format
    }

# Extract styles from ws_disp_det ("Samsol Dispatch Details")
ws_style_disp = wb_tmpl["Samsol Dispatch Details"]

disp_title_font = copy(ws_style_disp["A1"].font)
disp_title_align = copy(ws_style_disp["A1"].alignment)

disp_header_styles = {}
for col in range(1, 11):
    cell = ws_style_disp.cell(row=2, column=col)
    disp_header_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment)
    }

disp_data_styles = {}
for col in range(1, 11):
    cell = ws_style_disp.cell(row=3, column=col)
    disp_data_styles[col] = {
        'font': copy(cell.font),
        'fill': copy(cell.fill),
        'border': copy(cell.border),
        'alignment': copy(cell.alignment),
        'number_format': cell.number_format
    }

fill_alert = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

def is_samsol_pet_product(pid, name, customer):
    if pid is not None and (8000 <= pid < 9000):
        name_clean = " ".join(str(name).upper().split())
        allowed_names = {
            "BT-120 ML YELLOW",
            "BT-200 ML YELLOW",
            "PET BOTTLE LARGE (200 ML) YELLOW",
            "PET BOTTLE LARGE 200ML WHITE",
            "PET BOTTLE COCONUT OIL (200ML) WHITE",
            "PET BOTTLE SMALL (120ML) YELLOW"
        }
        def strip_all(s):
            return s.replace(" ", "").replace("-", "").replace("(", "").replace(")", "").upper()
            
        stripped_allowed = {strip_all(x) for x in allowed_names}
        stripped_name = strip_all(name_clean)
        
        if stripped_name in stripped_allowed:
            return True
            
        if customer and 'samsol' in str(customer).lower():
            return True
            
        return False
    return False

# 2. Parse Product Catalog
print("Parsing product catalog...")
xls_cat = pd.ExcelFile(catalog_path)
df_cat = pd.read_excel(xls_cat, sheet_name="Product_Catalog", skiprows=1)

catalog_map = {}
samsol_products_list = []

# Register extra manual catalog mapping for Coconut Oil PET Bottle
catalog_map["PET BOTTLE COCONUT OIL (200ML) WHITE"] = {
    'pid': 8025,
    'customer': 'Samsol International Private Limited',
    'name': 'PET BOTTLE COCONUT OIL (200ML) WHITE',
    'dia': '200 ml'
}

# Add Coconut Oil to products list directly
samsol_products_list.append({
    'name': 'PET BOTTLE COCONUT OIL (200ML) WHITE',
    'pid': 8025,
    'customer': 'Samsol International Private Limited',
    'dia': '200 ml'
})

for idx, row in df_cat.iterrows():
    name = str(row['Product Name']).strip()
    pid = row['Product ID']
    cust = str(row['Customer']).strip()
    dia = row['Dia (mm)']
    if pd.notna(pid):
        pid = int(pid)
    catalog_map[name.upper()] = {'pid': pid, 'customer': cust, 'name': name, 'dia': dia}
    
    c_lower = cust.lower()
    is_samsol_tube = (pid is not None and (pid < 8000 or pid >= 9000)) and ('samsol' in c_lower or 'abid' in c_lower or 'mateen' in c_lower) and ('hola' not in c_lower and 'hh' not in c_lower)
    is_samsol_pet = (pid is not None and (8000 <= pid < 9000)) and is_samsol_pet_product(pid, name, cust)
    
    if is_samsol_tube or is_samsol_pet:
        samsol_products_list.append({
            'name': name,
            'pid': pid,
            'customer': cust,
            'dia': dia
        })

samsol_products_list = sorted(samsol_products_list, key=lambda x: x['name'])
samsol_tubes = [p for p in samsol_products_list if p['pid'] is not None and (p['pid'] < 8000 or p['pid'] >= 9000)]
samsol_pet = [p for p in samsol_products_list if p['pid'] is not None and (8000 <= p['pid'] < 9000)]

# 3. Parse Tube production
print("Parsing Tube production raw data...")
df_p = pd.read_excel(prod_path, header=None)
prod_records = []
unidentified_prod = []

for idx, row in df_p.iterrows():
    row_val = row.dropna().tolist()
    if not row_val:
        continue
    dt = row[1]
    if pd.isna(dt) or not (isinstance(dt, pd.Timestamp) or hasattr(dt, 'date')):
        continue
    
    ref_num = row[0]
    date_val = pd.Timestamp(dt)
    pof_num = row[5]
    name_raw = str(row[6]).strip()
    dia_raw = str(row[8]).strip()
    
    normal_good = get_float(row[9])
    ot_good = get_float(row[12])
    normal_wastage = get_float(row[10])
    ot_wastage = get_float(row[13])
    
    total_good = int(normal_good + ot_good)
    total_wastage = int(normal_wastage + ot_wastage)
    
    name_key = name_raw.lower().strip()
    match = update_production.ALIASES.get((name_key, dia_raw))
    if not match:
        match = update_production.ALIASES.get((name_raw.lower().rstrip(), dia_raw))
        
    resolved_name = None
    resolved_pid = None
    resolved_cust = None
    
    if match:
        resolved_name, resolved_pid = match
        if resolved_pid in update_production.PID_TO_CUSTOMER:
            resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
        elif resolved_name.upper() in catalog_map:
            resolved_cust = catalog_map[resolved_name.upper()]['customer']
    else:
        name_upper = name_raw.upper()
        if name_upper in catalog_map:
            resolved_name = catalog_map[name_upper]['name']
            resolved_pid = catalog_map[name_upper]['pid']
            resolved_cust = catalog_map[name_upper]['customer']
            if resolved_pid in update_production.PID_TO_CUSTOMER:
                resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
                
    if resolved_name:
        cat_info = catalog_map.get(resolved_name.upper())
        cat_cust = cat_info['customer'] if cat_info else resolved_cust
        c_lower = str(cat_cust).lower()
        is_samsol = cat_cust and ('samsol' in c_lower or 'abid' in c_lower) and ('hola' not in c_lower and 'hh' not in c_lower)
        
        if is_samsol:
            prod_records.append({
                'Date': date_val.to_pydatetime(),
                'Month': datetime(date_val.year, date_val.month, 1),
                'Product Name': resolved_name,
                'Prod ID': resolved_pid,
                'Dia': dia_raw,
                'POF #': pof_num,
                'Ref. #': ref_num,
                'Produced Qty': total_good,
                'Wastage': total_wastage,
                'Catalog Customer': cat_cust,
                'Original Name': name_raw
            })
    else:
        unidentified_prod.append({
            'Source': 'Production',
            'Date': date_val.to_pydatetime(),
            'POF #': pof_num,
            'Ref. #': ref_num,
            'Product Name': name_raw,
            'Dia/Volume': dia_raw,
            'Qty': total_good,
            'Wastage/Repl. Qty': total_wastage,
            'Party Name (Act.)': 'N/A (No party in Production)'
        })

# 4. Parse PET production
print("Parsing PET production raw data...")
df_pet_p = pd.read_excel(pet_prod_path, sheet_name="Production Day wise", skiprows=2)

def map_pet_product(name_raw, volume_raw):
    n_lower = str(name_raw).lower().strip()
    v_lower = str(volume_raw).lower().strip()
    v_std = v_lower
    if '120' in v_lower: v_std = "120 ml"
    elif '200' in v_lower: v_std = "200 ml"
    elif '130' in v_lower: v_std = "130 ml"
    elif '150' in v_lower: v_std = "150 ml"
    elif '300' in v_lower: v_std = "300 ml"
    
    if "yellow small" in n_lower or (("yellow bottle" in n_lower or "yellow small" in n_lower) and "120" in v_std):
        return "PET BOTTLE SMALL (120ML) YELLOW", 8005
    elif "yellow large" in n_lower or (("yellow bottle" in n_lower or "yellow large" in n_lower) and "200" in v_std):
        return "PET BOTTLE LARGE (200 ML) YELLOW", 8006
    elif "white bottle" in n_lower and "200" in v_std:
        return "PET BOTTLE LARGE 200ML WHITE", 8007
    elif "black bottle" in n_lower and "200" in v_std:
        return "BLACK BOTTLE 200ML", 8008
    elif ("samsol black bottle" in n_lower or "black bottle" in n_lower) and "120" in v_std:
        return "BLACK SMALL BOTTLE 120ML", 8011
    elif "white bottle" in n_lower and "120" in v_std:
        return "WHITE SMALL BOTTLE 120ML", 8012
    elif "trp bottle" in n_lower and "200" in v_std:
        return "BT-200ML MUSTARD OIL (TRANSPARENT)", 8014
    elif "coconut" in n_lower:
        return "PET BOTTLE COCONUT OIL (200ML) WHITE", 8025
    else:
        name_upper = n_lower.upper()
        if name_upper in catalog_map:
            return catalog_map[name_upper]['name'], catalog_map[name_upper]['pid']
        return None, None

for idx, row in df_pet_p.iterrows():
    machine_raw = str(row['Machines']).strip()
    if machine_raw != "PF Machine":
        continue
    
    cust_raw = str(row['Customer']).strip()
    c_lower = cust_raw.lower()
    is_samsol_cust = ('samsol' in c_lower or 'abid' in c_lower or 'mateen' in c_lower) and ('hola' not in c_lower and 'hh' not in c_lower)
    if not is_samsol_cust:
        continue
        
    dt = row['Date']
    if pd.isna(dt):
        continue
    date_val = pd.Timestamp(dt)
    
    name_raw = str(row['Product Name']).strip()
    vol_raw = str(row['Dia(mm)/Volume']).strip()
    
    good_qty = int(get_float(row['Good Production']))
    wastage = int(get_float(row['Wastage']))
    
    resolved_name, resolved_pid = map_pet_product(name_raw, vol_raw)
    
    if resolved_name:
        cat_info = catalog_map.get(resolved_name.upper())
        cat_cust = cat_info['customer'] if cat_info else cust_raw
        
        if is_samsol_pet_product(resolved_pid, resolved_name, cat_cust):
            prod_records.append({
                'Date': date_val.to_pydatetime(),
                'Month': datetime(date_val.year, date_val.month, 1),
                'Product Name': resolved_name,
                'Prod ID': resolved_pid,
                'Dia': vol_raw,
                'POF #': 'N/A',
                'Ref. #': 'N/A',
                'Produced Qty': good_qty,
                'Wastage': wastage,
                'Catalog Customer': cat_cust,
                'Original Name': name_raw
            })
    else:
        unidentified_prod.append({
            'Source': 'Production (PET)',
            'Date': date_val.to_pydatetime(),
            'POF #': 'N/A',
            'Ref. #': 'N/A',
            'Product Name': name_raw,
            'Dia/Volume': vol_raw,
            'Qty': good_qty,
            'Wastage/Repl. Qty': wastage,
            'Party Name (Act.)': 'N/A (No party in Production)'
        })

# 5. Parse Tube dispatches
print("Parsing Tube dispatch raw data...")
df_d = pd.read_excel(disp_path, header=None)
disp_records = []
current_prod = None
SKIP_PREFIXES = ('dispatch report', 'month :', 'no.')
SKIP_EXACT = {'end of file'}

for idx, row in df_d.iterrows():
    col0 = row[0]
    if isinstance(col0, str) and col0.strip():
        c0 = col0.strip()
        c0_lower = c0.lower()
        if c0_lower in SKIP_EXACT or any(c0_lower.startswith(p) for p in SKIP_PREFIXES) or 'grand total' in c0_lower:
            continue
        if pd.isna(row[1]):
            current_prod = c0
            continue
            
    if current_prod:
        try:
            int(float(col0))
            date_val = pd.Timestamp(row[2])
            qty = get_float(row[7])
            repl_qty = get_float(row[8])
            party = str(row[12]).strip()
            pof = row[3]
            ref_num = row[1]
            dia_raw = str(row[5]).strip()
            
            resolved_name = update_dispatch.NAME_FIXES.get(current_prod, current_prod)
            resolved_pid = None
            resolved_cust = None
            
            name_upper = resolved_name.strip().upper()
            if name_upper in catalog_map:
                catalog_entry = catalog_map[name_upper]
                resolved_name = catalog_entry['name']
                resolved_pid = catalog_entry['pid']
                resolved_cust = catalog_entry['customer']
                if resolved_pid in update_production.PID_TO_CUSTOMER:
                    resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
            else:
                resolved_name = None
                
            if resolved_name:
                cat_info = catalog_map.get(resolved_name.upper())
                cat_cust = cat_info['customer'] if cat_info else resolved_cust
                c_lower = str(cat_cust).lower()
                is_samsol = cat_cust and ('samsol' in c_lower or 'abid' in c_lower) and ('hola' not in c_lower and 'hh' not in c_lower)
                
                if is_samsol:
                    disp_records.append({
                        'Date': date_val.to_pydatetime(),
                        'Month': datetime(date_val.year, date_val.month, 1),
                        'Product Name': resolved_name,
                        'Prod ID': resolved_pid,
                        'Dia': dia_raw,
                        'POF #': pof,
                        'Ref. #': ref_num,
                        'Dispatched Qty': int(qty),
                        'Replacement Qty': int(repl_qty) if pd.notna(repl_qty) else 0,
                        'Party Name (Act.)': party,
                        'Catalog Customer': cat_cust,
                        'Original Name': current_prod
                    })
            else:
                unidentified_prod.append({
                    'Source': 'Dispatch',
                    'Date': date_val.to_pydatetime(),
                    'POF #': pof,
                    'Ref. #': ref_num,
                    'Product Name': current_prod,
                    'Dia/Volume': dia_raw,
                    'Qty': int(qty) if pd.notna(qty) else 0,
                    'Wastage/Repl. Qty': int(repl_qty) if pd.notna(repl_qty) else 0,
                    'Party Name (Act.)': party if pd.notna(party) and party != 'nan' else 'N/A'
                })
        except (ValueError, TypeError):
            pass

# 6. Parse PET dispatches from dispatch pet nov to jul.xls
print("Parsing PET dispatch raw data from dispatch pet nov to jul.xls...")
df_pet_d = pd.read_excel(pet_disp_path, header=None)
current_pet_prod = None

# Custom mapping for PET products specifically to map any description to correct catalog entry
def map_pet_product_disp(name_raw):
    n_clean = str(name_raw).strip()
    n_lower = n_clean.lower()
    
    # 8005: PET BOTTLE SMALL (120ML) YELLOW
    if "yellow small" in n_lower or "small (120ml) yellow" in n_lower or "small (120 ml) yellow" in n_lower or n_lower == "pet bottle small (120ml) yellow":
        return "PET BOTTLE SMALL (120ML) YELLOW", 8005
    # 8006: PET BOTTLE LARGE (200 ML) YELLOW
    elif "yellow large" in n_lower or "large (200 ml) yellow" in n_lower or "large (200ml) yellow" in n_lower or n_lower == "pet bottle large (200 ml) yellow":
        return "PET BOTTLE LARGE (200 ML) YELLOW", 8006
    # 8007: PET BOTTLE LARGE 200ML WHITE
    elif "large 200ml white" in n_lower or n_lower == "pet bottle large 200ml white":
        return "PET BOTTLE LARGE 200ML WHITE", 8007
    # 8008: BLACK BOTTLE 200ML
    elif "black bottle 200ml" in n_lower or n_lower == "black bottle 200ml":
        return "BLACK BOTTLE 200ML", 8008
    # 8011: BLACK SMALL BOTTLE 120ML
    elif "black small bottle 120ml" in n_lower or n_lower == "black small bottle 120ml":
        return "BLACK SMALL BOTTLE 120ML", 8011
    # 8012: WHITE SMALL BOTTLE 120ML
    elif "white small bottle 120ml" in n_lower or n_lower == "white small bottle 120ml":
        return "WHITE SMALL BOTTLE 120ML", 8012
    # 8014: BT-200ML MUSTARD OIL (TRANSPARENT)
    elif "mustard" in n_lower or "bt-200ml mustard oil" in n_lower or "mustard oil" in n_lower:
        return "BT-200ML MUSTARD OIL (TRANSPARENT)", 8014
    # 8021: BT-120 ML YELLOW
    elif "bt-120 ml yellow" in n_lower or "bt-120ml yellow" in n_lower:
        return "BT-120 ML YELLOW", 8021
    # 8022: BT-200 ML YELLOW
    elif "bt-200 ml yellow" in n_lower or "bt-200ml yellow" in n_lower:
        return "BT-200 ML YELLOW", 8022
    # 8023: BT-130 ML TRANSPARENT
    elif "bt-130 ml transparent" in n_lower:
        return "BT-130 ML TRANSPARENT", 8023
    # 8017: BT-300 ML TRANSPARENT
    elif "bt-300 ml transparent" in n_lower:
        return "BT-300 ML TRANSPARENT", 8017
    elif "coconut" in n_lower:
        return "PET BOTTLE COCONUT OIL (200ML) WHITE", 8025
    else:
        n_upper = n_clean.upper()
        if n_upper in catalog_map:
            return catalog_map[n_upper]['name'], catalog_map[n_upper]['pid']
        return None, None

for idx, row in df_pet_d.iterrows():
    col0 = row[0]
    if isinstance(col0, str) and col0.strip():
        c0 = col0.strip()
        c0_lower = c0.lower()
        if c0_lower in SKIP_EXACT or any(c0_lower.startswith(p) for p in SKIP_PREFIXES) or 'grand total' in c0_lower:
            continue
        if pd.isna(row[1]):
            current_pet_prod = c0
            continue
            
    if current_pet_prod:
        try:
            int(float(col0))
            date_val = pd.Timestamp(row[2])
            qty = get_float(row[7])
            repl_qty = get_float(row[8])
            party = str(row[12]).strip()
            pof = row[3]
            ref_num = row[1]
            dia_raw = str(row[5]).strip()
            resolved_name, resolved_pid = map_pet_product_disp(current_pet_prod)
            
            if resolved_name:
                cat_info = catalog_map.get(resolved_name.upper())
                cat_cust = cat_info['customer'] if cat_info else party
                is_samsol = is_samsol_pet_product(resolved_pid, resolved_name, cat_cust)
                if is_samsol:
                    disp_records.append({
                        'Date': date_val.to_pydatetime(),
                        'Month': datetime(date_val.year, date_val.month, 1),
                        'Product Name': resolved_name,
                        'Prod ID': resolved_pid,
                        'Dia': cat_info['dia'] if cat_info else dia_raw,
                        'POF #': pof,
                        'Ref. #': ref_num,
                        'Dispatched Qty': int(qty),
                        'Replacement Qty': int(repl_qty) if pd.notna(repl_qty) else 0,
                        'Party Name (Act.)': party,
                        'Catalog Customer': cat_cust,
                        'Original Name': current_pet_prod
                    })
            else:
                unidentified_prod.append({
                    'Source': 'Dispatch (PET)',
                    'Date': date_val.to_pydatetime(),
                    'POF #': pof,
                    'Ref. #': ref_num,
                    'Product Name': current_pet_prod,
                    'Dia/Volume': dia_raw,
                    'Qty': int(qty) if pd.notna(qty) else 0,
                    'Wastage/Repl. Qty': int(repl_qty) if pd.notna(repl_qty) else 0,
                    'Party Name (Act.)': party if pd.notna(party) and party != 'nan' else 'N/A'
                })
        except (ValueError, TypeError):
            pass

df_p_samsol = pd.DataFrame(prod_records)
df_d_samsol = pd.DataFrame(disp_records)

# Find all months sorted
months_sorted = sorted(list(set(df_p_samsol['Month'].dropna().tolist() + df_d_samsol['Month'].dropna().tolist())))


# 7. Modify Workbook sheets
print("Modifying workbook from template...")
wb = openpyxl.load_workbook(template_path)

# --- A. Populating Samsol Production Logging ---
print("Populating Samsol Production Logging...")
ws_prod = wb["Samsol Production Logging"]
max_r = ws_prod.max_row
if max_r >= 3:
    ws_prod.delete_rows(3, max_r - 3 + 10)

ws_prod.views.sheetView[0].showGridLines = True
ws_prod.cell(row=1, column=1, value="Samsol Production Log Details (Nov 2025 - Jul 2026)")
ws_prod.cell(row=1, column=1).font = prod_title_font
ws_prod.cell(row=1, column=1).alignment = prod_title_align

# Header (Row 2) - Add Column J for Prod ID
cell_j_header = ws_prod.cell(row=2, column=10, value="Prod ID")
for prop in ['font', 'fill', 'border', 'alignment']:
    setattr(cell_j_header, prop, copy(getattr(ws_style_prod.cell(row=2, column=9), prop)))

prod_records_sorted = sorted(prod_records, key=lambda x: x['Date'])
row_idx = 3
for rec in prod_records_sorted:
    ws_prod.cell(row=row_idx, column=1, value=rec['Date'])
    ws_prod.cell(row=row_idx, column=2, value=rec['Month'])
    ws_prod.cell(row=row_idx, column=3, value=rec['Product Name'])
    ws_prod.cell(row=row_idx, column=4, value=rec['Dia'])
    ws_prod.cell(row=row_idx, column=5, value=rec['POF #'])
    ws_prod.cell(row=row_idx, column=6, value=rec['Ref. #'])
    ws_prod.cell(row=row_idx, column=7, value=rec['Produced Qty'])
    ws_prod.cell(row=row_idx, column=8, value=rec['Wastage'])
    ws_prod.cell(row=row_idx, column=9, value=rec['Catalog Customer'])
    ws_prod.cell(row=row_idx, column=10, value=rec['Prod ID'])
    
    for col in range(1, 10):
        cell = ws_prod.cell(row=row_idx, column=col)
        cell.font = copy(prod_data_styles[col]['font'])
        cell.fill = copy(prod_data_styles[col]['fill'])
        cell.border = copy(prod_data_styles[col]['border'])
        cell.alignment = copy(prod_data_styles[col]['alignment'])
        cell.number_format = prod_data_styles[col]['number_format']
        
        if col == 1:
            cell.number_format = 'yyyy-mm-dd'
        elif col == 2:
            cell.number_format = 'mmm-yy'
            
    cell_j = ws_prod.cell(row=row_idx, column=10)
    cell_j.font = copy(prod_data_styles[9]['font'])
    cell_j.fill = copy(prod_data_styles[9]['fill'])
    cell_j.border = copy(prod_data_styles[9]['border'])
    cell_j.alignment = align_center
    cell_j.number_format = '0'
    row_idx += 1

# Total Row
ws_prod.cell(row=row_idx, column=1, value="Total").alignment = align_center
tot_font = Font(name="Segoe UI", size=10, bold=True)
thin_side = Side(style='thin', color='D3D3D3')
double_side = Side(style='double', color='000000')
total_border = Border(top=thin_side, bottom=double_side)

for col in range(1, 11):
    cell = ws_prod.cell(row=row_idx, column=col)
    cell.font = tot_font
    cell.border = total_border
    if col in [7, 8]:
        cell.alignment = align_right
    else:
        cell.alignment = align_center

ws_prod.cell(row=row_idx, column=7, value=f"=SUM(G3:G{row_idx-1})").number_format = '#,##0'
ws_prod.cell(row=row_idx, column=8, value=f"=SUM(H3:H{row_idx-1})").number_format = '#,##0'
ws_prod.column_dimensions['J'].width = 12


# --- B. Populating Samsol Dispatch Details ---
print("Populating Samsol Dispatch Details...")
ws_disp = wb["Samsol Dispatch Details"]
max_r = ws_disp.max_row
if max_r >= 3:
    ws_disp.delete_rows(3, max_r - 3 + 10)

ws_disp.views.sheetView[0].showGridLines = True
ws_disp.cell(row=1, column=1, value="Samsol Dispatch Log Details (Nov 2025 - Jul 2026)")
ws_disp.cell(row=1, column=1).font = disp_title_font
ws_disp.cell(row=1, column=1).alignment = disp_title_align

# Header (Row 2) - Add Column K for Prod ID
cell_k_header = ws_disp.cell(row=2, column=11, value="Prod ID")
for prop in ['font', 'fill', 'border', 'alignment']:
    setattr(cell_k_header, prop, copy(getattr(ws_style_disp.cell(row=2, column=10), prop)))

disp_records_sorted = sorted(disp_records, key=lambda x: x['Date'])
row_idx_dd = 3
for rec in disp_records_sorted:
    ws_disp.cell(row=row_idx_dd, column=1, value=rec['Date'])
    ws_disp.cell(row=row_idx_dd, column=2, value=rec['Month'])
    ws_disp.cell(row=row_idx_dd, column=3, value=rec['Product Name'])
    ws_disp.cell(row=row_idx_dd, column=4, value=rec['Dia'])
    ws_disp.cell(row=row_idx_dd, column=5, value=rec['POF #'])
    ws_disp.cell(row=row_idx_dd, column=6, value=rec['Ref. #'])
    ws_disp.cell(row=row_idx_dd, column=7, value=rec['Dispatched Qty'])
    ws_disp.cell(row=row_idx_dd, column=8, value=rec['Replacement Qty'])
    ws_disp.cell(row=row_idx_dd, column=9, value=rec['Party Name (Act.)'])
    ws_disp.cell(row=row_idx_dd, column=10, value=rec['Catalog Customer'])
    ws_disp.cell(row=row_idx_dd, column=11, value=rec['Prod ID'])
    
    for col in range(1, 11):
        cell = ws_disp.cell(row=row_idx_dd, column=col)
        cell.font = copy(disp_data_styles[col]['font'])
        cell.fill = copy(disp_data_styles[col]['fill'])
        cell.border = copy(disp_data_styles[col]['border'])
        cell.alignment = copy(disp_data_styles[col]['alignment'])
        cell.number_format = disp_data_styles[col]['number_format']
        
        if col == 1:
            cell.number_format = 'yyyy-mm-dd'
        elif col == 2:
            cell.number_format = 'mmm-yy'
            
    cell_k = ws_disp.cell(row=row_idx_dd, column=11)
    cell_k.font = copy(disp_data_styles[10]['font'])
    cell_k.fill = copy(disp_data_styles[10]['fill'])
    cell_k.border = copy(disp_data_styles[10]['border'])
    cell_k.alignment = align_center
    cell_k.number_format = '0'
    row_idx_dd += 1

# Total Row
ws_disp.cell(row=row_idx_dd, column=1, value="Total").alignment = align_center
for col in range(1, 12):
    cell = ws_disp.cell(row=row_idx_dd, column=col)
    cell.font = tot_font
    cell.border = total_border
    if col in [7, 8]:
        cell.alignment = align_right
    else:
        cell.alignment = align_center

ws_disp.cell(row=row_idx_dd, column=7, value=f"=SUM(G3:G{row_idx_dd-1})").number_format = '#,##0'
ws_disp.cell(row=row_idx_dd, column=8, value=f"=SUM(H3:H{row_idx_dd-1})").number_format = '#,##0'
ws_disp.column_dimensions['K'].width = 12


# --- C. Populating Monthly Summary ---
print("Populating Monthly Summary...")
ws_sum = wb["Monthly Summary"]
max_r = ws_sum.max_row
if max_r >= 1:
    ws_sum.delete_rows(1, max_r + 10)

ws_sum.views.sheetView[0].showGridLines = True
ws_sum.cell(row=1, column=1, value="Samsol Monthly Summaries").font = sum_title_font
ws_sum.cell(row=1, column=1).alignment = sum_title_align

# Section 1: Tubes
ws_sum.cell(row=3, column=1, value="1. Samsol Tubes Monthly Summary").font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
headers_sum = ["Month", "Produced Qty", "Dispatched Qty", "Net Difference"]
for col in range(1, 5):
    cell = ws_sum.cell(row=4, column=col, value=headers_sum[col-1])
    cell.font = copy(sum_header_styles[col]['font'])
    cell.fill = copy(sum_header_styles[col]['fill'])
    cell.border = copy(sum_header_styles[col]['border'])
    cell.alignment = copy(sum_header_styles[col]['alignment'])

row_idx_sum = 5
for m_date in months_sorted:
    cell_m = ws_sum.cell(row=row_idx_sum, column=1, value=m_date)
    cell_m.font = copy(sum_data_styles[1]['font'])
    cell_m.fill = copy(sum_data_styles[1]['fill'])
    cell_m.border = copy(sum_data_styles[1]['border'])
    cell_m.alignment = copy(sum_data_styles[1]['alignment'])
    cell_m.number_format = 'mmm-yy'
    
    c_p = ws_sum.cell(row=row_idx_sum, column=2, value=f"=SUMIFS('Samsol Production Logging'!$G:$G, 'Samsol Production Logging'!$B:$B, $A{row_idx_sum}, 'Samsol Production Logging'!$J:$J, \"<8000\") + SUMIFS('Samsol Production Logging'!$G:$G, 'Samsol Production Logging'!$B:$B, $A{row_idx_sum}, 'Samsol Production Logging'!$J:$J, \">=9000\")")
    c_d = ws_sum.cell(row=row_idx_sum, column=3, value=f"=SUMIFS('Samsol Dispatch Details'!$G:$G, 'Samsol Dispatch Details'!$B:$B, $A{row_idx_sum}, 'Samsol Dispatch Details'!$K:$K, \"<8000\") + SUMIFS('Samsol Dispatch Details'!$G:$G, 'Samsol Dispatch Details'!$B:$B, $A{row_idx_sum}, 'Samsol Dispatch Details'!$K:$K, \">=9000\")")
    c_df = ws_sum.cell(row=row_idx_sum, column=4, value=f"=B{row_idx_sum}-C{row_idx_sum}")
    
    for col, val in enumerate([c_p, c_d, c_df], start=2):
        val.font = copy(sum_data_styles[col]['font'])
        val.fill = copy(sum_data_styles[col]['fill'])
        val.border = copy(sum_data_styles[col]['border'])
        val.alignment = copy(sum_data_styles[col]['alignment'])
        val.number_format = sum_data_styles[col]['number_format']
    row_idx_sum += 1

# Total Tubes Row
cell_tot_lbl = ws_sum.cell(row=row_idx_sum, column=1, value="Grand Total")
cell_tot_lbl.font = copy(sum_total_styles[1]['font'])
cell_tot_lbl.fill = copy(sum_total_styles[1]['fill'])
cell_tot_lbl.border = copy(sum_total_styles[1]['border'])
cell_tot_lbl.alignment = align_center

tot_p = ws_sum.cell(row=row_idx_sum, column=2, value=f"=SUM(B5:B{row_idx_sum-1})")
tot_d = ws_sum.cell(row=row_idx_sum, column=3, value=f"=SUM(C5:C{row_idx_sum-1})")
tot_df = ws_sum.cell(row=row_idx_sum, column=4, value=f"=SUM(D5:D{row_idx_sum-1})")

for col, val in enumerate([tot_p, tot_d, tot_df], start=2):
    val.font = copy(sum_total_styles[col]['font'])
    val.fill = copy(sum_total_styles[col]['fill'])
    val.border = copy(sum_total_styles[col]['border'])
    val.alignment = copy(sum_total_styles[col]['alignment'])
    val.number_format = sum_total_styles[col]['number_format']

# Section 2: PET
row_idx_sum += 3
ws_sum.cell(row=row_idx_sum-2, column=1, value="2. Samsol PET Monthly Summary").font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")

# Headers for PET Summary
for col in range(1, 5):
    cell = ws_sum.cell(row=row_idx_sum-1, column=col, value=headers_sum[col-1])
    cell.font = copy(sum_header_styles[col]['font'])
    cell.fill = copy(sum_header_styles[col]['fill'])
    cell.border = copy(sum_header_styles[col]['border'])
    cell.alignment = copy(sum_header_styles[col]['alignment'])

start_pet_row = row_idx_sum
for m_date in months_sorted:
    cell_m = ws_sum.cell(row=row_idx_sum, column=1, value=m_date)
    cell_m.font = copy(sum_data_styles[1]['font'])
    cell_m.fill = copy(sum_data_styles[1]['fill'])
    cell_m.border = copy(sum_data_styles[1]['border'])
    cell_m.alignment = copy(sum_data_styles[1]['alignment'])
    cell_m.number_format = 'mmm-yy'
    
    c_p = ws_sum.cell(row=row_idx_sum, column=2, value=f"=SUMIFS('Samsol Production Logging'!$G:$G, 'Samsol Production Logging'!$B:$B, $A{row_idx_sum}, 'Samsol Production Logging'!$J:$J, \">=8000\", 'Samsol Production Logging'!$J:$J, \"<9000\")")
    c_d = ws_sum.cell(row=row_idx_sum, column=3, value=f"=SUMIFS('Samsol Dispatch Details'!$G:$G, 'Samsol Dispatch Details'!$B:$B, $A{row_idx_sum}, 'Samsol Dispatch Details'!$K:$K, \">=8000\", 'Samsol Dispatch Details'!$K:$K, \"<9000\")")
    c_df = ws_sum.cell(row=row_idx_sum, column=4, value=f"=B{row_idx_sum}-C{row_idx_sum}")
    
    for col, val in enumerate([c_p, c_d, c_df], start=2):
        val.font = copy(sum_data_styles[col]['font'])
        val.fill = copy(sum_data_styles[col]['fill'])
        val.border = copy(sum_data_styles[col]['border'])
        val.alignment = copy(sum_data_styles[col]['alignment'])
        val.number_format = sum_data_styles[col]['number_format']
    row_idx_sum += 1

# Total PET Row
cell_tot_lbl = ws_sum.cell(row=row_idx_sum, column=1, value="Grand Total")
cell_tot_lbl.font = copy(sum_total_styles[1]['font'])
cell_tot_lbl.fill = copy(sum_total_styles[1]['fill'])
cell_tot_lbl.border = copy(sum_total_styles[1]['border'])
cell_tot_lbl.alignment = align_center

tot_p = ws_sum.cell(row=row_idx_sum, column=2, value=f"=SUM(B{start_pet_row}:B{row_idx_sum-1})")
tot_d = ws_sum.cell(row=row_idx_sum, column=3, value=f"=SUM(C{start_pet_row}:C{row_idx_sum-1})")
tot_df = ws_sum.cell(row=row_idx_sum, column=4, value=f"=SUM(D{start_pet_row}:D{row_idx_sum-1})")

for col, val in enumerate([tot_p, tot_d, tot_df], start=2):
    val.font = copy(sum_total_styles[col]['font'])
    val.fill = copy(sum_total_styles[col]['fill'])
    val.border = copy(sum_total_styles[col]['border'])
    val.alignment = copy(sum_total_styles[col]['alignment'])
    val.number_format = sum_total_styles[col]['number_format']


# --- D. Populating Interactive Dashboard ---
print("Populating Interactive Dashboard...")
ws_dash = wb["Interactive Dashboard"]
max_r = ws_dash.max_row
if max_r >= 1:
    ws_dash.delete_rows(1, max_r + 10)

ws_dash.views.sheetView[0].showGridLines = True
ws_dash.cell(row=1, column=1, value="Samsol Interactive Dashboard").font = dash_title_font
ws_dash.cell(row=1, column=1).alignment = dash_title_align

# Section 1 Title: Row 3
ws_dash.cell(row=3, column=1, value="1. Samsol Tubes Dashboard").font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")

# Headers (Row 4 Columns A to G)
headers_dash = ["Customer Name", "Product Name", "Dia", "Prod ID", "Produced Qty", "Dispatched Qty", "Net Difference"]
for col_idx, h in enumerate(headers_dash, start=1):
    cell = ws_dash.cell(row=4, column=col_idx, value=h)
    cell.font = copy(dash_header_styles[col_idx]['font'])
    cell.fill = copy(dash_header_styles[col_idx]['fill'])
    cell.border = copy(dash_header_styles[col_idx]['border'])
    cell.alignment = copy(dash_header_styles[col_idx]['alignment'])

# Write Tubes products
row_idx = 5
for prod in samsol_tubes:
    ws_dash.cell(row=row_idx, column=1, value=prod['customer']).alignment = align_left
    ws_dash.cell(row=row_idx, column=2, value=prod['name']).alignment = align_left
    ws_dash.cell(row=row_idx, column=3, value=prod['dia']).alignment = align_center
    ws_dash.cell(row=row_idx, column=4, value=prod['pid']).alignment = align_center
    
    prod_terms = [f"SUMIFS('Samsol Production Logging'!$G:$G, 'Samsol Production Logging'!$C:$C, $B{row_idx}, 'Samsol Production Logging'!$B:$B, $I${m_idx}) * IF($J${m_idx}=\"Y\", 1, 0)" for m_idx in range(5, 5 + len(months_sorted))]
    disp_terms = [f"SUMIFS('Samsol Dispatch Details'!$G:$G, 'Samsol Dispatch Details'!$C:$C, $B{row_idx}, 'Samsol Dispatch Details'!$B:$B, $I${m_idx}) * IF($J${m_idx}=\"Y\", 1, 0)" for m_idx in range(5, 5 + len(months_sorted))]
    
    ws_dash.cell(row=row_idx, column=5, value="=" + " + ".join(prod_terms))
    ws_dash.cell(row=row_idx, column=6, value="=" + " + ".join(disp_terms))
    ws_dash.cell(row=row_idx, column=7, value=f"=E{row_idx}-F{row_idx}")
    
    for col in range(1, 8):
        cell = ws_dash.cell(row=row_idx, column=col)
        cell.font = copy(dash_data_styles[col]['font'])
        cell.fill = copy(dash_data_styles[col]['fill'])
        cell.border = copy(dash_data_styles[col]['border'])
        if col not in [1, 2]:
            cell.alignment = copy(dash_data_styles[col]['alignment'])
        cell.number_format = dash_data_styles[col]['number_format']
    row_idx += 1

# Total Tubes Row
cell_tot_lbl = ws_dash.cell(row=row_idx, column=1, value="Grand Total")
cell_tot_lbl.font = copy(dash_total_styles[1]['font'])
cell_tot_lbl.fill = copy(dash_total_styles[1]['fill'])
cell_tot_lbl.border = copy(dash_total_styles[1]['border'])
cell_tot_lbl.alignment = align_left

for col in range(2, 5):
    cell = ws_dash.cell(row=row_idx, column=col)
    cell.font = copy(dash_total_styles[col]['font'])
    cell.fill = copy(dash_total_styles[col]['fill'])
    cell.border = copy(dash_total_styles[col]['border'])

tot_p = ws_dash.cell(row=row_idx, column=5, value=f"=SUM(E5:E{row_idx-1})")
tot_d = ws_dash.cell(row=row_idx, column=6, value=f"=SUM(F5:F{row_idx-1})")
tot_df = ws_dash.cell(row=row_idx, column=7, value=f"=SUM(G5:G{row_idx-1})")

for col, val in enumerate([tot_p, tot_d, tot_df], start=5):
    val.font = copy(dash_total_styles[col]['font'])
    val.fill = copy(dash_total_styles[col]['fill'])
    val.border = copy(dash_total_styles[col]['border'])
    val.alignment = copy(dash_total_styles[col]['alignment'])
    val.number_format = dash_total_styles[col]['number_format']

# Section 2 Title: Row row_idx+2
row_idx += 2
ws_dash.cell(row=row_idx, column=1, value="2. Samsol PET Dashboard").font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")

# Headers (Row row_idx+1 Columns A to G)
row_idx += 1
for col_idx, h in enumerate(headers_dash, start=1):
    cell = ws_dash.cell(row=row_idx, column=col_idx, value=h)
    cell.font = copy(dash_header_styles[col_idx]['font'])
    cell.fill = copy(dash_header_styles[col_idx]['fill'])
    cell.border = copy(dash_header_styles[col_idx]['border'])
    cell.alignment = copy(dash_header_styles[col_idx]['alignment'])

# Write PET products
start_pet_dash = row_idx + 1
row_idx += 1
for prod in samsol_pet:
    ws_dash.cell(row=row_idx, column=1, value=prod['customer']).alignment = align_left
    ws_dash.cell(row=row_idx, column=2, value=prod['name']).alignment = align_left
    ws_dash.cell(row=row_idx, column=3, value=prod['dia']).alignment = align_center
    ws_dash.cell(row=row_idx, column=4, value=prod['pid']).alignment = align_center
    
    prod_terms = [f"SUMIFS('Samsol Production Logging'!$G:$G, 'Samsol Production Logging'!$C:$C, $B{row_idx}, 'Samsol Production Logging'!$B:$B, $I${m_idx}) * IF($J${m_idx}=\"Y\", 1, 0)" for m_idx in range(5, 5 + len(months_sorted))]
    disp_terms = [f"SUMIFS('Samsol Dispatch Details'!$G:$G, 'Samsol Dispatch Details'!$C:$C, $B{row_idx}, 'Samsol Dispatch Details'!$B:$B, $I${m_idx}) * IF($J${m_idx}=\"Y\", 1, 0)" for m_idx in range(5, 5 + len(months_sorted))]
    
    ws_dash.cell(row=row_idx, column=5, value="=" + " + ".join(prod_terms))
    ws_dash.cell(row=row_idx, column=6, value="=" + " + ".join(disp_terms))
    ws_dash.cell(row=row_idx, column=7, value=f"=E{row_idx}-F{row_idx}")
    
    for col in range(1, 8):
        cell = ws_dash.cell(row=row_idx, column=col)
        cell.font = copy(dash_data_styles[col]['font'])
        cell.fill = copy(dash_data_styles[col]['fill'])
        cell.border = copy(dash_data_styles[col]['border'])
        if col not in [1, 2]:
            cell.alignment = copy(dash_data_styles[col]['alignment'])
        cell.number_format = dash_data_styles[col]['number_format']
    row_idx += 1

# Total PET Row
cell_tot_lbl = ws_dash.cell(row=row_idx, column=1, value="Grand Total")
cell_tot_lbl.font = copy(dash_total_styles[1]['font'])
cell_tot_lbl.fill = copy(dash_total_styles[1]['fill'])
cell_tot_lbl.border = copy(dash_total_styles[1]['border'])
cell_tot_lbl.alignment = align_left

for col in range(2, 5):
    cell = ws_dash.cell(row=row_idx, column=col)
    cell.font = copy(dash_total_styles[col]['font'])
    cell.fill = copy(dash_total_styles[col]['fill'])
    cell.border = copy(dash_total_styles[col]['border'])

tot_p = ws_dash.cell(row=row_idx, column=5, value=f"=SUM(E{start_pet_dash}:E{row_idx-1})")
tot_d = ws_dash.cell(row=row_idx, column=6, value=f"=SUM(F{start_pet_dash}:F{row_idx-1})")
tot_df = ws_dash.cell(row=row_idx, column=7, value=f"=SUM(G{start_pet_dash}:G{row_idx-1})")

for col, val in enumerate([tot_p, tot_d, tot_df], start=5):
    val.font = copy(dash_total_styles[col]['font'])
    val.fill = copy(dash_total_styles[col]['fill'])
    val.border = copy(dash_total_styles[col]['border'])
    val.alignment = copy(dash_total_styles[col]['alignment'])
    val.number_format = dash_total_styles[col]['number_format']

# Month Selection Panel (Columns 9 and 10, Rows 4 to 13)
cell_m_header = ws_dash.cell(row=4, column=9, value="Month")
cell_m_header.font = copy(panel_header_styles[9]['font'])
cell_m_header.fill = copy(panel_header_styles[9]['fill'])
cell_m_header.border = copy(panel_header_styles[9]['border'])
cell_m_header.alignment = copy(panel_header_styles[9]['alignment'])

cell_s_header = ws_dash.cell(row=4, column=10, value="Select (Y/N)")
cell_s_header.font = copy(panel_header_styles[10]['font'])
cell_s_header.fill = copy(panel_header_styles[10]['fill'])
cell_s_header.border = copy(panel_header_styles[10]['border'])
cell_s_header.alignment = copy(panel_header_styles[10]['alignment'])

for idx, m_date in enumerate(months_sorted, start=5):
    cell_m = ws_dash.cell(row=idx, column=9, value=m_date)
    cell_m.font = copy(panel_data_styles[9]['font'])
    cell_m.fill = copy(panel_data_styles[9]['fill'])
    cell_m.border = copy(panel_data_styles[9]['border'])
    cell_m.alignment = copy(panel_data_styles[9]['alignment'])
    cell_m.number_format = 'mmm-yy'
    
    cell_s = ws_dash.cell(row=idx, column=10, value="Y")
    cell_s.font = copy(panel_data_styles[10]['font'])
    cell_s.fill = copy(panel_data_styles[10]['fill'])
    cell_s.border = copy(panel_data_styles[10]['border'])
    cell_s.alignment = copy(panel_data_styles[10]['alignment'])
    cell_s.number_format = panel_data_styles[10]['number_format']

# Dropdown list validation for Selection
dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
ws_dash.add_data_validation(dv_yn)
for r in range(5, 5 + len(months_sorted)):
    dv_yn.add(ws_dash.cell(row=r, column=10))

ws_dash.column_dimensions['H'].width = 3


# --- E. Populating Product Monthly Breakdown ---
print("Populating Product Monthly Breakdown...")
if "Product Monthly Breakdown" in wb.sheetnames:
    ws_pb = wb["Product Monthly Breakdown"]
    ws_pb.delete_rows(1, ws_pb.max_row+10)
else:
    ws_pb = wb.create_sheet("Product Monthly Breakdown")

ws_pb.views.sheetView[0].showGridLines = True
ws_pb.cell(row=2, column=2, value="Monthly Production & Dispatch Breakdown by Product").font = sum_title_font

headers_pb = ["Product Name", "Customer (Catalog)", "Month", "Produced Qty", "Dispatched Qty", "Net Difference"]
for col_idx, h in enumerate(headers_pb, start=2):
    cell = ws_pb.cell(row=4, column=col_idx, value=h)
    col_style_idx = 1 if col_idx < 5 else col_idx - 3
    cell.font = copy(sum_header_styles[col_style_idx]['font'])
    cell.fill = copy(sum_header_styles[col_style_idx]['fill'])
    cell.border = copy(sum_header_styles[col_style_idx]['border'])
    cell.alignment = copy(sum_header_styles[col_style_idx]['alignment'])

p_prod = df_p_samsol.groupby(['Product Name', 'Month'])['Produced Qty'].sum().reset_index()
d_prod = df_d_samsol.groupby(['Product Name', 'Month'])['Dispatched Qty'].sum().reset_index()
df_pb_data = pd.merge(p_prod, d_prod, on=['Product Name', 'Month'], how='outer').fillna(0)
df_pb_data['Produced Qty'] = df_pb_data['Produced Qty'].astype(int)
df_pb_data['Dispatched Qty'] = df_pb_data['Dispatched Qty'].astype(int)
df_pb_data = df_pb_data.sort_values(['Product Name', 'Month'])

row_idx_pb = 5
for idx, row in df_pb_data.iterrows():
    c_name = ws_pb.cell(row=row_idx_pb, column=2, value=row['Product Name'])
    cust_val = catalog_map.get(row['Product Name'].upper(), {}).get('customer', 'Unknown')
    c_cust = ws_pb.cell(row=row_idx_pb, column=3, value=cust_val)
    c_month = ws_pb.cell(row=row_idx_pb, column=4, value=row['Month'])
    
    c_prod = ws_pb.cell(row=row_idx_pb, column=5, value=f"=SUMIFS('Samsol Production Logging'!$G:$G, 'Samsol Production Logging'!$C:$C, $B{row_idx_pb}, 'Samsol Production Logging'!$B:$B, $D{row_idx_pb})")
    c_disp = ws_pb.cell(row=row_idx_pb, column=6, value=f"=SUMIFS('Samsol Dispatch Details'!$G:$G, 'Samsol Dispatch Details'!$C:$C, $B{row_idx_pb}, 'Samsol Dispatch Details'!$B:$B, $D{row_idx_pb})")
    c_diff = ws_pb.cell(row=row_idx_pb, column=7, value=f"=E{row_idx_pb}-F{row_idx_pb}")
    
    c_name.alignment = align_left
    c_cust.alignment = align_left
    c_month.number_format = 'mmm-yy'
    
    for col_idx, cell in enumerate([c_name, c_cust, c_month, c_prod, c_disp, c_diff], start=2):
        col_style_idx = 1 if col_idx < 5 else col_idx - 3
        cell.font = copy(sum_data_styles[col_style_idx]['font'])
        cell.fill = copy(sum_data_styles[col_style_idx]['fill'])
        cell.border = copy(sum_data_styles[col_style_idx]['border'])
        if col_idx in [5, 6, 7]:
            cell.alignment = align_center
            cell.number_format = '#,##0'
    row_idx_pb += 1

# Total row
cell_tot_lbl = ws_pb.cell(row=row_idx_pb, column=2, value="Total")
cell_tot_lbl.font = copy(sum_total_styles[1]['font'])
cell_tot_lbl.fill = copy(sum_total_styles[1]['fill'])
cell_tot_lbl.border = copy(sum_total_styles[1]['border'])
cell_tot_lbl.alignment = align_left

for col in range(3, 5):
    cell = ws_pb.cell(row=row_idx_pb, column=col)
    cell.font = copy(sum_total_styles[1]['font'])
    cell.fill = copy(sum_total_styles[1]['fill'])
    cell.border = copy(sum_total_styles[1]['border'])

tot_p = ws_pb.cell(row=row_idx_pb, column=5, value=f"=SUM(E5:E{row_idx_pb-1})")
tot_d = ws_pb.cell(row=row_idx_pb, column=6, value=f"=SUM(F5:F{row_idx_pb-1})")
tot_df = ws_pb.cell(row=row_idx_pb, column=7, value=f"=SUM(G5:G{row_idx_pb-1})")

for col, val in enumerate([tot_p, tot_d, tot_df], start=5):
    val.font = copy(sum_total_styles[col-3]['font'])
    val.fill = copy(sum_total_styles[col-3]['fill'])
    val.border = copy(sum_total_styles[col-3]['border'])
    val.alignment = align_center
    val.number_format = '#,##0'


# --- F. Populating Unidentified Products ---
print("Populating Unidentified Products...")
if "Unidentified Products" in wb.sheetnames:
    ws_un = wb["Unidentified Products"]
    ws_un.delete_rows(1, ws_un.max_row+10)
else:
    ws_un = wb.create_sheet("Unidentified Products")

ws_un.views.sheetView[0].showGridLines = True
ws_un.cell(row=2, column=2, value="Products Lacking Product Catalog Mapping").font = sum_title_font
ws_un.cell(row=3, column=2, value="Verify these names to check if any belong to Samsol. Highlighted in light red.").font = Font(name="Segoe UI", size=9, italic=True, color="7F8C8D")

headers_un = ["Source", "Date", "POF #", "Ref. #", "Product Name", "Dia/Volume", "Qty", "Wastage/Repl. Qty", "Party Name (Act.)"]
for col_idx, h in enumerate(headers_un, start=2):
    cell = ws_un.cell(row=5, column=col_idx, value=h)
    cell.font = copy(sum_header_styles[min(col_idx-1, 4)]['font'])
    cell.fill = copy(sum_header_styles[min(col_idx-1, 4)]['fill'])
    cell.border = copy(sum_header_styles[min(col_idx-1, 4)]['border'])
    cell.alignment = copy(sum_header_styles[min(col_idx-1, 4)]['alignment'])

row_idx_un = 6
unidentified_prod = sorted(unidentified_prod, key=lambda x: (x['Source'], x['Date']))
for item in unidentified_prod:
    ws_un.cell(row=row_idx_un, column=2, value=item['Source'])
    ws_un.cell(row=row_idx_un, column=3, value=item['Date'])
    ws_un.cell(row=row_idx_un, column=4, value=item['POF #'])
    ws_un.cell(row=row_idx_un, column=5, value=item['Ref. #'])
    ws_un.cell(row=row_idx_un, column=6, value=item['Product Name'])
    ws_un.cell(row=row_idx_un, column=7, value=item['Dia/Volume'])
    ws_un.cell(row=row_idx_un, column=8, value=item['Qty'])
    ws_un.cell(row=row_idx_un, column=9, value=item['Wastage/Repl. Qty'])
    ws_un.cell(row=row_idx_un, column=10, value=item['Party Name (Act.)'])
    
    for col_idx in range(2, 11):
        cell = ws_un.cell(row=row_idx_un, column=col_idx)
        cell.font = copy(sum_data_styles[min(col_idx-1, 4)]['font'])
        cell.fill = fill_alert
        cell.border = copy(sum_data_styles[min(col_idx-1, 4)]['border'])
        cell.alignment = align_center
        
        if col_idx == 3:
            cell.number_format = 'yyyy-mm-dd'
        elif col_idx in [4, 5]:
            cell.number_format = '0'
        elif col_idx in [8, 9]:
            cell.number_format = '#,##0'
        elif col_idx in [6, 10]:
            cell.alignment = align_left
    row_idx_un += 1


# --- G. Auto-fit widths for New Sheets ONLY ---
print("Auto-fitting widths for new sheets...")
for sheet_name in ["Product Monthly Breakdown", "Unidentified Products"]:
    ws = wb[sheet_name]
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2, 3] and cell.value and len(str(cell.value)) > 25:
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Save Workbook
print(f"Saving workbook to {output_path}...")
wb.save(output_path)
print("Saved successfully!")
