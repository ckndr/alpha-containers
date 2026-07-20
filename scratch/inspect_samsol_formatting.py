import openpyxl

samsol_path = "D:/Alpha/Tubex Records/Samsol_Production_and_Dispatch.xlsx"
wb = openpyxl.load_workbook(samsol_path)
ws = wb["Monthly Summary"]

print("=== Col Widths ===")
for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    print(f"Col {col} width: {ws.column_dimensions[col].width}")

print("\n=== Row Heights ===")
for r in range(1, 15):
    print(f"Row {r} height: {ws.row_dimensions[r].height}")

print("\n=== Detailed Cell Format (Row 2, Column A to D) ===")
for c in range(1, 5):
    cell = ws.cell(row=2, column=c)
    print(f"Cell {cell.coordinate}: val={cell.value}")
    print(f"  font: name={cell.font.name}, size={cell.font.size}, bold={cell.font.bold}, color={cell.font.color.value if cell.font.color else None}")
    print(f"  fill: type={cell.fill.fill_type}, fgColor={cell.fill.fgColor.value if cell.fill.fgColor else None}")
    print(f"  border: left={cell.border.left.style}, right={cell.border.right.style}, top={cell.border.top.style}, bottom={cell.border.bottom.style}")
    print(f"  alignment: horiz={cell.alignment.horizontal}, vert={cell.alignment.vertical}")

print("\n=== Detailed Cell Format (Row 3, Column A to D) ===")
for c in range(1, 5):
    cell = ws.cell(row=3, column=c)
    print(f"Cell {cell.coordinate}: val={cell.value}")
    print(f"  font: name={cell.font.name}, size={cell.font.size}, bold={cell.font.bold}, color={cell.font.color.value if cell.font.color else None}")
    print(f"  fill: type={cell.fill.fill_type}, fgColor={cell.fill.fgColor.value if cell.fill.fgColor else None}")
    print(f"  border: left={cell.border.left.style}, right={cell.border.right.style}, top={cell.border.top.style}, bottom={cell.border.bottom.style}")
    print(f"  alignment: horiz={cell.alignment.horizontal}, vert={cell.alignment.vertical}")
    print(f"  num_format: {cell.number_format}")

print("\n=== Detailed Cell Format (Row 12, Column A to D) ===")
for c in range(1, 5):
    cell = ws.cell(row=12, column=c)
    print(f"Cell {cell.coordinate}: val={cell.value}")
    print(f"  font: name={cell.font.name}, size={cell.font.size}, bold={cell.font.bold}, color={cell.font.color.value if cell.font.color else None}")
    print(f"  fill: type={cell.fill.fill_type}, fgColor={cell.fill.fgColor.value if cell.fill.fgColor else None}")
    print(f"  border: left={cell.border.left.style}, right={cell.border.right.style}, top={cell.border.top.style}, bottom={cell.border.bottom.style}")
    print(f"  alignment: horiz={cell.alignment.horizontal}, vert={cell.alignment.vertical}")
    print(f"  num_format: {cell.number_format}")
