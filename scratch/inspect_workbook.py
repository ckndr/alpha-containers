import openpyxl

wb = openpyxl.load_workbook("d:\\Alpha\\Tubex_July26.xlsx")
print("Sheets:", wb.sheetnames)

if "MRP" in wb.sheetnames:
    ws_mrp = wb["MRP"]
    print("MRP sheet dimensions:", ws_mrp.dimensions)
    # Let's print the first 20 rows of MRP
    for r in range(1, 150):
        row_vals = [ws_mrp.cell(row=r, column=c).value for c in range(1, 15)]
        if any(row_vals):
            print(f"Row {r:03d}: {row_vals}")

if "Dashboard" in wb.sheetnames:
    ws_dash = wb["Dashboard"]
    print("Dashboard dimensions:", ws_dash.dimensions)
    for r in range(1, 100):
        row_vals = [ws_dash.cell(row=r, column=c).value for c in range(1, 15)]
        if any(row_vals):
            print(f"Row {r:03d}: {row_vals}")
