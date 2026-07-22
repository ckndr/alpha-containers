import openpyxl
from copy import copy

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

def is_numeric_item_id(val):
    if isinstance(val, int):
        return True
    if isinstance(val, str):
        val_clean = val.strip()
        if val_clean.isdigit():
            return True
    return False

def main():
    file_path = "d:\\Alpha\\Tubex_July26.xlsx"
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    # 1. Update Tubex_Dashboard sheet
    print("Updating Tubex_Dashboard sheet...")
    ws_dash = wb["Tubex_Dashboard"]
    
    # Let's find two empty rows at the bottom of Dashboard (e.g. rows 65 & 66)
    # We will write new rows there
    ws_dash.cell(row=65, column=2, value="PET")
    ws_dash.cell(row=65, column=3, value="Samsol International Private Limited")
    ws_dash.cell(row=65, column=4, value="PET BOTTLE SMALL (120 ML) YELLOW")
    ws_dash.cell(row=65, column=5, value="120 ml")
    ws_dash.cell(row=65, column=6, value=8016)
    ws_dash.cell(row=65, column=7, value=30000)
    
    ws_dash.cell(row=66, column=2, value="PET")
    ws_dash.cell(row=66, column=3, value="Alpha Labs PVT LTD")
    ws_dash.cell(row=66, column=4, value="PET BOTTLE (150ML)TRANSPARENT BODY MIST")
    ws_dash.cell(row=66, column=5, value="150 ml")
    ws_dash.cell(row=66, column=6, value=8020)
    ws_dash.cell(row=66, column=7, value=125000)
    
    # 2. Update MRP sheet
    print("Updating MRP sheet...")
    ws_mrp = wb["MRP"]
    
    # We insert 2 rows after row 101 (so new rows will be 102 and 103)
    print("Inserting 2 rows at row 102...")
    ws_mrp.insert_rows(102, 2)
    
    # Copy styles from row 101 to the new row 102 and 103
    for col in range(1, ws_mrp.max_column + 1):
        copy_style(ws_mrp.cell(row=101, column=col), ws_mrp.cell(row=102, column=col))
        copy_style(ws_mrp.cell(row=101, column=col), ws_mrp.cell(row=103, column=col))
        
    # Populate new row 102 (Samsol small 120 ml yellow PET order)
    ws_mrp.cell(row=102, column=1, value="120 ml")
    ws_mrp.cell(row=102, column=2, value="Samsol International Private Limited")
    ws_mrp.cell(row=102, column=3, value="PET BOTTLE SMALL (120 ML) YELLOW")
    ws_mrp.cell(row=102, column=4, value=8016)
    ws_mrp.cell(row=102, column=5, value=345)
    ws_mrp.cell(row=102, column=6, value="=INDEX(Tubex_Dashboard!$G$11:$G$61,MATCH(MRP!D102,Tubex_Dashboard!$F$11:$F$61,0))")
    ws_mrp.cell(row=102, column=7, value="=INDEX(Tubex_Dashboard!$H$11:$H$100,MATCH(MRP!D102,Tubex_Dashboard!$F$11:$F$100,0))")
    ws_mrp.cell(row=102, column=8, value="=F102-G102")
    ws_mrp.cell(row=102, column=9, value=None)
    
    # Populate new row 103 (Alpha Labs combined 346 & 347 PET order)
    ws_mrp.cell(row=103, column=1, value="150 ml")
    ws_mrp.cell(row=103, column=2, value="Alpha Labs PVT LTD")
    ws_mrp.cell(row=103, column=3, value="PET BOTTLE (150ML)TRANSPARENT BODY MIST")
    ws_mrp.cell(row=103, column=4, value=8020)
    ws_mrp.cell(row=103, column=5, value="346 & 347") # Combined JOFs
    ws_mrp.cell(row=103, column=6, value="=INDEX(Tubex_Dashboard!$G$11:$G$61,MATCH(MRP!D103,Tubex_Dashboard!$F$11:$F$61,0))")
    ws_mrp.cell(row=103, column=7, value="=INDEX(Tubex_Dashboard!$H$11:$H$100,MATCH(MRP!D103,Tubex_Dashboard!$F$11:$F$100,0))")
    ws_mrp.cell(row=103, column=8, value="=F103-G103")
    ws_mrp.cell(row=103, column=9, value=None)
    
    # 3. Update Total Row formulas (now at row 104)
    print("Updating PET Total Row (Row 104) formulas...")
    ws_mrp.cell(row=104, column=6, value="=SUM(F96:F103)")
    ws_mrp.cell(row=104, column=7, value="=SUM(G96:G103)")
    ws_mrp.cell(row=104, column=8, value="=SUM(H96:H103)")
    
    # 4. Update PET Material Plan (now at rows 107 to 113)
    print("Updating PET Material Plan (Rows 107 to 113)...")
    for r in range(107, 114):
        item_id = ws_mrp.cell(row=r, column=1).value
        if not is_numeric_item_id(item_id):
            continue
        # Column E: Required Qty
        ws_mrp.cell(row=r, column=5, value=f"=SUMPRODUCT((TableBOM[Item ID]=A{r})*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %])*SUMIF($D$96:$D$103,TableBOM[Product ID],$H$96:$H$103)/1000)")
        # Column F: Current Stock
        ws_mrp.cell(row=r, column=6, value=f"=IFERROR(INDEX(TableInventory[Store Balance],MATCH(A{r},TableInventory[Item ID],0)),0)+IFERROR(INDEX(TableInventory[Work In Process],MATCH(A{r},TableInventory[Item ID],0)),0)")
        # Column G: Surplus / (Deficit)
        ws_mrp.cell(row=r, column=7, value=f"=F{r}-E{r}")
        
        # Column H: Product Name(s) formula (concatenating all active orders: rows 96 to 103)
        inner_parts_h = [
            f'IF((COUNTIFS(TableBOM[Product ID],$D${x},TableBOM[Item ID],$A{r})>0)*($H${x}>0),$C${x}&\", \",\"\")'
            for x in range(96, 104)
        ]
        concat_h = " & ".join(inner_parts_h)
        formula_h = f'=IF(LEN({concat_h})>1,LEFT({concat_h},LEN({concat_h})-2),\"\")'
        ws_mrp.cell(row=r, column=8, value=formula_h)
        
        # Column I: Status formula
        ws_mrp.cell(row=r, column=9, value=f'=IF(E{r}=0,"Not needed",IF(G{r}<0,"SHORTAGE",IF(G{r}<F{r}*0.1,"LOW","OK")))')
        
    # 5. Update Ink Table (now at rows 117 to 148)
    print("Updating Ink Table (Rows 117 to 148)...")
    for r in range(117, 149):
        item_id = ws_mrp.cell(row=r, column=1).value
        if not is_numeric_item_id(item_id):
            continue
        # Column E: Avg Monthly Usage formula (referencing the updated range $D$96:$D$103 and $H$96:$H$103)
        ws_mrp.cell(row=r, column=5, value=f"=SUMPRODUCT((TableBOM[Item ID]=A{r})*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %])*SUMIF($D$96:$D$103,TableBOM[Product ID],$H$96:$H$103)/1000)")
        # Column H: Current Stock (indexed from inventory)
        ws_mrp.cell(row=r, column=8, value=f"=IFERROR(INDEX(TableInventory[Store Balance],MATCH(A{r},TableInventory[Item ID],0)),0)+IFERROR(INDEX(TableInventory[Work In Process],MATCH(A{r},TableInventory[Item ID],0)),0)")
        # Column F: Days of Stock Left
        ws_mrp.cell(row=r, column=6, value=f"=IF(E{r}=0,0,ROUND(H{r}/E{r}*30,1))")
        # Column G: Status
        ws_mrp.cell(row=r, column=7, value=f'=IF(E{r}=0,"Not needed",IF(H{r}-E{r}<0,"SHORTAGE",IF(H{r}-E{r}<E{r},"LOW","OK")))')
        
    # Save workbook
    print(f"Saving workbook to {file_path}...")
    wb.save(file_path)
    print("Workbook successfully updated!")

if __name__ == "__main__":
    main()
