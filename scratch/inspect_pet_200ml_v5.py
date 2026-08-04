import openpyxl

wb = openpyxl.load_workbook("D:/Alpha/Tubex_July26.xlsx")

# Check ALL tables in ALL sheets
for sn in wb.sheetnames:
    ws = wb[sn]
    if ws.tables:
        for table in ws.tables.values():
            print(f"Sheet: {sn}, Table: {table.name}, Ref: {table.ref}")

# Check the header row for Product_Catalog (row 2 since row 1 is title)
print("\n=== Product_Catalog header details ===")
ws = wb["Product_Catalog"]
for r in range(1, 4):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"  Row {r}: {row_vals}")

# Check the header row for BOM (row 2 since row 1 is title)
print("\n=== BOM header details ===")
ws = wb["BOM"]
for r in range(1, 4):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"  Row {r}: {row_vals}")

# What product IDs exist in the 8xxx range?
print("\n=== All product IDs in 8xxx range (catalog) ===")
ws = wb["Product_Catalog"]
ids_8xxx = set()
for r in range(2, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if isinstance(val, (int, float)) and 8000 <= val < 9000:
        ids_8xxx.add(int(val))
print(f"  IDs: {sorted(ids_8xxx)}")

print("\n=== All product IDs in 8xxx range (BOM) ===")
ws = wb["BOM"]
ids_8xxx_bom = set()
for r in range(2, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if isinstance(val, (int, float)) and 8000 <= val < 9000:
        ids_8xxx_bom.add(int(val))
print(f"  IDs: {sorted(ids_8xxx_bom)}")

print("\n=== All product IDs in 8xxx range (dashboard) ===")
ws = wb["Tubex_Dashboard"]
ids_8xxx_dash = set()
for r in range(2, ws.max_row + 1):
    val = ws.cell(row=r, column=6).value
    if isinstance(val, (int, float)) and 8000 <= val < 9000:
        ids_8xxx_dash.add(int(val))
print(f"  IDs: {sorted(ids_8xxx_dash)}")
