import openpyxl

wb = openpyxl.load_workbook("D:/Alpha/Tubex_July26.xlsx")

# 1. Product Catalog - full details
print("="*80)
print("PRODUCT CATALOG - Full structure")
print("="*80)
ws = wb["Product_Catalog"]
# Print ALL headers (row 1)
header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print(f"Header (cols 1-{ws.max_column}): {header}")

# Find last product ID and row
print(f"\nTotal rows: {ws.max_row}")
last_rows = []
for r in range(ws.max_row, max(1, ws.max_row-10), -1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    if any(row_vals):
        last_rows.append((r, row_vals))
print("Last rows with data:")
for r, vals in reversed(last_rows):
    print(f"  Row {r}: {vals}")

# Print all PET rows (not just 200ml)
print("\nAll PET rows in catalog:")
for r in range(2, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    row_str = str(row_vals).upper()
    if "PET" in row_str:
        print(f"  Row {r}: {row_vals}")

# Check if COCONUT already exists
print("\nCOCONUT rows in catalog:")
for r in range(2, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    row_str = str(row_vals).upper()
    if "COCONUT" in row_str:
        print(f"  Row {r}: {row_vals}")

# 2. BOM - full details for PET BOTTLE LARGE 200ML WHITE (8007)
print("\n" + "="*80)
print("BOM - Full structure")
print("="*80)
ws = wb["BOM"]
# Print header rows
for r in range(1, 5):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"  Row {r}: {row_vals}")

# Last rows
print(f"\nTotal rows: {ws.max_row}")
for r in range(max(1, ws.max_row-5), ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    if any(row_vals):
        print(f"  Row {r}: {row_vals}")

# All PET 200ML white (8007) BOM entries
print("\nBOM for PET BOTTLE LARGE 200ML WHITE (8007):")
for r in range(2, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val == 8007:
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"  Row {r}: {row_vals}")

# All PET MUSTARD OIL 200ML (8014) BOM entries
print("\nBOM for PET BOTTLE MUSTARD OIL (200ML) (8014):")
for r in range(2, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if val == 8014:
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"  Row {r}: {row_vals}")

# 3. Dashboard - full details around PET section
print("\n" + "="*80)
print("DASHBOARD - Full structure")
print("="*80)
ws = wb["Tubex_Dashboard"]
# Print rows 1-10 for header structure
for r in range(1, 12):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"  Row {r}: {row_vals}")

# Print all PET rows
print("\nAll PET rows in dashboard:")
for r in range(4, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(15, ws.max_column + 1))]
    row_str = str(row_vals).upper()
    if "PET" in row_str:
        print(f"  Row {r}: {row_vals}")

# Last rows of dashboard
print(f"\nTotal rows: {ws.max_row}")
for r in range(max(1, ws.max_row-5), ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(15, ws.max_column + 1))]
    if any(row_vals):
        print(f"  Row {r}: {row_vals}")

# Check max product ID
print("\nMax product IDs:")
ws = wb["Product_Catalog"]
max_id = 0
for r in range(2, ws.max_row + 1):
    val = ws.cell(row=r, column=1).value
    if isinstance(val, (int, float)) and val > max_id:
        max_id = val
print(f"  Max Product ID in catalog: {max_id}")

max_sub = 0
for r in range(2, ws.max_row + 1):
    val = ws.cell(row=r, column=2).value
    if isinstance(val, (int, float)) and val > max_sub:
        max_sub = val
print(f"  Max Sub-ID in catalog: {max_sub}")
