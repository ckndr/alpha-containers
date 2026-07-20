import openpyxl

wb = openpyxl.load_workbook("D:/Alpha/Tubex_July26.xlsx", data_only=True)
ws = wb["Tubex_Dashboard"]

print("=== Searching for Samsol in Tubex_Dashboard ===")
for r in range(1, ws.max_row+1):
    row_vals = [cell.value for cell in ws[r]]
    row_str = " | ".join([str(v) for v in row_vals if v is not None])
    if "samsol" in row_str.lower():
        print(f"Row {r:2d}: {row_str}")
