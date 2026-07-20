import pandas as pd
import openpyxl

file_path = "D:/Alpha/Tubex Records/Samsol_Production_and_Dispatch.xlsx"
print("Reading generated Excel:", file_path)

wb = openpyxl.load_workbook(file_path, data_only=False)
print("Sheet names:", wb.sheetnames)

# Let's inspect the first sheet (Monthly Summary)
ws_sum = wb["Monthly Summary"]
print("\n=== Monthly Summary Sheet Layout ===")
for r in range(1, 40):
    row_vals = [ws_sum.cell(row=r, column=c).value for c in range(1, 7)]
    if any(row_vals):
        print(f"Row {r:2d}: {row_vals}")

# Let's count rows in other sheets
for name in wb.sheetnames:
    if name != "Monthly Summary":
        ws = wb[name]
        print(f"Sheet '{name}': {ws.max_row} rows (including headers)")
