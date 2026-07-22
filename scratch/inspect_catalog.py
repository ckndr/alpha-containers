import openpyxl

wb = openpyxl.load_workbook("d:\\Alpha\\Tubex_July26.xlsx")
if "Product_Catalog" in wb.sheetnames:
    ws = wb["Product_Catalog"]
    print("=== PRODUCT CATALOG ===")
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        if any(row_vals):
            print(f"Row {r:03d}: {row_vals}")
else:
    print("Product_Catalog sheet not found!")
