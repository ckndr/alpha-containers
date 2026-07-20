import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

sys.path.append("D:/Alpha/Scripts")
import update_production
import update_dispatch

catalog_path = "D:/Alpha/Tubex_July26.xlsx"
prod_path = "D:/Alpha/Tubex Records/production nov to jul.xls"
disp_path = "D:/Alpha/Tubex Records/dispatch nov to jul.xls"
output_path = "D:/Alpha/Tubex Records/Samsol_Production_and_Dispatch.xlsx"

# 1. Load Product Catalog
xls_cat = pd.ExcelFile(catalog_path)
df_cat = pd.read_excel(xls_cat, sheet_name="Product_Catalog", skiprows=1)

# Catalog mapping
catalog_map = {}
for idx, row in df_cat.iterrows():
    name = str(row['Product Name']).strip().upper()
    pid = row['Product ID']
    cust = str(row['Customer']).strip()
    if pd.notna(pid):
        pid = int(pid)
    catalog_map[name] = {'pid': pid, 'customer': cust, 'name': row['Product Name']}

# 2. Parse Production File
print("Parsing Production...")
df_p = pd.read_excel(prod_path, header=None)
prod_records = []
unidentified_prod = []  # List of dicts for unidentified sheet

for idx, row in df_p.iterrows():
    row_val = row.dropna().tolist()
    if not row_val:
        continue
    dt = row[1]
    if pd.isna(dt) or not (isinstance(dt, pd.Timestamp) or hasattr(dt, 'date')):
        continue
    
    ref_num = row[0]
    date_val = pd.Timestamp(dt)
    pof_num = row[5]
    name_raw = str(row[6]).strip()
    dia_raw = str(row[8]).strip()
    
    def get_float(val):
        if pd.isna(val): return 0.0
        try: return float(val)
        except: return 0.0
        
    normal_good = get_float(row[9])
    ot_good = get_float(row[12])
    normal_wastage = get_float(row[10])
    ot_wastage = get_float(row[13])
    
    total_good = int(normal_good + ot_good)
    total_wastage = int(normal_wastage + ot_wastage)
    
    # Resolve product
    name_key = name_raw.lower().strip()
    match = update_production.ALIASES.get((name_key, dia_raw))
    if match is None:
        match = update_production.ALIASES.get((name_raw.lower().rstrip(), dia_raw))
        
    resolved_name = None
    resolved_pid = None
    resolved_cust = None
    
    if match:
        resolved_name, resolved_pid = match
        if resolved_pid in update_production.PID_TO_CUSTOMER:
            resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
        elif resolved_name.upper() in catalog_map:
            resolved_cust = catalog_map[resolved_name.upper()]['customer']
    else:
        name_upper = name_raw.upper()
        if name_upper in catalog_map:
            resolved_name = catalog_map[name_upper]['name']
            resolved_pid = catalog_map[name_upper]['pid']
            resolved_cust = catalog_map[name_upper]['customer']
            if resolved_pid in update_production.PID_TO_CUSTOMER:
                resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
                
    if resolved_name:
        # Determine definitive customer using Product_Catalog sheet
        cat_info = catalog_map.get(resolved_name.upper())
        cat_cust = cat_info['customer'] if cat_info else resolved_cust
        
        is_samsol = cat_cust and 'samsol' in str(cat_cust).lower()
        
        # Only log if it's a Samsol product
        if is_samsol:
            rec = {
                'Date': date_val,
                'Month': date_val.strftime("%Y-%m"),
                'Product Name': resolved_name,
                'Dia': dia_raw,
                'POF #': pof_num,
                'Ref. #': ref_num,
                'Produced Qty': total_good,
                'Wastage': total_wastage,
                'Catalog Customer': cat_cust,
                'Original Name': name_raw
            }
            prod_records.append(rec)
    else:
        # Unidentified
        unidentified_prod.append({
            'Source': 'Production',
            'Date': date_val,
            'POF #': pof_num,
            'Ref. #': ref_num,
            'Product Name': name_raw,
            'Dia/Volume': dia_raw,
            'Qty': total_good,
            'Wastage/Repl. Qty': total_wastage,
            'Party Name (Act.)': 'N/A (No party in Production)'
        })

# 3. Parse Dispatch File
print("Parsing Dispatch...")
df_d = pd.read_excel(disp_path, header=None)
disp_records = []
current_prod = None
SKIP_PREFIXES = ('dispatch report', 'month :', 'no.')
SKIP_EXACT = {'end of file'}

for idx, row in df_d.iterrows():
    col0 = row[0]
    if isinstance(col0, str) and col0.strip():
        c0 = col0.strip()
        c0_lower = c0.lower()
        if c0_lower in SKIP_EXACT or any(c0_lower.startswith(p) for p in SKIP_PREFIXES) or 'grand total' in c0_lower:
            continue
        if pd.isna(row[1]):
            current_prod = c0
            continue
            
    if current_prod:
        try:
            int(float(col0)) # check if it's a record row
            date_val = pd.Timestamp(row[2])
            qty = get_float(row[7]) # Disp. Qty
            repl_qty = get_float(row[8]) # Repl. Qty
            party = str(row[12]).strip()
            pof = row[3]
            ref_num = row[1] # Pk Id / Dly Id
            dia_raw = str(row[5]).strip()
            
            # Resolve product
            resolved_name = update_dispatch.NAME_FIXES.get(current_prod, current_prod)
            resolved_pid = None
            resolved_cust = None
            
            name_upper = resolved_name.strip().upper()
            if name_upper in catalog_map:
                catalog_entry = catalog_map[name_upper]
                resolved_name = catalog_entry['name']
                resolved_pid = catalog_entry['pid']
                resolved_cust = catalog_entry['customer']
                if resolved_pid in update_production.PID_TO_CUSTOMER:
                    resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
            else:
                resolved_name = None
                
            if resolved_name:
                cat_info = catalog_map.get(resolved_name.upper())
                cat_cust = cat_info['customer'] if cat_info else resolved_cust
                
                is_samsol = cat_cust and 'samsol' in str(cat_cust).lower()
                
                if is_samsol:
                    rec = {
                        'Date': date_val,
                        'Month': date_val.strftime("%Y-%m"),
                        'Product Name': resolved_name,
                        'Dia': dia_raw,
                        'POF #': pof,
                        'Ref. #': ref_num,
                        'Dispatched Qty': int(qty),
                        'Replacement Qty': int(repl_qty) if pd.notna(repl_qty) else 0,
                        'Party Name (Act.)': party,
                        'Catalog Customer': cat_cust,
                        'Original Name': current_prod
                    }
                    disp_records.append(rec)
            else:
                # Unidentified product in dispatch
                unidentified_prod.append({
                    'Source': 'Dispatch',
                    'Date': date_val,
                    'POF #': pof,
                    'Ref. #': ref_num,
                    'Product Name': current_prod,
                    'Dia/Volume': dia_raw,
                    'Qty': int(qty) if pd.notna(qty) else 0,
                    'Wastage/Repl. Qty': int(repl_qty) if pd.notna(repl_qty) else 0,
                    'Party Name (Act.)': party if pd.notna(party) and party != 'nan' else 'N/A'
                })
        except (ValueError, TypeError):
            pass

df_p_samsol = pd.DataFrame(prod_records)
df_d_samsol = pd.DataFrame(disp_records)

# 4. Generate summary dataframe
p_month = df_p_samsol.groupby('Month')['Produced Qty'].sum().reset_index()
d_month = df_d_samsol.groupby('Month')['Dispatched Qty'].sum().reset_index()

sum_official = pd.merge(p_month, d_month, on='Month', how='outer').fillna(0)
sum_official['Produced Qty'] = sum_official['Produced Qty'].astype(int)
sum_official['Dispatched Qty'] = sum_official['Dispatched Qty'].astype(int)
sum_official['Difference'] = sum_official['Produced Qty'] - sum_official['Dispatched Qty']
sum_official = sum_official.sort_values('Month')

# 5. Product summary month-by-month
p_prod = df_p_samsol.groupby(['Product Name', 'Month'])['Produced Qty'].sum().reset_index()
d_prod = df_d_samsol.groupby(['Product Name', 'Month'])['Dispatched Qty'].sum().reset_index()

df_prod_sum = pd.merge(p_prod, d_prod, on=['Product Name', 'Month'], how='outer').fillna(0)
df_prod_sum['Produced Qty'] = df_prod_sum['Produced Qty'].astype(int)
df_prod_sum['Dispatched Qty'] = df_prod_sum['Dispatched Qty'].astype(int)
df_prod_sum['Difference'] = df_prod_sum['Produced Qty'] - df_prod_sum['Dispatched Qty']
df_prod_sum = df_prod_sum.sort_values(['Product Name', 'Month'])

# Add Catalog Customer to the Product Summary
df_prod_sum['Customer'] = df_prod_sum['Product Name'].apply(lambda x: catalog_map[x.upper()]['customer'] if x.upper() in catalog_map else "Unknown")
df_prod_sum = df_prod_sum[['Product Name', 'Customer', 'Month', 'Produced Qty', 'Dispatched Qty', 'Difference']]

# 6. Writing using openpyxl for beautiful formatting
wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styles
font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E79")
font_section = Font(name="Segoe UI", size=12, bold=True, color="2C3E50")
font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
font_data = Font(name="Segoe UI", size=10)
font_total = Font(name="Segoe UI", size=10, bold=True)
font_note = Font(name="Segoe UI", size=9, italic=True, color="7F8C8D")

fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
fill_zebra = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
fill_total = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
fill_alert = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

thin_border_side = Side(style='thin', color='D3D3D3')
border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
border_total = Border(top=Side(style='thin', color='1F4E79'), bottom=Side(style='double', color='1F4E79'))

# ----------------- Sheet 1: Monthly Summary -----------------
ws_sum = wb.create_sheet("Monthly Summary")
ws_sum.views.sheetView[0].showGridLines = True

# Title Block
ws_sum.cell(row=2, column=2, value="Samsol Production & Dispatch Report").font = font_title
ws_sum.cell(row=3, column=2, value="Period: Nov 2025 - Jul 2026 | Source: D:\\Alpha\\Tubex Records").font = font_note

# Table 1: Official Samsol
ws_sum.cell(row=5, column=2, value="Samsol Monthly Summary").font = font_section

headers = ["Month", "Produced Qty", "Dispatched Qty", "Net Difference"]
for col_idx, h in enumerate(headers, start=2):
    cell = ws_sum.cell(row=7, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_data

current_row = 8
for idx, row in sum_official.iterrows():
    fill_to_use = fill_zebra if current_row % 2 == 0 else PatternFill(fill_type=None)
    
    c_month = ws_sum.cell(row=current_row, column=2, value=row['Month'])
    c_month.font = font_data
    c_month.alignment = align_center
    c_month.border = border_data
    c_month.fill = fill_to_use
    
    c_prod = ws_sum.cell(row=current_row, column=3, value=row['Produced Qty'])
    c_prod.font = font_data
    c_prod.number_format = '#,##0'
    c_prod.alignment = align_right
    c_prod.border = border_data
    c_prod.fill = fill_to_use
    
    c_disp = ws_sum.cell(row=current_row, column=4, value=row['Dispatched Qty'])
    c_disp.font = font_data
    c_disp.number_format = '#,##0'
    c_disp.alignment = align_right
    c_disp.border = border_data
    c_disp.fill = fill_to_use
    
    c_diff = ws_sum.cell(row=current_row, column=5, value=f"=C{current_row}-D{current_row}")
    c_diff.font = font_data
    c_diff.number_format = '#,##0'
    c_diff.alignment = align_right
    c_diff.border = border_data
    c_diff.fill = fill_to_use
    
    current_row += 1
    
c_tot_lbl = ws_sum.cell(row=current_row, column=2, value="Grand Total")
c_tot_lbl.font = font_total
c_tot_lbl.alignment = align_center
c_tot_lbl.border = border_total
c_tot_lbl.fill = fill_total

c_tot_prod = ws_sum.cell(row=current_row, column=3, value=f"=SUM(C8:C{current_row-1})")
c_tot_prod.font = font_total
c_tot_prod.number_format = '#,##0'
c_tot_prod.alignment = align_right
c_tot_prod.border = border_total
c_tot_prod.fill = fill_total

c_tot_disp = ws_sum.cell(row=current_row, column=4, value=f"=SUM(D8:D{current_row-1})")
c_tot_disp.font = font_total
c_tot_disp.number_format = '#,##0'
c_tot_disp.alignment = align_right
c_tot_disp.border = border_total
c_tot_disp.fill = fill_total

c_tot_diff = ws_sum.cell(row=current_row, column=5, value=f"=SUM(E8:E{current_row-1})")
c_tot_diff.font = font_total
c_tot_diff.number_format = '#,##0'
c_tot_diff.alignment = align_right
c_tot_diff.border = border_total
c_tot_diff.fill = fill_total

ws_sum.cell(row=current_row+2, column=2, value="Note: This table strictly includes products matching Customer = 'Samsol International Private Limited' in Tubex_July26.xlsx.").font = font_note

# ----------------- Sheet 2: Product Summary -----------------
ws_prod_sum = wb.create_sheet("Product Monthly Breakdown")
ws_prod_sum.views.sheetView[0].showGridLines = True

ws_prod_sum.cell(row=2, column=2, value="Monthly Production & Dispatch Breakdown by Product").font = font_section

headers_ps = ["Product Name", "Customer (Catalog)", "Month", "Produced Qty", "Dispatched Qty", "Net Difference"]
for col_idx, h in enumerate(headers_ps, start=2):
    cell = ws_prod_sum.cell(row=4, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_data

row_idx_ps = 5
for idx, row in df_prod_sum.iterrows():
    fill_to_use = fill_zebra if row_idx_ps % 2 == 0 else PatternFill(fill_type=None)
    
    ws_prod_sum.cell(row=row_idx_ps, column=2, value=row['Product Name']).font = font_data
    ws_prod_sum.cell(row=row_idx_ps, column=2).alignment = align_left
    ws_prod_sum.cell(row=row_idx_ps, column=2).border = border_data
    ws_prod_sum.cell(row=row_idx_ps, column=2).fill = fill_to_use
    
    ws_prod_sum.cell(row=row_idx_ps, column=3, value=row['Customer']).font = font_data
    ws_prod_sum.cell(row=row_idx_ps, column=3).alignment = align_left
    ws_prod_sum.cell(row=row_idx_ps, column=3).border = border_data
    ws_prod_sum.cell(row=row_idx_ps, column=3).fill = fill_to_use
    
    ws_prod_sum.cell(row=row_idx_ps, column=4, value=row['Month']).font = font_data
    ws_prod_sum.cell(row=row_idx_ps, column=4).alignment = align_center
    ws_prod_sum.cell(row=row_idx_ps, column=4).border = border_data
    ws_prod_sum.cell(row=row_idx_ps, column=4).fill = fill_to_use
    
    c_prod = ws_prod_sum.cell(row=row_idx_ps, column=5, value=row['Produced Qty'])
    c_prod.font = font_data
    c_prod.number_format = '#,##0'
    c_prod.alignment = align_right
    c_prod.border = border_data
    c_prod.fill = fill_to_use
    
    c_disp = ws_prod_sum.cell(row=row_idx_ps, column=6, value=row['Dispatched Qty'])
    c_disp.font = font_data
    c_disp.number_format = '#,##0'
    c_disp.alignment = align_right
    c_disp.border = border_data
    c_disp.fill = fill_to_use
    
    c_diff = ws_prod_sum.cell(row=row_idx_ps, column=7, value=f"=E{row_idx_ps}-F{row_idx_ps}")
    c_diff.font = font_data
    c_diff.number_format = '#,##0'
    c_diff.alignment = align_right
    c_diff.border = border_data
    c_diff.fill = fill_to_use
    
    row_idx_ps += 1

# Total row
c_tot_lbl = ws_prod_sum.cell(row=row_idx_ps, column=2, value="Total")
c_tot_lbl.font = font_total
c_tot_lbl.alignment = align_left
c_tot_lbl.border = border_total
c_tot_lbl.fill = fill_total

ws_prod_sum.cell(row=row_idx_ps, column=3, value="").border = border_total
ws_prod_sum.cell(row=row_idx_ps, column=3).fill = fill_total
ws_prod_sum.cell(row=row_idx_ps, column=4, value="").border = border_total
ws_prod_sum.cell(row=row_idx_ps, column=4).fill = fill_total

c_tot_prod = ws_prod_sum.cell(row=row_idx_ps, column=5, value=f"=SUM(E5:E{row_idx_ps-1})")
c_tot_prod.font = font_total
c_tot_prod.number_format = '#,##0'
c_tot_prod.alignment = align_right
c_tot_prod.border = border_total
c_tot_prod.fill = fill_total

c_tot_disp = ws_prod_sum.cell(row=row_idx_ps, column=6, value=f"=SUM(F5:F{row_idx_ps-1})")
c_tot_disp.font = font_total
c_tot_disp.number_format = '#,##0'
c_tot_disp.alignment = align_right
c_tot_disp.border = border_total
c_tot_disp.fill = fill_total

c_tot_diff = ws_prod_sum.cell(row=row_idx_ps, column=7, value=f"=SUM(G5:G{row_idx_ps-1})")
c_tot_diff.font = font_total
c_tot_diff.number_format = '#,##0'
c_tot_diff.alignment = align_right
c_tot_diff.border = border_total
c_tot_diff.fill = fill_total


# ----------------- Sheet 3: Production Details -----------------
ws_p_det = wb.create_sheet("Samsol Production Details")
ws_p_det.views.sheetView[0].showGridLines = True

ws_p_det.cell(row=2, column=2, value="Samsol Production Log Details (Nov 2025 - Jul 2026)").font = font_section

headers_p_det = ["Date", "Month", "Product Name", "Dia (mm)", "POF #", "Ref. #", "Produced Qty", "Wastage", "Catalog Customer"]
for col_idx, h in enumerate(headers_p_det, start=2):
    cell = ws_p_det.cell(row=4, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_data

df_p_samsol_det = df_p_samsol.sort_values('Date')
row_idx_pd = 5
for idx, row in df_p_samsol_det.iterrows():
    fill_to_use = fill_zebra if row_idx_pd % 2 == 0 else PatternFill(fill_type=None)
    
    ws_p_det.cell(row=row_idx_pd, column=2, value=row['Date'].strftime("%Y-%m-%d")).font = font_data
    ws_p_det.cell(row=row_idx_pd, column=2).alignment = align_center
    ws_p_det.cell(row=row_idx_pd, column=2).border = border_data
    ws_p_det.cell(row=row_idx_pd, column=2).fill = fill_to_use
    
    ws_p_det.cell(row=row_idx_pd, column=3, value=row['Month']).font = font_data
    ws_p_det.cell(row=row_idx_pd, column=3).alignment = align_center
    ws_p_det.cell(row=row_idx_pd, column=3).border = border_data
    ws_p_det.cell(row=row_idx_pd, column=3).fill = fill_to_use
    
    ws_p_det.cell(row=row_idx_pd, column=4, value=row['Product Name']).font = font_data
    ws_p_det.cell(row=row_idx_pd, column=4).alignment = align_left
    ws_p_det.cell(row=row_idx_pd, column=4).border = border_data
    ws_p_det.cell(row=row_idx_pd, column=4).fill = fill_to_use
    
    ws_p_det.cell(row=row_idx_pd, column=5, value=row['Dia']).font = font_data
    ws_p_det.cell(row=row_idx_pd, column=5).alignment = align_center
    ws_p_det.cell(row=row_idx_pd, column=5).border = border_data
    ws_p_det.cell(row=row_idx_pd, column=5).fill = fill_to_use
    
    ws_p_det.cell(row=row_idx_pd, column=6, value=row['POF #']).font = font_data
    ws_p_det.cell(row=row_idx_pd, column=6).alignment = align_center
    ws_p_det.cell(row=row_idx_pd, column=6).border = border_data
    ws_p_det.cell(row=row_idx_pd, column=6).fill = fill_to_use
    
    ws_p_det.cell(row=row_idx_pd, column=7, value=row['Ref. #']).font = font_data
    ws_p_det.cell(row=row_idx_pd, column=7).alignment = align_center
    ws_p_det.cell(row=row_idx_pd, column=7).border = border_data
    ws_p_det.cell(row=row_idx_pd, column=7).fill = fill_to_use
    
    c_prod = ws_p_det.cell(row=row_idx_pd, column=8, value=row['Produced Qty'])
    c_prod.font = font_data
    c_prod.number_format = '#,##0'
    c_prod.alignment = align_right
    c_prod.border = border_data
    c_prod.fill = fill_to_use
    
    c_wast = ws_p_det.cell(row=row_idx_pd, column=9, value=row['Wastage'])
    c_wast.font = font_data
    c_wast.number_format = '#,##0'
    c_wast.alignment = align_right
    c_wast.border = border_data
    c_wast.fill = fill_to_use
    
    ws_p_det.cell(row=row_idx_pd, column=10, value=row['Catalog Customer']).font = font_data
    ws_p_det.cell(row=row_idx_pd, column=10).alignment = align_left
    ws_p_det.cell(row=row_idx_pd, column=10).border = border_data
    ws_p_det.cell(row=row_idx_pd, column=10).fill = fill_to_use
    
    row_idx_pd += 1

# Total row
ws_p_det.cell(row=row_idx_pd, column=2, value="Total").font = font_total
ws_p_det.cell(row=row_idx_pd, column=2).border = border_total
ws_p_det.cell(row=row_idx_pd, column=2).fill = fill_total

for c in range(3, 8):
    ws_p_det.cell(row=row_idx_pd, column=c, value="").border = border_total
    ws_p_det.cell(row=row_idx_pd, column=c).fill = fill_total

c_tot_prod = ws_p_det.cell(row=row_idx_pd, column=8, value=f"=SUM(H5:H{row_idx_pd-1})")
c_tot_prod.font = font_total
c_tot_prod.number_format = '#,##0'
c_tot_prod.alignment = align_right
c_tot_prod.border = border_total
c_tot_prod.fill = fill_total

c_tot_wast = ws_p_det.cell(row=row_idx_pd, column=9, value=f"=SUM(I5:I{row_idx_pd-1})")
c_tot_wast.font = font_total
c_tot_wast.number_format = '#,##0'
c_tot_wast.alignment = align_right
c_tot_wast.border = border_total
c_tot_wast.fill = fill_total

ws_p_det.cell(row=row_idx_pd, column=10, value="").border = border_total
ws_p_det.cell(row=row_idx_pd, column=10).fill = fill_total


# ----------------- Sheet 4: Dispatch Details -----------------
ws_d_det = wb.create_sheet("Samsol Dispatch Details")
ws_d_det.views.sheetView[0].showGridLines = True

ws_d_det.cell(row=2, column=2, value="Samsol Dispatch Log Details (Nov 2025 - Jul 2026)").font = font_section

headers_d_det = ["Date", "Month", "Product Name", "Dia (mm)", "POF #", "Ref. #", "Dispatched Qty", "Replacement Qty", "Party Name (Act.)", "Catalog Customer"]
for col_idx, h in enumerate(headers_d_det, start=2):
    cell = ws_d_det.cell(row=4, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_data

df_d_samsol_det = df_d_samsol.sort_values('Date')
row_idx_dd = 5
for idx, row in df_d_samsol_det.iterrows():
    fill_to_use = fill_zebra if row_idx_dd % 2 == 0 else PatternFill(fill_type=None)
    
    ws_d_det.cell(row=row_idx_dd, column=2, value=row['Date'].strftime("%Y-%m-%d")).font = font_data
    ws_d_det.cell(row=row_idx_dd, column=2).alignment = align_center
    ws_d_det.cell(row=row_idx_dd, column=2).border = border_data
    ws_d_det.cell(row=row_idx_dd, column=2).fill = fill_to_use
    
    ws_d_det.cell(row=row_idx_dd, column=3, value=row['Month']).font = font_data
    ws_d_det.cell(row=row_idx_dd, column=3).alignment = align_center
    ws_d_det.cell(row=row_idx_dd, column=3).border = border_data
    ws_d_det.cell(row=row_idx_dd, column=3).fill = fill_to_use
    
    ws_d_det.cell(row=row_idx_dd, column=4, value=row['Product Name']).font = font_data
    ws_d_det.cell(row=row_idx_dd, column=4).alignment = align_left
    ws_d_det.cell(row=row_idx_dd, column=4).border = border_data
    ws_d_det.cell(row=row_idx_dd, column=4).fill = fill_to_use
    
    ws_d_det.cell(row=row_idx_dd, column=5, value=row['Dia']).font = font_data
    ws_d_det.cell(row=row_idx_dd, column=5).alignment = align_center
    ws_d_det.cell(row=row_idx_dd, column=5).border = border_data
    ws_d_det.cell(row=row_idx_dd, column=5).fill = fill_to_use
    
    ws_d_det.cell(row=row_idx_dd, column=6, value=row['POF #']).font = font_data
    ws_d_det.cell(row=row_idx_dd, column=6).alignment = align_center
    ws_d_det.cell(row=row_idx_dd, column=6).border = border_data
    ws_d_det.cell(row=row_idx_dd, column=6).fill = fill_to_use
    
    ws_d_det.cell(row=row_idx_dd, column=7, value=row['Ref. #']).font = font_data
    ws_d_det.cell(row=row_idx_dd, column=7).alignment = align_center
    ws_d_det.cell(row=row_idx_dd, column=7).border = border_data
    ws_d_det.cell(row=row_idx_dd, column=7).fill = fill_to_use
    
    c_disp = ws_d_det.cell(row=row_idx_dd, column=8, value=row['Dispatched Qty'])
    c_disp.font = font_data
    c_disp.number_format = '#,##0'
    c_disp.alignment = align_right
    c_disp.border = border_data
    c_disp.fill = fill_to_use
    
    c_repl = ws_d_det.cell(row=row_idx_dd, column=9, value=row['Replacement Qty'])
    c_repl.font = font_data
    c_repl.number_format = '#,##0'
    c_repl.alignment = align_right
    c_repl.border = border_data
    c_repl.fill = fill_to_use
    
    ws_d_det.cell(row=row_idx_dd, column=10, value=row['Party Name (Act.)']).font = font_data
    ws_d_det.cell(row=row_idx_dd, column=10).alignment = align_left
    ws_d_det.cell(row=row_idx_dd, column=10).border = border_data
    ws_d_det.cell(row=row_idx_dd, column=10).fill = fill_to_use
    
    ws_d_det.cell(row=row_idx_dd, column=11, value=row['Catalog Customer']).font = font_data
    ws_d_det.cell(row=row_idx_dd, column=11).alignment = align_left
    ws_d_det.cell(row=row_idx_dd, column=11).border = border_data
    ws_d_det.cell(row=row_idx_dd, column=11).fill = fill_to_use
    
    row_idx_dd += 1

# Total row
ws_d_det.cell(row=row_idx_dd, column=2, value="Total").font = font_total
ws_d_det.cell(row=row_idx_dd, column=2).border = border_total
ws_d_det.cell(row=row_idx_dd, column=2).fill = fill_total

for c in range(3, 8):
    ws_d_det.cell(row=row_idx_dd, column=c, value="").border = border_total
    ws_d_det.cell(row=row_idx_dd, column=c).fill = fill_total

c_tot_disp = ws_d_det.cell(row=row_idx_dd, column=8, value=f"=SUM(H5:H{row_idx_dd-1})")
c_tot_disp.font = font_total
c_tot_disp.number_format = '#,##0'
c_tot_disp.alignment = align_right
c_tot_disp.border = border_total
c_tot_disp.fill = fill_total

c_tot_repl = ws_d_det.cell(row=row_idx_dd, column=9, value=f"=SUM(I5:I{row_idx_dd-1})")
c_tot_repl.font = font_total
c_tot_repl.number_format = '#,##0'
c_tot_repl.alignment = align_right
c_tot_repl.border = border_total
c_tot_repl.fill = fill_total

ws_d_det.cell(row=row_idx_dd, column=10, value="").border = border_total
ws_d_det.cell(row=row_idx_dd, column=10).fill = fill_total
ws_d_det.cell(row=row_idx_dd, column=11, value="").border = border_total
ws_d_det.cell(row=row_idx_dd, column=11).fill = fill_total


# ----------------- Sheet 5: Unidentified Products -----------------
ws_un = wb.create_sheet("Unidentified Products")
ws_un.views.sheetView[0].showGridLines = True

ws_un.cell(row=2, column=2, value="Products Lacking Product Catalog Mapping").font = font_section
ws_un.cell(row=3, column=2, value="Verify these names to check if any belong to Samsol. Highlighted in light red.").font = font_note

headers_un = ["Source", "Date", "POF #", "Ref. #", "Product Name", "Dia/Volume", "Qty", "Wastage/Repl. Qty", "Party Name (Act.)"]
for col_idx, h in enumerate(headers_un, start=2):
    cell = ws_un.cell(row=5, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_data

row_idx_un = 6
# Sort unidentified products by source and date
unidentified_prod = sorted(unidentified_prod, key=lambda x: (x['Source'], x['Date']))
for item in unidentified_prod:
    ws_un.cell(row=row_idx_un, column=2, value=item['Source']).font = font_data
    ws_un.cell(row=row_idx_un, column=2).alignment = align_center
    ws_un.cell(row=row_idx_un, column=2).border = border_data
    ws_un.cell(row=row_idx_un, column=2).fill = fill_alert
    
    ws_un.cell(row=row_idx_un, column=3, value=item['Date'].strftime("%Y-%m-%d")).font = font_data
    ws_un.cell(row=row_idx_un, column=3).alignment = align_center
    ws_un.cell(row=row_idx_un, column=3).border = border_data
    ws_un.cell(row=row_idx_un, column=3).fill = fill_alert
    
    ws_un.cell(row=row_idx_un, column=4, value=item['POF #']).font = font_data
    ws_un.cell(row=row_idx_un, column=4).alignment = align_center
    ws_un.cell(row=row_idx_un, column=4).border = border_data
    ws_un.cell(row=row_idx_un, column=4).fill = fill_alert
    
    ws_un.cell(row=row_idx_un, column=5, value=item['Ref. #']).font = font_data
    ws_un.cell(row=row_idx_un, column=5).alignment = align_center
    ws_un.cell(row=row_idx_un, column=5).border = border_data
    ws_un.cell(row=row_idx_un, column=5).fill = fill_alert
    
    ws_un.cell(row=row_idx_un, column=6, value=item['Product Name']).font = Font(name="Segoe UI", size=10, bold=True)
    ws_un.cell(row=row_idx_un, column=6).alignment = align_left
    ws_un.cell(row=row_idx_un, column=6).border = border_data
    ws_un.cell(row=row_idx_un, column=6).fill = fill_alert
    
    ws_un.cell(row=row_idx_un, column=7, value=item['Dia/Volume']).font = font_data
    ws_un.cell(row=row_idx_un, column=7).alignment = align_center
    ws_un.cell(row=row_idx_un, column=7).border = border_data
    ws_un.cell(row=row_idx_un, column=7).fill = fill_alert
    
    c_qty = ws_un.cell(row=row_idx_un, column=8, value=item['Qty'])
    c_qty.font = font_data
    c_qty.number_format = '#,##0'
    c_qty.alignment = align_right
    c_qty.border = border_data
    c_qty.fill = fill_alert
    
    c_wast = ws_un.cell(row=row_idx_un, column=9, value=item['Wastage/Repl. Qty'])
    c_wast.font = font_data
    c_wast.number_format = '#,##0'
    c_wast.alignment = align_right
    c_wast.border = border_data
    c_wast.fill = fill_alert
    
    ws_un.cell(row=row_idx_un, column=10, value=item['Party Name (Act.)']).font = font_data
    ws_un.cell(row=row_idx_un, column=10).alignment = align_left
    ws_un.cell(row=row_idx_un, column=10).border = border_data
    ws_un.cell(row=row_idx_un, column=10).fill = fill_alert
    
    row_idx_un += 1

# Auto-fit column widths
for ws in wb.worksheets:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row in [1, 2, 3] and cell.value and len(str(cell.value)) > 25:
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
                
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Save the workbook
print(f"Saving workbook to {output_path}...")
wb.save(output_path)
print("Saved successfully!")
