import openpyxl

samsol_path = "D:/Alpha/Tubex Records/Samsol_Production_and_Dispatch.xlsx"
tubex_path = "D:/Alpha/Tubex_July26.xlsx"

# 1. Inspect user changes in Samsol sheet
print("=== Inspecting Samsol Sheet Formatting ===")
wb_s = openpyxl.load_workbook(samsol_path)
ws_s = wb_s["Monthly Summary"]

print("Row 2, Col 2 value:", ws_s.cell(row=2, column=2).value)
cell_sample = ws_s.cell(row=7, column=2) # Header cell
print("Header cell font:", cell_sample.font.name, "size:", cell_sample.font.size, "bold:", cell_sample.font.bold, "color:", cell_sample.font.color.rgb if cell_sample.font.color else None)
print("Header cell fill:", cell_sample.fill.fill_type, "fgColor:", cell_sample.fill.fgColor.rgb if cell_sample.fill.fgColor else None)

cell_data = ws_s.cell(row=8, column=2) # Data cell
print("Data cell value:", cell_data.value)
print("Data cell font:", cell_data.font.name, "size:", cell_data.font.size, "color:", cell_data.font.color.rgb if cell_data.font.color else None)
print("Data cell fill:", cell_data.fill.fill_type, "fgColor:", cell_data.fill.fgColor.rgb if cell_data.fill.fgColor else None)
print("Data cell alignment:", cell_data.alignment.horizontal, cell_data.alignment.vertical)

# Let's dump the first 15 rows of Samsol Monthly Summary
print("\nSamsol Monthly Summary Grid:")
for r in range(1, 16):
    row_vals = []
    for c in range(1, 10):
        val = ws_s.cell(row=r, column=c).value
        row_vals.append(val if val is not None else "")
    if any(row_vals):
        print(f"Row {r:2d}: {row_vals}")

# 2. Inspect Tubex Dashboard
print("\n=== Inspecting Tubex Dashboard Sheet ===")
wb_t = openpyxl.load_workbook(tubex_path)
ws_t = wb_t["Tubex_Dashboard"]

# Let's inspect rows 1 to 20 of Tubex Dashboard
print("Tubex Dashboard Grid:")
for r in range(1, 25):
    row_vals = []
    for c in range(1, 15):
        val = ws_t.cell(row=r, column=c).value
        row_vals.append(val if val is not None else "")
    if any(row_vals):
        print(f"Row {r:2d}: {row_vals}")
