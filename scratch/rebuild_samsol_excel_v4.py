import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
from datetime import datetime
import sys

sys.path.append("D:/Alpha/Scripts")
import update_production
import update_dispatch

catalog_path = "D:/Alpha/Tubex_July26.xlsx"
prod_path = "D:/Alpha/Tubex Records/production nov to jul.xls"
pet_prod_path = "D:/Alpha/Tubex Records/Production report Jan-2026 till Date.xlsx"
pet_disp_path = "D:/Alpha/Tubex Records/Samsol PET Orders.xlsx"
disp_path = "D:/Alpha/Tubex Records/dispatch nov to jul.xls"
output_path = "D:/Alpha/Tubex Records/Samsol_Production_and_Dispatch.xlsx"

# Alignments
align_left = Alignment(horizontal="left", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# Helper function to convert nan/empty cells to float
def get_float(val):
    if pd.isna(val): return 0.0
    try: return float(val)
    except: return 0.0

# Date helper for PET orders
def clean_date_str(date_str):
    if pd.isna(date_str):
        return None
    d_str = str(date_str).strip()
    for suffix in ['th', 'rd', 'nd', 'st']:
        d_str = d_str.replace(suffix, "")
    try:
        return datetime.strptime(d_str, "%d %B %Y")
    except Exception as e:
        print(f"Error parsing date {date_str}: {e}")
        return None

# 1. Load the user's workbook to extract styling
print("Loading user workbook to extract styling...")
wb = openpyxl.load_workbook(output_path)
ws_style = wb["Monthly Summary"]

# Extract styles from A1 (Title), A2 (Header), A3 (Data), and A12 (Total)
title_font = copy(ws_style["A1"].font)
title_align = copy(ws_style["A1"].alignment)

header_font = copy(ws_style["A2"].font)
header_fill = copy(ws_style["A2"].fill)
header_border = copy(ws_style["A2"].border)
header_align = copy(ws_style["A2"].alignment)

data_font = copy(ws_style["A3"].font)
data_fill = copy(ws_style["A3"].fill)
data_border = copy(ws_style["A3"].border)
data_align = copy(ws_style["A3"].alignment)

total_font = copy(ws_style["A12"].font)
total_fill = copy(ws_style["A12"].fill)
total_border = copy(ws_style["A12"].border)
total_align = copy(ws_style["A12"].alignment)

fill_alert = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

# Helper function to apply styles to a cell
def style_cell(cell, style_type="data", num_format=None):
    if style_type == "title":
        cell.font = title_font
        cell.alignment = title_align
    elif style_type == "header":
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align
    elif style_type == "data":
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
    elif style_type == "total":
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        cell.alignment = total_align
        
    if num_format:
        cell.number_format = num_format


# 2. Parse master catalog
print("Parsing master catalog...")
xls_cat = pd.ExcelFile(catalog_path)
df_cat = pd.read_excel(xls_cat, sheet_name="Product_Catalog", skiprows=1)

catalog_map = {}
samsol_tubes = []
samsol_pet = []

for idx, row in df_cat.iterrows():
    name = str(row['Product Name']).strip()
    pid = row['Product ID']
    cust = str(row['Customer']).strip()
    dia = row['Dia (mm)']
    if pd.notna(pid):
        pid = int(pid)
    catalog_map[name.upper()] = {'pid': pid, 'customer': cust, 'name': name, 'dia': dia}
    
    if 'samsol' in cust.lower() or 'abid' in cust.lower():
        prod_entry = {
            'name': name,
            'pid': pid,
            'customer': cust,
            'dia': dia
        }
        if pd.notna(pid) and pid >= 8000:
            samsol_pet.append(prod_entry)
        else:
            samsol_tubes.append(prod_entry)

samsol_tubes = sorted(samsol_tubes, key=lambda x: x['name'])
samsol_pet = sorted(samsol_pet, key=lambda x: x['name'])


# 3. Parse Tube production (nov to jul.xls)
print("Parsing Tube production raw data...")
df_p = pd.read_excel(prod_path, header=None)
prod_records = []
unidentified_prod = []

for idx, row in df_p.iterrows():
    row_val = row.dropna().tolist()
    if not row_val:
        continue
    dt = row[1]
    if pd.isna(dt) or not (isinstance(dt, pd.Timestamp) or hasattr(dt, 'date')):
        continue
    
    ref_num = row[0]
    date_val = pd.Timestamp(dt)
    pof_num = row[5]
    name_raw = str(row[6]).strip()
    dia_raw = str(row[8]).strip()
    
    normal_good = get_float(row[9])
    ot_good = get_float(row[12])
    normal_wastage = get_float(row[10])
    ot_wastage = get_float(row[13])
    
    total_good = int(normal_good + ot_good)
    total_wastage = int(normal_wastage + ot_wastage)
    
    name_key = name_raw.lower().strip()
    match = update_production.ALIASES.get((name_key, dia_raw))
    if not match:
        match = update_production.ALIASES.get((name_raw.lower().rstrip(), dia_raw))
        
    resolved_name = None
    resolved_pid = None
    resolved_cust = None
    
    if match:
        resolved_name, resolved_pid = match
        if resolved_pid in update_production.PID_TO_CUSTOMER:
            resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
        elif resolved_name.upper() in catalog_map:
            resolved_cust = catalog_map[resolved_name.upper()]['customer']
    else:
        name_upper = name_raw.upper()
        if name_upper in catalog_map:
            resolved_name = catalog_map[name_upper]['name']
            resolved_pid = catalog_map[name_upper]['pid']
            resolved_cust = catalog_map[name_upper]['customer']
            if resolved_pid in update_production.PID_TO_CUSTOMER:
                resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
                
    if resolved_name:
        cat_info = catalog_map.get(resolved_name.upper())
        cat_cust = cat_info['customer'] if cat_info else resolved_cust
        
        is_samsol = cat_cust and ('samsol' in str(cat_cust).lower() or 'abid' in str(cat_cust).lower())
        
        if is_samsol:
            prod_records.append({
                'Date': date_val.to_pydatetime(),
                'Month': datetime(date_val.year, date_val.month, 1),
                'Product Name': resolved_name,
                'Prod ID': resolved_pid,
                'Dia': dia_raw,
                'POF #': pof_num,
                'Ref. #': ref_num,
                'Produced Qty': total_good,
                'Wastage': total_wastage,
                'Catalog Customer': cat_cust,
                'Original Name': name_raw
            })
    else:
        unidentified_prod.append({
            'Source': 'Production',
            'Date': date_val.to_pydatetime(),
            'POF #': pof_num,
            'Ref. #': ref_num,
            'Product Name': name_raw,
            'Dia/Volume': dia_raw,
            'Qty': total_good,
            'Wastage/Repl. Qty': total_wastage,
            'Party Name (Act.)': 'N/A (No party in Production)'
        })


# 4. Parse PET production (Production report Jan-2026 till Date.xlsx)
print("Parsing PET production raw data...")
df_pet_p = pd.read_excel(pet_prod_path, sheet_name="Production Day wise", skiprows=2)

def map_pet_product(name_raw, volume_raw):
    n_lower = str(name_raw).lower().strip()
    v_lower = str(volume_raw).lower().strip()
    v_std = v_lower
    if '120' in v_lower: v_std = "120 ml"
    elif '200' in v_lower: v_std = "200 ml"
    elif '130' in v_lower: v_std = "130 ml"
    elif '150' in v_lower: v_std = "150 ml"
    elif '300' in v_lower: v_std = "300 ml"
    
    if "yellow small" in n_lower or (("yellow bottle" in n_lower or "yellow small" in n_lower) and "120" in v_std):
        return "PET BOTTLE SMALL (120ML) YELLOW", 8005
    elif "yellow large" in n_lower or (("yellow bottle" in n_lower or "yellow large" in n_lower) and "200" in v_std):
        return "PET BOTTLE LARGE (200 ML) YELLOW", 8006
    elif "white bottle" in n_lower and "200" in v_std:
        return "PET BOTTLE LARGE 200ML WHITE", 8007
    elif "black bottle" in n_lower and "200" in v_std:
        return "BLACK BOTTLE 200ML", 8008
    elif ("samsol black bottle" in n_lower or "black bottle" in n_lower) and "120" in v_std:
        return "BLACK SMALL BOTTLE 120ML", 8011
    elif "white bottle" in n_lower and "120" in v_std:
        return "WHITE SMALL BOTTLE 120ML", 8012
    elif "trp bottle" in n_lower and "200" in v_std:
        return "BT-200ML MUSTARD OIL (TRANSPARENT)", 8014
    else:
        name_upper = n_lower.upper()
        if name_upper in catalog_map:
            return catalog_map[name_upper]['name'], catalog_map[name_upper]['pid']
        return None, None

for idx, row in df_pet_p.iterrows():
    machine_raw = str(row['Machines']).strip()
    if machine_raw != "PF Machine":
        continue
    
    cust_raw = str(row['Customer']).strip()
    is_samsol_cust = 'samsol' in cust_raw.lower() or 'abid' in cust_raw.lower()
    if not is_samsol_cust:
        continue
        
    dt = row['Date']
    if pd.isna(dt):
        continue
    date_val = pd.Timestamp(dt)
    
    name_raw = str(row['Product Name']).strip()
    vol_raw = str(row['Dia(mm)/Volume']).strip()
    
    good_qty = int(get_float(row['Good Production']))
    wastage = int(get_float(row['Wastage']))
    
    resolved_name, resolved_pid = map_pet_product(name_raw, vol_raw)
    
    if resolved_name:
        cat_info = catalog_map.get(resolved_name.upper())
        cat_cust = cat_info['customer'] if cat_info else cust_raw
        
        prod_records.append({
            'Date': date_val.to_pydatetime(),
            'Month': datetime(date_val.year, date_val.month, 1),
            'Product Name': resolved_name,
            'Prod ID': resolved_pid,
            'Dia': vol_raw,
            'POF #': 'N/A',
            'Ref. #': 'N/A',
            'Produced Qty': good_qty,
            'Wastage': wastage,
            'Catalog Customer': cat_cust,
            'Original Name': name_raw
        })
    else:
        unidentified_prod.append({
            'Source': 'Production (PET)',
            'Date': date_val.to_pydatetime(),
            'POF #': 'N/A',
            'Ref. #': 'N/A',
            'Product Name': name_raw,
            'Dia/Volume': vol_raw,
            'Qty': good_qty,
            'Wastage/Repl. Qty': wastage,
            'Party Name (Act.)': 'N/A (No party in Production)'
        })


# 5. Parse Tube dispatches (dispatch nov to jul.xls)
print("Parsing Tube dispatch raw data...")
df_d = pd.read_excel(disp_path, header=None)
disp_records = []
current_prod = None
SKIP_PREFIXES = ('dispatch report', 'month :', 'no.')
SKIP_EXACT = {'end of file'}

for idx, row in df_d.iterrows():
    col0 = row[0]
    if isinstance(col0, str) and col0.strip():
        c0 = col0.strip()
        c0_lower = c0.lower()
        if c0_lower in SKIP_EXACT or any(c0_lower.startswith(p) for p in SKIP_PREFIXES) or 'grand total' in c0_lower:
            continue
        if pd.isna(row[1]):
            current_prod = c0
            continue
            
    if current_prod:
        try:
            int(float(col0))
            date_val = pd.Timestamp(row[2])
            qty = get_float(row[7])
            repl_qty = get_float(row[8])
            party = str(row[12]).strip()
            pof = row[3]
            ref_num = row[1]
            dia_raw = str(row[5]).strip()
            
            resolved_name = update_dispatch.NAME_FIXES.get(current_prod, current_prod)
            resolved_pid = None
            resolved_cust = None
            
            name_upper = resolved_name.strip().upper()
            if name_upper in catalog_map:
                catalog_entry = catalog_map[name_upper]
                resolved_name = catalog_entry['name']
                resolved_pid = catalog_entry['pid']
                resolved_cust = catalog_entry['customer']
                if resolved_pid in update_production.PID_TO_CUSTOMER:
                    resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
            else:
                resolved_name = None
                
            if resolved_name:
                cat_info = catalog_map.get(resolved_name.upper())
                cat_cust = cat_info['customer'] if cat_info else resolved_cust
                is_samsol = cat_cust and ('samsol' in str(cat_cust).lower() or 'abid' in str(cat_cust).lower())
                
                if is_samsol:
                    disp_records.append({
                        'Date': date_val.to_pydatetime(),
                        'Month': datetime(date_val.year, date_val.month, 1),
                        'Product Name': resolved_name,
                        'Prod ID': resolved_pid,
                        'Dia': dia_raw,
                        'POF #': pof,
                        'Ref. #': ref_num,
                        'Dispatched Qty': int(qty),
                        'Replacement Qty': int(repl_qty) if pd.notna(repl_qty) else 0,
                        'Party Name (Act.)': party,
                        'Catalog Customer': cat_cust,
                        'Original Name': current_prod
                    })
            else:
                unidentified_prod.append({
                    'Source': 'Dispatch',
                    'Date': date_val.to_pydatetime(),
                    'POF #': pof,
                    'Ref. #': ref_num,
                    'Product Name': current_prod,
                    'Dia/Volume': dia_raw,
                    'Qty': int(qty) if pd.notna(qty) else 0,
                    'Wastage/Repl. Qty': int(repl_qty) if pd.notna(repl_qty) else 0,
                    'Party Name (Act.)': party if pd.notna(party) and party != 'nan' else 'N/A'
                })
        except (ValueError, TypeError):
            pass


# 6. Parse PET dispatches (Samsol PET Orders.xlsx)
print("Parsing PET dispatch raw data...")
df_pet_d = pd.read_excel(pet_disp_path)

def map_order_product(name_raw):
    n_clean = str(name_raw).strip()
    n_upper = n_clean.upper()
    if n_upper in catalog_map:
        return catalog_map[n_upper]['name'], catalog_map[n_upper]['pid']
    
    if "BT-120 ML YELLOW" in n_upper or "BT-120ML YELLOW" in n_upper:
        return "BT-120 ML YELLOW", 8021
    elif "BT-200 ML YELLOW" in n_upper or "BT-200ML YELLOW" in n_upper:
        return "BT-200 ML YELLOW", 8022
    elif "BT-130 ML TRANSPARENT" in n_upper:
        return "BT-130 ML TRANSPARENT", 8023
    elif "BT-300 ML TRANSPARENT" in n_upper:
        return "BT-300 ML TRANSPARENT", 8017
    else:
        return n_clean, None

for idx, row in df_pet_d.iterrows():
    name_raw = row['Product Name']
    if pd.isna(name_raw) or str(name_raw).strip().upper() == 'TOTAL':
        continue
    
    disp_qty = row['Dispatched']
    if pd.isna(disp_qty) or disp_qty <= 0:
        continue
        
    cust_raw = str(row['Customer']).strip()
    dt_raw = row['Date']
    date_val = clean_date_str(dt_raw)
    if date_val is None:
        continue
        
    resolved_name, resolved_pid = map_order_product(name_raw)
    
    if resolved_name and resolved_pid:
        cat_info = catalog_map.get(resolved_name.upper())
        cat_cust = cat_info['customer'] if cat_info else cust_raw
        
        disp_records.append({
            'Date': date_val,
            'Month': datetime(date_val.year, date_val.month, 1),
            'Product Name': resolved_name,
            'Prod ID': resolved_pid,
            'Dia': cat_info['dia'] if cat_info else 'N/A',
            'POF #': 'N/A',
            'Ref. #': 'N/A',
            'Dispatched Qty': int(disp_qty),
            'Replacement Qty': 0,
            'Party Name (Act.)': cust_raw,
            'Catalog Customer': cat_cust,
            'Original Name': name_raw
        })
    else:
        unidentified_prod.append({
            'Source': 'Dispatch (PET)',
            'Date': date_val,
            'POF #': 'N/A',
            'Ref. #': 'N/A',
            'Product Name': name_raw,
            'Dia/Volume': 'N/A',
            'Qty': int(disp_qty),
            'Wastage/Repl. Qty': 0,
            'Party Name (Act.)': cust_raw
        })

df_p_samsol = pd.DataFrame(prod_records)
df_d_samsol = pd.DataFrame(disp_records)

# Unique sorted months list
months_sorted = sorted(list(set(df_p_samsol['Month'].dropna().tolist() + df_d_samsol['Month'].dropna().tolist())))


# 7. Create Sheet 1: Interactive Dashboard (Tubes & PET tables side-by-side or stacked)
print("Creating Interactive Dashboard...")
if "Interactive Dashboard" in wb.sheetnames:
    ws_dash = wb["Interactive Dashboard"]
    ws_dash.delete_rows(1, ws_dash.max_row+10)
else:
    ws_dash = wb.create_sheet("Interactive Dashboard", 0)

ws_dash.views.sheetView[0].showGridLines = True

# Title
ws_dash.cell(row=2, column=2, value="Samsol Interactive Dashboard").font = title_font
ws_dash.cell(row=2, column=2).alignment = title_align

# Instructions
ws_dash.cell(row=4, column=2, value="Toggle months 'Y' or 'N' in the selection panel on the right to dynamically filter both tables below:").font = Font(name="Segoe UI", size=9.5, italic=True, color="555555")

# Month Selection Panel Headers (Columns J and K)
cell_jh = ws_dash.cell(row=5, column=10, value="Month")
style_cell(cell_jh, "header")
cell_kh = ws_dash.cell(row=5, column=11, value="Select (Y/N)")
style_cell(cell_kh, "header")

# Month selectors
for idx, m_date in enumerate(months_sorted, start=6):
    c_m = ws_dash.cell(row=idx, column=10, value=m_date)
    style_cell(c_m, "data", "mmm-yy")
    
    c_sel = ws_dash.cell(row=idx, column=11, value="Y")
    style_cell(c_sel, "data")
    c_sel.font = Font(name="Segoe UI", size=10, bold=True)

# Y/N dropdown validation
dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
ws_dash.add_data_validation(dv_yn)
for idx in range(6, 6 + len(months_sorted)):
    dv_yn.add(ws_dash.cell(row=idx, column=11))

# TABLE 1: Tubes Dashboard
ws_dash.cell(row=5, column=2, value="1. Samsol Tubes Dashboard").font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
headers_dash = ["Customer Name", "Product Name", "Dia", "Prod ID", "Produced Qty", "Dispatched Qty", "Net Difference"]
for col_idx, h in enumerate(headers_dash, start=2):
    cell = ws_dash.cell(row=7, column=col_idx, value=h)
    style_cell(cell, "header")

row_idx = 8
for prod in samsol_tubes:
    c_cust = ws_dash.cell(row=row_idx, column=2, value=prod['customer'])
    style_cell(c_cust, "data")
    c_cust.alignment = align_left
    
    c_name = ws_dash.cell(row=row_idx, column=3, value=prod['name'])
    style_cell(c_name, "data")
    c_name.alignment = align_left
    
    c_dia = ws_dash.cell(row=row_idx, column=4, value=prod['dia'])
    style_cell(c_dia, "data")
    
    c_pid = ws_dash.cell(row=row_idx, column=5, value=prod['pid'])
    style_cell(c_pid, "data", "0")
    
    # Produced formula: Note that sum range is now Column I (Produced Qty) in Details
    prod_terms = []
    disp_terms = []
    for m_idx in range(6, 6 + len(months_sorted)):
        prod_terms.append(f"SUMIFS('Samsol Production Details'!$I:$I, 'Samsol Production Details'!$D:$D, $C{row_idx}, 'Samsol Production Details'!$C:$C, $J${m_idx}) * IF($K${m_idx}=\"Y\", 1, 0)")
        disp_terms.append(f"SUMIFS('Samsol Dispatch Details'!$I:$I, 'Samsol Dispatch Details'!$D:$D, $C{row_idx}, 'Samsol Dispatch Details'!$C:$C, $J${m_idx}) * IF($K${m_idx}=\"Y\", 1, 0)")
        
    ws_dash.cell(row=row_idx, column=6, value="=" + " + ".join(prod_terms))
    style_cell(ws_dash.cell(row=row_idx, column=6), "data", "#,##0")
    
    ws_dash.cell(row=row_idx, column=7, value="=" + " + ".join(disp_terms))
    style_cell(ws_dash.cell(row=row_idx, column=7), "data", "#,##0")
    
    ws_dash.cell(row=row_idx, column=8, value=f"=F{row_idx}-G{row_idx}")
    style_cell(ws_dash.cell(row=row_idx, column=8), "data", "#,##0")
    row_idx += 1

# Tubes Grand Total
ws_dash.cell(row=row_idx, column=2, value="Grand Total").alignment = align_left
for c in range(2, 6):
    style_cell(ws_dash.cell(row=row_idx, column=c), "total")
style_cell(ws_dash.cell(row=row_idx, column=6, value=f"=SUM(F8:F{row_idx-1})"), "total", "#,##0")
style_cell(ws_dash.cell(row=row_idx, column=7, value=f"=SUM(G8:G{row_idx-1})"), "total", "#,##0")
style_cell(ws_dash.cell(row=row_idx, column=8, value=f"=SUM(H8:H{row_idx-1})"), "total", "#,##0")


# TABLE 2: PET Dashboard
row_idx += 3
ws_dash.cell(row=row_idx-2, column=2, value="2. Samsol PET Dashboard").font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
for col_idx, h in enumerate(["Customer Name", "Product Name", "Volume", "Prod ID", "Produced Qty", "Dispatched Qty", "Net Difference"], start=2):
    cell = ws_dash.cell(row=row_idx-1, column=col_idx, value=h)
    style_cell(cell, "header")

start_pet_row = row_idx
for prod in samsol_pet:
    c_cust = ws_dash.cell(row=row_idx, column=2, value=prod['customer'])
    style_cell(c_cust, "data")
    c_cust.alignment = align_left
    
    c_name = ws_dash.cell(row=row_idx, column=3, value=prod['name'])
    style_cell(c_name, "data")
    c_name.alignment = align_left
    
    c_dia = ws_dash.cell(row=row_idx, column=4, value=prod['dia'])
    style_cell(c_dia, "data")
    
    c_pid = ws_dash.cell(row=row_idx, column=5, value=prod['pid'])
    style_cell(c_pid, "data", "0")
    
    prod_terms = []
    disp_terms = []
    for m_idx in range(6, 6 + len(months_sorted)):
        prod_terms.append(f"SUMIFS('Samsol Production Details'!$I:$I, 'Samsol Production Details'!$D:$D, $C{row_idx}, 'Samsol Production Details'!$C:$C, $J${m_idx}) * IF($K${m_idx}=\"Y\", 1, 0)")
        disp_terms.append(f"SUMIFS('Samsol Dispatch Details'!$I:$I, 'Samsol Dispatch Details'!$D:$D, $C{row_idx}, 'Samsol Dispatch Details'!$C:$C, $J${m_idx}) * IF($K${m_idx}=\"Y\", 1, 0)")
        
    ws_dash.cell(row=row_idx, column=6, value="=" + " + ".join(prod_terms))
    style_cell(ws_dash.cell(row=row_idx, column=6), "data", "#,##0")
    
    ws_dash.cell(row=row_idx, column=7, value="=" + " + ".join(disp_terms))
    style_cell(ws_dash.cell(row=row_idx, column=7), "data", "#,##0")
    
    ws_dash.cell(row=row_idx, column=8, value=f"=F{row_idx}-G{row_idx}")
    style_cell(ws_dash.cell(row=row_idx, column=8), "data", "#,##0")
    row_idx += 1

# PET Grand Total
ws_dash.cell(row=row_idx, column=2, value="Grand Total").alignment = align_left
for c in range(2, 6):
    style_cell(ws_dash.cell(row=row_idx, column=c), "total")
style_cell(ws_dash.cell(row=row_idx, column=6, value=f"=SUM(F{start_pet_row}:F{row_idx-1})"), "total", "#,##0")
style_cell(ws_dash.cell(row=row_idx, column=7, value=f"=SUM(G{start_pet_row}:G{row_idx-1})"), "total", "#,##0")
style_cell(ws_dash.cell(row=row_idx, column=8, value=f"=SUM(H{start_pet_row}:H{row_idx-1})"), "total", "#,##0")


# 8. Re-create Sheet 2: Monthly Summary with 2 sections (Tubes and PET)
print("Updating Monthly Summary sheet...")
if "Monthly Summary" in wb.sheetnames:
    ws_sum = wb["Monthly Summary"]
    ws_sum.delete_rows(1, ws_sum.max_row+10)
else:
    ws_sum = wb.create_sheet("Monthly Summary")

ws_sum.views.sheetView[0].showGridLines = True

# Title row
ws_sum.cell(row=1, column=1, value="Samsol Monthly Summaries").font = title_font
ws_sum.cell(row=1, column=1).alignment = title_align

# Section 1: Tubes
ws_sum.cell(row=3, column=1, value="1. Samsol Tubes Monthly Summary").font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
headers_sum = ["Month", "Produced Qty", "Dispatched Qty", "Net Difference"]
for col_idx, h in enumerate(headers_sum, start=1):
    cell = ws_sum.cell(row=5, column=col_idx, value=h)
    style_cell(cell, "header")

row_idx_sum = 6
for m_date in months_sorted:
    c_m = ws_sum.cell(row=row_idx_sum, column=1, value=m_date)
    style_cell(c_m, "data", "mmm-yy")
    
    # Tubes condition: Prod ID < 8000 (Column E in details sheets is Prod ID)
    c_p = ws_sum.cell(row=row_idx_sum, column=2, value=f"=SUMIFS('Samsol Production Details'!$I:$I, 'Samsol Production Details'!$C:$C, $A{row_idx_sum}, 'Samsol Production Details'!$E:$E, \"<8000\")")
    style_cell(c_p, "data", "#,##0")
    
    c_d = ws_sum.cell(row=row_idx_sum, column=3, value=f"=SUMIFS('Samsol Dispatch Details'!$I:$I, 'Samsol Dispatch Details'!$C:$C, $A{row_idx_sum}, 'Samsol Dispatch Details'!$E:$E, \"<8000\")")
    style_cell(c_d, "data", "#,##0")
    
    c_df = ws_sum.cell(row=row_idx_sum, column=4, value=f"=B{row_idx_sum}-C{row_idx_sum}")
    style_cell(c_df, "data", "#,##0")
    row_idx_sum += 1

# Section 1 Total
ws_sum.cell(row=row_idx_sum, column=1, value="Grand Total").alignment = align_center
style_cell(ws_sum.cell(row=row_idx_sum, column=1), "total")
style_cell(ws_sum.cell(row=row_idx_sum, column=2, value=f"=SUM(B6:B{row_idx_sum-1})"), "total", "#,##0")
style_cell(ws_sum.cell(row=row_idx_sum, column=3, value=f"=SUM(C6:C{row_idx_sum-1})"), "total", "#,##0")
style_cell(ws_sum.cell(row=row_idx_sum, column=4, value=f"=SUM(D6:D{row_idx_sum-1})"), "total", "#,##0")


# Section 2: PET
row_idx_sum += 3
ws_sum.cell(row=row_idx_sum-2, column=1, value="2. Samsol PET Monthly Summary").font = Font(name="Segoe UI", size=12, bold=True, color="1F4E79")
for col_idx, h in enumerate(headers_sum, start=1):
    cell = ws_sum.cell(row=row_idx_sum-1, column=col_idx, value=h)
    style_cell(cell, "header")

start_pet_sum_row = row_idx_sum
for m_date in months_sorted:
    c_m = ws_sum.cell(row=row_idx_sum, column=1, value=m_date)
    style_cell(c_m, "data", "mmm-yy")
    
    # PET condition: Prod ID >= 8000 (Column E in details sheets)
    c_p = ws_sum.cell(row=row_idx_sum, column=2, value=f"=SUMIFS('Samsol Production Details'!$I:$I, 'Samsol Production Details'!$C:$C, $A{row_idx_sum}, 'Samsol Production Details'!$E:$E, \">=8000\")")
    style_cell(c_p, "data", "#,##0")
    
    c_d = ws_sum.cell(row=row_idx_sum, column=3, value=f"=SUMIFS('Samsol Dispatch Details'!$I:$I, 'Samsol Dispatch Details'!$C:$C, $A{row_idx_sum}, 'Samsol Dispatch Details'!$E:$E, \">=8000\")")
    style_cell(c_d, "data", "#,##0")
    
    c_df = ws_sum.cell(row=row_idx_sum, column=4, value=f"=B{row_idx_sum}-C{row_idx_sum}")
    style_cell(c_df, "data", "#,##0")
    row_idx_sum += 1

# Section 2 Total
ws_sum.cell(row=row_idx_sum, column=1, value="Grand Total").alignment = align_center
style_cell(ws_sum.cell(row=row_idx_sum, column=1), "total")
style_cell(ws_sum.cell(row=row_idx_sum, column=2, value=f"=SUM(B{start_pet_sum_row}:B{row_idx_sum-1})"), "total", "#,##0")
style_cell(ws_sum.cell(row=row_idx_sum, column=3, value=f"=SUM(C{start_pet_sum_row}:C{row_idx_sum-1})"), "total", "#,##0")
style_cell(ws_sum.cell(row=row_idx_sum, column=4, value=f"=SUM(D{start_pet_sum_row}:D{row_idx_sum-1})"), "total", "#,##0")


# 9. Re-create Sheet 3: Product Monthly Breakdown
print("Updating Product Monthly Breakdown sheet...")
if "Product Monthly Breakdown" in wb.sheetnames:
    ws_pb = wb["Product Monthly Breakdown"]
    ws_pb.delete_rows(1, ws_pb.max_row+10)
else:
    ws_pb = wb.create_sheet("Product Monthly Breakdown")

ws_pb.views.sheetView[0].showGridLines = True
ws_pb.cell(row=2, column=2, value="Monthly Production & Dispatch Breakdown by Product").font = title_font

# Headers
headers_pb = ["Product Name", "Customer (Catalog)", "Month", "Produced Qty", "Dispatched Qty", "Net Difference"]
for col_idx, h in enumerate(headers_pb, start=2):
    cell = ws_pb.cell(row=4, column=col_idx, value=h)
    style_cell(cell, "header")

# Generate Product breakdown data
p_prod = df_p_samsol.groupby(['Product Name', 'Month'])['Produced Qty'].sum().reset_index()
d_prod = df_d_samsol.groupby(['Product Name', 'Month'])['Dispatched Qty'].sum().reset_index()
df_pb_data = pd.merge(p_prod, d_prod, on=['Product Name', 'Month'], how='outer').fillna(0)
df_pb_data['Produced Qty'] = df_pb_data['Produced Qty'].astype(int)
df_pb_data['Dispatched Qty'] = df_pb_data['Dispatched Qty'].astype(int)
df_pb_data = df_pb_data.sort_values(['Product Name', 'Month'])

row_idx_pb = 5
for idx, row in df_pb_data.iterrows():
    c_name = ws_pb.cell(row=row_idx_pb, column=2, value=row['Product Name'])
    style_cell(c_name, "data")
    c_name.alignment = align_left
    
    cust_val = catalog_map[row['Product Name'].upper()]['customer']
    c_cust = ws_pb.cell(row=row_idx_pb, column=3, value=cust_val)
    style_cell(c_cust, "data")
    c_cust.alignment = align_left
    
    c_month = ws_pb.cell(row=row_idx_pb, column=4, value=row['Month'])
    style_cell(c_month, "data", "mmm-yy")
    
    c_prod = ws_pb.cell(row=row_idx_pb, column=5, value=row['Produced Qty'])
    style_cell(c_prod, "data", "#,##0")
    
    c_disp = ws_pb.cell(row=row_idx_pb, column=6, value=row['Dispatched Qty'])
    style_cell(c_disp, "data", "#,##0")
    
    c_diff = ws_pb.cell(row=row_idx_pb, column=7, value=f"=E{row_idx_pb}-F{row_idx_pb}")
    style_cell(c_diff, "data", "#,##0")
    
    row_idx_pb += 1

# Total Row
ws_pb.cell(row=row_idx_pb, column=2, value="Total").alignment = align_left
for c in range(2, 5):
    style_cell(ws_pb.cell(row=row_idx_pb, column=c), "total")

style_cell(ws_pb.cell(row=row_idx_pb, column=5, value=f"=SUM(E5:E{row_idx_pb-1})"), "total", "#,##0")
style_cell(ws_pb.cell(row=row_idx_pb, column=6, value=f"=SUM(F5:F{row_idx_pb-1})"), "total", "#,##0")
style_cell(ws_pb.cell(row=row_idx_pb, column=7, value=f"=SUM(G5:G{row_idx_pb-1})"), "total", "#,##0")


# 10. Re-create Sheet 4: Samsol Production Details (with Prod ID column)
print("Updating Samsol Production Details sheet...")
if "Samsol Production Details" in wb.sheetnames:
    ws_pd = wb["Samsol Production Details"]
    ws_pd.delete_rows(1, ws_pd.max_row+10)
else:
    ws_pd = wb.create_sheet("Samsol Production Details")

ws_pd.views.sheetView[0].showGridLines = True
ws_pd.cell(row=2, column=2, value="Samsol Production Log Details (Nov 2025 - Jul 2026)").font = title_font

headers_pd = ["Date", "Month", "Product Name", "Prod ID", "Dia/Volume", "POF #", "Ref. #", "Produced Qty", "Wastage", "Catalog Customer"]
for col_idx, h in enumerate(headers_pd, start=2):
    cell = ws_pd.cell(row=4, column=col_idx, value=h)
    style_cell(cell, "header")

df_pd_data = df_p_samsol.sort_values('Date')
row_idx_pd = 5
for idx, row in df_pd_data.iterrows():
    ws_pd.cell(row=row_idx_pd, column=2, value=row['Date'])
    style_cell(ws_pd.cell(row=row_idx_pd, column=2), "data", "yyyy-mm-dd")
    
    ws_pd.cell(row=row_idx_pd, column=3, value=row['Month'])
    style_cell(ws_pd.cell(row=row_idx_pd, column=3), "data", "mmm-yy")
    
    c_name = ws_pd.cell(row=row_idx_pd, column=4, value=row['Product Name'])
    style_cell(c_name, "data")
    c_name.alignment = align_left
    
    # Prod ID
    ws_pd.cell(row=row_idx_pd, column=5, value=row['Prod ID'])
    style_cell(ws_pd.cell(row=row_idx_pd, column=5), "data", "0")
    
    ws_pd.cell(row=row_idx_pd, column=6, value=row['Dia'])
    style_cell(ws_pd.cell(row=row_idx_pd, column=6), "data")
    
    ws_pd.cell(row=row_idx_pd, column=7, value=row['POF #'])
    style_cell(ws_pd.cell(row=row_idx_pd, column=7), "data", "@")
    
    ws_pd.cell(row=row_idx_pd, column=8, value=row['Ref. #'])
    style_cell(ws_pd.cell(row=row_idx_pd, column=8), "data", "@")
    
    c_prod = ws_pd.cell(row=row_idx_pd, column=9, value=row['Produced Qty'])
    style_cell(c_prod, "data", "#,##0")
    
    c_wast = ws_pd.cell(row=row_idx_pd, column=10, value=row['Wastage'])
    style_cell(c_wast, "data", "#,##0")
    
    c_cust = ws_pd.cell(row=row_idx_pd, column=11, value=row['Catalog Customer'])
    style_cell(c_cust, "data")
    c_cust.alignment = align_left
    
    row_idx_pd += 1

# Total row
ws_pd.cell(row=row_idx_pd, column=2, value="Total").alignment = align_left
for c in range(2, 9):
    style_cell(ws_pd.cell(row=row_idx_pd, column=c), "total")
style_cell(ws_pd.cell(row=row_idx_pd, column=9, value=f"=SUM(I5:I{row_idx_pd-1})"), "total", "#,##0")
style_cell(ws_pd.cell(row=row_idx_pd, column=10, value=f"=SUM(J5:J{row_idx_pd-1})"), "total", "#,##0")
style_cell(ws_pd.cell(row=row_idx_pd, column=11), "total")


# 11. Re-create Sheet 5: Samsol Dispatch Details (with Prod ID column)
print("Updating Samsol Dispatch Details sheet...")
if "Samsol Dispatch Details" in wb.sheetnames:
    ws_dd = wb["Samsol Dispatch Details"]
    ws_dd.delete_rows(1, ws_dd.max_row+10)
else:
    ws_dd = wb.create_sheet("Samsol Dispatch Details")

ws_dd.views.sheetView[0].showGridLines = True
ws_dd.cell(row=2, column=2, value="Samsol Dispatch Log Details (Nov 2025 - Jul 2026)").font = title_font

headers_dd = ["Date", "Month", "Product Name", "Prod ID", "Dia/Volume", "POF #", "Ref. #", "Dispatched Qty", "Replacement Qty", "Party Name (Act.)", "Catalog Customer"]
for col_idx, h in enumerate(headers_dd, start=2):
    cell = ws_dd.cell(row=4, column=col_idx, value=h)
    style_cell(cell, "header")

df_dd_data = df_d_samsol.sort_values('Date')
row_idx_dd = 5
for idx, row in df_dd_data.iterrows():
    ws_dd.cell(row=row_idx_dd, column=2, value=row['Date'])
    style_cell(ws_dd.cell(row=row_idx_dd, column=2), "data", "yyyy-mm-dd")
    
    ws_dd.cell(row=row_idx_dd, column=3, value=row['Month'])
    style_cell(ws_dd.cell(row=row_idx_dd, column=3), "data", "mmm-yy")
    
    c_name = ws_dd.cell(row=row_idx_dd, column=4, value=row['Product Name'])
    style_cell(c_name, "data")
    c_name.alignment = align_left
    
    # Prod ID
    ws_dd.cell(row=row_idx_dd, column=5, value=row['Prod ID'])
    style_cell(ws_dd.cell(row=row_idx_dd, column=5), "data", "0")
    
    ws_dd.cell(row=row_idx_dd, column=6, value=row['Dia'])
    style_cell(ws_dd.cell(row=row_idx_dd, column=6), "data")
    
    ws_dd.cell(row=row_idx_dd, column=7, value=row['POF #'])
    style_cell(ws_dd.cell(row=row_idx_dd, column=7), "data", "@")
    
    ws_dd.cell(row=row_idx_dd, column=8, value=row['Ref. #'])
    style_cell(ws_dd.cell(row=row_idx_dd, column=8), "data", "@")
    
    c_disp = ws_dd.cell(row=row_idx_dd, column=9, value=row['Dispatched Qty'])
    style_cell(c_disp, "data", "#,##0")
    
    c_repl = ws_dd.cell(row=row_idx_dd, column=10, value=row['Replacement Qty'])
    style_cell(c_repl, "data", "#,##0")
    
    c_party = ws_dd.cell(row=row_idx_dd, column=11, value=row['Party Name (Act.)'])
    style_cell(c_party, "data")
    c_party.alignment = align_left
    
    c_cust = ws_dd.cell(row=row_idx_dd, column=12, value=row['Catalog Customer'])
    style_cell(c_cust, "data")
    c_cust.alignment = align_left
    
    row_idx_dd += 1

# Total row
ws_dd.cell(row=row_idx_dd, column=2, value="Total").alignment = align_left
for c in range(2, 9):
    style_cell(ws_dd.cell(row=row_idx_dd, column=c), "total")
style_cell(ws_dd.cell(row=row_idx_dd, column=9, value=f"=SUM(I5:I{row_idx_dd-1})"), "total", "#,##0")
style_cell(ws_dd.cell(row=row_idx_dd, column=10, value=f"=SUM(J5:J{row_idx_dd-1})"), "total", "#,##0")
style_cell(ws_dd.cell(row=row_idx_dd, column=11), "total")
style_cell(ws_dd.cell(row=row_idx_dd, column=12), "total")


# 12. Re-create Sheet 6: Unidentified Products
print("Updating Unidentified Products sheet...")
if "Unidentified Products" in wb.sheetnames:
    ws_un = wb["Unidentified Products"]
    ws_un.delete_rows(1, ws_un.max_row+10)
else:
    ws_un = wb.create_sheet("Unidentified Products")

ws_un.views.sheetView[0].showGridLines = True
ws_un.cell(row=2, column=2, value="Products Lacking Product Catalog Mapping").font = title_font

headers_un = ["Source", "Date", "POF #", "Ref. #", "Product Name", "Dia/Volume", "Qty", "Wastage/Repl. Qty", "Party Name (Act.)"]
for col_idx, h in enumerate(headers_un, start=2):
    cell = ws_un.cell(row=4, column=col_idx, value=h)
    style_cell(cell, "header")

row_idx_un = 5
unidentified_prod = sorted(unidentified_prod, key=lambda x: (x['Source'], x['Date']))
for item in unidentified_prod:
    for c in range(2, 11):
        cell = ws_un.cell(row=row_idx_un, column=c)
        style_cell(cell, "data")
        cell.fill = fill_alert
        
    ws_un.cell(row=row_idx_un, column=2, value=item['Source'])
    ws_un.cell(row=row_idx_un, column=3, value=item['Date']).number_format = 'yyyy-mm-dd'
    ws_un.cell(row=row_idx_un, column=4, value=item['POF #']).number_format = '0'
    ws_un.cell(row=row_idx_un, column=5, value=item['Ref. #']).number_format = '0'
    
    ws_un.cell(row=row_idx_un, column=6, value=item['Product Name']).font = Font(name="Segoe UI", size=10, bold=True)
    ws_un.cell(row=row_idx_un, column=6).alignment = align_left
    
    ws_un.cell(row=row_idx_un, column=7, value=item['Dia/Volume'])
    
    c_qty = ws_un.cell(row=row_idx_un, column=8, value=item['Qty'])
    c_qty.number_format = '#,##0'
    
    c_wast = ws_un.cell(row=row_idx_un, column=9, value=item['Wastage/Repl. Qty'])
    c_wast.number_format = '#,##0'
    
    c_party = ws_un.cell(row=row_idx_un, column=10, value=item['Party Name (Act.)'])
    c_party.alignment = align_left
    
    row_idx_un += 1


# Auto-fit column widths for all sheets
print("Auto-fitting column widths...")
for ws in wb.worksheets:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row in [1, 2, 3] and cell.value and len(str(cell.value)) > 25:
                continue
            if col_letter == 'Z':
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
                
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Save the workbook
print(f"Saving workbook to {output_path}...")
wb.save(output_path)
print("Saved successfully!")
