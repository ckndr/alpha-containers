import openpyxl

wb = openpyxl.load_workbook("d:\\Alpha\\Tubex_July26.xlsx")
ws = wb["BOM"]
print("BOM dimensions:", ws.dimensions)
with open("d:\\Alpha\\scratch\\inspected_bom_all.txt", "w", encoding="utf-8") as f:
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 12)]
        if any(val is not None for val in row_vals):
            f.write(f"Row {r:03d}: {row_vals}\n")
print("Done inspecting BOM!")
