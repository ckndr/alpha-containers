import openpyxl

wb = openpyxl.load_workbook("d:\\Alpha\\Tubex_July26.xlsx")
ws = wb["MRP"]
for r in range(115, 149):
    print(f"Row {r:03d}: A={ws.cell(row=r, column=1).value}, E={ws.cell(row=r, column=5).value}, F={ws.cell(row=r, column=6).value}, G={ws.cell(row=r, column=7).value}, H={ws.cell(row=r, column=8).value}")
