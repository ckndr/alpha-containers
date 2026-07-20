import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
from datetime import datetime
import sys

sys.path.append("D:/Alpha/Scripts")
import update_production
import update_dispatch

catalog_path = "D:/Alpha/Tubex_July26.xlsx"
prod_path = "D:/Alpha/Tubex Records/production nov to jul.xls"
disp_path = "D:/Alpha/Tubex Records/dispatch nov to jul.xls"
output_path = "D:/Alpha/Tubex Records/Samsol_Production_and_Dispatch.xlsx"

# Alignments
align_left = Alignment(horizontal="left", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# Helper function to convert nan/empty cells to float
def get_float(val):
    if pd.isna(val): return 0.0
    try: return float(val)
    except: return 0.0

# 1. Load the user's workbook to extract styling
print("Loading user workbook to extract styling...")
wb = openpyxl.load_workbook(output_path)
ws_style = wb["Monthly Summary"]

# Extract styles from A1 (Title), A2 (Header), A3 (Data), and A12 (Total)
title_font = copy(ws_style["A1"].font)
title_align = copy(ws_style["A1"].alignment)

header_font = copy(ws_style["A2"].font)
header_fill = copy(ws_style["A2"].fill)
header_border = copy(ws_style["A2"].border)
header_align = copy(ws_style["A2"].alignment)

data_font = copy(ws_style["A3"].font)
data_fill = copy(ws_style["A3"].fill)
data_border = copy(ws_style["A3"].border)
data_align = copy(ws_style["A3"].alignment)

total_font = copy(ws_style["A12"].font)
total_fill = copy(ws_style["A12"].fill)
total_border = copy(ws_style["A12"].border)
total_align = copy(ws_style["A12"].alignment)

# Background fill alert (for unidentified sheet)
fill_alert = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

# Helper function to apply styles to a cell
def style_cell(cell, style_type="data", num_format=None):
    if style_type == "title":
        cell.font = title_font
        cell.alignment = title_align
    elif style_type == "header":
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align
    elif style_type == "data":
        cell.font = data_font
        cell.fill = data_fill
        cell.border = data_border
        cell.alignment = data_align
    elif style_type == "total":
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        cell.alignment = total_align
        
    if num_format:
        cell.number_format = num_format

# Ensure grid lines and mmm-yy formatting on Monthly Summary Column A
print("Formatting Monthly Summary sheet...")
ws_style.views.sheetView[0].showGridLines = True
style_cell(ws_style["A1"], "title")
# Headers (Row 2)
for col in range(1, 5):
    style_cell(ws_style.cell(row=2, column=col), "header")
# Data (Rows 3 to 11)
for row in range(3, 12):
    cell_month = ws_style.cell(row=row, column=1)
    if isinstance(cell_month.value, str):
        try:
            cell_month.value = datetime.strptime(cell_month.value, "%Y-%m-%d")
        except:
            pass
    style_cell(cell_month, "data", "mmm-yy")
    
    for col in range(2, 5):
        style_cell(ws_style.cell(row=row, column=col), "data", "#,##0")
# Totals (Row 12)
for col in range(1, 5):
    style_cell(ws_style.cell(row=12, column=col), "total", "#,##0" if col > 1 else None)


# 2. Parse raw data logs (Samsol products only)
print("Parsing raw data...")
xls_cat = pd.ExcelFile(catalog_path)
df_cat = pd.read_excel(xls_cat, sheet_name="Product_Catalog", skiprows=1)

catalog_map = {}
samsol_products_list = []
for idx, row in df_cat.iterrows():
    name = str(row['Product Name']).strip()
    pid = row['Product ID']
    cust = str(row['Customer']).strip()
    dia = row['Dia (mm)']
    if pd.notna(pid):
        pid = int(pid)
    catalog_map[name.upper()] = {'pid': pid, 'customer': cust, 'name': name, 'dia': dia}
    
    if 'samsol' in cust.lower():
        samsol_products_list.append({
            'name': name,
            'pid': pid,
            'customer': cust,
            'dia': dia
        })

samsol_products_list = sorted(samsol_products_list, key=lambda x: x['name'])

# Parse Production
df_p = pd.read_excel(prod_path, header=None)
prod_records = []
unidentified_prod = []

for idx, row in df_p.iterrows():
    row_val = row.dropna().tolist()
    if not row_val:
        continue
    dt = row[1]
    if pd.isna(dt) or not (isinstance(dt, pd.Timestamp) or hasattr(dt, 'date')):
        continue
    
    ref_num = row[0]
    date_val = pd.Timestamp(dt)
    pof_num = row[5]
    name_raw = str(row[6]).strip()
    dia_raw = str(row[8]).strip()
    
    normal_good = get_float(row[9])
    ot_good = get_float(row[12])
    normal_wastage = get_float(row[10])
    ot_wastage = get_float(row[13])
    
    total_good = int(normal_good + ot_good)
    total_wastage = int(normal_wastage + ot_wastage)
    
    name_key = name_raw.lower().strip()
    match = update_production.ALIASES.get((name_key, dia_raw))
    if match is None:
        match = update_production.ALIASES.get((name_raw.lower().rstrip(), dia_raw))
        
    resolved_name = None
    resolved_pid = None
    resolved_cust = None
    
    if match:
        resolved_name, resolved_pid = match
        if resolved_pid in update_production.PID_TO_CUSTOMER:
            resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
        elif resolved_name.upper() in catalog_map:
            resolved_cust = catalog_map[resolved_name.upper()]['customer']
    else:
        name_upper = name_raw.upper()
        if name_upper in catalog_map:
            resolved_name = catalog_map[name_upper]['name']
            resolved_pid = catalog_map[name_upper]['pid']
            resolved_cust = catalog_map[name_upper]['customer']
            if resolved_pid in update_production.PID_TO_CUSTOMER:
                resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
                
    if resolved_name:
        cat_info = catalog_map.get(resolved_name.upper())
        cat_cust = cat_info['customer'] if cat_info else resolved_cust
        
        is_samsol = cat_cust and 'samsol' in str(cat_cust).lower()
        
        if is_samsol:
            prod_records.append({
                'Date': date_val.to_pydatetime(),
                'Month': datetime(date_val.year, date_val.month, 1),
                'Product Name': resolved_name,
                'Dia': dia_raw,
                'POF #': pof_num,
                'Ref. #': ref_num,
                'Produced Qty': total_good,
                'Wastage': total_wastage,
                'Catalog Customer': cat_cust,
                'Original Name': name_raw
            })
    else:
        unidentified_prod.append({
            'Source': 'Production',
            'Date': date_val.to_pydatetime(),
            'POF #': pof_num,
            'Ref. #': ref_num,
            'Product Name': name_raw,
            'Dia/Volume': dia_raw,
            'Qty': total_good,
            'Wastage/Repl. Qty': total_wastage,
            'Party Name (Act.)': 'N/A (No party in Production)'
        })

# Parse Dispatch
df_d = pd.read_excel(disp_path, header=None)
disp_records = []
current_prod = None
SKIP_PREFIXES = ('dispatch report', 'month :', 'no.')
SKIP_EXACT = {'end of file'}

for idx, row in df_d.iterrows():
    col0 = row[0]
    if isinstance(col0, str) and col0.strip():
        c0 = col0.strip()
        c0_lower = c0.lower()
        if c0_lower in SKIP_EXACT or any(c0_lower.startswith(p) for p in SKIP_PREFIXES) or 'grand total' in c0_lower:
            continue
        if pd.isna(row[1]):
            current_prod = c0
            continue
            
    if current_prod:
        try:
            int(float(col0))
            date_val = pd.Timestamp(row[2])
            qty = get_float(row[7])
            repl_qty = get_float(row[8])
            party = str(row[12]).strip()
            pof = row[3]
            ref_num = row[1]
            dia_raw = str(row[5]).strip()
            
            resolved_name = update_dispatch.NAME_FIXES.get(current_prod, current_prod)
            resolved_pid = None
            resolved_cust = None
            
            name_upper = resolved_name.strip().upper()
            if name_upper in catalog_map:
                catalog_entry = catalog_map[name_upper]
                resolved_name = catalog_entry['name']
                resolved_pid = catalog_entry['pid']
                resolved_cust = catalog_entry['customer']
                if resolved_pid in update_production.PID_TO_CUSTOMER:
                    resolved_cust = update_production.PID_TO_CUSTOMER[resolved_pid]
            else:
                resolved_name = None
                
            if resolved_name:
                cat_info = catalog_map.get(resolved_name.upper())
                cat_cust = cat_info['customer'] if cat_info else resolved_cust
                is_samsol = cat_cust and 'samsol' in str(cat_cust).lower()
                
                if is_samsol:
                    disp_records.append({
                        'Date': date_val.to_pydatetime(),
                        'Month': datetime(date_val.year, date_val.month, 1),
                        'Product Name': resolved_name,
                        'Dia': dia_raw,
                        'POF #': pof,
                        'Ref. #': ref_num,
                        'Dispatched Qty': int(qty),
                        'Replacement Qty': int(repl_qty) if pd.notna(repl_qty) else 0,
                        'Party Name (Act.)': party,
                        'Catalog Customer': cat_cust,
                        'Original Name': current_prod
                    })
            else:
                unidentified_prod.append({
                    'Source': 'Dispatch',
                    'Date': date_val.to_pydatetime(),
                    'POF #': pof,
                    'Ref. #': ref_num,
                    'Product Name': current_prod,
                    'Dia/Volume': dia_raw,
                    'Qty': int(qty) if pd.notna(qty) else 0,
                    'Wastage/Repl. Qty': int(repl_qty) if pd.notna(repl_qty) else 0,
                    'Party Name (Act.)': party if pd.notna(party) and party != 'nan' else 'N/A'
                })
        except (ValueError, TypeError):
            pass

df_p_samsol = pd.DataFrame(prod_records)
df_d_samsol = pd.DataFrame(disp_records)

# Months sorted (Nov-25 to Jul-26)
months_sorted = sorted(list(set(df_p_samsol['Month'].dropna().tolist() + df_d_samsol['Month'].dropna().tolist())))


# 3. Create the first sheet: Interactive Dashboard
print("Creating Interactive Dashboard sheet...")
if "Interactive Dashboard" in wb.sheetnames:
    ws_dash = wb["Interactive Dashboard"]
    ws_dash.delete_rows(1, ws_dash.max_row+10)
else:
    ws_dash = wb.create_sheet("Interactive Dashboard", 0)

ws_dash.views.sheetView[0].showGridLines = True

# Title Row
ws_dash.cell(row=2, column=2, value="Samsol Interactive Dashboard").font = title_font
ws_dash.cell(row=2, column=2).alignment = title_align

# Spacer Column I width
ws_dash.column_dimensions['I'].width = 3

# Instruction row
ws_dash.cell(row=4, column=2, value="Toggle months 'Y' or 'N' in the selection panel on the right to dynamically filter the table:").font = Font(name="Segoe UI", size=9.5, italic=True, color="555555")

# Month Selection Panel Headers (Columns J and K)
cell_jh = ws_dash.cell(row=5, column=10, value="Month")
style_cell(cell_jh, "header")
cell_kh = ws_dash.cell(row=5, column=11, value="Select (Y/N)")
style_cell(cell_kh, "header")

# Write Month Selection List in J6:J14 and dropdowns in K6:K14
for idx, m_date in enumerate(months_sorted, start=6):
    c_m = ws_dash.cell(row=idx, column=10, value=m_date)
    style_cell(c_m, "data", "mmm-yy")
    
    c_sel = ws_dash.cell(row=idx, column=11, value="Y") # Default to all selected (Y)
    style_cell(c_sel, "data")
    c_sel.font = Font(name="Segoe UI", size=10, bold=True)

# Add dropdown validation (Y/N) for K6:K14
dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
ws_dash.add_data_validation(dv_yn)
for idx in range(6, 6 + len(months_sorted)):
    dv_yn.add(ws_dash.cell(row=idx, column=11))

# Write Dashboard Table Headers (Columns B to H)
headers_dash = ["Customer Name", "Product Name", "Dia", "Prod ID", "Produced Qty", "Dispatched Qty", "Net Difference"]
for col_idx, h in enumerate(headers_dash, start=2):
    cell = ws_dash.cell(row=5, column=col_idx, value=h)
    style_cell(cell, "header")

# Write Samsol Products with multi-month sum formulas
# Formula structure: SUMIFS(..., Month, J6) * IF(K6="Y",1,0) + SUMIFS(..., Month, J7) * IF(K7="Y",1,0) + ...
row_idx = 6
for prod in samsol_products_list:
    c_cust = ws_dash.cell(row=row_idx, column=2, value=prod['customer'])
    style_cell(c_cust, "data")
    c_cust.alignment = align_left
    
    c_name = ws_dash.cell(row=row_idx, column=3, value=prod['name'])
    style_cell(c_name, "data")
    c_name.alignment = align_left
    
    c_dia = ws_dash.cell(row=row_idx, column=4, value=prod['dia'])
    style_cell(c_dia, "data")
    
    c_pid = ws_dash.cell(row=row_idx, column=5, value=prod['pid'])
    style_cell(c_pid, "data", "0")
    
    # Build Production Formula summing up each month if toggled 'Y'
    prod_terms = []
    disp_terms = []
    for m_idx in range(6, 6 + len(months_sorted)):
        prod_terms.append(f"SUMIFS('Samsol Production Details'!$H:$H, 'Samsol Production Details'!$D:$D, $C{row_idx}, 'Samsol Production Details'!$C:$C, $J${m_idx}) * IF($K${m_idx}=\"Y\", 1, 0)")
        disp_terms.append(f"SUMIFS('Samsol Dispatch Details'!$H:$H, 'Samsol Dispatch Details'!$D:$D, $C{row_idx}, 'Samsol Dispatch Details'!$C:$C, $J${m_idx}) * IF($K${m_idx}=\"Y\", 1, 0)")
        
    formula_prod = "=" + " + ".join(prod_terms)
    formula_disp = "=" + " + ".join(disp_terms)
    
    c_prod = ws_dash.cell(row=row_idx, column=6, value=formula_prod)
    style_cell(c_prod, "data", "#,##0")
    
    c_disp = ws_dash.cell(row=row_idx, column=7, value=formula_disp)
    style_cell(c_disp, "data", "#,##0")
    
    # Difference formula
    c_diff = ws_dash.cell(row=row_idx, column=8, value=f"=F{row_idx}-G{row_idx}")
    style_cell(c_diff, "data", "#,##0")
    
    row_idx += 1

# Write Grand Total row for main dashboard table
ws_dash.cell(row=row_idx, column=2, value="Grand Total").alignment = align_left
for c in range(2, 6):
    style_cell(ws_dash.cell(row=row_idx, column=c), "total")

style_cell(ws_dash.cell(row=row_idx, column=6, value=f"=SUM(F6:F{row_idx-1})"), "total", "#,##0")
style_cell(ws_dash.cell(row=row_idx, column=7, value=f"=SUM(G6:G{row_idx-1})"), "total", "#,##0")
style_cell(ws_dash.cell(row=row_idx, column=8, value=f"=SUM(H6:H{row_idx-1})"), "total", "#,##0")


# 4. Re-create Sheet 3: Product Monthly Breakdown
print("Updating Product Monthly Breakdown sheet...")
if "Product Monthly Breakdown" in wb.sheetnames:
    ws_pb = wb["Product Monthly Breakdown"]
    ws_pb.delete_rows(1, ws_pb.max_row+10)
else:
    ws_pb = wb.create_sheet("Product Monthly Breakdown")

ws_pb.views.sheetView[0].showGridLines = True
ws_pb.cell(row=2, column=2, value="Monthly Production & Dispatch Breakdown by Product").font = title_font

# Headers
headers_pb = ["Product Name", "Customer (Catalog)", "Month", "Produced Qty", "Dispatched Qty", "Net Difference"]
for col_idx, h in enumerate(headers_pb, start=2):
    cell = ws_pb.cell(row=4, column=col_idx, value=h)
    style_cell(cell, "header")

# Generate Product breakdown data
p_prod = df_p_samsol.groupby(['Product Name', 'Month'])['Produced Qty'].sum().reset_index()
d_prod = df_d_samsol.groupby(['Product Name', 'Month'])['Dispatched Qty'].sum().reset_index()
df_pb_data = pd.merge(p_prod, d_prod, on=['Product Name', 'Month'], how='outer').fillna(0)
df_pb_data['Produced Qty'] = df_pb_data['Produced Qty'].astype(int)
df_pb_data['Dispatched Qty'] = df_pb_data['Dispatched Qty'].astype(int)
df_pb_data = df_pb_data.sort_values(['Product Name', 'Month'])

row_idx_pb = 5
for idx, row in df_pb_data.iterrows():
    c_name = ws_pb.cell(row=row_idx_pb, column=2, value=row['Product Name'])
    style_cell(c_name, "data")
    c_name.alignment = align_left
    
    cust_val = catalog_map[row['Product Name'].upper()]['customer']
    c_cust = ws_pb.cell(row=row_idx_pb, column=3, value=cust_val)
    style_cell(c_cust, "data")
    c_cust.alignment = align_left
    
    c_month = ws_pb.cell(row=row_idx_pb, column=4, value=row['Month'])
    style_cell(c_month, "data", "mmm-yy")
    
    c_prod = ws_pb.cell(row=row_idx_pb, column=5, value=row['Produced Qty'])
    style_cell(c_prod, "data", "#,##0")
    
    c_disp = ws_pb.cell(row=row_idx_pb, column=6, value=row['Dispatched Qty'])
    style_cell(c_disp, "data", "#,##0")
    
    c_diff = ws_pb.cell(row=row_idx_pb, column=7, value=f"=E{row_idx_pb}-F{row_idx_pb}")
    style_cell(c_diff, "data", "#,##0")
    
    row_idx_pb += 1

# Total Row
ws_pb.cell(row=row_idx_pb, column=2, value="Total").alignment = align_left
for c in range(2, 5):
    style_cell(ws_pb.cell(row=row_idx_pb, column=c), "total")

style_cell(ws_pb.cell(row=row_idx_pb, column=5, value=f"=SUM(E5:E{row_idx_pb-1})"), "total", "#,##0")
style_cell(ws_pb.cell(row=row_idx_pb, column=6, value=f"=SUM(F5:F{row_idx_pb-1})"), "total", "#,##0")
style_cell(ws_pb.cell(row=row_idx_pb, column=7, value=f"=SUM(G5:G{row_idx_pb-1})"), "total", "#,##0")


# 5. Re-create Sheet 4: Samsol Production Details
print("Updating Samsol Production Details sheet...")
if "Samsol Production Details" in wb.sheetnames:
    ws_pd = wb["Samsol Production Details"]
    ws_pd.delete_rows(1, ws_pd.max_row+10)
else:
    ws_pd = wb.create_sheet("Samsol Production Details")

ws_pd.views.sheetView[0].showGridLines = True
ws_pd.cell(row=2, column=2, value="Samsol Production Log Details (Nov 2025 - Jul 2026)").font = title_font

headers_pd = ["Date", "Month", "Product Name", "Dia (mm)", "POF #", "Ref. #", "Produced Qty", "Wastage", "Catalog Customer"]
for col_idx, h in enumerate(headers_pd, start=2):
    cell = ws_pd.cell(row=4, column=col_idx, value=h)
    style_cell(cell, "header")

df_pd_data = df_p_samsol.sort_values('Date')
row_idx_pd = 5
for idx, row in df_pd_data.iterrows():
    c_date = ws_pd.cell(row=row_idx_pd, column=2, value=row['Date'])
    style_cell(c_date, "data", "yyyy-mm-dd")
    
    c_month = ws_pd.cell(row=row_idx_pd, column=3, value=row['Month'])
    style_cell(c_month, "data", "mmm-yy")
    
    c_name = ws_pd.cell(row=row_idx_pd, column=4, value=row['Product Name'])
    style_cell(c_name, "data")
    c_name.alignment = align_left
    
    c_dia = ws_pd.cell(row=row_idx_pd, column=5, value=row['Dia'])
    style_cell(c_dia, "data")
    
    c_pof = ws_pd.cell(row=row_idx_pd, column=6, value=row['POF #'])
    style_cell(c_pof, "data", "0")
    
    c_ref = ws_pd.cell(row=row_idx_pd, column=7, value=row['Ref. #'])
    style_cell(c_ref, "data", "0")
    
    c_prod = ws_pd.cell(row=row_idx_pd, column=8, value=row['Produced Qty'])
    style_cell(c_prod, "data", "#,##0")
    
    c_wast = ws_pd.cell(row=row_idx_pd, column=9, value=row['Wastage'])
    style_cell(c_wast, "data", "#,##0")
    
    c_cust = ws_pd.cell(row=row_idx_pd, column=10, value=row['Catalog Customer'])
    style_cell(c_cust, "data")
    c_cust.alignment = align_left
    
    row_idx_pd += 1

# Total row
ws_pd.cell(row=row_idx_pd, column=2, value="Total").alignment = align_left
for c in range(2, 8):
    style_cell(ws_pd.cell(row=row_idx_pd, column=c), "total")

style_cell(ws_pd.cell(row=row_idx_pd, column=8, value=f"=SUM(H5:H{row_idx_pd-1})"), "total", "#,##0")
style_cell(ws_pd.cell(row=row_idx_pd, column=9, value=f"=SUM(I5:I{row_idx_pd-1})"), "total", "#,##0")
style_cell(ws_pd.cell(row=row_idx_pd, column=10), "total")


# 6. Re-create Sheet 5: Samsol Dispatch Details
print("Updating Samsol Dispatch Details sheet...")
if "Samsol Dispatch Details" in wb.sheetnames:
    ws_dd = wb["Samsol Dispatch Details"]
    ws_dd.delete_rows(1, ws_dd.max_row+10)
else:
    ws_dd = wb.create_sheet("Samsol Dispatch Details")

ws_dd.views.sheetView[0].showGridLines = True
ws_dd.cell(row=2, column=2, value="Samsol Dispatch Log Details (Nov 2025 - Jul 2026)").font = title_font

headers_dd = ["Date", "Month", "Product Name", "Dia (mm)", "POF #", "Ref. #", "Dispatched Qty", "Replacement Qty", "Party Name (Act.)", "Catalog Customer"]
for col_idx, h in enumerate(headers_dd, start=2):
    cell = ws_dd.cell(row=4, column=col_idx, value=h)
    style_cell(cell, "header")

df_dd_data = df_d_samsol.sort_values('Date')
row_idx_dd = 5
for idx, row in df_dd_data.iterrows():
    c_date = ws_dd.cell(row=row_idx_dd, column=2, value=row['Date'])
    style_cell(c_date, "data", "yyyy-mm-dd")
    
    c_month = ws_dd.cell(row=row_idx_dd, column=3, value=row['Month'])
    style_cell(c_month, "data", "mmm-yy")
    
    c_name = ws_dd.cell(row=row_idx_dd, column=4, value=row['Product Name'])
    style_cell(c_name, "data")
    c_name.alignment = align_left
    
    c_dia = ws_dd.cell(row=row_idx_dd, column=5, value=row['Dia'])
    style_cell(c_dia, "data")
    
    c_pof = ws_dd.cell(row=row_idx_dd, column=6, value=row['POF #'])
    style_cell(c_pof, "data", "0")
    
    c_ref = ws_dd.cell(row=row_idx_dd, column=7, value=row['Ref. #'])
    style_cell(c_ref, "data", "0")
    
    c_disp = ws_dd.cell(row=row_idx_dd, column=8, value=row['Dispatched Qty'])
    style_cell(c_disp, "data", "#,##0")
    
    c_repl = ws_dd.cell(row=row_idx_dd, column=9, value=row['Replacement Qty'])
    style_cell(c_repl, "data", "#,##0")
    
    c_party = ws_dd.cell(row=row_idx_dd, column=10, value=row['Party Name (Act.)'])
    style_cell(c_party, "data")
    c_party.alignment = align_left
    
    c_cust = ws_dd.cell(row=row_idx_dd, column=11, value=row['Catalog Customer'])
    style_cell(c_cust, "data")
    c_cust.alignment = align_left
    
    row_idx_dd += 1

# Total row
ws_dd.cell(row=row_idx_dd, column=2, value="Total").alignment = align_left
for c in range(2, 8):
    style_cell(ws_dd.cell(row=row_idx_dd, column=c), "total")

style_cell(ws_dd.cell(row=row_idx_dd, column=8, value=f"=SUM(H5:H{row_idx_dd-1})"), "total", "#,##0")
style_cell(ws_dd.cell(row=row_idx_dd, column=9, value=f"=SUM(I5:I{row_idx_dd-1})"), "total", "#,##0")
style_cell(ws_dd.cell(row=row_idx_dd, column=10), "total")
style_cell(ws_dd.cell(row=row_idx_dd, column=11), "total")


# 7. Re-create Sheet 6: Unidentified Products
print("Updating Unidentified Products sheet...")
if "Unidentified Products" in wb.sheetnames:
    ws_un = wb["Unidentified Products"]
    ws_un.delete_rows(1, ws_un.max_row+10)
else:
    ws_un = wb.create_sheet("Unidentified Products")

ws_un.views.sheetView[0].showGridLines = True
ws_un.cell(row=2, column=2, value="Products Lacking Product Catalog Mapping").font = title_font

headers_un = ["Source", "Date", "POF #", "Ref. #", "Product Name", "Dia/Volume", "Qty", "Wastage/Repl. Qty", "Party Name (Act.)"]
for col_idx, h in enumerate(headers_un, start=2):
    cell = ws_un.cell(row=4, column=col_idx, value=h)
    style_cell(cell, "header")

row_idx_un = 5
unidentified_prod = sorted(unidentified_prod, key=lambda x: (x['Source'], x['Date']))
for item in unidentified_prod:
    for c in range(2, 11):
        cell = ws_un.cell(row=row_idx_un, column=c)
        style_cell(cell, "data")
        cell.fill = fill_alert
        
    ws_un.cell(row=row_idx_un, column=2, value=item['Source'])
    ws_un.cell(row=row_idx_un, column=3, value=item['Date']).number_format = 'yyyy-mm-dd'
    ws_un.cell(row=row_idx_un, column=4, value=item['POF #']).number_format = '0'
    ws_un.cell(row=row_idx_un, column=5, value=item['Ref. #']).number_format = '0'
    
    ws_un.cell(row=row_idx_un, column=6, value=item['Product Name']).font = Font(name="Segoe UI", size=10, bold=True)
    ws_un.cell(row=row_idx_un, column=6).alignment = align_left
    
    ws_un.cell(row=row_idx_un, column=7, value=item['Dia/Volume'])
    
    c_qty = ws_un.cell(row=row_idx_un, column=8, value=item['Qty'])
    c_qty.number_format = '#,##0'
    
    c_wast = ws_un.cell(row=row_idx_un, column=9, value=item['Wastage/Repl. Qty'])
    c_wast.number_format = '#,##0'
    
    c_party = ws_un.cell(row=row_idx_un, column=10, value=item['Party Name (Act.)'])
    c_party.alignment = align_left
    
    row_idx_un += 1


# Auto-fit column widths for all sheets
print("Auto-fitting column widths...")
for ws in wb.worksheets:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            if cell.row in [1, 2, 3] and cell.value and len(str(cell.value)) > 25:
                continue
            if col_letter == 'Z':
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
                
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Save the workbook
print(f"Saving workbook to {output_path}...")
wb.save(output_path)
print("Saved successfully!")
