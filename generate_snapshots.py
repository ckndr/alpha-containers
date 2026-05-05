"""
Alpha Containers - Snapshot Generator v1
=========================================
PURPOSE:
  Reads the Inventory sheet from the latest AlphaContainers*.xlsx and
  generates two PNG images for the daily WhatsApp production group:

  Image 1: SLUGS INVENTORY  (slug rows only, includes WIP + Pieces columns)
  Image 2: INVENTORY         (Lacquer through Thinner, store columns only)

  Images are saved in the same folder as the script with today's date
  in the filename, e.g.:
    snapshot_slugs_29Apr.png
    snapshot_inventory_29Apr.png

USAGE:
  Double-click Generate_Snapshots.bat
  Or: python generate_snapshots.py

NO EXCEL CHANGES: this script is read-only. It never modifies any file.
"""

import os
import re
import glob
from datetime import date
from openpyxl import load_workbook

import warnings
warnings.filterwarnings("ignore", message=".*Data Validation.*")
warnings.filterwarnings("ignore", message=".*extension.*")

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    print("  ERROR: Pillow not installed. Run:")
    print("    pip install Pillow")
    raise


# -----------------------------------------------------------------------
# VISUAL DESIGN
# -----------------------------------------------------------------------
# Colors (RGB tuples)
C_NAVY        = (31,  56, 100)     # #1F3864 -- header banner
C_COL_HEADER  = (46,  84, 150)     # #2E5496 -- column headers
C_ROW_ALT     = (235, 243, 255)    # light blue alternating row
C_ROW_PLAIN   = (255, 255, 255)    # white row
C_TEXT_WHITE  = (255, 255, 255)
C_TEXT_DARK   = (30,  30,  30)
C_TEXT_BLUE   = (31,  56, 100)     # for category values
C_GRID        = (180, 200, 230)    # cell border
C_ZERO        = (0,   0,   0)      # normal zero
C_ZERO_HILITE = (0,   100, 180)    # coloured zero (matching screenshot)
C_PIECES_BG   = (20,  40,  80)     # dark header for Pieces column
C_DASH        = (120, 120, 120)    # "-" values


# -----------------------------------------------------------------------
# LAYOUT
# -----------------------------------------------------------------------
SCALE         = 2          # render at 2x then downsample for sharpness
FONT_SIZE     = 14 * SCALE
HEADER_SIZE   = 15 * SCALE
TITLE_SIZE    = 16 * SCALE
PAD_V         = 7  * SCALE   # vertical cell padding
PAD_H         = 8  * SCALE   # horizontal cell padding
ROW_H         = (FONT_SIZE + PAD_V * 2)
COL_HDR_H     = int(ROW_H * 1.6)
TITLE_H       = int(ROW_H * 1.7)


def load_font(size, bold=False):
    """Load a font, fallback to PIL default."""
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "C:/Windows/Fonts/segoeui.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    bold_candidates = [c for c in candidates if 'Bold' in c or 'bold' in c]
    reg_candidates  = [c for c in candidates if 'Bold' not in c and 'bold' not in c]
    search = bold_candidates + reg_candidates if bold else reg_candidates + bold_candidates
    for path in search:
        if os.path.exists(path):
            try:
                return ImageFont.truetype(path, size)
            except Exception:
                continue
    return ImageFont.load_default()


def text_w(draw, text, font):
    try:
        bb = draw.textbbox((0, 0), text, font=font)
        return bb[2] - bb[0]
    except Exception:
        return draw.textlength(text, font=font)


def fmt_val(v):
    """Format a cell value for display."""
    if v is None or v == '':
        return ''
    s = str(v).strip()
    if s == '-':
        return '-'
    try:
        f = float(s)
        if f == int(f):
            return '{:,}'.format(int(f))
        return '{:,.1f}'.format(f)
    except (ValueError, TypeError):
        return s


def val_color(v, col_name=''):
    """Return text color based on value."""
    if v is None or v == '' or v == '-':
        return C_DASH if v == '-' else C_TEXT_DARK
    try:
        f = float(str(v))
        if f == 0:
            return C_ZERO_HILITE
    except (ValueError, TypeError):
        pass
    return C_TEXT_DARK


def draw_table(title, col_headers, col_widths, col_aligns, rows, row_colors,
               special_last_col=False):
    """
    Render one table as a PIL Image.
    col_aligns: list of 'L'/'R'/'C' per column
    special_last_col: True -> last column gets dark navy header
    Returns Image (at SCALE*x, caller downsamples)
    """
    font_reg  = load_font(FONT_SIZE, bold=False)
    font_bold = load_font(FONT_SIZE, bold=True)
    font_hdr  = load_font(HEADER_SIZE, bold=True)
    font_ttl  = load_font(TITLE_SIZE, bold=True)

    n_cols    = len(col_headers)
    total_w   = sum(col_widths)
    n_rows    = len(rows)
    total_h   = TITLE_H + COL_HDR_H + n_rows * ROW_H + SCALE * 4  # bottom pad

    img  = Image.new('RGB', (total_w, total_h), C_ROW_PLAIN)
    draw = ImageDraw.Draw(img)

    # ---- Title banner ----
    draw.rectangle([0, 0, total_w, TITLE_H], fill=C_NAVY)
    tw = text_w(draw, title, font_ttl)
    draw.text(((total_w - tw) // 2, (TITLE_H - TITLE_SIZE) // 2),
              title, fill=C_TEXT_WHITE, font=font_ttl)

    # ---- Column headers ----
    y = TITLE_H
    x = 0
    for ci, (hdr, w) in enumerate(zip(col_headers, col_widths)):
        bg = C_PIECES_BG if (special_last_col and ci == n_cols - 1) else C_COL_HEADER
        draw.rectangle([x, y, x + w, y + COL_HDR_H], fill=bg)
        # draw right border
        draw.line([x + w - 1, y, x + w - 1, y + COL_HDR_H], fill=C_GRID, width=1)
        # centre text vertically + horizontally
        tw = text_w(draw, hdr, font_hdr)
        # handle multi-line headers (split on '\n')
        lines = hdr.split('\n')
        line_h = HEADER_SIZE + PAD_V
        total_text_h = len(lines) * line_h
        ty = y + (COL_HDR_H - total_text_h) // 2
        for li, line in enumerate(lines):
            lw = text_w(draw, line, font_hdr)
            tx = x + (w - lw) // 2
            draw.text((tx, ty + li * line_h), line, fill=C_TEXT_WHITE, font=font_hdr)
        x += w

    # bottom border of header
    draw.line([0, TITLE_H + COL_HDR_H, total_w, TITLE_H + COL_HDR_H],
              fill=C_GRID, width=SCALE)

    # ---- Data rows ----
    for ri, (row_vals, row_bg) in enumerate(zip(rows, row_colors)):
        y = TITLE_H + COL_HDR_H + ri * ROW_H
        x = 0
        for ci, (val, w, align) in enumerate(zip(row_vals, col_widths, col_aligns)):
            draw.rectangle([x, y, x + w, y + ROW_H], fill=row_bg)
            draw.line([x + w - 1, y, x + w - 1, y + ROW_H], fill=C_GRID, width=1)  # right border
            draw.line([x, y + ROW_H - 1, x + w, y + ROW_H - 1], fill=C_GRID, width=1)  # bottom

            s     = fmt_val(val)
            color = val_color(val)

            # Category column: use navy
            if ci == 1 and s not in ('', '-'):
                color = C_TEXT_BLUE

            font = font_reg
            tw   = text_w(draw, s, font)

            if align == 'R':
                tx = x + w - PAD_H - tw
            elif align == 'C':
                tx = x + (w - tw) // 2
            else:
                tx = x + PAD_H

            ty = y + (ROW_H - FONT_SIZE) // 2
            draw.text((tx, ty), s, fill=color, font=font)
            x += w

    return img


# -----------------------------------------------------------------------
def read_inventory(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Inventory']

    month_name = date.today().strftime('%B').upper()
    current_year = date.today().year

    # Detect month from A1 if present
    a1 = str(ws['A1'].value or '')
    m = re.search(r'([A-Z]+)\s+MONTH', a1, re.IGNORECASE)
    if m:
        month_name = m.group(1).upper()

    rows_all = []
    for r in range(3, ws.max_row + 1):
        item_id  = ws.cell(r, 1).value
        cat      = ws.cell(r, 2).value
        name     = ws.cell(r, 3).value
        uom      = ws.cell(r, 4).value
        opening  = ws.cell(r, 5).value
        received = ws.cell(r, 6).value
        issued   = ws.cell(r, 7).value
        store_bal= ws.cell(r, 8).value
        wip      = ws.cell(r, 9).value
        pieces   = ws.cell(r, 10).value
        if item_id is None:
            continue
        rows_all.append({
            'row': r, 'id': item_id, 'cat': cat, 'name': name, 'uom': uom,
            'opening': opening, 'received': received, 'issued': issued,
            'store_bal': store_bal, 'wip': wip, 'pieces': pieces,
        })

    return rows_all, month_name


def make_slug_image(rows_all, month_name):
    slugs = [r for r in rows_all if r['cat'] == 'Slug']
    title = "SLUGS INVENTORY %s MONTH TO DATE" % month_name

    col_headers = [
        'Item ID', 'Category', 'Item Name\n(ERP)', 'UOM',
        'Opening', 'Received\nfrom Vendor', 'Issued to\nProduction',
        'Store\nBalance', 'Work In\nProcess', 'Pieces Can Be\nProduced\n(Store + Floor)'
    ]
    col_aligns  = ['R', 'C', 'L', 'C', 'R', 'R', 'R', 'R', 'R', 'R']
    col_widths  = [w * SCALE for w in [70, 74, 175, 52, 68, 105, 105, 80, 80, 172]]

    data_rows  = []
    row_colors = []
    for i, r in enumerate(slugs):
        data_rows.append([
            r['id'], r['cat'], r['name'], r['uom'],
            r['opening'], r['received'], r['issued'],
            r['store_bal'], r['wip'], r['pieces'],
        ])
        row_colors.append(C_ROW_ALT if i % 2 else C_ROW_PLAIN)

    img = draw_table(title, col_headers, col_widths, col_aligns,
                     data_rows, row_colors, special_last_col=True)
    return img


def make_inventory_image(rows_all, month_name):
    STOP_AFTER = 'Thinner'
    inv_rows = []
    for r in rows_all:
        if r['cat'] == 'Slug':
            continue
        inv_rows.append(r)
        if r['cat'] == STOP_AFTER:
            # keep adding until category changes
            pass

    # Cut at first category after Thinner
    cut = len(inv_rows)
    found_thinner = False
    for i, r in enumerate(inv_rows):
        if r['cat'] == STOP_AFTER:
            found_thinner = True
        elif found_thinner:
            cut = i
            break
    inv_rows = inv_rows[:cut]

    title = "INVENTORY %s MONTH TO DATE" % month_name

    col_headers = [
        'Item ID', 'Category', 'Item Name (ERP)', 'UOM',
        'Opening', 'Received\nfrom Vendor', 'Issued to\nProduction', 'Store\nBalance'
    ]
    col_aligns  = ['R', 'C', 'L', 'C', 'R', 'R', 'R', 'R']
    col_widths  = [w * SCALE for w in [62, 85, 255, 50, 68, 95, 95, 78]]

    data_rows  = []
    row_colors = []
    for i, r in enumerate(inv_rows):
        data_rows.append([
            r['id'], r['cat'], r['name'], r['uom'],
            r['opening'], r['received'], r['issued'], r['store_bal'],
        ])
        row_colors.append(C_ROW_ALT if i % 2 else C_ROW_PLAIN)

    img = draw_table(title, col_headers, col_widths, col_aligns,
                     data_rows, row_colors, special_last_col=False)
    return img


# -----------------------------------------------------------------------
def main():
    SEP = "=" * 55

    print("")
    print(SEP)
    print("  Alpha Containers - Snapshot Generator v1")
    print(SEP)
    print("")

    folder_path = os.path.dirname(os.path.abspath(__file__))
    files = glob.glob(os.path.join(folder_path, "AlphaContainers*.xlsx"))
    if not files:
        print("  ERROR: No AlphaContainers*.xlsx found.")
        return
    excel_path = sorted(files)[-1]
    print("  Reading: " + os.path.basename(excel_path))

    print("")
    print("  Reading Inventory sheet...")
    rows_all, month_name = read_inventory(excel_path)
    slug_count = sum(1 for r in rows_all if r['cat'] == 'Slug')
    print("  Slugs: %d rows | Month: %s" % (slug_count, month_name))

    date_str = date.today().strftime('%d%b')  # e.g. 29Apr

    print("")
    print("[1/2] Generating slug snapshot...")
    slug_img_hires = make_slug_image(rows_all, month_name)
    # Downsample for crisp final image
    final_w = slug_img_hires.width // SCALE
    final_h = slug_img_hires.height // SCALE
    slug_img = slug_img_hires.resize((final_w, final_h), Image.LANCZOS)
    slug_path = os.path.join(folder_path, "snapshot_slugs_%s.png" % date_str)
    slug_img.save(slug_path, dpi=(150, 150))
    print("  Saved: " + os.path.basename(slug_path))

    print("")
    print("[2/2] Generating inventory snapshot...")
    inv_img_hires = make_inventory_image(rows_all, month_name)
    final_w = inv_img_hires.width // SCALE
    final_h = inv_img_hires.height // SCALE
    inv_img = inv_img_hires.resize((final_w, final_h), Image.LANCZOS)
    inv_path = os.path.join(folder_path, "snapshot_inventory_%s.png" % date_str)
    inv_img.save(inv_path, dpi=(150, 150))
    print("  Saved: " + os.path.basename(inv_path))

    print("")
    print("  Done. Send these two files to the WhatsApp group:")
    print("    " + os.path.basename(slug_path))
    print("    " + os.path.basename(inv_path))
    print(SEP)


if __name__ == "__main__":
    main()
