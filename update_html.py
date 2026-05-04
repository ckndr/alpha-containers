"""
update_html.py
Alpha Containers — Dashboard HTML Auto-Updater
------------------------------------------------
Reads AlphaContainers_vX_XX.xlsx → calculates live KPIs, MTD production,
downtime, and order compliance → injects into AlphaContainers_App.html.

Run manually or add to Run All Updates.bat:
    python update_html.py

Author: Sikander / Claude
Version: 1.0
"""

import os, re, json, glob
from datetime import datetime, date

# ── PATH SETUP ──────────────────────────────────────────────
DIR   = os.path.dirname(os.path.abspath(__file__))

# Find latest AlphaContainers Excel file
excel_pattern = os.path.join(DIR, 'AlphaContainers_v*.xlsx')
excel_files   = sorted(glob.glob(excel_pattern))
if not excel_files:
    raise FileNotFoundError(f"No AlphaContainers_v*.xlsx found in {DIR}")
EXCEL_PATH = excel_files[-1]

HTML_PATH = os.path.join(DIR, 'AlphaContainers_App.html')

print(f"Reading:  {os.path.basename(EXCEL_PATH)}")
print(f"Updating: {os.path.basename(HTML_PATH)}")

# ── LOAD EXCEL ───────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    raise ImportError("openpyxl not installed. Run: pip install openpyxl --break-system-packages")

wb      = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws_pl   = wb['Production_Log']
ws_db   = wb['Tubex_Dashboard']

# ── MONTH FILTER (current month) ────────────────────────────
now        = datetime.now()
cur_month  = now.month
cur_year   = now.year
month_name = now.strftime('%B %Y')

# ── PRODUCTION LOG HEADERS (row 2) ──────────────────────────
# Cols: A=Date B=Machine C=Customer D=Product E=Dia F=PID
#       G=Target H=Good I=Reject J=Waste%
#       K=Mech L=Elec M=MatShort N=Changeover O=Operations
#       P=PowerShutdown Q=GasShutdown R=WorkersShort

# ── READ PRODUCTION LOG ─────────────────────────────────────
mtd_by_pid    = {}   # pid → total good qty (MTD, printing, no varnish)
yest_tube     = 0
yest_pet      = 0
latest_date   = None

downtime_cols = {
    'Mechanical':        11,   # col K
    'Electrical':        12,   # col L
    'Material Shortage': 13,   # col M
    'Changeover':        14,   # col N
    'Operations':        15,   # col O
    'Power Shutdown':    16,   # col P
    'Gas Shutdown':      17,   # col Q
    'Workers Shortage':  18,   # col R
}
dt_totals = {k: 0.0 for k in downtime_cols}

for row in ws_pl.iter_rows(min_row=3, values_only=True):
    row_date  = row[0]   # col A
    machine   = row[1]   # col B
    prod_name = row[3]   # col D
    pid       = row[5]   # col F
    good_qty  = row[7]   # col H

    if not row_date or not machine: continue
    if not isinstance(row_date, (date, datetime)): continue

    row_dt = row_date if isinstance(row_date, datetime) else datetime(row_date.year, row_date.month, row_date.day)

    if row_dt.month != cur_month or row_dt.year != cur_year: continue

    # Track latest date for "yesterday"
    if latest_date is None or row_dt.date() > latest_date:
        latest_date = row_dt.date()

    # Downtime (sum regardless of machine)
    for cat, col_idx in downtime_cols.items():
        val = row[col_idx - 1]  # convert to 0-based
        if val and isinstance(val, (int, float)):
            dt_totals[cat] += float(val)

    if not pid or not good_qty: continue
    good_qty = int(good_qty) if good_qty else 0

    is_print = str(machine).upper().startswith('PRINT') or str(machine).upper().startswith('PLINE')
    is_pet   = str(machine).upper().startswith('PF') or str(machine).upper().startswith('PET')
    is_varn  = '(VARNISH)' in str(prod_name).upper()

    # MTD by PID
    if (is_print and not is_varn) or is_pet:
        pid = int(pid)
        mtd_by_pid[pid] = mtd_by_pid.get(pid, 0) + good_qty

# Yesterday's production
if latest_date:
    for row in ws_pl.iter_rows(min_row=3, values_only=True):
        row_date = row[0]
        machine  = row[1]
        prod_name= row[3]
        pid      = row[5]
        good_qty = row[7]
        if not row_date or not machine: continue
        if not isinstance(row_date, (date, datetime)): continue
        row_dt = row_date if isinstance(row_date, datetime) else datetime(row_date.year, row_date.month, row_date.day)
        if row_dt.date() != latest_date: continue
        if not good_qty: continue
        good_qty = int(good_qty)
        is_print = str(machine).upper().startswith('PRINT') or str(machine).upper().startswith('PLINE')
        is_pet   = str(machine).upper().startswith('PF') or str(machine).upper().startswith('PET')
        is_varn  = '(VARNISH)' in str(prod_name).upper()
        if is_print and not is_varn:
            yest_tube += good_qty
        elif is_pet:
            yest_pet += good_qty

# Total MTD
tube_mtd = sum(v for k,v in mtd_by_pid.items() if k < 8000)
pet_mtd  = sum(v for k,v in mtd_by_pid.items() if k >= 8000)

print(f"\nKPIs for {month_name}:")
print(f"  Tube MTD: {tube_mtd:,}  |  Yesterday: {yest_tube:,}")
print(f"  PET MTD:  {pet_mtd:,}   |  Yesterday: {yest_pet:,}")

# ── READ ACTIVE ORDERS FROM DASHBOARD ───────────────────────
TUBE_ROWS = range(11, 19)   # rows 11–18
PET_ROWS  = range(21, 27)   # rows 21–26

def read_orders(row_range, is_pet=False):
    orders = []
    for r in row_range:
        product  = ws_db.cell(row=r, column=4).value
        customer = ws_db.cell(row=r, column=3).value
        dia      = ws_db.cell(row=r, column=5).value
        pid      = ws_db.cell(row=r, column=6).value
        ordered  = ws_db.cell(row=r, column=7).value
        if not pid or not product: continue
        pid_int  = int(pid)
        produced = mtd_by_pid.get(pid_int, 0)
        dia_str  = (str(dia) + ('ml' if is_pet else 'mm')) if dia else '—'
        orders.append({
            'pid':      pid_int,
            'product':  str(product),
            'customer': str(customer) if customer else '—',
            'dia':      dia_str,
            'ordered':  int(ordered) if ordered else 0,
            'produced': produced,
            'dispatch': 0,
        })
    return orders

tube_orders = read_orders(TUBE_ROWS, is_pet=False)
pet_orders  = read_orders(PET_ROWS,  is_pet=True)

print(f"\nTube orders: {len(tube_orders)}")
for o in tube_orders:
    print(f"  PID {o['pid']} — {o['product'][:30]} | ordered={o['ordered']:,} | produced={o['produced']:,}")
print(f"\nPET orders: {len(pet_orders)}")

# ── BUILD DATA OBJECT ────────────────────────────────────────
DOWNTIME_ICONS = {
    'Mechanical':'🔧', 'Electrical':'⚡', 'Workers Shortage':'👷',
    'Power Shutdown':'🔌', 'Material Shortage':'📦',
    'Gas Shutdown':'🔥', 'Changeover':'🔄', 'Operations':'⚙️',
}

dash_data = {
    'month':       month_name,
    'lastUpdated': f"Updated {now.strftime('%d-%b-%Y %H:%M')} from {os.path.basename(EXCEL_PATH)}",
    'kpi': {
        'tubeYest': yest_tube,
        'tubeMTD':  tube_mtd,
        'petYest':  yest_pet,
        'petMTD':   pet_mtd,
    },
    'downtime': [
        {'cat': cat, 'icon': DOWNTIME_ICONS.get(cat,'⏱'), 'hrs': round(dt_totals[cat], 2)}
        for cat in downtime_cols
    ],
    'tubeOrders': tube_orders,
    'petOrders':  pet_orders,
}

# ── INJECT INTO HTML ─────────────────────────────────────────
if not os.path.exists(HTML_PATH):
    raise FileNotFoundError(f"HTML not found: {HTML_PATH}")

with open(HTML_PATH, 'r', encoding='utf-8') as f:
    html = f.read()

# Find DATA_START / DATA_END markers
marker_start = '/* DATA_START */'
marker_end   = '/* DATA_END */'
pos_start    = html.find(marker_start)
pos_end      = html.find(marker_end)

if pos_start == -1 or pos_end == -1:
    raise RuntimeError("DATA_START / DATA_END markers not found in HTML. File may be corrupted.")

# Build new data block
new_data = f"{marker_start}\nconst DASH_DATA = {json.dumps(dash_data, indent=2)};\n{marker_end}"

# Also update the generated timestamp in the comment above DATA_START
ts_old = re.search(r'Last generated: .+', html)
if ts_old:
    html = html.replace(ts_old.group(), f"Last generated: {now.strftime('%d-%b-%Y %H:%M')}")

# Replace data block
html = html[:pos_start] + new_data + html[pos_end + len(marker_end):]

with open(HTML_PATH, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\n✓ HTML updated successfully → {os.path.basename(HTML_PATH)}")
print(f"  Open in Chrome on PC or Android to view dashboard.")
print(f"  On Android: tap ⋮ menu → 'Add to Home Screen' to install as app.")
