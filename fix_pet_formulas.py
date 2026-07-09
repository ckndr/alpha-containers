import openpyxl

wb = openpyxl.load_workbook('Tubex_July26.xlsx')
sheet = wb['MRP']

def get_pet_formula(row_idx):
    # Pet Products are in rows 98 to 103
    parts = []
    for i in range(98, 104):
        part = f'IF((COUNTIFS(TableBOM[Product ID],$D${i},TableBOM[Item ID],$A{row_idx})>0)*($H${i}>0),$C${i}&", ","")'
        parts.append(part)
    
    concatenated = ' & '.join(parts)
    
    formula = f'=IF(LEN({concatenated})>1,LEFT({concatenated},LEN({concatenated})-2),"")'
    return formula

count = 0
for row_idx in range(107, sheet.max_row + 1):
    cell = sheet.cell(row=row_idx, column=8) # Column H
    if isinstance(cell.value, str) and cell.value.startswith('=IF(LEN('):
        cell.value = get_pet_formula(row_idx)
        count += 1
        
wb.save('Tubex_July26.xlsx')
print(f"Pet Requirement Formulas updated successfully for {count} rows.")
