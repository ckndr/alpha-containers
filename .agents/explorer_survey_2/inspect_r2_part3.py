import openpyxl

print("=== Detail check on R2-05: Inks in Aerosol_BOM vs Job Card ===")
wb_jc = openpyxl.load_workbook('d:/Alpha/Aerosol/Aerosol_Job_Card.xlsx', data_only=False)
ws_abom = wb_jc['Aerosol_BOM']
print("Aerosol_BOM max rows:", ws_abom.max_row)
for r in range(1, min(ws_abom.max_row+1, 35)):
    row_v = [ws_abom.cell(r, c).value for c in range(1, 16)]
    if any(row_v):
        print(f"BOM Row {r:2d}: {row_v[:10]}")

print("\n=== Products_DB in Aerosol_Job_Card.xlsx ===")
ws_pdb = wb_jc['Products_DB']
for r in range(1, ws_pdb.max_row+1):
    print(f"DB Row {r}: {[ws_pdb.cell(r, c).value for c in range(1, ws_pdb.max_column+1)]}")

print("\n=== Detail check on R2-07: Tubex BOM and MRP formulas ===")
wb_tubex = openpyxl.load_workbook('d:/Alpha/Tubex_Aug26.xlsx', data_only=False)
ws_tbom = wb_tubex['BOM']
print("BOM columns:", [ws_tbom.cell(2, c).value for c in range(1, 15)])
for r in range(3, 8):
    print(f"Row {r}:", [ws_tbom.cell(r, c).value for c in range(1, 12)])

ws_tmrp = wb_tubex['MRP']
print("MRP Header (Row 2):", [ws_tmrp.cell(2, c).value for c in range(1, 15)])
for r in range(3, 8):
    print(f"MRP Row {r}:", [ws_tmrp.cell(r, c).value for c in range(1, 12)])
print("MRP Raw material requirements (e.g. Rows 100-105):")
for r in range(100, 110):
    if ws_tmrp.cell(r, 1).value is not None:
        print(f"MRP RM Row {r}: Item={ws_tmrp.cell(r,1).value}, E{r}={ws_tmrp.cell(r,5).value}, F{r}={ws_tmrp.cell(r,6).value}, G{r}={ws_tmrp.cell(r,7).value}")

print("\n=== Detail check on R2-08, R2-09, R2-10 in Production.xlsx ===")
wb_p = openpyxl.load_workbook('d:/Alpha/Production.xlsx', data_only=False)
print("Production.xlsx sheet names:", wb_p.sheetnames)
# Check Dashbord
ws_d = wb_p['Dashbord']
for r in range(1, 20):
    vals = [ws_d.cell(r, c).value for c in range(1, ws_d.max_column+1)]
    if any(vals):
        print(f"Dashbord Row {r}: {vals[:10]}")

# Check Sheet3
ws_s3 = wb_p['Sheet3']
for r in range(1, 10):
    vals = [ws_s3.cell(r, c).value for c in range(1, 15)]
    if any(vals):
        print(f"Sheet3 Row {r}: {vals[:8]}")

# Check Production Day wise
ws_pdw = wb_p['Production Day wise']
print("Production Day wise header (Row 2):", [ws_pdw.cell(2, c).value for c in range(1, ws_pdw.max_column+1)])
for r in range(3, 8):
    print(f"PDW Row {r}: K={ws_pdw.cell(r,11).value}, L={ws_pdw.cell(r,12).value}, M={ws_pdw.cell(r,13).value}, N={ws_pdw.cell(r,14).value}")

print("\n=== Check Pending order references across workbooks ===")
import glob
for f in glob.glob('d:/Alpha/**/*.xlsx', recursive=True):
    if '~$' in f: continue
    try:
        w = openpyxl.load_workbook(f, read_only=True)
        for s in w.sheetnames:
            if '01-05' in s or 'pending' in s.lower() or 'order' in s.lower():
                print(f"Found sheet '{s}' in {f}")
        w.close()
    except Exception as e:
        pass
