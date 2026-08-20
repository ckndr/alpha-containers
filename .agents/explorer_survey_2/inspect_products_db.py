import openpyxl

wb = openpyxl.load_workbook('d:/Alpha/Aerosol/Aerosol_Job_Card.xlsx', data_only=False)
ws_pdb = wb['Products_DB']
for r in range(1, ws_pdb.max_row+1):
    vals = [ws_pdb.cell(r, c).value for c in range(1, ws_pdb.max_column+1)]
    print(f"Row {r:2d}: {vals}")
