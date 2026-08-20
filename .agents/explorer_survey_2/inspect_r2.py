import openpyxl
import os

print("=== R2-01: Tubex_Aug26.xlsx Tubex_Dashboard G12:G56 ===")
wb = openpyxl.load_workbook('d:/Alpha/Tubex_Aug26.xlsx', data_only=False)
ws = wb['Tubex_Dashboard']
for r in [12, 13, 20, 30, 40, 50, 56]:
    print(f"G{r}: {ws.cell(r, 7).value}")

print("\n=== R2-02: Tubex_Aug26.xlsx Product_Catalog J50:P55 ===")
ws = wb['Product_Catalog']
for r in range(50, 56):
    print(f"Row {r} (Product={ws.cell(r,2).value}): J{r}={ws.cell(r, 10).value[:60]}... P{r}={ws.cell(r, 16).value[:60]}...")

print("\n=== R2-03: Aerosol/Aerosol BOM.xlsx Theoretical BOM K6:K7 ===")
wb_aero_bom = openpyxl.load_workbook('d:/Alpha/Aerosol/Aerosol BOM.xlsx', data_only=False)
ws_tb = wb_aero_bom['Theoretical BOM']
print(f"Row 6 Material: {ws_tb['C6'].value}, J6={ws_tb['J6'].value}, K6={ws_tb['K6'].value}, L6={ws_tb['L6'].value}")
print(f"Row 7 Material: {ws_tb['C7'].value}, J7={ws_tb['J7'].value}, K7={ws_tb['K7'].value}, L7={ws_tb['L7'].value}")

print("\n=== R2-04 & R2-05: Aerosol/Aerosol_Job_Card.xlsx ===")
wb_aero_jc = openpyxl.load_workbook('d:/Alpha/Aerosol/Aerosol_Job_Card.xlsx', data_only=False)
ws_jc = wb_aero_jc['Job Card']
print("Header cells: B7 =", ws_jc['B7'].value, "B8 =", ws_jc['B8'].value, "D8 =", ws_jc['D8'].value)
for r in range(12, 37):
    mat_cat = ws_jc[f'A{r}'].value
    desc = ws_jc[f'B{r}'].value
    req = ws_jc[f'E{r}'].value
    if mat_cat or desc or req:
        print(f"Row {r:2d}: Cat={mat_cat} | Desc={desc} | ReqFormula={req}")

print("\n=== R2-06 & R2-15: Tubex_Aug26.xlsx Inventory J3:J111 & J63 ===")
ws_inv = wb['Inventory']
for r in [3, 4, 5, 10, 20, 63, 100]:
    print(f"Row {r} (Item={ws_inv.cell(r, 1).value}, Desc={ws_inv.cell(r, 2).value}): J{r}={ws_inv.cell(r, 10).value}")

print("\n=== R2-07: Scrap Model in Tubex_Aug26.xlsx BOM vs Aerosol BOM.xlsx ===")
ws_bom = wb['BOM']
print("Tubex BOM Header:", [ws_bom.cell(2, c).value for c in range(1, 10)])
for r in range(3, 8):
    print(f"Tubex BOM Row {r}: PID={ws_bom.cell(r, 1).value}, Cat={ws_bom.cell(r, 3).value}, Per1000={ws_bom.cell(r, 6).value}, Scrap%={ws_bom.cell(r, 7).value}, GrossPer1000={ws_bom.cell(r, 8).value}")
print("Aerosol BOM Header:", [ws_tb.cell(5, c).value for c in range(1, 14)])
for r in range(6, 10):
    print(f"Aero BOM Row {r}: Cat={ws_tb.cell(r, 2).value}, Net={ws_tb.cell(r, 10).value}, Scrap={ws_tb.cell(r, 11).value}, Gross={ws_tb.cell(r, 12).value}")

print("\n=== R2-08, R2-09, R2-10: Production.xlsx ===")
wb_prod = openpyxl.load_workbook('d:/Alpha/Production.xlsx', data_only=False)
print("Production.xlsx sheets:", wb_prod.sheetnames)
if 'Summary 14-08-2026' in wb_prod.sheetnames:
    ws_s = wb_prod['Summary 14-08-2026']
    print("Summary 14-08-2026 B11:B13 =", ws_s['B11'].value, ws_s['B12'].value, ws_s['B13'].value)
    print("Summary 14-08-2026 B22:B24 =", ws_s['B22'].value, ws_s['B23'].value, ws_s['B24'].value)
if 'Production Day wise' in wb_prod.sheetnames:
    ws_pdw = wb_prod['Production Day wise']
    print("Production Day wise N1 =", ws_pdw['N1'].value)
    print("Production Day wise N3 =", ws_pdw['N3'].value)
    print("Production Day wise N4 =", ws_pdw['N4'].value)
if 'Sheet3' in wb_prod.sheetnames:
    ws_s3 = wb_prod['Sheet3']
    print("Sheet3 J3 =", ws_s3['J3'].value)
    print("Sheet3 L3 =", ws_s3['L3'].value)

print("\n=== R2-11: Aerosol/Tubex_v10_30.xlsx MRP F118:G121 ===")
wb_v10 = openpyxl.load_workbook('d:/Alpha/Aerosol/Tubex_v10_30.xlsx', data_only=False)
ws_v10_mrp = wb_v10['MRP']
for r in range(118, 122):
    print(f"MRP Row {r}: D={ws_v10_mrp.cell(r,4).value}, E={ws_v10_mrp.cell(r,5).value}, F={ws_v10_mrp.cell(r,6).value}, G={ws_v10_mrp.cell(r,7).value}")

print("\n=== R2-12: August_Plan.xlsx August Plan PET K10:M10 ===")
wb_plan = openpyxl.load_workbook('d:/Alpha/August_Plan.xlsx', data_only=False)
ws_pet_plan = wb_plan['August Plan PET']
print("Rows 6-10 in August Plan PET:")
for r in range(6, 12):
    print(f"Row {r:2d}: B={ws_pet_plan.cell(r, 2).value} | K={ws_pet_plan.cell(r, 11).value} | L={ws_pet_plan.cell(r, 12).value} | M={ws_pet_plan.cell(r, 13).value}")

print("\n=== R2-13: Tubex_Aug26.xlsx FG Stock I4:I20 ===")
ws_fg = wb['FG Stock']
for r in range(4, 15):
    print(f"FG Stock Row {r}: B={ws_fg.cell(r, 2).value}, C={ws_fg.cell(r, 3).value}, I={ws_fg.cell(r, 9).value}")

print("\n=== R2-14: Tubex_Aug26.xlsx Tubex_Dashboard N7:N18 ===")
ws_dash = wb['Tubex_Dashboard']
for r in range(6, 20):
    print(f"Row {r:2d}: M={ws_dash.cell(r, 13).value}, N={ws_dash.cell(r, 14).value}, O={ws_dash.cell(r, 15).value}")

print("\n=== R2-16: Check Pending.xlsx or pending order calculations ===")
for path in ['d:/Alpha/Pending.xlsx', 'd:/Alpha/Tubex Records/Samsol PET Orders.xlsx']:
    if os.path.exists(path):
        wb_p = openpyxl.load_workbook(path, data_only=False)
        print(f"{path} sheets:", wb_p.sheetnames)
