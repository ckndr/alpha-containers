import openpyxl

wb = openpyxl.load_workbook('d:/Alpha/Aerosol/Aerosol_Job_Card.xlsx', data_only=False)
ws_jc = wb['Job Card']
print("=== Job Card all rows ===")
for r in range(1, ws_jc.max_row+1):
    vals = [ws_jc.cell(r, c).value for c in range(1, 8)]
    if any(vals):
        print(f"Row {r:2d}: {vals}")

ws_abom = wb['Aerosol_BOM']
print("\n=== Aerosol_BOM all rows ===")
for r in range(1, ws_abom.max_row+1):
    vals = [ws_abom.cell(r, c).value for c in range(1, ws_abom.max_column+1)]
    if any(vals):
        print(f"Row {r:2d}: {vals}")
