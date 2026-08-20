import openpyxl

wb = openpyxl.load_workbook('d:/Alpha/Tubex_Aug26.xlsx', data_only=False)
ws = wb['Tubex_Dashboard']
print("Tubex_Dashboard dimensions:", ws.max_row, ws.max_column)
for r in range(1, 35):
    row_vals = [ws.cell(r, c).value for c in range(1, 16)]
    if any(row_vals):
        print(f"Row {r:2d}: {row_vals}")
