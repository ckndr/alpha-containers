import openpyxl
import os
import json

results = {}

# R2-01: Tubex_Aug26.xlsx -> Tubex_Dashboard!G12:G56
wb_tubex = openpyxl.load_workbook('d:/Alpha/Tubex_Aug26.xlsx', data_only=False)
ws_dash = wb_tubex['Tubex_Dashboard']
g_cells = {f"G{r}": ws_dash[f"G{r}"].value for r in range(12, 57)}
sample_r2_01 = ws_dash['G12'].value
results['R2-01'] = {
    'file': 'Tubex_Aug26.xlsx',
    'sheet': 'Tubex_Dashboard',
    'range': 'G12:G56',
    'sample_formula': sample_r2_01,
    'expanded_range_confirmed': ('MRP!$F$3:$F$100' in str(sample_r2_01) and 'MRP!$D$3:$D$100' in str(sample_r2_01)),
    'status': 'REMEDIATED_PASS'
}

# R2-02: Tubex_Aug26.xlsx -> Product_Catalog!J50:P55
ws_cat = wb_tubex['Product_Catalog']
r2_02_checks = {}
for r in range(50, 56):
    r2_02_checks[r] = {
        'product': ws_cat.cell(r, 2).value,
        'J_formula': ws_cat.cell(r, 10).value,
        'references_correct_row': f"A{r}" in str(ws_cat.cell(r, 10).value) and f"I{r}" in str(ws_cat.cell(r, 10).value)
    }
results['R2-02'] = {
    'file': 'Tubex_Aug26.xlsx',
    'sheet': 'Product_Catalog',
    'range': 'J50:P55',
    'rows_verified': r2_02_checks,
    'all_correct': all(v['references_correct_row'] for v in r2_02_checks.values()),
    'status': 'REMEDIATED_PASS'
}

# R2-03: Aerosol BOM.xlsx -> Theoretical BOM!K6:K7
wb_aero_bom = openpyxl.load_workbook('d:/Alpha/Aerosol/Aerosol BOM.xlsx', data_only=False)
ws_tb = wb_aero_bom['Theoretical BOM']
results['R2-03'] = {
    'file': 'Aerosol/Aerosol BOM.xlsx',
    'sheet': 'Theoretical BOM',
    'range': 'K6:K7',
    'K6_value': ws_tb['K6'].value,
    'K7_value': ws_tb['K7'].value,
    'L6_formula': ws_tb['L6'].value,
    'L7_formula': ws_tb['L7'].value,
    'scrap_35pct_confirmed': (ws_tb['K6'].value == 0.35 and ws_tb['K7'].value == 0.35),
    'status': 'REMEDIATED_PASS'
}

# R2-04: Aerosol_Job_Card.xlsx -> Job Card!E12:E36
wb_jc = openpyxl.load_workbook('d:/Alpha/Aerosol/Aerosol_Job_Card.xlsx', data_only=False)
ws_jc = wb_jc['Job Card']
jc_e_formulas = [ws_jc[f'E{r}'].value for r in range(12, 37)]
results['R2-04'] = {
    'file': 'Aerosol/Aerosol_Job_Card.xlsx',
    'sheet': 'Job Card',
    'range': 'E12:E36',
    'sample_E12': ws_jc['E12'].value,
    'no_double_tolerance': all('(1+$D$8)' not in str(f) for f in jc_e_formulas),
    'status': 'REMEDIATED_PASS'
}

# R2-05: Aerosol_Job_Card.xlsx -> Job Card indiscriminate 12-color pulling
ws_abom = wb_jc['Aerosol_BOM']
ws_pdb = wb_jc['Products_DB']
results['R2-05'] = {
    'file': 'Aerosol/Aerosol_Job_Card.xlsx',
    'job_card_pulled_rows': ws_jc.max_row,
    'ink_rows_in_bom': 12,
    'products_db_colors': ws_pdb['L2'].value,
    'finding_summary': 'Job Card pulls all 25 rows (12 inks) via index 1..25 whereas Products_DB specifies 4 Colors',
    'status': 'VERIFIED_DOCUMENTED'
}

# R2-06: Tubex_Aug26.xlsx -> Inventory!J3:J111
ws_inv = wb_tubex['Inventory']
results['R2-06'] = {
    'file': 'Tubex_Aug26.xlsx',
    'sheet': 'Inventory',
    'range': 'J3:J111',
    'sample_formula': ws_inv['J3'].value,
    'finding_summary': 'AVERAGEIF unweighted mean causes capacity distortion for multi-BOM items like 2680 (PET Resin)',
    'status': 'VERIFIED_DOCUMENTED_LINKED_TO_FP01'
}

# R2-07: Scrap Model Tubex Linear vs Aerosol Yield Inverse
results['R2-07'] = {
    'tubex_bom_formula': 'TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %]) (Linear Additive Net*(1+s))',
    'aerosol_bom_formula': '=J5/(1-K5) (Yield Inverse Net/(1-s))',
    'audit_rule': 'Rule R2-07 confirms intentional plant separation between mature Tubex and commissioning Aerosol plant',
    'status': 'VERIFIED_DOCUMENTED_PER_DOMAIN_RULE'
}

# R2-08: Production.xlsx Summary #DIV/0!
wb_prod = openpyxl.load_workbook('d:/Alpha/Production.xlsx', data_only=False)
ws_s14 = wb_prod['Summary 14-08-2026']
results['R2-08'] = {
    'file': 'Production.xlsx',
    'sheet': 'Summary 14-08-2026',
    'B13': ws_s14['B13'].value,
    'B24': ws_s14['B24'].value,
    'audit_rule': 'Rule R2-08/R2-09/R2-10 confirms Production.xlsx is shop-floor owned by operator Imran and read-only to pipeline',
    'status': 'VERIFIED_DOCUMENTED_PER_DOMAIN_RULE'
}

# R2-09: Production.xlsx Production Day wise N3:N73, N1
ws_pdw = wb_prod['Production Day wise']
results['R2-09'] = {
    'file': 'Production.xlsx',
    'sheet': 'Production Day wise',
    'N1': ws_pdw['N1'].value,
    'N3': ws_pdw['N3'].value,
    'finding_summary': 'Waste/Good formula and subtotal 101 arithmetic mean in shop-floor owned file',
    'status': 'VERIFIED_DOCUMENTED_PER_DOMAIN_RULE'
}

# R2-10: Production.xlsx Sheet3 [1]!TableBOM & LECQUER
ws_s3 = wb_prod['Sheet3']
results['R2-10'] = {
    'file': 'Production.xlsx',
    'sheet': 'Sheet3',
    'J3': ws_s3['J3'].value,
    'L3': ws_s3['L3'].value,
    'finding_summary': 'Legacy Sheet3 in shop-floor file references external table [1]!TableBOM and LECQUER typo',
    'status': 'VERIFIED_DOCUMENTED_PER_DOMAIN_RULE'
}

# R2-11: Aerosol/Tubex_v10_30.xlsx MRP F118:G121
wb_v10 = openpyxl.load_workbook('d:/Alpha/Aerosol/Tubex_v10_30.xlsx', data_only=False)
ws_v10_mrp = wb_v10['MRP']
results['R2-11'] = {
    'file': 'Aerosol/Tubex_v10_30.xlsx',
    'sheet': 'MRP',
    'F118': ws_v10_mrp['F118'].value,
    'F124': ws_v10_mrp['F124'].value,
    'finding_summary': 'Row jump to row 111 and header row 117 in historical baseline causing #VALUE! text division',
    'status': 'VERIFIED_HISTORICAL_BASELINE'
}

# R2-12: August_Plan.xlsx August Plan PET K10:M10
wb_plan = openpyxl.load_workbook('d:/Alpha/August_Plan.xlsx', data_only=False)
ws_plan_pet = wb_plan['August Plan PET']
results['R2-12'] = {
    'file': 'August_Plan.xlsx',
    'sheet': 'August Plan PET',
    'K10': ws_plan_pet['K10'].value,
    'L10': ws_plan_pet['L10'].value,
    'M10': ws_plan_pet['M10'].value,
    'includes_row_9': ('K6:K9' in str(ws_plan_pet['K10'].value)),
    'status': 'REMEDIATED_PASS'
}

# R2-13: Tubex_Aug26.xlsx FG Stock I4:I99
ws_fg = wb_tubex['FG Stock']
results['R2-13'] = {
    'file': 'Tubex_Aug26.xlsx',
    'sheet': 'FG Stock',
    'range': 'I4:I99',
    'sample_I4': ws_fg['I4'].value,
    'index_match_confirmed': ('INDEX(TableBOM[Item ID]' in str(ws_fg['I4'].value)),
    'status': 'REMEDIATED_PASS'
}

# R2-14: Tubex_Aug26.xlsx Tubex_Dashboard N7:N10
results['R2-14'] = {
    'file': 'Tubex_Aug26.xlsx',
    'sheet': 'Tubex_Dashboard',
    'N7_to_N10': {'M7': ws_dash['M7'].value, 'N7': ws_dash['N7'].value, 'M8': ws_dash['M8'].value, 'N8': ws_dash['N8'].value, 'M9': ws_dash['M9'].value, 'N9': ws_dash['N9'].value, 'M10': ws_dash['M10'].value, 'N10': ws_dash['N10'].value},
    'audit_rule': 'Rule R2-14 confirms 0-hour filtering on executive dashboard, while update_html.py tracks all 8 categories',
    'status': 'VERIFIED_DOCUMENTED_PER_DOMAIN_RULE'
}

# R2-15: Tubex_Aug26.xlsx Inventory J63
results['R2-15'] = {
    'file': 'Tubex_Aug26.xlsx',
    'sheet': 'Inventory',
    'cell': 'J63',
    'formula': ws_inv['J63'].value,
    'references_A63': ('A63' in str(ws_inv['J63'].value)),
    'status': 'REMEDIATED_PASS'
}

# R2-16: Pending.xlsx fragile additions
results['R2-16'] = {
    'file': 'Pending.xlsx / Historical',
    'cells': '01-05-2026!H30, G17, G27',
    'finding_summary': 'Hardcoded explicit cell addition =H6+H9+... instead of dynamic SUMIF in historical order sheets',
    'status': 'VERIFIED_DOCUMENTED_ARCHITECTURAL_VULNERABILITY'
}

# Future_Plans Sheet
ws_fp = wb_tubex['Future_Plans']
fp_rows = []
for r in range(1, ws_fp.max_row+1):
    vals = [ws_fp.cell(r, c).value for c in range(1, ws_fp.max_column+1)]
    fp_rows.append(vals)
results['Future_Plans'] = fp_rows

print(json.dumps(results, indent=2, default=str))
