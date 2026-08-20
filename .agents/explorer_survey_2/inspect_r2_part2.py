import openpyxl

print("=== R2-04 & R2-05: Aerosol_Job_Card.xlsx ===")
wb_aero_jc = openpyxl.load_workbook('d:/Alpha/Aerosol/Aerosol_Job_Card.xlsx', data_only=False)
ws_jc = wb_aero_jc['Job Card']
for r in range(11, 26):
    row_data = [ws_jc.cell(r, c).value for c in range(1, 8)]
    print(f"Row {r:2d}: {row_data}")

print("\n=== R2-06 & R2-15: Tubex_Aug26.xlsx Inventory ===")
wb = openpyxl.load_workbook('d:/Alpha/Tubex_Aug26.xlsx', data_only=False)
ws_inv = wb['Inventory']
for r in range(3, 10):
    print(f"Row {r} (Item={ws_inv.cell(r, 1).value}, Desc={ws_inv.cell(r, 2).value}): J{r}={ws_inv.cell(r, 10).value}")
print(f"Row 63 (Item={ws_inv.cell(63, 1).value}, Desc={ws_inv.cell(63, 2).value}): J63={ws_inv.cell(63, 10).value}")
print(f"Row 64 (Item={ws_inv.cell(64, 1).value}, Desc={ws_inv.cell(64, 2).value}): J64={ws_inv.cell(64, 10).value}")

print("\n=== Inventory Column J across all rows ===")
anomalies = []
for r in range(3, ws_inv.max_row + 1):
    item_id = ws_inv.cell(r, 1).value
    formula = str(ws_inv.cell(r, 10).value)
    if item_id is not None:
        expected_ref = f"A{r}"
        if expected_ref not in formula and formula != "None" and formula != "-":
            anomalies.append((r, item_id, formula))
print(f"Inventory Column J row anomalies count: {len(anomalies)}")
for a in anomalies:
    print(f"  Anomaly at Row {a[0]}: Item {a[1]}, formula: {a[2]}")

