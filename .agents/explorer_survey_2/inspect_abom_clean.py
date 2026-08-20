import openpyxl

wb = openpyxl.load_workbook('d:/Alpha/Aerosol/Aerosol_Job_Card.xlsx', data_only=False)
ws_abom = wb['Aerosol_BOM']
print("=== Aerosol_BOM rows ===")
for r in range(1, ws_abom.max_row+1):
    vals = [ws_abom.cell(r, c).value for c in range(1, ws_abom.max_column+1)]
    if any(vals):
        print(f"Row {r:2d}: Key={vals[0]} | PID={vals[1]} | Group={vals[6]} | Item={vals[7]} | Name={vals[8]} | Net={vals[10]} | Scrap={vals[11]} | Gross={vals[12]}")
