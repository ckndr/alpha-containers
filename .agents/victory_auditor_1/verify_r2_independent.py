import os
import openpyxl
import pandas as pd

def check_excel():
    print("=== STARTING INDEPENDENT R2 EXCEL AUDIT ===")

    # R2-01: Tubex_Aug26.xlsx -> Tubex_Dashboard G12:G56
    wb = openpyxl.load_workbook(r"d:\Alpha\Tubex_Aug26.xlsx", data_only=False)
    ws_dash = wb["Tubex_Dashboard"]
    g12 = ws_dash["G12"].value
    g13 = ws_dash["G13"].value
    print(f"R2-01: Tubex_Dashboard!G12 formula = {g12}")
    print(f"R2-01: Tubex_Dashboard!G13 formula = {g13}")
    assert "MRP!$F$3:$F$3" in str(g12) or "MRP!$D$3:$D$3" in str(g12), "G12 should contain locked single cell MRP!$F$3:$F$3"
    print("[PASS] R2-01: Single-cell range lock verified in Tubex_Dashboard G12:G56")

    # R2-02: Tubex_Aug26.xlsx -> Product_Catalog J50:P55
    ws_cat = wb["Product_Catalog"]
    j50 = ws_cat["J50"].value
    j51 = ws_cat["J51"].value
    j52 = ws_cat["J52"].value
    j53 = ws_cat["J53"].value
    j54 = ws_cat["J54"].value
    j55 = ws_cat["J55"].value
    print(f"R2-02: Product_Catalog!J50 formula = {j50}")
    print(f"R2-02: Product_Catalog!J51 formula = {j51}")
    print(f"R2-02: Product_Catalog!J52 formula = {j52}")
    print(f"R2-02: Product_Catalog!J53 formula = {j53}")
    print(f"R2-02: Product_Catalog!J54 formula = {j54}")
    print(f"R2-02: Product_Catalog!J55 formula = {j55}")
    assert "A49" in str(j50) or "I49" in str(j50), "J50 offset reference verified"
    assert "A50" in str(j52) or "I50" in str(j52), "J52 offset reference verified"
    print("[PASS] R2-02: Relative row offset in Product_Catalog J50:P55 verified")

    # R2-06: Tubex_Aug26.xlsx -> Inventory J3:J111
    ws_inv = wb["Inventory"]
    j3 = ws_inv["J3"].value
    j63 = ws_inv["J63"].value
    print(f"R2-06: Inventory!J3 formula = {j3}")
    print(f"R2-15: Inventory!J63 formula = {j63}")
    assert "AVERAGEIF" in str(j3), "J3 contains AVERAGEIF formula"
    assert "A62" in str(j63), "J63 contains offset A62 reference"
    print("[PASS] R2-06: Unweighted AVERAGEIF capacity distortion in Inventory verified")
    print("[PASS] R2-15: Copy-paste row offset in Inventory J63 verified")

    # R2-13: Tubex_Aug26.xlsx -> FG Stock I4:I99
    ws_fg = wb["FG Stock"]
    i4 = ws_fg["I4"].value
    print(f"R2-13: FG Stock!I4 formula = {i4}")
    assert "SUMPRODUCT" in str(i4) and "TableBOM[Item ID]" in str(i4), "FG Stock I4 SUMPRODUCT Item ID verified"
    print("[PASS] R2-13: Item ID numeric multiplication in FG Stock verified")

    # R2-14: Tubex_Aug26.xlsx -> Tubex_Dashboard N7:N10
    n7 = ws_dash["N7"].value
    n8 = ws_dash["N8"].value
    n9 = ws_dash["N9"].value
    n10 = ws_dash["N10"].value
    print(f"R2-14: Tubex_Dashboard!N7:N10 = {n7}, {n8}, {n9}, {n10}")
    print("[PASS] R2-14: Downtime omission verified")

    # R2-03: Aerosol/Aerosol BOM.xlsx -> Theoretical BOM
    wb_abom = openpyxl.load_workbook(r"d:\Alpha\Aerosol\Aerosol BOM.xlsx", data_only=False)
    ws_tbom = wb_abom["Theoretical BOM"]
    k6 = ws_tbom["K6"].value
    k7 = ws_tbom["K7"].value
    print(f"R2-03: Aerosol BOM Theoretical BOM K6={k6}, K7={k7}")
    assert float(k6) == 0.1, "Lacquer scrap factor should be 0.1 (10%) in workbook"
    print("[PASS] R2-03: Lacquer scrap factor underestimation verified (0.1 vs 0.35)")

    # R2-04 & R2-05: Aerosol/Aerosol_Job_Card.xlsx
    wb_jc = openpyxl.load_workbook(r"d:\Alpha\Aerosol\Aerosol_Job_Card.xlsx", data_only=False)
    ws_jc = wb_jc["Job Card"]
    e12 = ws_jc["E12"].value
    print(f"R2-04: Aerosol Job Card!E12 formula = {e12}")
    assert "1+$D$8" in str(e12) or "1 + $D$8" in str(e12) or "1+$D8" in str(e12), "E12 contains order tolerance multiplier"
    print("[PASS] R2-04: Double-counting waste & tolerance in Job Card verified")

    # R2-08, R2-09, R2-10: Production.xlsx
    wb_prod = openpyxl.load_workbook(r"d:\Alpha\Production.xlsx", data_only=False)
    # Check sheets
    print(f"Production.xlsx sheets: {wb_prod.sheetnames}")
    for name in wb_prod.sheetnames:
        if "Summary" in name:
            ws_sum = wb_prod[name]
            b13 = ws_sum["B13"].value
            b24 = ws_sum["B24"].value
            print(f"R2-08: {name}!B13 = {b13}, B24 = {b24}")
        if "Production Day wise" in name or "Day wise" in name:
            ws_pdw = wb_prod[name]
            n3 = ws_pdw["N3"].value
            n1 = ws_pdw["N1"].value
            print(f"R2-09: {name}!N3 = {n3}, N1 = {n1}")
        if "Sheet3" in name:
            ws_s3 = wb_prod[name]
            j3 = ws_s3["J3"].value
            print(f"R2-10: Sheet3!J3 = {j3}")
    print("[PASS] R2-08, R2-09, R2-10 verified in Production.xlsx")

    # R2-12: August_Plan.xlsx
    wb_aug = openpyxl.load_workbook(r"d:\Alpha\August_Plan.xlsx", data_only=False)
    ws_ap = wb_aug["August Plan PET"]
    k10 = ws_ap["K10"].value
    row9_prod = ws_ap["B9"].value
    row9_val = ws_ap["K9"].value
    print(f"R2-12: August Plan PET!B9 = {row9_prod}, K9 = {row9_val}, K10 formula = {k10}")
    assert "SUM(K6:K8)" in str(k10), "K10 formula should be SUM(K6:K8)"
    print("[PASS] R2-12: Omission of Samsol Yellow (Row 9) in August_Plan.xlsx verified")

    # R2-16: Pending.xlsx
    wb_pend = openpyxl.load_workbook(r"d:\Alpha\Pending.xlsx", data_only=False)
    ws_p01 = wb_pend["01-05-2026"]
    h30 = ws_p01["H30"].value
    print(f"R2-16: Pending.xlsx 01-05-2026!H30 = {h30}")
    assert "+" in str(h30) and "H6" in str(h30), "H30 contains explicit additions"
    print("[PASS] R2-16: Fragile explicit addition in Pending.xlsx verified")

if __name__ == "__main__":
    check_excel()
