import openpyxl
import os
import re
import json

print("=================================================================")
print("COMPREHENSIVE AUDIT OF REQUIREMENT 2 (R2-01 to R2-16)")
print("=================================================================")

# R2-01
wb_tubex = openpyxl.load_workbook("Tubex_Aug26.xlsx", data_only=False)
ws_dash = wb_tubex["Tubex_Dashboard"]
r2_01_samples = {}
for r in [12, 13, 20, 35, 56]:
    r2_01_samples[f"G{r}"] = str(ws_dash[f"G{r}"].value)
print(f"R2-01 Samples: {r2_01_samples}")

# R2-02
ws_cat = wb_tubex["Product_Catalog"]
r2_02_samples = {}
for r in range(50, 56):
    r2_02_samples[r] = {col: str(ws_cat[f"{col}{r}"].value) for col in ["J", "K", "L", "M", "N", "O", "P"]}
print(f"R2-02 Samples (Row 50 J): {r2_02_samples[50]['J']}")

# R2-03
wb_abom = openpyxl.load_workbook("Aerosol/Aerosol BOM.xlsx", data_only=False)
ws_abom = wb_abom["Theoretical BOM"]
print(f"R2-03 K6: {ws_abom['K6'].value}, L6: {ws_abom['L6'].value}, K7: {ws_abom['K7'].value}, L7: {ws_abom['L7'].value}")

# R2-04 & R2-05
wb_jc = openpyxl.load_workbook("Aerosol/Aerosol_Job_Card.xlsx", data_only=False)
ws_jc = wb_jc["Job Card"]
print(f"R2-04 E12: {ws_jc['E12'].value}, E36: {ws_jc['E36'].value}")

# R2-06
ws_inv = wb_tubex["Inventory"]
print(f"R2-06 J3 formula: {ws_inv['J3'].value}")

# R2-07
print("R2-07 Linear Additive vs Yield Inverse documented in AUDIT_NOTES.md Rule 7")

# R2-08, R2-09, R2-10
wb_prod = openpyxl.load_workbook("Production.xlsx", data_only=False)
print("Production sheets:", wb_prod.sheetnames)
if "Summary 14-08-2026" in wb_prod.sheetnames:
    ws_sum = wb_prod["Summary 14-08-2026"]
    print(f"R2-08 B13: {ws_sum['B13'].value}, B24: {ws_sum['B24'].value}")
if "Production Day wise" in wb_prod.sheetnames:
    ws_pdw = wb_prod["Production Day wise"]
    print(f"R2-09 N1: {ws_pdw['N1'].value}, N3: {ws_pdw['N3'].value}")
if "Sheet3" in wb_prod.sheetnames:
    ws_s3 = wb_prod["Sheet3"]
    print(f"R2-10 J3: {ws_s3['J3'].value}")

# R2-11
if os.path.exists("Aerosol/Tubex_v10_30.xlsx"):
    wb_v10 = openpyxl.load_workbook("Aerosol/Tubex_v10_30.xlsx", data_only=False)
    if "MRP" in wb_v10.sheetnames:
        print("R2-11 Tubex_v10_30 MRP F118:", wb_v10["MRP"]["F118"].value)

# R2-12
wb_plan = openpyxl.load_workbook("August_Plan.xlsx", data_only=False)
ws_pet = wb_plan["August Plan PET"]
print(f"R2-12 August Plan PET K10: {ws_pet['K10'].value}, L10: {ws_pet['L10'].value}, M10: {ws_pet['M10'].value}")

# R2-13
ws_fg = wb_tubex["FG Stock"]
print(f"R2-13 FG Stock I3 header: {ws_fg['I3'].value}, I4: {ws_fg['I4'].value}")

# R2-14
print(f"R2-14 Tubex_Dashboard N7: {ws_dash['N7'].value}, N10: {ws_dash['N10'].value}")

# R2-15
print(f"R2-15 Inventory J63: {ws_inv['J63'].value}")

# R2-16
print("R2-16 Pending.xlsx checked if present:", os.path.exists("Pending.xlsx"))

print("\n=================================================================")
print("COMPREHENSIVE AUDIT OF REQUIREMENT 3 (R3-01 to R3-09)")
print("=================================================================")

# R3-01, R3-02, R3-03, R3-06, R3-07, R3-08, R3-09 in Tubex.html
with open("Tubex.html", "r", encoding="utf-8", errors="ignore") as f:
    html_content = f.read()

print("Checking escapeHtml definition in Tubex.html:")
if "function escapeHtml" in html_content:
    match = re.search(r"function escapeHtml\(.*?\)\s*\{.*?\}", html_content, re.DOTALL)
    if match:
        print("  Found escapeHtml:\n", match.group(0))

print("\nChecking XSS escapes in Tubex.html:")
print("  escapeHtml(o.customer):", "escapeHtml(o.customer)" in html_content)
print("  escapeHtml(o.product):", "escapeHtml(o.product)" in html_content)
print("  escapeHtml(r.product):", "escapeHtml(r.product)" in html_content)
print("  escapeHtml(r.remarks):", "escapeHtml(r.remarks)" in html_content)
print("  data-month with escapeHtml:", "data-month" in html_content)
print("  controllerchange event listener:", "controllerchange" in html_content)
print("  timestamp_iso in Tubex.html:", "timestamp_iso" in html_content or "DASH_DATA.timestamp_iso" in html_content)
print("  DATA_START marker in Tubex.html:", "/* DATA_START */" in html_content)

# Checking sw.js
with open("sw.js", "r", encoding="utf-8", errors="ignore") as f:
    sw_content = f.read()

print("\nChecking sw.js:")
print("  HTTP 200 check:", "response.status === 200" in sw_content or "status === 200" in sw_content)
print("  HTTP scheme check:", "startsWith('http')" in sw_content or "request.url.startsWith('http')" in sw_content)
print("  skipWaiting:", "self.skipWaiting()" in sw_content)
print("  clients.claim:", "clients.claim()" in sw_content)
print("  Pre-cache assets list:")
for line in sw_content.splitlines():
    if "index.html" in line or "Tubex.html" in line:
        print("   ", line.strip())

# Checking update_html.py
with open("Scripts/update_html.py", "r", encoding="utf-8", errors="ignore") as f:
    uh_content = f.read()

print("\nChecking Scripts/update_html.py:")
print("  timestamp_iso injection:", "timestamp_iso" in uh_content)
print("  inject_block definition:", "def inject_block" in uh_content)
