import openpyxl
import os

print("=== CHECKING R2-01: Tubex_Aug26.xlsx Tubex_Dashboard G12:G56 ===")
wb = openpyxl.load_workbook("Tubex_Aug26.xlsx", data_only=False)
ws = wb["Tubex_Dashboard"]
for r in [12, 13, 20, 35, 56]:
    print(f"G{r}: {ws[f'G{r}'].value}")

print("\n=== CHECKING R2-02: Tubex_Aug26.xlsx Product_Catalog J50:P55 ===")
ws_cat = wb["Product_Catalog"]
for r in range(50, 56):
    print(f"Row {r} (A={ws_cat[f'A{r}'].value}, I={ws_cat[f'I{r}'].value}):")
    for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P']:
        print(f"  {col}{r}: {ws_cat[f'{col}{r}'].value}")

print("\n=== CHECKING R2-13: Tubex_Aug26.xlsx FG Stock I4:I15 ===")
ws_fg = wb["FG Stock"]
for r in range(4, 10):
    print(f"I{r}: {ws_fg[f'I{r}'].value}")

print("\n=== CHECKING R2-15: Tubex_Aug26.xlsx Inventory J62:J65 ===")
ws_inv = wb["Inventory"]
for r in range(62, 66):
    print(f"J{r}: {ws_inv[f'J{r}'].value}")

print("\n=== CHECKING R2-14: Tubex_Aug26.xlsx Tubex_Dashboard M7:O10, M14:O18 ===")
for r in range(7, 11):
    print(f"Row {r}: M={ws[f'M{r}'].value}, N={ws[f'N{r}'].value}, O={ws[f'O{r}'].value}")
for r in range(14, 19):
    print(f"Row {r}: M={ws[f'M{r}'].value}, N={ws[f'N{r}'].value}, O={ws[f'O{r}'].value}")

print("\n=== CHECKING R2-12: August_Plan.xlsx August Plan PET K10:M10 ===")
if os.path.exists("August_Plan.xlsx"):
    wb_aug = openpyxl.load_workbook("August_Plan.xlsx", data_only=False)
    ws_ap = wb_aug["August Plan PET"]
    for col in ["K", "L", "M"]:
        print(f"{col}10: {ws_ap[f'{col}10'].value}")
    for r in range(6, 11):
        print(f"Row {r}: Product={ws_ap[f'B{r}'].value}, K={ws_ap[f'K{r}'].value}, L={ws_ap[f'L{r}'].value}, M={ws_ap[f'M{r}'].value}")

print("\n=== CHECKING Aerosol BOM & Job Card ===")
if os.path.exists("Aerosol/Aerosol BOM.xlsx"):
    wb_abom = openpyxl.load_workbook("Aerosol/Aerosol BOM.xlsx", data_only=False)
    ws_abom = wb_abom["Theoretical BOM"]
    print("Aerosol BOM K6:", ws_abom["K6"].value, "L6:", ws_abom["L6"].value)
    print("Aerosol BOM K7:", ws_abom["K7"].value, "L7:", ws_abom["L7"].value)

if os.path.exists("Aerosol/Aerosol_Job_Card.xlsx"):
    wb_jc = openpyxl.load_workbook("Aerosol/Aerosol_Job_Card.xlsx", data_only=False)
    ws_jc = wb_jc["Job Card"]
    print("Job Card E12:", ws_jc["E12"].value)
    print("Job Card E36:", ws_jc["E36"].value)
