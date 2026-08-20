import sys
import io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

wb = openpyxl.load_workbook("Tubex_Aug26.xlsx", data_only=True)
print("Sheetnames:", wb.sheetnames)
if "Future_Plans" in wb.sheetnames:
    ws_fp = wb["Future_Plans"]
    print("Dimensions:", ws_fp.dimensions)
    for r in range(1, 45):
        row_vals = [f"{ws_fp.cell(r, c).value}" for c in range(1, 10) if ws_fp.cell(r, c).value is not None]
        if row_vals:
            print(f"Row {r:2d}: " + " | ".join(row_vals))
