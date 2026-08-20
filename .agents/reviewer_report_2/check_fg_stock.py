import openpyxl

wb = openpyxl.load_workbook("Tubex_Aug26.xlsx", data_only=False)
ws_fg = wb["FG Stock"]
print("FG Stock dimensions:", ws_fg.dimensions)
for r in range(1, 10):
    row_vals = [f"{openpyxl.utils.get_column_letter(c)}{r}: {ws_fg.cell(r, c).value}" for c in range(1, 15)]
    print(f"Row {r}: " + " | ".join(row_vals[:6]))
    print("        " + " | ".join(row_vals[6:]))
