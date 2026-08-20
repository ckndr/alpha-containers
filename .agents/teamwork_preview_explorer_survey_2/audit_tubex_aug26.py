import openpyxl
import json
import os

wb_f = openpyxl.load_workbook('Tubex_Aug26.xlsx', data_only=False)
wb_v = openpyxl.load_workbook('Tubex_Aug26.xlsx', data_only=True)

out = []
def log(s):
    out.append(s)

log("================================================================================")
log("FORENSIC ANALYSIS: Tubex_Aug26.xlsx")
log("================================================================================")

for sname in wb_f.sheetnames:
    ws_f = wb_f[sname]
    ws_v = wb_v[sname]
    log(f"\n### SHEET: {sname} (Rows: {ws_f.max_row}, Cols: {ws_f.max_column})")
    
    # Read headers
    for r in range(1, min(6, ws_f.max_row + 1)):
        row_vals = [ws_f.cell(row=r, column=c).value for c in range(1, min(ws_f.max_column + 1, 35))]
        if any(v is not None for v in row_vals):
            log(f"Row {r} (Header/Data): {[str(v) if v is not None else '' for v in row_vals]}")
            
    # Check tables
    if hasattr(ws_f, 'tables') and ws_f.tables:
        log(f"Tables in {sname}:")
        for tname in ws_f.tables:
            tbl = ws_f.tables[tname]
            ref_str = getattr(tbl, 'ref', str(tbl))
            log(f"  Table: {tname}, Ref: {ref_str}")

    # Inspect cell formulas and inconsistencies
    # Group formulas by column to detect pattern breaks
    for c in range(1, ws_f.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(c)
        formulas = []
        constants = []
        errors = []
        for r in range(1, ws_f.max_row + 1):
            cell_f = ws_f.cell(row=r, column=c)
            cell_v = ws_v.cell(row=r, column=c)
            vf = str(cell_f.value) if cell_f.value is not None else ''
            vv = str(cell_v.value) if cell_v.value is not None else ''
            
            if vf.startswith('='):
                formulas.append((r, vf, vv))
            elif vf != '':
                constants.append((r, vf))
            
            if any(err in vv for err in ['#REF!', '#VALUE!', '#NAME?', '#DIV/0!', '#N/A', '#NULL!', '#NUM!']) or any(err in vf for err in ['#REF!', '#VALUE!', '#NAME?', '#DIV/0!', '#N/A']):
                errors.append((r, vf, vv))
                
        if errors:
            log(f"  [ERROR IN COL {col_letter}] {len(errors)} error cells:")
            for r, vf, vv in errors[:10]:
                log(f"    Cell {col_letter}{r}: Formula='{vf}', CachedVal='{vv}'")
                
        if formulas:
            log(f"  Col {col_letter}: {len(formulas)} formulas, {len(constants)} constants.")
            log(f"    Sample formula (row {formulas[0][0]}): {formulas[0][1]} -> Val: {formulas[0][2]}")
            if len(formulas) > 1:
                log(f"    Sample formula (row {formulas[-1][0]}): {formulas[-1][1]} -> Val: {formulas[-1][2]}")
            # Check if there are gaps where constants or empties interrupt formulas in middle of table
            if len(formulas) > 5 and len(constants) > 0:
                # Find if constants are between formula rows
                first_f_row = formulas[0][0]
                last_f_row = formulas[-1][0]
                interspersed = [cr for cr, cv in constants if first_f_row < cr < last_f_row]
                if interspersed:
                    log(f"    [WARNING: FORMULA INTERRUPT] Constants intermixed with formulas at rows: {interspersed}")

report_text = "\n".join(out)
with open(r'd:\Alpha\.agents\teamwork_preview_explorer_survey_2\tubex_aug26_analysis.txt', 'w', encoding='utf-8') as f:
    f.write(report_text)
print("Finished Tubex_Aug26 analysis.")
