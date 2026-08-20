import openpyxl
import re

files = [
    'Tubex_Aug26.xlsx',
    'Production.xlsx',
    'Pending.xlsx',
    'August_Plan.xlsx',
    'Aerosol/Aerosol BOM.xlsx',
    'Aerosol/Aerosol Raw Materials.xlsx',
    'Aerosol/Aerosol_Job_Card.xlsx',
    'Aerosol/Aerosol_Production_Entry.xlsx',
    'Aerosol/Tubex_v10_30.xlsx',
    'PET_SKUs.xlsx',
    'Pet Format.xlsx'
]

out = []
def log(s):
    out.append(s)

log("================================================================================")
log("EXACT RELATIVE ROW OFFSET & FORMULA DISCREPANCY AUDIT")
log("================================================================================")

for fpath in files:
    wb = openpyxl.load_workbook(fpath, data_only=False)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v and str(v).startswith('='):
                    form = str(v)
                    # Find all local un-anchored cell references like A12, F12, $A12, F$12 (not $A$12)
                    # Exclude sheet names and table names
                    # Match pattern: ([A-Z]+)(\d+) or (\$?[A-Z]+)(\$?\d+)
                    refs = re.findall(r'(?<![A-Za-z0-9_])(\$?[A-Z]{1,3})(\$?)(\d+)(?![A-Za-z0-9_])', form)
                    for col_part, row_anchor, row_num in refs:
                        # If row is not anchored with $, check if row_num != r
                        if row_anchor == '':
                            rn = int(row_num)
                            # If this is inside a row-by-row table and the reference is on the same row, rn should be r
                            # If rn is r-1, r-2, r+1, flag for inspection (excluding headers like row 1, 2 or summary rows)
                            if abs(rn - r) in [1, 2, 3] and rn > 2 and r > 3:
                                # Check if it's an offset reference
                                log(f"[{fpath} :: {sname}!{cell.coordinate}] Formula: '{form}' | Refers to row {rn} (current row is {r})")

with open(r'd:\Alpha\.agents\teamwork_preview_explorer_survey_2\formula_offset_audit.txt', 'w', encoding='utf-8') as f:
    f.write("\n".join(out))

print("Formula offset audit complete.")
