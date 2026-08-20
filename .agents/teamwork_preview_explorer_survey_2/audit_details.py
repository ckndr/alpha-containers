import openpyxl

def audit_details():
    out = []
    
    # Check Tubex_Aug26 Tubex_Dashboard formulas in detail
    wb = openpyxl.load_workbook('Tubex_Aug26.xlsx', data_only=False)
    ws_dash = wb['Tubex_Dashboard']
    out.append("=== TUBEX_AUG26: Tubex_Dashboard Analysis ===")
    for r in range(10, 68):
        c_b = ws_dash.cell(row=r, column=2).value # Type (TUBE/PET)
        c_f = ws_dash.cell(row=r, column=6).value # Product ID
        c_g = ws_dash.cell(row=r, column=7).value # Requirement
        c_h = ws_dash.cell(row=r, column=8).value # MTD Production
        c_i = ws_dash.cell(row=r, column=9).value # Balance
        c_j = ws_dash.cell(row=r, column=10).value # % Complete
        out.append(f"Row {r:2d}: Type={c_b!s:6} | PID={c_f!s:6} | Req={c_g!s:40} | Prod={c_h!s:40}")

    # Check MRP Sheet in Tubex_Aug26
    ws_mrp = wb['MRP']
    out.append("\n=== TUBEX_AUG26: MRP Sheet Analysis ===")
    for r in range(1, 20):
        vals = [ws_mrp.cell(row=r, column=c).value for c in range(1, 12)]
        out.append(f"Row {r:2d}: {vals}")
        
    out.append("\n=== TUBEX_AUG26: MRP Sheet Rows 90 to 106 ===")
    for r in range(90, min(ws_mrp.max_row + 1, 107)):
        vals = [ws_mrp.cell(row=r, column=c).value for c in range(1, 12)]
        out.append(f"Row {r:2d}: {vals}")

    # Check Production.xlsx sheets
    wb_prod = openpyxl.load_workbook('Production.xlsx', data_only=False)
    out.append("\n=== PRODUCTION.XLSX: Dashbord Sheet ===")
    ws_pdash = wb_prod['Dashbord']
    for r in range(1, 25):
        vals = [ws_pdash.cell(row=r, column=c).value for c in range(1, 15)]
        out.append(f"Row {r:2d}: {vals}")

    # Check Production.xlsx Summary 14-08-2026
    ws_psum = wb_prod['Summary 14-08-2026']
    out.append("\n=== PRODUCTION.XLSX: Summary 14-08-2026 ===")
    for r in range(1, 26):
        vals = [ws_psum.cell(row=r, column=c).value for c in range(1, 10)]
        out.append(f"Row {r:2d}: {vals}")

    # Check Pending.xlsx
    wb_pend = openpyxl.load_workbook('Pending.xlsx', data_only=False)
    ws_pend = wb_pend['01-05-2026']
    out.append("\n=== PENDING.XLSX: 01-05-2026 ===")
    for r in range(1, ws_pend.max_row + 1):
        vals = [ws_pend.cell(row=r, column=c).value for c in range(1, ws_pend.max_column + 1)]
        out.append(f"Row {r:2d}: {vals}")

    # Check August_Plan.xlsx
    wb_plan = openpyxl.load_workbook('August_Plan.xlsx', data_only=False)
    ws_atp = wb_plan['August Tubes Plan']
    out.append("\n=== AUGUST_PLAN.XLSX: August Tubes Plan ===")
    for r in range(1, 20):
        vals = [ws_atp.cell(row=r, column=c).value for c in range(1, 16)]
        out.append(f"Row {r:2d}: {vals}")

    ws_app = wb_plan['August Plan PET']
    out.append("\n=== AUGUST_PLAN.XLSX: August Plan PET ===")
    for r in range(1, 20):
        vals = [ws_app.cell(row=r, column=c).value for c in range(1, 15)]
        out.append(f"Row {r:2d}: {vals}")

    with open(r'd:\Alpha\.agents\teamwork_preview_explorer_survey_2\detailed_audit_dump.txt', 'w', encoding='utf-8') as f:
        f.write("\n".join(out))
    print("Detailed audit dump written.")

audit_details()
