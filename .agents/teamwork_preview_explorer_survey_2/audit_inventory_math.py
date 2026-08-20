import openpyxl

wb = openpyxl.load_workbook('Tubex_Aug26.xlsx', data_only=True)
ws_inv = wb['Inventory']

print("=== INVENTORY BALANCE RECONCILIATION AUDIT (Tubex_Aug26) ===")
mismatches = 0
for r in range(3, ws_inv.max_row + 1):
    item_id = ws_inv.cell(row=r, column=1).value
    item_name = ws_inv.cell(row=r, column=3).value
    opening = ws_inv.cell(row=r, column=5).value or 0
    received = ws_inv.cell(row=r, column=6).value or 0
    issued = ws_inv.cell(row=r, column=7).value or 0
    balance = ws_inv.cell(row=r, column=8).value
    wip = ws_inv.cell(row=r, column=9).value or 0
    pcs_can_produce = ws_inv.cell(row=r, column=10).value
    
    try:
        calc_bal = float(opening) + float(received) - float(issued)
        if balance is not None and balance != '':
            bal_num = float(balance)
            if abs(calc_bal - bal_num) > 0.001:
                print(f"Row {r} (Item {item_id} - {item_name}): Opening({opening}) + Rec({received}) - Iss({issued}) = {calc_bal} != Store Balance({balance})")
                mismatches += 1
    except Exception as e:
        pass

print(f"Total Stock Balance Mismatches: {mismatches}")
