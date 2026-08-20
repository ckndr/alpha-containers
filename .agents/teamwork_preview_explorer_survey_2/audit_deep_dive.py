import openpyxl
import re
import json
import os

files = [
    'Tubex_Aug26.xlsx',
    'Production.xlsx',
    'Pending.xlsx',
    'August_Plan.xlsx',
    'Aerosol/Aerosol BOM.xlsx',
    'Aerosol/Aerosol Raw Materials.xlsx',
    'Aerosol/Aerosol_Job_Card.xlsx',
    'Aerosol/Aerosol_Production_Entry.xlsx',
    'Aerosol/Tubex_v10_30.xlsx',
    'PET_SKUs.xlsx',
    'Pet Format.xlsx'
]

# Real external link pattern (contains .xls or path or bracket followed by sheet name)
real_ext_pattern = re.compile(r"(\[.*?\.xlsx?\]|'[^']*\[.*?\][^']*'|'?[A-Za-z]:\\[^']*'?)", re.IGNORECASE)
error_pattern = re.compile(r'#(REF!|VALUE!|NAME\?|DIV/0!|N/A|NULL!|NUM!)')

def analyze_workbook(fpath):
    print(f"\n=======================================================")
    print(f"ANALYZING: {fpath}")
    print(f"=======================================================")
    
    wb_f = openpyxl.load_workbook(fpath, data_only=False)
    wb_v = openpyxl.load_workbook(fpath, data_only=True)
    
    for sname in wb_f.sheetnames:
        ws_f = wb_f[sname]
        ws_v = wb_v[sname]
        print(f"\n--- Sheet: {sname} (Rows: {ws_f.max_row}, Cols: {ws_f.max_column}) ---")
        
        # Check headers (first 3 rows)
        headers = []
        for r in range(1, min(4, ws_f.max_row + 1)):
            row_vals = [str(ws_f.cell(row=r, column=c).value or '') for c in range(1, min(ws_f.max_column + 1, 30))]
            if any(row_vals):
                print(f"  Header Row {r}: {row_vals[:15]}")
        
        # Audit cells
        broken_formulas = []
        cached_errors = []
        external_refs = []
        formulas_by_col = {}
        empty_or_hardcoded_in_formula_cols = {}
        
        for r in range(1, ws_f.max_row + 1):
            for c in range(1, ws_f.max_column + 1):
                cell_f = ws_f.cell(row=r, column=c)
                cell_v = ws_v.cell(row=r, column=c)
                
                vf = str(cell_f.value) if cell_f.value is not None else ''
                vv = str(cell_v.value) if cell_v.value is not None else ''
                
                if vf.startswith('='):
                    if c not in formulas_by_col:
                        formulas_by_col[c] = []
                    formulas_by_col[c].append((r, vf))
                    
                    if error_pattern.search(vf):
                        broken_formulas.append((cell_f.coordinate, vf))
                    
                    # Check if true external link
                    if '.xls' in vf.lower() or ':\\' in vf:
                        external_refs.append((cell_f.coordinate, vf))
                        
                if error_pattern.match(vv):
                    cached_errors.append((cell_f.coordinate, vv, vf))
        
        if broken_formulas:
            print(f"  [!] BROKEN FORMULAS IN TEXT ({len(broken_formulas)}):")
            for coord, form in broken_formulas[:10]:
                print(f"      {coord}: {form}")
        else:
            print("  [OK] No broken formula strings (#REF!, #DIV/0!, etc. in formula syntax)")
            
        if cached_errors:
            print(f"  [!] CACHED ERROR VALUES ({len(cached_errors)}):")
            for coord, val, form in cached_errors[:15]:
                print(f"      {coord}: Cached={val} | Formula/Value={form}")
        else:
            print("  [OK] No cached error values")
            
        if external_refs:
            print(f"  [!] EXTERNAL FILE REFERENCES ({len(external_refs)}):")
            for coord, form in external_refs[:10]:
                print(f"      {coord}: {form}")
        else:
            print("  [OK] No external workbook file links")

        # Column formula pattern summary
        print("  Column Formula Summary:")
        for col_idx, flist in sorted(formulas_by_col.items()):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            header_name = ws_f.cell(row=1, column=col_idx).value or ws_f.cell(row=2, column=col_idx).value or f"Col_{col_letter}"
            sample_formula = flist[0][1]
            rows_with_formula = [r for r, f in flist]
            print(f"    Col {col_letter} ({header_name}): {len(flist)} formulas (e.g., row {flist[0][0]}: {sample_formula})")

for f in files:
    analyze_workbook(f)
