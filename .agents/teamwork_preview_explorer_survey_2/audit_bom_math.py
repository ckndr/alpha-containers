import openpyxl
import json

out = []
def log(s):
    out.append(s)

log("================================================================================")
log("DEEP FORENSIC AUDIT: BOM & SCRAP FACTOR MATHEMATICAL MODELS")
log("================================================================================")

# 1. Aerosol BOM.xlsx
log("\n--- 1. Aerosol/Aerosol BOM.xlsx ---")
wb = openpyxl.load_workbook('Aerosol/Aerosol BOM.xlsx', data_only=False)
ws = wb['Theoretical BOM']
log("Theoretical BOM Rows:")
for r in range(2, ws.max_row + 1):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    log(f"Row {r}: {row_vals}")

ws_calc = wb['Req. Calculator']
log("\nReq. Calculator Rows:")
for r in range(1, ws_calc.max_row + 1):
    row_vals = [ws_calc.cell(row=r, column=c).value for c in range(1, ws_calc.max_column + 1)]
    log(f"Row {r}: {row_vals}")

# 2. Aerosol_Job_Card.xlsx
log("\n--- 2. Aerosol/Aerosol_Job_Card.xlsx ---")
wb_jc = openpyxl.load_workbook('Aerosol/Aerosol_Job_Card.xlsx', data_only=False)
ws_jc = wb_jc['Job Card']
log("Job Card Rows:")
for r in range(1, ws_jc.max_row + 1):
    row_vals = [ws_jc.cell(row=r, column=c).value for c in range(1, ws_jc.max_column + 1)]
    if any(v is not None for v in row_vals):
        log(f"Row {r}: {row_vals}")

ws_abom = wb_jc['Aerosol_BOM']
log("\nAerosol_BOM Sheet in Job Card Rows:")
for r in range(1, ws_abom.max_row + 1):
    row_vals = [ws_abom.cell(row=r, column=c).value for c in range(1, ws_abom.max_column + 1)]
    log(f"Row {r}: {row_vals}")

# 3. Aerosol_Production_Entry.xlsx
log("\n--- 3. Aerosol/Aerosol_Production_Entry.xlsx ---")
wb_pe = openpyxl.load_workbook('Aerosol/Aerosol_Production_Entry.xlsx', data_only=False)
ws_pe = wb_pe['Data_Entry']
log("Data_Entry Sample Rows:")
for r in range(1, 10):
    row_vals = [ws_pe.cell(row=r, column=c).value for c in range(1, ws_pe.max_column + 1)]
    log(f"Row {r}: {row_vals}")

ws_pm = wb_pe['Product_Master']
log("Product_Master Sample Rows:")
for r in range(1, 10):
    row_vals = [ws_pm.cell(row=r, column=c).value for c in range(1, ws_pm.max_column + 1)]
    log(f"Row {r}: {row_vals}")

# 4. Tubex_Aug26.xlsx BOM Sheet & MRP Sheet
log("\n--- 4. Tubex_Aug26.xlsx BOM Sheet Analysis ---")
wb_t = openpyxl.load_workbook('Tubex_Aug26.xlsx', data_only=False)
ws_tbom = wb_t['BOM']
log("BOM sample rows:")
for r in range(2, 20):
    row_vals = [ws_tbom.cell(row=r, column=c).value for c in range(1, ws_tbom.max_column + 1)]
    log(f"Row {r}: {row_vals}")

# Check scrap percentages across all BOM rows in Tubex_Aug26
log("\nDistinct Scrap % values in Tubex_Aug26 BOM:")
scrap_vals = {}
uom_vals = {}
cat_vals = {}
for r in range(3, ws_tbom.max_row + 1):
    pid = ws_tbom.cell(row=r, column=1).value
    cat = ws_tbom.cell(row=r, column=6).value
    uom = ws_tbom.cell(row=r, column=9).value
    per1000 = ws_tbom.cell(row=r, column=10).value
    scrap = ws_tbom.cell(row=r, column=12).value
    
    scrap_vals[scrap] = scrap_vals.get(scrap, 0) + 1
    uom_vals[uom] = uom_vals.get(uom, 0) + 1
    cat_vals[cat] = cat_vals.get(cat, 0) + 1

log(f"Scrap values count: {scrap_vals}")
log(f"UOM values count: {uom_vals}")
log(f"Material Categories count: {cat_vals}")

with open(r'd:\Alpha\.agents\teamwork_preview_explorer_survey_2\bom_math_analysis.txt', 'w', encoding='utf-8') as f:
    f.write("\n".join(out))

print("Finished BOM math analysis.")
