import openpyxl
import re

files = [
    'Tubex_Aug26.xlsx',
    'Production.xlsx',
    'Pending.xlsx',
    'August_Plan.xlsx',
    'Aerosol/Aerosol BOM.xlsx',
    'Aerosol/Aerosol Raw Materials.xlsx',
    'Aerosol/Aerosol_Job_Card.xlsx',
    'Aerosol/Aerosol_Production_Entry.xlsx',
    'Aerosol/Tubex_v10_30.xlsx',
    'PET_SKUs.xlsx',
    'Pet Format.xlsx'
]

out = []
def log(s):
    out.append(s)

log("================================================================================")
log("COMPREHENSIVE MATHEMATICAL & DATA INTEGRITY AUDIT")
log("================================================================================")

for fpath in files:
    wb_f = openpyxl.load_workbook(fpath, data_only=False)
    wb_v = openpyxl.load_workbook(fpath, data_only=True)
    
    log(f"\n==================================================")
    log(f"FILE: {fpath}")
    log(f"==================================================")
    
    for sname in wb_f.sheetnames:
        ws_f = wb_f[sname]
        ws_v = wb_v[sname]
        
        # 1. Column formula consistency analysis
        for c in range(1, ws_f.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(c)
            row_formulas = {}
            for r in range(1, ws_f.max_row + 1):
                val_f = ws_f.cell(row=r, column=c).value
                if val_f and str(val_f).startswith('='):
                    row_formulas[r] = str(val_f)
                    
            if len(row_formulas) > 3:
                # Analyze formula structure/pattern variations
                # Normalize relative row numbers
                patterns = {}
                for r, form in row_formulas.items():
                    # Replace current row number in formula with {R}
                    # E.g., A12 -> A{R}, F12 -> F{R}
                    norm = re.sub(rf'(?<!\d){r}(?!\d)', '{R}', form)
                    if norm not in patterns:
                        patterns[norm] = []
                    patterns[norm].append(r)
                
                if len(patterns) > 1:
                    log(f"  [FORMULA PATTERN ANOMALY] Sheet '{sname}' Col {col_letter}: {len(patterns)} distinct formula patterns:")
                    for pat, rlist in patterns.items():
                        log(f"    Pattern ({len(rlist)} rows, e.g. {rlist[:5]}): {pat}")

        # 2. Check for number stored as text in critical ID/qty columns
        type_mismatches = []
        for r in range(2, ws_f.max_row + 1):
            for c in range(1, ws_f.max_column + 1):
                cell_v = ws_v.cell(row=r, column=c)
                v = cell_v.value
                if isinstance(v, str):
                    v_str = v.strip()
                    if v_str.isdigit() and len(v_str) < 10:
                        header = str(ws_f.cell(row=1, column=c).value or ws_f.cell(row=2, column=c).value or '')
                        # Check if header implies numeric quantity or ID
                        if any(k in header.lower() for k in ['qty', 'quantity', 'balance', 'pcs', 'order', 'stock', 'sr']):
                            type_mismatches.append((cell_v.coordinate, header, v))
                            
        if type_mismatches:
            log(f"  [NUMERIC STORED AS TEXT] Sheet '{sname}': {len(type_mismatches)} instances (e.g. {type_mismatches[:5]})")

with open(r'd:\Alpha\.agents\teamwork_preview_explorer_survey_2\formula_integrity_check.txt', 'w', encoding='utf-8') as f:
    f.write("\n".join(out))

print("Formula integrity check finished.")
