import openpyxl
import re
import json
import os

files = [
    'Tubex_Aug26.xlsx',
    'Production.xlsx',
    'Pending.xlsx',
    'August_Plan.xlsx',
    'Aerosol/Aerosol BOM.xlsx',
    'Aerosol/Aerosol Raw Materials.xlsx',
    'Aerosol/Aerosol_Job_Card.xlsx',
    'Aerosol/Aerosol_Production_Entry.xlsx',
    'Aerosol/Tubex_v10_30.xlsx',
    'PET_SKUs.xlsx',
    'Pet Format.xlsx'
]

error_pattern = re.compile(r'#(REF!|VALUE!|NAME\?|DIV/0!|N/A|NULL!|NUM!)')
external_pattern = re.compile(r'\[.*?\]')

findings = {}

for fpath in files:
    try:
        print(f"Auditing {fpath}...")
        wb_f = openpyxl.load_workbook(fpath, data_only=False)
        wb_v = openpyxl.load_workbook(fpath, data_only=True)
        file_report = {
            'sheets': {},
            'broken_formulas': [],
            'cached_errors': [],
            'external_links': [],
            'volatile_formulas': [],
            'formula_inconsistencies': []
        }
        
        for sname in wb_f.sheetnames:
            ws_f = wb_f[sname]
            ws_v = wb_v[sname]
            sheet_report = {
                'max_row': ws_f.max_row,
                'max_col': ws_f.max_column,
                'formula_count': 0
            }
            
            for r in range(1, ws_f.max_row + 1):
                for c in range(1, ws_f.max_column + 1):
                    cell_f = ws_f.cell(row=r, column=c)
                    cell_v = ws_v.cell(row=r, column=c)
                    
                    val_f = str(cell_f.value) if cell_f.value is not None else ''
                    val_v = str(cell_v.value) if cell_v.value is not None else ''
                    
                    is_formula = val_f.startswith('=')
                    if is_formula:
                        sheet_report['formula_count'] += 1
                        
                        # Check error patterns in formula text
                        m_err = error_pattern.search(val_f)
                        if m_err:
                            file_report['broken_formulas'].append({
                                'sheet': sname,
                                'cell': cell_f.coordinate,
                                'formula': val_f,
                                'error_match': m_err.group(0)
                            })
                            
                        # Check external references
                        m_ext = external_pattern.search(val_f)
                        if m_ext:
                            file_report['external_links'].append({
                                'sheet': sname,
                                'cell': cell_f.coordinate,
                                'formula': val_f,
                                'external_target': m_ext.group(0)
                            })
                            
                        # Check volatile functions
                        volatile_list = ['INDIRECT(', 'OFFSET(', 'TODAY(', 'NOW(', 'RAND(', 'RANDBETWEEN(']
                        if any(fn in val_f.upper() for fn in volatile_list):
                            file_report['volatile_formulas'].append({
                                'sheet': sname,
                                'cell': cell_f.coordinate,
                                'formula': val_f
                            })
                    
                    # Check cached error values
                    if error_pattern.match(val_v):
                        file_report['cached_errors'].append({
                            'sheet': sname,
                            'cell': cell_f.coordinate,
                            'cached_value': val_v,
                            'formula': val_f if is_formula else None
                        })
            
            file_report['sheets'][sname] = sheet_report
            
        findings[fpath] = file_report
    except Exception as e:
        findings[fpath] = {'error': str(e)}

out_file = os.path.join(os.path.dirname(__file__), 'raw_audit_findings.json')
with open(out_file, 'w', encoding='utf-8') as f:
    json.dump(findings, f, indent=2)

print(f"Audit completed. Summary written to {out_file}")
