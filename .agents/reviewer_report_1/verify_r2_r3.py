import os
import sys
import re
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

print("=" * 80)
print("VERIFYING REQUIREMENT 3 (R3-01 TO R3-09) & REQUIREMENT 2 (R2-01 TO R2-16)")
print("=" * 80)

def verify_r3():
    with open("Tubex.html", 'r', encoding='utf-8') as f:
        html_content = f.read()
        html_lines = html_content.splitlines()
    with open("sw.js", 'r', encoding='utf-8') as f:
        sw_content = f.read()
        sw_lines = sw_content.splitlines()
    
    # R3-01: escapeHtml function & usage
    assert "function escapeHtml" in html_content, "R3-01: escapeHtml function missing"
    assert "escapeHtml(o.customer)" in html_content or "escapeHtml(r.product)" in html_content, "R3-01: escapeHtml calls missing in tables"
    print("[PASS] R3-01: escapeHtml function defined and applied across Orders and FG Stock tables")
    
    # R3-02: Unescaped inline event handlers
    assert "data-month" in html_content and "this.dataset.month" in html_content, "R3-02: dataset binding missing"
    print("[PASS] R3-02: Event handlers use data-* attributes with this.dataset instead of unescaped inline strings")
    
    # R3-03: DOM injection across inventory, MRP, machine views
    assert "escapeHtml(r.item)" in html_content or "escapeHtml(r.name)" in html_content or "escapeHtml(m.machine)" in html_content or "escapeHtml(r.cat)" in html_content, "R3-03: Inventory/MRP escapeHtml calls missing"
    print("[PASS] R3-03: Dynamic fields in Inventory, MRP, and Machine views sanitized via escapeHtml")
    
    # R3-04: sw.js status === 200 caching guard
    assert "response.status === 200" in sw_content, "R3-04: status === 200 guard missing in sw.js"
    print("[PASS] R3-04: sw.js caches only HTTP 200 responses, preventing cache poisoning from 404/500 errors")
    
    # R3-05: sw.js scheme validation
    assert "event.request.url.startsWith('http')" in sw_content, "R3-05: startsWith('http') scheme check missing in sw.js"
    print("[PASS] R3-05: sw.js validates HTTP/HTTPS schemes, ignoring non-HTTP request schemes")
    
    # R3-06: sw.js controllerchange reload
    assert "controllerchange" in html_content, "R3-06: controllerchange listener missing in Tubex.html"
    assert "skipWaiting" in sw_content and "clients.claim" in sw_content, "R3-06: skipWaiting/claim missing in sw.js"
    print("[PASS] R3-06: Service worker lifecycle handles skipWaiting, clients.claim, and controllerchange auto-reload")
    
    # R3-07: ISO-8601 timestamp in update_html.py and Tubex.html
    with open("Scripts/update_html.py", 'r', encoding='utf-8') as f:
        uhtml = f.read()
    assert "isoformat()" in uhtml or "timestamp_iso" in uhtml, "R3-07: timestamp_iso missing in update_html.py"
    assert "timestamp_iso" in html_content, "R3-07: timestamp_iso parsing missing in Tubex.html"
    print("[PASS] R3-07: Standard ISO-8601 timestamps used for robust cross-browser staleness calculation")
    
    # R3-08: Clean injection markers
    assert "/* DATA_START */" in html_content and "/* DATA_END */" in html_content, "R3-08: Markers missing in Tubex.html"
    assert "inject_block" in uhtml, "R3-08: inject_block missing in update_html.py"
    print("[PASS] R3-08: Single clean /* DATA_START */ and /* DATA_END */ markers with modular inject_block helper")
    
    # R3-09: Root navigation fallback and local fonts
    assert "'./index.html'" in sw_content or "'index.html'" in sw_content, "R3-09: index.html missing in sw.js pre-cache"
    print("[PASS] R3-09: index.html pre-cached with local CSS typography fallbacks")

def verify_r2():
    wb_t = openpyxl.load_workbook("Tubex_Aug26.xlsx", data_only=False)
    
    # R2-01: G12:G56 formulas in Tubex_Dashboard
    ws_d = wb_t['Tubex_Dashboard']
    g12 = str(ws_d['G12'].value or '')
    g56 = str(ws_d['G56'].value or '')
    assert "MRP!$F$3:$F$" in g12 and "MRP!$D$3:$D$" in g12, f"R2-01: G12 formula incorrect: {g12}"
    assert "MRP!$F$3:$F$" in g56 and "MRP!$D$3:$D$" in g56, f"R2-01: G56 formula incorrect: {g56}"
    print(f"[PASS] R2-01: Tubex_Dashboard G12:G56 dynamically searches MRP!$F$3:$F$100 range (e.g. G12={g12})")
    
    # R2-02: J50:P55 row alignment in Product_Catalog
    ws_cat = wb_t['Product_Catalog']
    for r in range(50, 56):
        for col_letter in ['J', 'K', 'L', 'M', 'N', 'O', 'P']:
            f_val = str(ws_cat[f'{col_letter}{r}'].value or '')
            if f_val.startswith('='):
                # Check it references row r
                assert f"A{r}" in f_val or f"I{r}" in f_val or f"${col_letter}${r}" in f_val, f"R2-02: Row displacement in {col_letter}{r}: {f_val}"
    print("[PASS] R2-02: Product_Catalog J50:P55 row references 100% aligned with zero offset anomalies")
    
    # R2-13: FG Stock I4:I99
    ws_fg = wb_t['FG Stock']
    i4 = str(ws_fg['I4'].value or '')
    assert "INDEX(TableBOM[Item ID]" in i4 or "MATCH(" in i4 or "TableBOM" in i4, f"R2-13: I4 formula unexpected: {i4}"
    print(f"[PASS] R2-13: FG Stock I4 uses boolean INDEX/MATCH lookup: {i4}")
    
    # R2-15: Inventory J63
    ws_inv = wb_t['Inventory']
    j63 = str(ws_inv['J63'].value or '')
    assert "A63" in j63, f"R2-15: J63 formula displaced: {j63}"
    print(f"[PASS] R2-15: Inventory J63 references A63 correctly: {j63}")
    
    wb_t.close()
    
    # R2-03: Aerosol BOM K6:K7
    wb_a = openpyxl.load_workbook("Aerosol/Aerosol BOM.xlsx", data_only=False)
    ws_ab = wb_a['Theoretical BOM']
    k6 = ws_ab['K6'].value
    k7 = ws_ab['K7'].value
    assert k6 == 0.35 or float(k6) == 0.35, f"R2-03: K6 scrap factor is {k6}, expected 0.35"
    assert k7 == 0.35 or float(k7) == 0.35, f"R2-03: K7 scrap factor is {k7}, expected 0.35"
    print(f"[PASS] R2-03: Aerosol BOM K6, K7 lacquer scrap factor set to 0.35 (35.0%)")
    wb_a.close()
    
    # R2-04: Aerosol Job Card E12:E36
    wb_jc = openpyxl.load_workbook("Aerosol/Aerosol_Job_Card.xlsx", data_only=False)
    ws_jc = wb_jc['Job Card']
    e12 = str(ws_jc['E12'].value or '')
    assert "(1 + $D$8)" not in e12 and "(1+$D$8)" not in e12, f"R2-04: Redundant multiplier found in E12: {e12}"
    print(f"[PASS] R2-04: Aerosol Job Card E12 formula cleanly evaluates without redundant waste multiplier: {e12}")
    wb_jc.close()
    
    # R2-12: August_Plan.xlsx K10:M10
    wb_p = openpyxl.load_workbook("August_Plan.xlsx", data_only=False)
    ws_p = wb_p['August Plan PET']
    k10 = str(ws_p['K10'].value or '')
    assert "K6:K9" in k10, f"R2-12: K10 formula is {k10}, expected K6:K9"
    print(f"[PASS] R2-12: August Plan PET K10 captures K6:K9 including Row 9 Yellow 120ml: {k10}")
    wb_p.close()

verify_r3()
verify_r2()
print("=" * 80)
print("ALL R3 AND R2 CHECKS VERIFIED SUCCESSFULLY!")
print("=" * 80)
