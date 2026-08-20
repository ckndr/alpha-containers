import os
import sys
import glob
import openpyxl
import re

sys.stdout.reconfigure(encoding='utf-8')

print("=" * 80)
print("VERIFYING EXCEL WORKBOOKS FORMULA INTEGRITY (15 WORKBOOKS)")
print("=" * 80)

workbooks = [
    ("Active Production Model", "Tubex_Aug26.xlsx"),
    ("Production Planning", "August_Plan.xlsx"),
    ("PET SKU Reference", "PET_SKUs.xlsx"),
    ("PET Format Reference", "Pet Format.xlsx"),
    ("Master BOM Catalog", "Aerosol/Aerosol BOM.xlsx"),
    ("Material Stock Model", "Aerosol/Aerosol Raw Materials.xlsx"),
    ("Job Card Model", "Aerosol/Aerosol_Job_Card.xlsx"),
    ("Production Entry", "Aerosol/Aerosol_Production_Entry.xlsx"),
    ("Historical Archives", "Tubex Records/Dashboard_Archive.xlsx"),
    ("Historical Archives", "Tubex Records/Production_Archive.xlsx"),
    ("Historical Orders", "Tubex Records/Samsol PET Orders.xlsx"),
    ("Historical Production", "Tubex Records/Samsol_Production.xlsx"),
    ("Shop-Floor Input", "Production.xlsx"),
    ("Legacy Baseline (Closed)", "Aerosol/Tubex_v10_30.xlsx"),
    ("Legacy Archive (Closed)", "Tubex Records/Tubex_July26.xlsx")
]

error_keywords = ['#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A']

for category, rel_path in workbooks:
    full_path = os.path.join('d:\\Alpha', rel_path)
    if not os.path.exists(full_path):
        print(f"[MISSING] {category:25s} | {rel_path}")
        continue
    
    try:
        wb = openpyxl.load_workbook(full_path, data_only=False)
        sheet_count = len(wb.sheetnames)
        formula_count = 0
        active_errors = 0
        error_details = []
        
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows():
                for cell in row:
                    val = str(cell.value or '')
                    if val.startswith('='):
                        formula_count += 1
                        for err in error_keywords:
                            if err in val:
                                active_errors += 1
                                error_details.append(f"{sname}!{cell.coordinate}: {val}")
        
        wb.close()
        status = "PASS" if active_errors == 0 else f"{active_errors} ERRORS"
        print(f"[{status:7s}] {category:24s} | {rel_path:36s} | Sheets: {sheet_count:2d} | Formulas: {formula_count:5d} | Active Errors: {active_errors}")
        if error_details:
            for ed in error_details[:5]:
                print(f"          -> {ed}")
    except Exception as e:
        print(f"[ERROR]   {category:24s} | {rel_path:36s} | Failed to inspect: {e}")

print("=" * 80)
