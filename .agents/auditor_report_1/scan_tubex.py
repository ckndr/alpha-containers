import openpyxl

wb = openpyxl.load_workbook('d:/Alpha/Tubex_Aug26.xlsx', data_only=False)
wb_data = openpyxl.load_workbook('d:/Alpha/Tubex_Aug26.xlsx', data_only=True)

formula_count = 0
formula_errors = []
cached_errors = []

for sname in wb.sheetnames:
    ws = wb[sname]
    ws_d = wb_data[sname]
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                formula_count += 1
                for err in ['#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A']:
                    if err in val:
                        formula_errors.append((sname, cell.coordinate, val, err))
            c_d = ws_d[cell.coordinate]
            if str(c_d.value) in ['#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A']:
                cached_errors.append((sname, cell.coordinate, str(c_d.value)))

print(f"Sheets: {len(wb.sheetnames)}")
print(f"Total Formulas: {formula_count}")
print(f"Formula Errors: {len(formula_errors)}")
print(f"Cached Value Errors: {len(cached_errors)}")
if formula_errors:
    print("Formula Errors found:", formula_errors)
if cached_errors:
    print("Cached Value Errors found:", cached_errors)
