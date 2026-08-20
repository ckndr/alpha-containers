import glob
import os
import openpyxl

files = [
    "d:/Alpha/Tubex_Aug26.xlsx",
    "d:/Alpha/August_Plan.xlsx",
    "d:/Alpha/Aerosol/Aerosol BOM.xlsx",
    "d:/Alpha/Aerosol/Aerosol_Job_Card.xlsx",
    "d:/Alpha/Aerosol/Aerosol Raw Materials.xlsx",
    "d:/Alpha/Aerosol/Aerosol_Production_Entry.xlsx",
    "d:/Alpha/PET_SKUs.xlsx",
    "d:/Alpha/Pet Format.xlsx",
    "d:/Alpha/Tubex Records/Dashboard_Archive.xlsx",
    "d:/Alpha/Tubex Records/Production_Archive.xlsx",
    "d:/Alpha/Tubex Records/Samsol PET Orders.xlsx",
    "d:/Alpha/Tubex Records/Samsol_Production_and_Dispatch.xlsx"
]

print("="*80)
print("  MULTI-WORKBOOK EXHAUSTIVE FORMULA & DATA SCAN")
print("="*80)

for fpath in files:
    if not os.path.exists(fpath):
        continue
    wb = openpyxl.load_workbook(fpath, data_only=False)
    wb_d = openpyxl.load_workbook(fpath, data_only=True)
    f_count = 0
    f_errs = []
    c_errs = []
    for s in wb.sheetnames:
        ws = wb[s]
        ws_d = wb_d[s]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith('='):
                    f_count += 1
                    for err in ['#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A']:
                        if err in v:
                            f_errs.append((s, c.coordinate, v, err))
                cd = ws_d[c.coordinate]
                if str(cd.value) in ['#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A']:
                    c_errs.append((s, c.coordinate, str(cd.value)))
    print(f"{os.path.basename(fpath):35s} | Sheets: {len(wb.sheetnames):2d} | Formulas: {f_count:5d} | Formula Errs: {len(f_errs)} | Cached Errs: {len(c_errs)}")
