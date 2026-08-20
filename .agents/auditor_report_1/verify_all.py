import os
import sys
import glob
import py_compile
import subprocess
import openpyxl

def banner(title):
    print("\n" + "="*80)
    print(f"  {title}")
    print("="*80)

def check_compilation():
    banner("CHECK 1: PYTHON SCRIPT COMPILATION (Scripts/*.py and Root)")
    py_files = glob.glob("d:/Alpha/Scripts/*.py") + glob.glob("d:/Alpha/*.py")
    print(f"Found {len(py_files)} Python files to compile.")
    compiled_ok = 0
    errors = []
    for f in py_files:
        try:
            py_compile.compile(f, doraise=True)
            compiled_ok += 1
        except Exception as e:
            errors.append((f, str(e)))
    print(f"Compiled successfully: {compiled_ok}/{len(py_files)}")
    if errors:
        for f, err in errors:
            print(f"  [ERROR] {f}: {err}")
    else:
        print("  [PASS] All Python scripts compiled with zero syntax/indentation errors.")
    return len(errors) == 0

def check_excel_formulas():
    banner("CHECK 2: CROSS-WORKBOOK FORMULA INTEGRITY SCAN")
    workbooks = [
        "d:/Alpha/Tubex_Aug26.xlsx",
        "d:/Alpha/August_Plan.xlsx",
        "d:/Alpha/PET_SKUs.xlsx",
        "d:/Alpha/Pet Format.xlsx",
        "d:/Alpha/Aerosol/Aerosol BOM.xlsx",
        "d:/Alpha/Aerosol/Aerosol Raw Materials.xlsx",
        "d:/Alpha/Aerosol/Aerosol_Job_Card.xlsx",
        "d:/Alpha/Aerosol/Aerosol_Production_Entry.xlsx",
        "d:/Alpha/Production.xlsx"
    ] + glob.glob("d:/Alpha/Tubex Records/*.xlsx")

    results = []
    for wb_path in workbooks:
        if not os.path.exists(wb_path):
            print(f"  [SKIP] Not found: {wb_path}")
            continue
        try:
            wb = openpyxl.load_workbook(wb_path, data_only=False)
            wb_data = openpyxl.load_workbook(wb_path, data_only=True)
            formula_count = 0
            formula_errors = []
            cached_errors = []
            
            for sname in wb.sheetnames:
                ws = wb[sname]
                ws_data = wb_data[sname] if sname in wb_data.sheetnames else None
                for row in ws.iter_rows():
                    for cell in row:
                        val = cell.value
                        if isinstance(val, str) and val.startswith('='):
                            formula_count += 1
                            for err in ['#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A']:
                                if err in val:
                                    formula_errors.append((sname, cell.coordinate, val, err))
                        if ws_data:
                            c_data = ws_data[cell.coordinate]
                            dval = str(c_data.value) if c_data.value is not None else ""
                            if dval in ['#REF!', '#VALUE!', '#DIV/0!', '#NAME?', '#N/A']:
                                cached_errors.append((sname, cell.coordinate, dval))
                                
            results.append({
                "path": wb_path,
                "sheets": len(wb.sheetnames),
                "formulas": formula_count,
                "formula_errors": formula_errors,
                "cached_errors": cached_errors
            })
            print(f"  Workbook: {os.path.basename(wb_path)} | Sheets: {len(wb.sheetnames)} | Formulas: {formula_count} | Formula Errors: {len(formula_errors)} | Cached Errors: {len(cached_errors)}")
            if formula_errors:
                for s, coord, fstr, err in formula_errors[:5]:
                    print(f"    Formula Error in {s}!{coord}: {fstr} ({err})")
            if cached_errors and "Production.xlsx" not in wb_path and "July26" not in wb_path:
                for s, coord, dval in cached_errors[:5]:
                    print(f"    Cached Error in {s}!{coord}: {dval}")
        except Exception as e:
            print(f"  [ERROR reading {wb_path}]: {e}")

def check_com_processes():
    banner("CHECK 3: EXCEL COM PROCESS MONITORING")
    cmd = ["powershell", "-Command", "(Get-Process EXCEL -ErrorAction SilentlyContinue).Count"]
    res = subprocess.run(cmd, capture_output=True, text=True)
    count = res.stdout.strip()
    if not count:
        count = "0"
    print(f"Active EXCEL.EXE processes currently running: {count}")
    return int(count)

if __name__ == '__main__':
    check_compilation()
    check_excel_formulas()
    check_com_processes()
