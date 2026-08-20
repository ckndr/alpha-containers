import openpyxl

wb = openpyxl.load_workbook('d:/Alpha/Aerosol/Aerosol_Job_Card.xlsx', data_only=False)
ws = wb['Aerosol_BOM']

print("Aerosol_BOM column mapping:")
for col_idx in range(1, 16):
    header = ws.cell(1, col_idx).value
    val_row2 = ws.cell(2, col_idx).value
    print(f"Col {col_idx}: Header = '{header}', Row 2 = '{val_row2}'")

# Check the proposed formula:
# VLOOKUP(..., 10) -> Col 10 is UOM ("kg")
# VLOOKUP(..., 11) -> Col 11 is Net Qty / 1000 (20)
# VLOOKUP(..., 12) -> Col 12 is Waste + Tolerance (0.1)
# VLOOKUP(..., 13) -> Col 13 is Gross Qty / 1000 (22.2222)
