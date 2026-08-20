import openpyxl

wb = openpyxl.load_workbook('d:/Alpha/Tubex_Aug26.xlsx', data_only=False)
ws_cat = wb['Product_Catalog']

print("Product_Catalog row 2 headers:")
for c in range(1, 18):
    print(f"Col {c} ({openpyxl.utils.get_column_letter(c)}): {ws_cat.cell(2, c).value}")

print("\nProduct_Catalog row 50 (PID 8009):")
for c in range(10, 17):
    print(f"Col {c} ({openpyxl.utils.get_column_letter(c)}): {ws_cat.cell(50, c).value}")
