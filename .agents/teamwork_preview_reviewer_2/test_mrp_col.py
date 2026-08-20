import openpyxl

wb = openpyxl.load_workbook('d:/Alpha/Tubex_Aug26.xlsx', data_only=False)
ws_mrp = wb['MRP']

print("MRP headers row 2:")
for c in range(1, 15):
    print(f"Col {c} ({openpyxl.utils.get_column_letter(c)}): {ws_mrp.cell(2, c).value}")

print("\nMRP row 3:")
for c in range(1, 15):
    print(f"Col {c} ({openpyxl.utils.get_column_letter(c)}): {ws_mrp.cell(3, c).value}")
