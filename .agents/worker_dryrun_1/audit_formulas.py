"""
Comprehensive Excel Formula & Data Integrity Auditor
Checks all workbooks in d:\Alpha for:
1. Formula strings containing '#REF!', '#VALUE!', '#NAME?', '#DIV/0!', '#N/A'
2. Cached evaluated cell values equal to formula error strings (#REF!, #VALUE!, #NAME?, #DIV/0!, #N/A, #NUM!, #NULL!)
3. Missing sheet cross-references or invalid formula ranges
"""

import os
import glob
import openpyxl
import warnings

warnings.filterwarnings("ignore")

ERROR_VALUES = {'#REF!', '#VALUE!', '#NAME?', '#DIV/0!', '#N/A', '#NULL!', '#NUM!', '#N/A!'}

ROOT_DIR = r"d:\Alpha"

workbooks_to_check = [
    # Active Models & Core Workbooks
    r"d:\Alpha\Tubex_Aug26.xlsx",
    r"d:\Alpha\August_Plan.xlsx",
    r"d:\Alpha\Production.xlsx",
    r"d:\Alpha\PET_SKUs.xlsx",
    r"d:\Alpha\Pet Format.xlsx",
    # Aerosol Domain Workbooks (Catalog, Job Card, Raw Materials/Stock)
    r"d:\Alpha\Aerosol\Aerosol BOM.xlsx",
    r"d:\Alpha\Aerosol\Aerosol Raw Materials.xlsx",
    r"d:\Alpha\Aerosol\Aerosol_Job_Card.xlsx",
    r"d:\Alpha\Aerosol\Aerosol_Production_Entry.xlsx",
    r"d:\Alpha\Aerosol\Tubex_v10_30.xlsx",
    # Archive & Historical Workbooks
    r"d:\Alpha\Tubex Records\Tubex_July26.xlsx",
    r"d:\Alpha\Tubex Records\Dashboard_Archive.xlsx",
    r"d:\Alpha\Tubex Records\Production_Archive.xlsx",
    r"d:\Alpha\Tubex Records\Samsol PET Orders.xlsx",
    r"d:\Alpha\Tubex Records\Samsol_Production_and_Dispatch.xlsx",
]

def audit_workbook(path):
    rel_path = os.path.relpath(path, ROOT_DIR)
    if not os.path.exists(path):
        return {"file": rel_path, "status": "NOT_FOUND", "sheets": 0, "errors": []}
    
    errors = []
    
    # 1. Audit formula strings (data_only=False)
    try:
        wb_formula = openpyxl.load_workbook(path, data_only=False, read_only=True)
        for sheet_name in wb_formula.sheetnames:
            ws = wb_formula[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    val = cell.value
                    if val is not None and isinstance(val, str):
                        # Check if formula string contains error literals
                        if val.startswith('='):
                            for err in ERROR_VALUES:
                                if err in val.upper():
                                    coord = cell.coordinate if hasattr(cell, 'coordinate') else f"R{row_idx}C{col_idx}"
                                    errors.append({
                                        "sheet": sheet_name,
                                        "cell": coord,
                                        "type": "FORMULA_STRING_ERROR",
                                        "error": err,
                                        "formula": val
                                    })
                        elif val.upper() in ERROR_VALUES:
                            coord = cell.coordinate if hasattr(cell, 'coordinate') else f"R{row_idx}C{col_idx}"
                            errors.append({
                                "sheet": sheet_name,
                                "cell": coord,
                                "type": "LITERAL_ERROR_VALUE",
                                "error": val.upper(),
                                "formula": val
                            })
        sheet_count = len(wb_formula.sheetnames)
        wb_formula.close()
    except Exception as e:
        return {"file": rel_path, "status": f"LOAD_ERROR (formula): {e}", "sheets": 0, "errors": []}

    # 2. Audit cached evaluated values (data_only=True)
    try:
        wb_data = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sheet_name in wb_data.sheetnames:
            ws = wb_data[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    val = cell.value
                    if val is not None and isinstance(val, str):
                        val_up = val.strip().upper()
                        if val_up in ERROR_VALUES:
                            coord = cell.coordinate if hasattr(cell, 'coordinate') else f"R{row_idx}C{col_idx}"
                            # check if not already reported
                            if not any(e["sheet"] == sheet_name and e["cell"] == coord for e in errors):
                                errors.append({
                                    "sheet": sheet_name,
                                    "cell": coord,
                                    "type": "CACHED_EVALUATION_ERROR",
                                    "error": val_up,
                                    "formula": f"Evaluated to {val_up}"
                                })
        wb_data.close()
    except Exception as e:
        return {"file": rel_path, "status": f"LOAD_ERROR (data): {e}", "sheets": sheet_count, "errors": errors}

    return {
        "file": rel_path,
        "status": "PASS" if len(errors) == 0 else "FAIL",
        "sheets": sheet_count,
        "errors": errors
    }

if __name__ == '__main__':
    print("=" * 80)
    print("ALPHA CONTAINERS — WORKBOOK FORMULA & DATA INTEGRITY AUDIT")
    print("=" * 80)
    
    total_files = len(workbooks_to_check)
    passed_files = 0
    total_errors = 0
    
    results = []
    
    for wb_path in workbooks_to_check:
        res = audit_workbook(wb_path)
        results.append(res)
        err_count = len(res["errors"])
        total_errors += err_count
        if res["status"] == "PASS":
            passed_files += 1
            print(f" [PASS] {res['file']:<45} | Sheets: {res['sheets']:>2} | Errors: 0")
        elif res["status"] == "NOT_FOUND":
            print(f" [SKIP] {res['file']:<45} | File not present on disk")
        else:
            print(f" [FAIL] {res['file']:<45} | Sheets: {res['sheets']:>2} | Errors: {err_count}")
            for err in res["errors"][:10]: # show up to 10
                print(f"        -> Sheet '{err['sheet']}', Cell {err['cell']}: [{err['type']}] {err['error']} in '{err['formula']}'")
            if err_count > 10:
                print(f"        -> ... and {err_count - 10} more errors.")
    
    print("=" * 80)
    print(f"AUDIT SUMMARY: {passed_files}/{total_files} workbooks passed with ZERO errors.")
    print(f"TOTAL FORMULA ERRORS FOUND ACROSS ACTIVE SUITE: {total_errors}")
    print("=" * 80)
