"""
Alpha Containers — Master Daily Pipeline Dry Run Harness
Executes Scripts/daily.py with --skip-prod --skip-wip --skip-git
Pipes empty newline to satisfy final 'Press Enter to close...' prompt
Monitors:
- Pre and post EXCEL.EXE processes
- Full stdout and stderr capture
- Generated Logs/ files (log, mismatches.log, error_summary.txt, etc.)
- Post-run formula integrity across Tubex_Aug26.xlsx
"""

import os
import sys
import time
import subprocess
import openpyxl

# Set console encoding to UTF-8
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

SCRIPTS_DIR = r"d:\Alpha\Scripts"
ROOT_DIR = r"d:\Alpha"
LOGS_DIR = r"d:\Alpha\Logs"

def safe_print(msg=""):
    try:
        print(msg)
    except UnicodeEncodeError:
        enc = getattr(sys.stdout, 'encoding', 'ascii') or 'ascii'
        print(msg.encode(enc, errors='replace').decode(enc))

def get_excel_processes():
    res = subprocess.run(
        ["powershell", "-NoProfile", "-Command", "@(Get-Process EXCEL -ErrorAction SilentlyContinue).Count"],
        capture_output=True,
        text=True
    )
    count_str = res.stdout.strip()
    return int(count_str) if count_str.isdigit() else 0

def run_daily_dry_run():
    safe_print("=" * 80)
    safe_print("ALPHA MASTER PIPELINE DRY RUN: daily.py --skip-prod --skip-wip --skip-git")
    safe_print("=" * 80)
    
    excel_before = get_excel_processes()
    safe_print(f"[*] Initial Active EXCEL.EXE Processes: {excel_before}")
    
    start_time = time.time()
    
    # Run daily.py non-interactively with piped input ("\n")
    proc = subprocess.Popen(
        [sys.executable, "daily.py", "--skip-prod", "--skip-wip", "--skip-git"],
        cwd=SCRIPTS_DIR,
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding='utf-8',
        errors='replace'
    )
    
    stdout, stderr = proc.communicate(input="\n")
    duration = time.time() - start_time
    exit_code = proc.returncode
    
    excel_after = get_excel_processes()
    safe_print(f"[*] Final Active EXCEL.EXE Processes:   {excel_after}")
    safe_print(f"[*] Dry Run Duration: {duration:.2f} seconds")
    safe_print(f"[*] Process Exit Code: {exit_code}")
    
    safe_print("\n" + "=" * 80)
    safe_print("DAILY PIPELINE STDOUT OUTPUT:")
    safe_print("=" * 80)
    safe_print(stdout)
    
    if stderr.strip():
        safe_print("\n" + "=" * 80)
        safe_print("DAILY PIPELINE STDERR OUTPUT:")
        safe_print("=" * 80)
        safe_print(stderr)
        
    safe_print("\n" + "=" * 80)
    safe_print("VERIFYING GENERATED LOG FILES & ARTIFACTS IN Logs/:")
    safe_print("=" * 80)
    for log_f in sorted(os.listdir(LOGS_DIR)):
        full_p = os.path.join(LOGS_DIR, log_f)
        sz = os.path.getsize(full_p)
        mtime = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(os.path.getmtime(full_p)))
        safe_print(f"  {log_f:<35} | Size: {sz:>8} bytes | Modified: {mtime}")
    
    # Check error_summary.txt content
    err_sum_path = os.path.join(LOGS_DIR, "error_summary.txt")
    if os.path.exists(err_sum_path):
        safe_print("\n--- CONTENT OF error_summary.txt ---")
        with open(err_sum_path, 'r', encoding='utf-8', errors='replace') as f:
            safe_print(f.read().strip())
            
    # Check formula integrity of Tubex_Aug26.xlsx after full pipeline
    safe_print("\n" + "=" * 80)
    safe_print("POST-PIPELINE FORMULA AUDIT: Tubex_Aug26.xlsx")
    safe_print("=" * 80)
    wb = openpyxl.load_workbook(os.path.join(ROOT_DIR, "Tubex_Aug26.xlsx"), data_only=False, read_only=True)
    formula_errors = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            for c_idx, cell in enumerate(row, start=1):
                val = cell.value
                if val and isinstance(val, str) and val.startswith('='):
                    for err in ['#REF!', '#VALUE!', '#NAME?', '#DIV/0!', '#N/A']:
                        if err in val.upper():
                            formula_errors.append((sheet_name, cell.coordinate, err, val))
    wb.close()
    
    if formula_errors:
        safe_print(f"[FAIL] Found {len(formula_errors)} formula errors in Tubex_Aug26.xlsx!")
        for s, c, e, f in formula_errors:
            safe_print(f"  Sheet {s}, Cell {c}: {e} in {f}")
    else:
        safe_print("[PASS] ZERO formula errors found across all sheets of Tubex_Aug26.xlsx after full pipeline execution!")
        
    return exit_code == 0 and excel_after <= excel_before and len(formula_errors) == 0

if __name__ == '__main__':
    ok = run_daily_dry_run()
    safe_print("\n" + "=" * 80)
    safe_print(f"OVERALL PIPELINE DRY RUN RESULT: {'PASSED (RELIABILITY ASSERTION VERIFIED)' if ok else 'FAILED'}")
    safe_print("=" * 80)
