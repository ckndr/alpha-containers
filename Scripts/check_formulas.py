import openpyxl

file_path = r'd:\Alpha\Tubex_v10_30.xlsx'
wb = openpyxl.load_workbook(file_path)  # Without data_only to see formulas
ws = wb['Inventory']

# Check if Store Balance cells are formulas or values
# Slug section rows 3-9 (1-indexed), Store Balance is column H (8)
print("Checking Inventory sheet - Slug section (Store Balance column):")
for row in range(1, 12):
    cell = ws.cell(row=row, column=8)
    cell_a = ws.cell(row=row, column=1)
    print(f"  Row {row}: Item={cell_a.value}, Store Balance={cell.value}, Type={type(cell.value).__name__}")

print("\nChecking Inventory sheet - Other materials section:")
for row in range(12, 25):
    cell = ws.cell(row=row, column=8)
    cell_a = ws.cell(row=row, column=1)
    print(f"  Row {row}: Item={cell_a.value}, Store Balance={cell.value}, Type={type(cell.value).__name__}")
