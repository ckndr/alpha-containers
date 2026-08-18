import os
import openpyxl

ALPHA_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RECORDS_DIR = os.path.join(ALPHA_DIR, "Tubex Records")
PRODUCTION_ARCHIVE = os.path.join(RECORDS_DIR, "Production_Archive.xlsx")

def normalize_product_name(prod_str):
    if not prod_str:
        return ""
    p_up = prod_str.upper().strip()
    if "MUSTARD" in p_up:
        return "PET BOTTLE MUSTARD OIL (200ML) TRANSPARENT"
    return prod_str.strip()

def get_key(cust, month, prod, ptype):
    return (cust, month, prod, ptype)

def extract_all_customer_records():
    """
    Reads all production and dispatch data purely from the central Production_Archive.xlsx.
    Returns a list of dicts with:
      customer, month, product, type, dia, produced, dispatched, balance, reject, good
    """
    records = {}
    
    if not os.path.exists(PRODUCTION_ARCHIVE):
        print(f"ERROR: {PRODUCTION_ARCHIVE} not found.")
        return []

    wb = openpyxl.load_workbook(PRODUCTION_ARCHIVE, data_only=True)
    
    # 1. Parse All Months (Production)
    if "All Months" in wb.sheetnames:
        ws_p = wb["All Months"]
        # Headers: Month(1), Date(2), Machine(3), Customer(4), Product Name(5), Dia(6), PID(7), Target(8), Good(9), Reject(10)
        for r in range(4, ws_p.max_row + 1):
            month_val = ws_p.cell(r, 1).value
            if not month_val:
                continue
                
            machine = ws_p.cell(r, 3).value or ""
            customer = ws_p.cell(r, 4).value or "Other Customers"
            product = ws_p.cell(r, 5).value or ""
            dia = ws_p.cell(r, 6).value or ""
            good = ws_p.cell(r, 9).value or 0
            reject = ws_p.cell(r, 10).value or 0
            
            prod_str = normalize_product_name(str(product))
            cust_str = str(customer).strip()
            month_str = str(month_val).strip()
            dia_str = str(dia).strip()
            mach_str = str(machine).strip()
            
            # Determine machine/product type
            count_production = False
            if mach_str.startswith("PF"):
                ptype = "PET"
                count_production = True
            elif mach_str.startswith("Print") and "(Varnish)" not in prod_str:
                ptype = "TUBE"
                count_production = True
            else:
                ptype = "PET" if ("PET" in prod_str.upper() or "BOTTLE" in prod_str.upper() or prod_str.upper().startswith("BT-")) else "TUBE"
            
            if count_production:
                try: good = int(good)
                except: good = 0
                try: reject = int(reject)
                except: reject = 0
                
                k = get_key(cust_str, month_str, prod_str, ptype)
                if k not in records:
                    records[k] = {"produced": 0, "dispatched": 0, "reject": 0, "dia": dia_str}
                    
                records[k]["produced"] += good
                records[k]["reject"] += reject

    # 2. Parse Dispatch_Log
    if "Dispatch_Log" in wb.sheetnames:
        ws_d = wb["Dispatch_Log"]
        # Headers: Month(1), Date(2), Type(3), Customer(4), Product Name(5), Dia(6), OrdQty(7), DispQty(8)
        for r in range(4, ws_d.max_row + 1):
            month_val = ws_d.cell(r, 1).value
            if not month_val:
                continue
                
            ptype = str(ws_d.cell(r, 3).value or "TUBE").strip()
            customer = ws_d.cell(r, 4).value or "Other Customers"
            product = ws_d.cell(r, 5).value or ""
            dia = ws_d.cell(r, 6).value or ""
            disp = ws_d.cell(r, 8).value or 0
            
            prod_str = normalize_product_name(str(product))
            cust_str = str(customer).strip()
            month_str = str(month_val).strip()
            dia_str = str(dia).strip()
            
            try: disp = int(disp)
            except: disp = 0
            
            k = get_key(cust_str, month_str, prod_str, ptype)
            if k not in records:
                records[k] = {"produced": 0, "dispatched": 0, "reject": 0, "dia": dia_str}
                
            records[k]["dispatched"] += disp

    wb.close()

    # Convert to list format
    flat_list = []
    for (cust, month, prod, ptype), d in records.items():
        flat_list.append({
            "customer": cust,
            "month": month,
            "product": prod,
            "type": ptype,
            "dia": d["dia"],
            "produced": d["produced"],
            "dispatched": d["dispatched"],
            "balance": d["produced"] - d["dispatched"],
            "reject": d["reject"],
            "good": d["produced"] # backwards compatible
        })

    # Filter out ghost records with all-zero values
    flat_list = [r for r in flat_list if r['produced'] != 0 or r['dispatched'] != 0 or r['reject'] != 0]

    return flat_list

if __name__ == "__main__":
    recs = extract_all_customer_records()
    print(f"Total Customer-Month-Product Records: {len(recs)}")
    
    july_pet = sum(r["produced"] for r in recs if r["month"] == "July 2026" and r["type"] == "PET")
    july_tube = sum(r["produced"] for r in recs if r["month"] == "July 2026" and r["type"] == "TUBE")
    print(f"July 2026 Total PET Produced: {july_pet:,}")
    print(f"July 2026 Total TUBE Produced: {july_tube:,}")
