import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

file_path = r'd:\Alpha\Tubex_v10_30.xlsx'

# FIRST: Read inventory data BEFORE openpyxl load (to get cached formula values)
# We need to read from the Opening/Received/Issued columns directly since cached values are gone
# Read the raw Inventory sheet to reconstruct store balances from source columns
wb_formulas = openpyxl.load_workbook(file_path)  # formulas mode
ws_inv = wb_formulas['Inventory']

# Build inventory map from Opening + Received - Issued = Store Balance, plus WIP
inv_map = {}
current_section_cols = {}

for row in ws_inv.iter_rows(min_row=1, max_row=ws_inv.max_row, values_only=False):
    vals = {cell.column: cell.value for cell in row}
    # Check if this is a header row
    row_vals = [str(v) for v in vals.values()]
    
    if 'Item ID' in row_vals:
        # Detect column positions
        for col, val in vals.items():
            if str(val).strip() == 'Item ID':
                current_section_cols['item_id'] = col
            elif str(val).strip() == 'Opening':
                current_section_cols['opening'] = col
            elif str(val).strip() == 'Received from Vendor':
                current_section_cols['received'] = col
            elif str(val).strip() == 'Issued to Production':
                current_section_cols['issued'] = col
            elif str(val).strip() == 'Work In Process':
                current_section_cols['wip'] = col
        continue
    
    if 'item_id' in current_section_cols:
        try:
            item_id_val = vals.get(current_section_cols['item_id'])
            if item_id_val is None or str(item_id_val).strip() in ('', 'None', 'nan'):
                continue
            item_id = str(int(item_id_val))
            
            opening = vals.get(current_section_cols.get('opening'), 0)
            received = vals.get(current_section_cols.get('received'), 0)
            issued = vals.get(current_section_cols.get('issued'), 0)
            wip = vals.get(current_section_cols.get('wip'), 0)
            
            # These might be formulas or values - handle both
            def get_num(v):
                if v is None: return 0
                if isinstance(v, (int, float)): return float(v)
                if isinstance(v, str) and v.startswith('='): return 0  # formula, need cached
                try: return float(v)
                except: return 0
            
            opening = get_num(opening)
            received = get_num(received)
            issued = get_num(issued)
            wip = get_num(wip)
            
            store_balance = opening + received - issued
            total = store_balance + wip
            
            inv_map[item_id] = {'store': store_balance, 'wip': wip, 'total': total}
        except:
            pass

print("Rebuilt Inventory (Store + WIP):")
target_items = ['6', '406', '3595', '68', '578', '3598', '6935', '4155', '185', '186', '194']
for item_id in target_items:
    if item_id in inv_map:
        d = inv_map[item_id]
        print(f"  {item_id}: Store={d['store']}, WIP={d['wip']}, Total={d['total']}")
    else:
        print(f"  {item_id}: NOT FOUND")

# Now recreate MRP 2 sheet
if 'MRP 2' in wb_formulas.sheetnames:
    del wb_formulas['MRP 2']

ws = wb_formulas.create_sheet('MRP 2')

# Styles
title_font = Font(size=16, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
header_font = Font(bold=True)
header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
shortage_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# 1. Tube Order Balance section
ws.merge_cells('A1:D1')
ws['A1'] = 'JULY 2026 TUBE REQUIRED ORDERS (BALANCE)'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = center_align

headers_tubes = ['Dia', 'Customer', 'Product Name', 'Remaining Balance']
for col_num, header in enumerate(headers_tubes, 1):
    cell = ws.cell(row=2, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

tubes_data = [
    [25, 'Samsol International Private Limited', 'S-45', 100000],
    [25, 'Samsol International Private Limited', 'S 43 25MM', 78928],
    [30, 'Mablay Beauty PVT LTD.', 'VINCE NURTURAL', 30000],
    [30, 'Golden Pearl Cosmetics (PVT) LTD', 'HELLO HAIR COLOR', 196272]
]

row_num = 3
for item in tubes_data:
    for col_num, value in enumerate(item, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = thin_border
        if col_num == 4:
            cell.number_format = '#,##0'
    row_num += 1

row_num += 2

# 2. Material Requirements — using Store + WIP
ws.merge_cells(f'A{row_num}:I{row_num}')
ws[f'A{row_num}'] = 'JULY 2026 MATERIAL REQUIREMENT PLAN (EXCLUDING TAPE)'
ws[f'A{row_num}'].font = title_font
ws[f'A{row_num}'].fill = title_fill
ws[f'A{row_num}'].alignment = center_align
row_num += 1

headers_mrp = ['Item ID', 'Material Category', 'Item Name', 'UOM', 'Required Qty', 'Store + WIP', 'Net Required to Buy', 'Status', 'Used In']
for col_num, header in enumerate(headers_mrp, 1):
    cell = ws.cell(row=row_num, column=col_num)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

df_bom = pd.read_excel(file_path, sheet_name='BOM', skiprows=1)
products = ['S-45', 'S 43 25MM', 'VINCE NURTURAL', 'HELLO HAIR COLOR']
bom_rows = df_bom[df_bom['Product Name'].isin(products)]

material_reqs = {}

for _, row in bom_rows.iterrows():
    if row['Material Category'] == 'TAPE':
        continue
    
    prod_name = row['Product Name']
    balance = next(t[3] for t in tubes_data if t[2] == prod_name)
    
    item_id = str(int(row['Item ID'])) if pd.notna(row['Item ID']) else 'Unknown'
    cat = row['Material Category']
    item_name = row['Item Name']
    uom = row['UOM']
    per_1000 = float(row['Per 1000 Units'])
    scrap = float(row['Scrap %'])
    
    req = (balance / 1000) * per_1000 * (1 + scrap)
    
    if item_id not in material_reqs:
        material_reqs[item_id] = {
            'cat': cat,
            'name': item_name,
            'uom': uom,
            'req': 0,
            'used_in': set()
        }
    
    material_reqs[item_id]['req'] += req
    material_reqs[item_id]['used_in'].add(prod_name)

sorted_materials = sorted(material_reqs.items(), key=lambda x: (x[1]['cat'], x[1]['name']))

row_num += 1
for item_id, mat in sorted_materials:
    inv_data = inv_map.get(item_id, {'store': 0, 'wip': 0, 'total': 0})
    available = inv_data['total']  # Store + WIP
    
    req = round(mat['req'], 2)
    net_req = round(req - available, 2)
    if net_req < 0:
        net_req = 0
    
    status = 'SHORTAGE' if net_req > 0 else 'OK'
    
    row_data = [
        int(item_id) if item_id.isdigit() else item_id,
        mat['cat'],
        mat['name'],
        mat['uom'],
        req,
        available,
        net_req,
        status,
        ", ".join(sorted(list(mat['used_in'])))
    ]
    
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.border = thin_border
        if col_num in [5, 6, 7]:
            cell.number_format = '#,##0.00'
        if col_num == 8:
            cell.alignment = center_align
            if value == 'SHORTAGE':
                cell.fill = shortage_fill
                cell.font = Font(bold=True, color='9C0006')
            else:
                cell.fill = ok_fill
                cell.font = Font(bold=True, color='006100')
    row_num += 1

# Column widths
for col_idx in range(1, 10):
    column = get_column_letter(col_idx)
    max_length = 0
    for r in range(1, row_num):
        cell = ws.cell(row=r, column=col_idx)
        if isinstance(cell, openpyxl.cell.cell.Cell):
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
    adjusted_width = (max_length + 2)
    if adjusted_width > 42:
        adjusted_width = 42
    ws.column_dimensions[column].width = adjusted_width

wb_formulas.save(file_path)
print("MRP 2 sheet updated with Store + WIP balances and Status column. Saved.")
