import openpyxl
import pandas as pd

file_path = r'd:\Alpha\Tubex_v10_30.xlsx'

# 1. Read MRP 2 sheet raw
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['MRP 2']

print("=" * 80)
print("MRP 2 SHEET - FULL DUMP")
print("=" * 80)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
    vals = []
    for cell in row:
        vals.append(f"{cell.value}")
    print(" | ".join(vals))

print("\n\n")

# 2. Cross-check against BOM
print("=" * 80)
print("CROSS-CHECK: BOM Per 1000 Units x Balance / 1000 x (1 + Scrap%)")
print("=" * 80)

df_bom = pd.read_excel(file_path, sheet_name='BOM', skiprows=1)
df_mrp = pd.read_excel(file_path, sheet_name='MRP', skiprows=1)

products = ['S-45', 'S 43 25MM', 'VINCE NURTURAL', 'HELLO HAIR COLOR']
balances = {'S-45': 100000, 'S 43 25MM': 78928, 'VINCE NURTURAL': 30000, 'HELLO HAIR COLOR': 196272}

bom_rows = df_bom[df_bom['Product Name'].isin(products)]

# Aggregate by Item ID (excluding TAPE)
material_reqs = {}
for _, row in bom_rows.iterrows():
    if row['Material Category'] == 'TAPE':
        continue
    prod_name = row['Product Name']
    balance = balances[prod_name]
    item_id = str(int(row['Item ID'])) if pd.notna(row['Item ID']) else '?'
    per_1000 = float(row['Per 1000 Units'])
    scrap = float(row['Scrap %'])
    req = (balance / 1000) * per_1000 * (1 + scrap)
    
    if item_id not in material_reqs:
        material_reqs[item_id] = {
            'cat': row['Material Category'],
            'name': row['Item Name'],
            'uom': row['UOM'],
            'req': 0,
            'used_in': set(),
            'breakdown': []
        }
    material_reqs[item_id]['req'] += req
    material_reqs[item_id]['used_in'].add(prod_name)
    material_reqs[item_id]['breakdown'].append(
        f"  {prod_name}: {balance}/1000 x {per_1000} x (1+{scrap}) = {req:.4f}"
    )

# 3. Get store balances
df_inv = pd.read_excel(file_path, sheet_name='Inventory', header=None)
inv_map = {}
item_id_col = None
store_bal_col = None
for _, row in df_inv.iterrows():
    for i, val in enumerate(row):
        if str(val).strip() == 'Item ID':
            item_id_col = i
        elif str(val).strip() == 'Store Balance':
            store_bal_col = i
    if item_id_col is not None and store_bal_col is not None:
        try:
            item_id = str(int(row[item_id_col]))
            store_bal = row[store_bal_col]
            inv_map[item_id] = store_bal
        except:
            pass

sorted_mats = sorted(material_reqs.items(), key=lambda x: (x[1]['cat'], x[1]['name']))

for item_id, mat in sorted_mats:
    store_bal = inv_map.get(item_id, 0)
    try:
        store_bal = float(store_bal)
    except:
        store_bal = 0
    net = max(mat['req'] - store_bal, 0)
    print(f"\nItem ID: {item_id} | {mat['cat']} | {mat['name']} | {mat['uom']}")
    for b in mat['breakdown']:
        print(b)
    print(f"  TOTAL Required: {mat['req']:.4f}")
    print(f"  Store Balance:  {store_bal}")
    print(f"  Net to Buy:     {net:.4f}")
    print(f"  Used In:        {', '.join(sorted(mat['used_in']))}")

# 4. Verify MRP orders match
print("\n\n")
print("=" * 80)
print("CROSS-CHECK: MRP Order Balances")
print("=" * 80)
mrp_rows = df_mrp[df_mrp['Product Name'].isin(products)]
for _, row in mrp_rows.iterrows():
    print(f"{row['Product Name']}: Required={row['Required Qty']}, Produced={row['Produced']}, Remaining={row['Remaining Balance']}")

# 5. Check if VINCE NURTURAL has carton in BOM
print("\n\nVINCE NURTURAL full BOM:")
vince_bom = df_bom[df_bom['Product Name'] == 'VINCE NURTURAL']
for _, row in vince_bom.iterrows():
    print(f"  {row['Material Category']}: {row['Item Name']} (ID: {row['Item ID']}) | Per 1000: {row['Per 1000 Units']} | Scrap: {row['Scrap %']}")
