import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def copy_style(src_cell, dest_cell):
    if src_cell.has_style:
        dest_cell.font = Font(
            name=src_cell.font.name,
            size=src_cell.font.size,
            bold=src_cell.font.bold,
            italic=src_cell.font.italic,
            charset=src_cell.font.charset,
            color=src_cell.font.color,
            underline=src_cell.font.underline,
            strike=src_cell.font.strike
        )
        if src_cell.fill and src_cell.fill.fill_type:
            dest_cell.fill = PatternFill(
                fill_type=src_cell.fill.fill_type,
                start_color=src_cell.fill.start_color,
                end_color=src_cell.fill.end_color
            )
        dest_cell.border = Border(
            left=src_cell.border.left,
            right=src_cell.border.right,
            top=src_cell.border.top,
            bottom=src_cell.border.bottom
        )
        dest_cell.alignment = Alignment(
            horizontal=src_cell.alignment.horizontal,
            vertical=src_cell.alignment.vertical,
            text_rotation=src_cell.alignment.text_rotation,
            wrap_text=src_cell.alignment.wrap_text,
            shrink_to_fit=src_cell.alignment.shrink_to_fit,
            indent=src_cell.alignment.indent
        )
        dest_cell.number_format = src_cell.number_format

# Load workbook
wb = openpyxl.load_workbook(r'd:\Alpha\Aerosol_BOM.xlsx')

# Sheet: BOM
ws_bom = wb['BOM']

bom_data = [
    [9001, 100, 'Aerosol Customer', 'Aerosol Can 45mm', 45, 'SCHEKOSOL INT BEIGE (400 4 902)', 'kg', 5.1, 'Lacquer', 0.1, 'Qty 5.1 kg/1000. Beige interior lacquer. Epoxy resin, solids 38%, wet film 41–48 g/m², dry film 10–12 g/m². Alternative to Gold lacquer. (Ref: TDS 400 4 902)'],
    [9001, 100, 'Aerosol Customer', 'Aerosol Can 45mm', 45, 'SCHEKOSOL CLEAR BC (422 9 918)', 'kg', 3.0, 'Base Coat', 0.1, 'Qty 3.0 kg/1000. Colourless basecoat for exterior. Polyester, solids 51%, wet film 25–31 g/m², dry film 10–12 µm. Alternative to White BC. (Ref: TDS 422 9 918)'],
    [9001, 100, 'Aerosol Customer', 'Aerosol Can 45mm', 45, 'SCHEKOSOL OPV SILKMATT (422 9 903)', 'kg', 1.5, 'Varnish', 0.1, 'Qty 1.5 kg/1000. Silk matt overprint varnish for exterior. Polyester, solids 47%, wet film 11–17 g/m², dry film 4–6 µm. Alternative to Glossy OPV. (Ref: TDS 422 9 903)']
]

for row_idx, data in enumerate(bom_data, start=10):
    for col_idx, val in enumerate(data, start=1):
        cell = ws_bom.cell(row=row_idx, column=col_idx, value=val)
        # Copy style from row 9 (same column)
        src_cell = ws_bom.cell(row=9, column=col_idx)
        copy_style(src_cell, cell)

# Sheet: Req. Calculator
ws_calc = wb['Req. Calculator']

# Insert 3 rows starting at row 11 (the original total row)
ws_calc.insert_rows(11, 3)

calc_data = [
    [8, '=BOM!F10', '=BOM!I10', '=BOM!G10', '=BOM!H10', '=BOM!J10', '=(D2/1000)*BOM!H10', '=(D2/1000)*BOM!H10*(1+BOM!J10)'],
    [9, '=BOM!F11', '=BOM!I11', '=BOM!G11', '=BOM!H11', '=BOM!J11', '=(D2/1000)*BOM!H11', '=(D2/1000)*BOM!H11*(1+BOM!J11)'],
    [10, '=BOM!F12', '=BOM!I12', '=BOM!G12', '=BOM!H12', '=BOM!J12', '=(D2/1000)*BOM!H12', '=(D2/1000)*BOM!H12*(1+BOM!J12)']
]

for i, data in enumerate(calc_data):
    row_idx = 11 + i
    for col_idx, val in enumerate(data, start=1):
        cell = ws_calc.cell(row=row_idx, column=col_idx, value=val)
        # Copy style from original row 10 (which is now row 10, i.e., index 10)
        src_cell = ws_calc.cell(row=10, column=col_idx)
        copy_style(src_cell, cell)

# Update TOTAL row (now row 14)
total_row_idx = 14
ws_calc.cell(row=total_row_idx, column=7, value='=SUM(G4:G13)')
ws_calc.cell(row=total_row_idx, column=8, value='=SUM(H4:H13)')

# Save the workbook
wb.save(r'd:\Alpha\Aerosol_BOM.xlsx')
print("Successfully updated Aerosol_BOM.xlsx!")
