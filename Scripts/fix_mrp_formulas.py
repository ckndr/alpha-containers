import openpyxl

wb = openpyxl.load_workbook('Tubex_July26.xlsx', data_only=False)
mrp = wb['MRP']

start_row = 14
end_row = 95

for r in range(start_row, end_row + 1):
    # If the row is empty, break or continue
    if mrp.cell(r, 1).value is None:
        continue
    
    # E
    mrp.cell(r, 5).value = f'=SUMPRODUCT((TableBOM[Item ID]=A{r})*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %])*SUMIF($D$3:$D$10,TableBOM[Product ID],$H$3:$H$10)/1000)'
    
    # F
    mrp.cell(r, 6).value = f'=IFERROR(INDEX(TableInventory[Store Balance],MATCH(A{r},TableInventory[Item ID],0)),0)+IFERROR(INDEX(TableInventory[Work In Process],MATCH(A{r},TableInventory[Item ID],0)),0)'
    
    # G
    mrp.cell(r, 7).value = f'=F{r}-E{r}'
    
    # H
    concat_parts = []
    for i in range(3, 11):
        concat_parts.append(f'IF((COUNTIFS(TableBOM[Product ID],$D${i},TableBOM[Item ID],$A{r})>0)*($H${i}>0),$C${i}&", ","")')
    concat_str = ' & '.join(concat_parts)
    mrp.cell(r, 8).value = f'=IF(LEN({concat_str})>1,LEFT({concat_str},LEN({concat_str})-2),"")'
    
    # I
    mrp.cell(r, 9).value = f'=IF(E{r}=0,"Not needed",IF(G{r}<0,"SHORTAGE",IF(G{r}<F{r}*0.1,"LOW","OK")))'
    
    # J
    mrp.cell(r, 10).value = f'=IF(B{r}<>"CAP","",SUMPRODUCT((ISNUMBER(SEARCH("Cap Not Available",\'FG Stock\'!$H$4:$H$100)))*(\'FG Stock\'!$I$4:$I$100=A{r})*\'FG Stock\'!$F$4:$F$100))'

wb.save('Tubex_July26.xlsx')
print('Formulas fixed for rows 14 to 95')
