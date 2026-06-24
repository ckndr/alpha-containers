import openpyxl

wb = openpyxl.load_workbook(r'd:\Alpha\Aerosol_BOM.xlsx')
ws = wb['Req. Calculator']

# Unmerge A11:F11 if it exists in merged_cells
if 'A11:F11' in [str(r) for r in ws.merged_cells.ranges]:
    print("Unmerging A11:F11...")
    ws.unmerge_cells('A11:F11')

# Set formulas/values for row 11
ws.cell(row=11, column=2, value='=BOM!F10')
ws.cell(row=11, column=3, value='=BOM!I10')
ws.cell(row=11, column=4, value='=BOM!G10')
ws.cell(row=11, column=5, value='=BOM!H10')
ws.cell(row=11, column=6, value='=BOM!J10')

# Merge A14:F14
print("Merging A14:F14...")
ws.merge_cells('A14:F14')

# Save workbook
wb.save(r'd:\Alpha\Aerosol_BOM.xlsx')
print("Successfully fixed Req. Calculator sheet!")
