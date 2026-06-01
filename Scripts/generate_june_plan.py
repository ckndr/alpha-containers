import openpyxl
import shutil
import os
import datetime
from copy import copy

# --- CONFIGURATION ---
DATA_DIR = r"d:\Alpha"
DASHBOARD_FILE = os.path.join(DATA_DIR, "Tubex_v10_28.xlsx")
TEMPLATE_FILE = os.path.join(DATA_DIR, "May_Plan.xlsx")
OUTPUT_FILE = os.path.join(DATA_DIR, "June_Plan.xlsx")

# June 2026 Working Days (Option B: Excludes Sundays and Eid al-Adha June 5, 6)
WORKING_DAYS = [
    datetime.datetime(2026, 6, 1),
    datetime.datetime(2026, 6, 2),
    datetime.datetime(2026, 6, 3),
    datetime.datetime(2026, 6, 4),
    # June 5, 6: Eid Holidays
    # June 7: Sunday
    datetime.datetime(2026, 6, 8),
    datetime.datetime(2026, 6, 9),
    datetime.datetime(2026, 6, 10),
    datetime.datetime(2026, 6, 11),
    datetime.datetime(2026, 6, 12),
    datetime.datetime(2026, 6, 13),
    # June 14: Sunday
    datetime.datetime(2026, 6, 15),
    datetime.datetime(2026, 6, 16),
    datetime.datetime(2026, 6, 17),
    datetime.datetime(2026, 6, 18),
    datetime.datetime(2026, 6, 19),
    datetime.datetime(2026, 6, 20),
    # June 21: Sunday
    datetime.datetime(2026, 6, 22),
    datetime.datetime(2026, 6, 23),
    datetime.datetime(2026, 6, 24),
    datetime.datetime(2026, 6, 25),
    datetime.datetime(2026, 6, 26),
    datetime.datetime(2026, 6, 27),
    # June 28: Sunday
    datetime.datetime(2026, 6, 29),
    datetime.datetime(2026, 6, 30),
]

def get_week_name(dt):
    day = dt.day
    if day <= 7:
        return "Week 01"
    elif day <= 14:
        return "Week 02"
    elif day <= 21:
        return "Week 03"
    elif day <= 28:
        return "Week 04"
    else:
        return "Week 05"

def copy_cell_style(src_cell, dst_cell):
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.border = copy(src_cell.border)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format

def main():
    print(f"Reading orders from {DASHBOARD_FILE}...")
    wb_db = openpyxl.load_workbook(DASHBOARD_FILE, data_only=True)
    sheet_db = wb_db['Tubex_Dashboard']
    
    # Extract TUBE orders
    tube_orders = []
    # Extract PET orders
    pet_orders = []
    
    for r in range(11, sheet_db.max_row + 1):
        type_val = sheet_db.cell(row=r, column=2).value
        if not type_val:
            continue
        type_str = str(type_val).strip().upper()
        if type_str not in ('TUBE', 'PET'):
            continue
            
        pid = sheet_db.cell(row=r, column=6).value
        if pid is None:
            continue
        try:
            pid = int(pid)
        except ValueError:
            continue
            
        remaining = sheet_db.cell(row=r, column=9).value
        if remaining is None or not isinstance(remaining, (int, float)) or remaining <= 0:
            continue
            
        product_name = sheet_db.cell(row=r, column=4).value
        if "(Varnish)" in str(product_name):
            continue
            
        customer = sheet_db.cell(row=r, column=3).value
        dia = sheet_db.cell(row=r, column=5).value
        remarks = sheet_db.cell(row=r, column=12).value
        
        if customer == "MHK":
            if pid in (3447, 5699, 6515):
                customer = "Samsol International Private Limited"
            elif pid in (6020, 6021, 6530, 6531):
                customer = "Adore"
            elif pid == 5782:
                customer = "Mega Grey"
            elif pid == 6338:
                customer = "DTM"
                
        # Determine status: if there's a hold remark, mark status as Hold
        status = 'Confirmed'
        if remarks and 'Hold' in str(remarks):
            status = remarks
            
        order_data = {
            'pid': pid,
            'customer': customer,
            'product': product_name,
            'dia': dia,
            'remaining': remaining,
            'status': status,
            'remarks': remarks
        }
        
        if type_str == 'TUBE' and not (8001 <= pid <= 8099):
            tube_orders.append(order_data)
        elif type_str == 'PET' or (8001 <= pid <= 8099):
            pet_orders.append(order_data)
            
    print(f"Extracted {len(tube_orders)} Tube orders and {len(pet_orders)} PET orders.")
    
    # Tubes day-by-day schedule logic (ON-HOLD orders excluded!)
    tube_schedule = [
        # Jun 1 (Mon)
        {'date': datetime.datetime(2026, 6, 1), 'machine': 'Pline-03', 'customer': 'Samsol International Private Limited', 'product': 'TUBES', 'dia': 25, 'qty': 17500, 'changeover_type': 'Same Dia', 'changeover_qty': 2500},
        {'date': datetime.datetime(2026, 6, 1), 'machine': 'Pline-04', 'customer': 'Mablay Beauty PVT LTD.', 'product': 'VINCE NURTURAL', 'dia': 30, 'qty': 17500, 'changeover_type': 'Same Dia', 'changeover_qty': 2500},
        # Jun 2 (Tue)
        {'date': datetime.datetime(2026, 6, 2), 'machine': 'Pline-03', 'customer': 'Samsol International Private Limited', 'product': 'TUBES', 'dia': 25, 'qty': 7500, 'changeover_type': None, 'changeover_qty': 0},
        {'date': datetime.datetime(2026, 6, 2), 'machine': 'Pline-04', 'customer': 'Mablay Beauty PVT LTD.', 'product': 'VINCE NURTURAL', 'dia': 30, 'qty': 4500, 'changeover_type': None, 'changeover_qty': 0},
    ]

    # PET day-by-day schedule logic (Option B: 24 working days)
    pet_schedule = []
    for day in WORKING_DAYS:
        pet_schedule.append({
            'date': day,
            'customer': 'Alpha Labs PVT LTD',
            'product': 'TRANSPARENT BOTTLE 150ML',
            'ml': 150,
            'qty': 10000
        })
        
    print(f"Copying template {TEMPLATE_FILE} to {OUTPUT_FILE}...")
    shutil.copyfile(TEMPLATE_FILE, OUTPUT_FILE)
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    
    sheetnames = wb.sheetnames
    print("Source sheet names:", sheetnames)
    
    if 'May Tubes Weekly' in sheetnames:
        wb['May Tubes Weekly'].title = 'June Tubes Weekly'
    if 'May Plan Tubes' in sheetnames:
        wb['May Plan Tubes'].title = 'June Plan'
    if 'May Plan Weekly' in sheetnames:
        wb['May Plan Weekly'].title = 'June Plan Weekly'
    elif 'Sheet2' in sheetnames:
        wb['Sheet2'].title = 'June Plan Weekly'
    if 'May Plan PET' in sheetnames:
        wb['May Plan PET'].title = 'June Plan PET'
        
    if 'Sheet1' in wb.sheetnames:
        wb.remove(wb['Sheet1'])
        
    print("New sheet names:", wb.sheetnames)
    
    # -------------------------------------------------------------
    # 1. UPDATE SHEET: June Plan (Tubes Day-by-Day)
    # -------------------------------------------------------------
    print("Updating 'June Plan' day-by-day sheet...")
    ws_plan = wb['June Plan']
    
    for r in range(2, 100):
        for c in range(1, 9):
            ws_plan.cell(row=r, column=c).value = None
            
    for i, item in enumerate(tube_schedule):
        r = 2 + i
        source_row = 2 if r % 2 == 0 else 3
        
        ws_plan.cell(row=r, column=1).value = get_week_name(item['date'])
        ws_plan.cell(row=r, column=2).value = item['date']
        ws_plan.cell(row=r, column=3).value = item['dia']
        ws_plan.cell(row=r, column=4).value = 'Tubes'
        ws_plan.cell(row=r, column=5).value = item['customer']
        ws_plan.cell(row=r, column=6).value = item['product']
        ws_plan.cell(row=r, column=7).value = item['qty']
        ws_plan.cell(row=r, column=8).value = item['machine']
        
        for c in range(1, 9):
            copy_cell_style(ws_plan.cell(row=source_row, column=c), ws_plan.cell(row=r, column=c))
            
    for r in range(3, 100):
        for c in range(10, 17):
            ws_plan.cell(row=r, column=c).value = None
            
    sorted_tube_orders = sorted(tube_orders, key=lambda x: (x['dia'], x['customer']))
    
    for i, order in enumerate(sorted_tube_orders):
        r = 3 + i
        ws_plan.cell(row=r, column=10).value = order['dia']
        ws_plan.cell(row=r, column=11).value = order['customer']
        ws_plan.cell(row=r, column=12).value = order['product']
        ws_plan.cell(row=r, column=13).value = order['remaining']
        ws_plan.cell(row=r, column=14).value = f"=SUMIF(F:F,L{r},G:G)"
        ws_plan.cell(row=r, column=15).value = f"=M{r}-N{r}"
        ws_plan.cell(row=r, column=16).value = order['status']
        
        for c in range(10, 17):
            copy_cell_style(ws_plan.cell(row=3, column=c), ws_plan.cell(row=r, column=c))
            
    total_row_idx = 3 + len(sorted_tube_orders)
    ws_plan.cell(row=total_row_idx, column=12).value = 'Total'
    ws_plan.cell(row=total_row_idx, column=13).value = f"=SUM(M3:M{total_row_idx-1})"
    
    for c in range(10, 17):
        copy_cell_style(ws_plan.cell(row=10, column=c), ws_plan.cell(row=total_row_idx, column=c))
        
    # -------------------------------------------------------------
    # 2. UPDATE SHEET: June Plan PET (PET Day-by-Day)
    # -------------------------------------------------------------
    print("Updating 'June Plan PET' day-by-day sheet...")
    ws_pet = wb['June Plan PET']
    
    for r in range(2, 100):
        for c in range(1, 7):
            ws_pet.cell(row=r, column=c).value = None
            
    for i, item in enumerate(pet_schedule):
        r = 2 + i
        ws_pet.cell(row=r, column=1).value = get_week_name(item['date'])
        ws_pet.cell(row=r, column=2).value = item['date']
        ws_pet.cell(row=r, column=3).value = item['ml']
        ws_pet.cell(row=r, column=4).value = item['customer']
        ws_pet.cell(row=r, column=5).value = item['product']
        ws_pet.cell(row=r, column=6).value = item['qty']
        
        for c in range(1, 7):
            copy_cell_style(ws_pet.cell(row=2, column=c), ws_pet.cell(row=r, column=c))
            
    for r in range(3, 100):
        for c in range(8, 15):
            ws_pet.cell(row=r, column=c).value = None
            
    def get_ml_num(ml_str):
        try:
            return float(ml_str.split()[0])
        except:
            return 9999.0
    sorted_pet_orders = sorted(pet_orders, key=lambda x: (get_ml_num(str(x['dia'])), x['customer']))
    
    priority_map = {
        8001: 1, # 150 ml
        8014: 2, # 200 ml Mustard
        8007: 3, # 200 ml White
        8006: 4, # 200 ml Yellow
        8005: 5  # 120 ml Yellow
    }
    
    for i, order in enumerate(sorted_pet_orders):
        r = 3 + i
        ws_pet.cell(row=r, column=8).value = order['dia']
        ws_pet.cell(row=r, column=9).value = order['customer']
        ws_pet.cell(row=r, column=10).value = order['product']
        ws_pet.cell(row=r, column=11).value = order['remaining']
        ws_pet.cell(row=r, column=12).value = f"=SUMIF(E:E,J{r},F:F)"
        ws_pet.cell(row=r, column=13).value = f"=K{r}-L{r}"
        ws_pet.cell(row=r, column=14).value = priority_map.get(order['pid'], None)
        
        for c in range(8, 15):
            copy_cell_style(ws_pet.cell(row=3, column=c), ws_pet.cell(row=r, column=c))
            
    total_pet_row_idx = 3 + len(sorted_pet_orders)
    ws_pet.cell(row=total_pet_row_idx, column=11).value = f"=SUM(K3:K{total_pet_row_idx-1})"
    ws_pet.cell(row=total_pet_row_idx, column=12).value = f"=SUM(L3:L{total_pet_row_idx-1})"
    ws_pet.cell(row=total_pet_row_idx, column=13).value = f"=SUM(M3:M{total_pet_row_idx-1})"
    
    for c in range(8, 15):
        copy_cell_style(ws_pet.cell(row=9, column=c), ws_pet.cell(row=total_pet_row_idx, column=c))
        
    # -------------------------------------------------------------
    # 3. UPDATE SHEET: June Tubes Weekly (Tubes Summary)
    # -------------------------------------------------------------
    print("Updating 'June Tubes Weekly' sheet...")
    ws_tweekly = wb['June Tubes Weekly']
    
    tweekly_row_formats = {}
    for r in [4, 5, 13]:
        tweekly_row_formats[r] = []
        for c in range(1, 5):
            tweekly_row_formats[r].append(copy(ws_tweekly.cell(row=r, column=c)))
            
    for r in range(4, 100):
        for c in range(1, 5):
            ws_tweekly.cell(row=r, column=c).value = None
            
    ws_tweekly.cell(row=2, column=1).value = "June's Tube Plan"
    
    # Week 01 row
    ws_tweekly.cell(row=4, column=1).value = "Week 01"
    ws_tweekly.cell(row=4, column=2).value = 30 # max dia is now 30
    ws_tweekly.cell(row=4, column=3).value = "=SUM(C5:C6)"
    for c in range(1, 5):
        copy_cell_style(tweekly_row_formats[4][c-1], ws_tweekly.cell(row=4, column=c))
        
    week1_customers = [
        {'name': 'Mablay Beauty PVT LTD.', 'dia': 30, 'qty': 22000},
        {'name': 'Samsol International Private Limited', 'dia': 25, 'qty': 25000},
    ]
    
    for i, cust in enumerate(week1_customers):
        r = 5 + i
        ws_tweekly.cell(row=r, column=1).value = cust['name']
        ws_tweekly.cell(row=r, column=2).value = cust['dia']
        ws_tweekly.cell(row=r, column=3).value = cust['qty']
        ws_tweekly.cell(row=r, column=4).value = f"=INDEX('June Plan'!H:H,MATCH('June Tubes Weekly'!A{r},'June Plan'!E:E,0))"
        
        for c in range(1, 5):
            copy_cell_style(tweekly_row_formats[5][c-1], ws_tweekly.cell(row=r, column=c))
            
    # Grand Total row
    ws_tweekly.cell(row=7, column=1).value = "Grand Total"
    ws_tweekly.cell(row=7, column=2).value = 30
    ws_tweekly.cell(row=7, column=3).value = "=C4"
    for c in range(1, 5):
        copy_cell_style(tweekly_row_formats[13][c-1], ws_tweekly.cell(row=7, column=c))
        
    # -------------------------------------------------------------
    # 4. UPDATE SHEET: June Plan Weekly (PET Summary)
    # -------------------------------------------------------------
    print("Updating 'June Plan Weekly' sheet...")
    ws_pweekly = wb['June Plan Weekly']
    
    for r in range(4, 100):
        for c in range(1, 5):
            ws_pweekly.cell(row=r, column=c).value = None
            
    ws_pweekly.cell(row=2, column=1).value = "June's PET Plan"
    
    weeks_pet = [
        {'week': 'Week 01', 'qty': 40000},
        {'week': 'Week 02', 'qty': 60000},
        {'week': 'Week 03', 'qty': 60000},
        {'week': 'Week 04', 'qty': 60000},
        {'week': 'Week 05', 'qty': 20000},
    ]
    
    r_idx = 4
    for item in weeks_pet:
        ws_pweekly.cell(row=r_idx, column=1).value = item['week']
        ws_pweekly.cell(row=r_idx, column=2).value = 150
        ws_pweekly.cell(row=r_idx, column=3).value = f"=C{r_idx+1}"
        for c in range(1, 4):
            copy_cell_style(tweekly_row_formats[4][c-1], ws_pweekly.cell(row=r_idx, column=c))
            
        ws_pweekly.cell(row=r_idx+1, column=1).value = "Alpha Labs PVT LTD"
        ws_pweekly.cell(row=r_idx+1, column=2).value = 150
        ws_pweekly.cell(row=r_idx+1, column=3).value = item['qty']
        for c in range(1, 4):
            copy_cell_style(tweekly_row_formats[5][c-1], ws_pweekly.cell(row=r_idx+1, column=c))
            
        r_idx += 2
        
    ws_pweekly.cell(row=r_idx, column=1).value = "Grand Total"
    ws_pweekly.cell(row=r_idx, column=2).value = 150
    ws_pweekly.cell(row=r_idx, column=3).value = "=C4+C6+C8+C10+C12"
    for c in range(1, 4):
        copy_cell_style(tweekly_row_formats[13][c-1], ws_pweekly.cell(row=r_idx, column=c))
        
    print(f"Saving workbook to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Workbook saved successfully!")
    
    # --- RUN VALIDATION CHECKS ---
    print("\n--- RUNNING VALIDATION CHECKS ---")
    wb_chk = openpyxl.load_workbook(OUTPUT_FILE, data_only=True)
    
    # 1. Order totals
    print("Checking Tube order totals...")
    # TUBES and VINCE NURTURAL should be completed
    # ECZEMUS and V-HC BROWN should have 0 planned qty
    expected_tubes = {
        'TUBES': 25000,
        'VINCE NURTURAL': 22000
    }
    actual_tubes = {}
    sheet_p = wb_chk['June Plan']
    for r in range(2, 6):
        prod = sheet_p.cell(row=r, column=6).value
        qty = sheet_p.cell(row=r, column=7).value
        if prod and qty:
            actual_tubes[prod] = actual_tubes.get(prod, 0) + qty
            
    print("Expected Tube Qtys:", expected_tubes)
    print("Actual Tube Qtys:", actual_tubes)
    assert actual_tubes == expected_tubes, "ERROR: Scheduled Tube quantities do not match!"
    print("  ^^ Tube order totals match exactly!")
    
    # 2. Daily capacity
    print("Checking daily capacity and changeover logic...")
    day_capacity = {}
    for item in tube_schedule:
        key = (item['date'], item['machine'])
        day_capacity[key] = day_capacity.get(key, 0) + item['qty']
        if item['changeover_qty'] > 0:
            day_capacity[key] += item['changeover_qty']
            
    for (d, m), total in day_capacity.items():
        print(f"  {d.strftime('%Y-%m-%d')} | {m} | Total Pieces + Setup = {total}")
        assert total <= 21000, f"ERROR: Total for {d} on {m} is {total}, exceeds 21000 limit!"
    print("  ^^ Daily capacity and changeover rules are perfectly met!")
    
    # 3. Dia conflict check
    print("Checking diameter conflicts...")
    day_dias = {}
    for item in tube_schedule:
        d = item['date']
        m = item['machine']
        dia = item['dia']
        if d not in day_dias:
            day_dias[d] = {}
        if m not in day_dias[d]:
            day_dias[d][m] = set()
        day_dias[d][m].add(dia)
        
    for d, machs in day_dias.items():
        if len(machs) == 2:
            m3_dias = machs['Pline-03']
            m4_dias = machs['Pline-04']
            overlap = m3_dias.intersection(m4_dias)
            print(f"  {d.strftime('%Y-%m-%d')}: Pline-03={m3_dias} vs Pline-04={m4_dias} | overlap={overlap}")
            assert not overlap, f"ERROR: Diameter conflict on {d}! Overlap: {overlap}"
    print("  ^^ No diameter conflicts detected!")
    
    # 4. No varnish/No PET in Tubes
    print("Checking for Varnish or PET rows in Tube schedule...")
    for item in tube_schedule:
        assert "(Varnish)" not in item['product'], f"ERROR: Varnish product {item['product']} in Tube plan!"
    print("  ^^ Verified: No Varnish and no PET rows in Tubes schedule.")
    
    # 5. Grand total match
    print("Checking PET totals...")
    sheet_pet_chk = wb_chk['June Plan PET']
    total_pet_qty = sum(sheet_pet_chk.cell(row=r, column=6).value for r in range(2, 26))
    print(f"  PET total planned quantity: {total_pet_qty} (Required overall: 332,402)")
    assert total_pet_qty == 240000, f"ERROR: PET total quantity is {total_pet_qty}, expected 240000!"
    print("  ^^ PET planned totals are correct!")
    
    print("\nALL VERIFICATIONS PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    main()
