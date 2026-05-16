"""
sort_dashboard.py
Alpha Containers — Dashboard Active/Inactive Auto-Sorter
---------------------------------------------------------
Scans every product row in Tubex_Dashboard, classifies each as active
(has orders, production, or dispatch) or inactive, then rewrites the
sheet so active rows sit at the top with no gaps, followed by TOTAL
rows, then inactive rows.  Layout is fully dynamic — sections expand
and shrink to fit the actual data.

Run manually or via Run_All_Updates.bat:
    python sort_dashboard.py

Author: Sikander / Claude (Opus)
Version: 1.0
"""

import os
import sys
import glob
from datetime import datetime, date
from copy import copy

import warnings
warnings.filterwarnings("ignore", message=".*Data Validation.*")
warnings.filterwarnings("ignore", message=".*extension.*")

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers

# ── PATH SETUP ──────────────────────────────────────────────
DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

excel_pattern = os.path.join(DIR, 'AlphaContainers_v*.xlsx')
excel_files   = sorted(glob.glob(excel_pattern))
if not excel_files:
    raise FileNotFoundError(f"No AlphaContainers_v*.xlsx found in {DIR}")
EXCEL_PATH = excel_files[-1]

print(f"Sort Dashboard: {os.path.basename(EXCEL_PATH)}")

# ── SAFETY CHECKS ───────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from alpha_checks import check_not_locked
check_not_locked(EXCEL_PATH)

# ── OPEN WORKBOOK ────────────────────────────────────────────
wb = load_workbook(EXCEL_PATH, data_only=False)
ws = wb['Tubex_Dashboard']
ws_pl = wb['Production_Log']

# ── COMPUTE MTD PRODUCTION FROM PRODUCTION_LOG ───────────────
# Since user only keeps current month in Production_Log, no month
# filter needed.  We just sum good qty by PID for Print/PET machines.
mtd_by_pid = {}
for row in ws_pl.iter_rows(min_row=3, values_only=True):
    machine   = row[1]   # col B
    prod_name = row[3]   # col D
    pid       = row[5]   # col F
    good_qty  = row[7]   # col H

    if not machine or not pid or not good_qty:
        continue

    mach_up  = str(machine).upper()
    is_print = mach_up.startswith('PRINT') or mach_up.startswith('PLINE')
    is_pet   = mach_up.startswith('PF') or mach_up.startswith('PET')
    is_varn  = '(VARNISH)' in str(prod_name).upper()

    if (is_print and not is_varn) or is_pet:
        try:
            pid_int = int(pid)
            mtd_by_pid[pid_int] = mtd_by_pid.get(pid_int, 0) + int(good_qty)
        except (TypeError, ValueError):
            pass

print(f"  Production Log: {len(mtd_by_pid)} PIDs with MTD production")

# ── READ ALL PRODUCT ROWS ───────────────────────────────────
# Scan rows 11 onward; collect every row with Type = TUBE or PET.
# Skip TOTAL rows (col D = "TOTAL") and blank rows.

DATA_COLS = range(2, 12)  # columns B(2) through K(11)

def read_product_row(ws, r):
    """Read a product row's data as a dict."""
    pid_raw = ws.cell(r, 6).value  # col F = Prod ID
    if pid_raw is None:
        return None
    try:
        pid_int = int(pid_raw)
    except (TypeError, ValueError):
        return None

    orders_raw   = ws.cell(r, 7).value   # col G
    dispatch_raw = ws.cell(r, 11).value  # col K

    # Parse orders (always a plain value)
    try:
        orders = int(orders_raw) if orders_raw else 0
    except (TypeError, ValueError):
        orders = 0

    # Parse dispatch (could be value or simple formula like =8232+24696)
    try:
        dispatch = int(dispatch_raw) if dispatch_raw else 0
    except (TypeError, ValueError):
        dispatch = 0

    produced = mtd_by_pid.get(pid_int, 0)
    is_active = (orders > 0) or (produced > 0) or (dispatch > 0)

    return {
        'type':     str(ws.cell(r, 2).value).strip().upper(),  # TUBE or PET
        'customer': ws.cell(r, 3).value,   # col C
        'product':  ws.cell(r, 4).value,   # col D
        'dia':      ws.cell(r, 5).value,   # col E
        'pid':      pid_int,               # col F
        'orders':   orders_raw,            # col G (keep original value)
        'dispatch_raw': ws.cell(r, 11).value,  # col K (preserve formula or value)
        'produced': produced,
        'is_active': is_active,
        'orig_row': r,
    }


all_tubes = []
all_pets  = []

for r in range(11, 200):
    type_val = ws.cell(r, 2).value   # col B
    name_val = ws.cell(r, 4).value   # col D

    if not type_val:
        # Could be TOTAL row or blank — skip
        continue

    type_str = str(type_val).strip().upper()
    if type_str not in ('TUBE', 'PET'):
        continue

    row_data = read_product_row(ws, r)
    if row_data is None:
        continue

    if type_str == 'TUBE':
        all_tubes.append(row_data)
    else:
        all_pets.append(row_data)

active_tubes   = [t for t in all_tubes if t['is_active']]
inactive_tubes = [t for t in all_tubes if not t['is_active']]
active_pets    = [p for p in all_pets  if p['is_active']]
inactive_pets  = [p for p in all_pets  if not p['is_active']]

# ── SORT BY DIAMETER (low → high) ───────────────────────────
# Tubes: dia is numeric (int/float like 16, 19, 20.5, 25, 30)
# PETs:  dia is string like "120 ml", "200 ml" — extract the number
import re

def dia_sort_key(row_data):
    """Extract numeric dia for sorting. Returns float."""
    dia = row_data.get('dia')
    if dia is None:
        return 9999.0
    if isinstance(dia, (int, float)):
        return float(dia)
    # String like "120 ml" — extract number
    m = re.search(r'(\d+(?:\.\d+)?)', str(dia))
    return float(m.group(1)) if m else 9999.0

active_tubes.sort(key=dia_sort_key)
inactive_tubes.sort(key=dia_sort_key)
active_pets.sort(key=dia_sort_key)
inactive_pets.sort(key=dia_sort_key)

print(f"  Tubes: {len(active_tubes)} active, {len(inactive_tubes)} inactive")
print(f"  PET:   {len(active_pets)} active, {len(inactive_pets)} inactive")

# ── SAVE TOTAL ROW FORMATTING ────────────────────────────────
# Find current TOTAL rows and copy their formatting for reuse.
def save_row_format(ws, r, cols):
    """Save font/fill/border/alignment/number_format for specified columns."""
    fmt = {}
    for c in cols:
        cell = ws.cell(r, c)
        fmt[c] = {
            'font':      copy(cell.font),
            'fill':      copy(cell.fill),
            'border':    copy(cell.border),
            'alignment': copy(cell.alignment),
            'number_format': cell.number_format,
        }
    return fmt

# Find current TOTAL rows
total_fmt = None
for r in range(11, 100):
    if ws.cell(r, 4).value == "TOTAL" and not ws.cell(r, 2).value:
        total_fmt = save_row_format(ws, r, DATA_COLS)
        break

# Also save a product row format (from row 11) as template
product_fmt = save_row_format(ws, 11, DATA_COLS)

# ── CALCULATE NEW LAYOUT ────────────────────────────────────
# Layout (no blank separators — TOTAL rows provide visual break):
#   Row 11              : first active tube (sorted by dia)
#   Row 10 + N_at       : last active tube
#   Row 11 + N_at       : TUBE TOTAL
#   Row 12 + N_at       : first active PET (sorted by ml)
#   Row 11 + N_at + N_ap: last active PET
#   Row 12 + N_at + N_ap: PET TOTAL
#   Row 13 + N_at + N_ap: first inactive tube
#   ...                 : inactive tubes, then inactive PETs

N_at = len(active_tubes)
N_ap = len(active_pets)
N_it = len(inactive_tubes)
N_ip = len(inactive_pets)

# Row positions (all 1-indexed)
FIRST_ROW = 11

tube_active_start = FIRST_ROW
tube_active_end   = FIRST_ROW + N_at - 1   # -1 because inclusive
tube_total_row    = FIRST_ROW + N_at

pet_active_start  = tube_total_row + 1
pet_active_end    = pet_active_start + N_ap - 1
pet_total_row     = pet_active_start + N_ap

inactive_tube_start = pet_total_row + 1
inactive_tube_end   = inactive_tube_start + N_it - 1

if N_it > 0:
    inactive_pet_start = inactive_tube_end + 1
else:
    inactive_pet_start = inactive_tube_start
inactive_pet_end    = inactive_pet_start + N_ip - 1

last_used_row = max(inactive_pet_end, pet_total_row)

print(f"\n  New layout:")
print(f"    Active TUBE : rows {tube_active_start}-{tube_active_end} ({N_at} items)")
print(f"    TUBE TOTAL  : row {tube_total_row}")
print(f"    Active PET  : rows {pet_active_start}-{pet_active_end} ({N_ap} items)")
print(f"    PET TOTAL   : row {pet_total_row}")
print(f"    Inactive TUBE: rows {inactive_tube_start}-{inactive_tube_end} ({N_it} items)")
print(f"    Inactive PET : rows {inactive_pet_start}-{inactive_pet_end} ({N_ip} items)")
print(f"    Last used row: {last_used_row}")

# ── FORMULA TEMPLATES ────────────────────────────────────────
# Simplified: no MONTH/YEAR filter (user keeps only current month in log)

TUBE_H_TPL = (
    '=SUMPRODUCT((Production_Log!$F$3:$F$8963=F{r})'
    '*(LEFT(Production_Log!$B$3:$B$8963,5)="Print")'
    '*(ISERROR(SEARCH("(Varnish)",Production_Log!$D$3:$D$8963)))'
    '*Production_Log!$H$3:$H$8963)'
)

PET_H_TPL = (
    '=SUMPRODUCT((Production_Log!$F$3:$F$8963=F{r})'
    '*Production_Log!$H$3:$H$8963)'
)

I_TPL = '=G{r}-H{r}'
J_TPL = '=IF(G{r}=0,"-",H{r}/G{r})'

# ── FORMATTING DEFINITIONS ───────────────────────────────────
# Product row formatting (from inspection of existing rows 11-24):
#   Font: Arial 10pt, not bold (except H=col8 and K=col11 which ARE bold)
#   Alignment: B=center, C=left, D=left, E=center, F-K=center (J=center)
#   Number format: B=General, C=General, D=#,##0, E=0.#, F=#,##0,
#                  G=#,##0, H=#,##0, I=#,##0, J=0%, K=#,##0

PRODUCT_FONT        = Font(name='Arial', size=10, bold=False)
PRODUCT_FONT_BOLD   = Font(name='Arial', size=10, bold=True)

PRODUCT_COL_STYLES = {
    2:  {'bold': False, 'nf': 'General',  'halign': 'center'},  # B = Type
    3:  {'bold': False, 'nf': 'General',  'halign': 'left'},    # C = Customer
    4:  {'bold': False, 'nf': '#,##0',    'halign': 'left'},    # D = Product Name
    5:  {'bold': False, 'nf': '0.#',      'halign': 'center'},  # E = Dia
    6:  {'bold': False, 'nf': '#,##0',    'halign': 'center'},  # F = Prod ID
    7:  {'bold': False, 'nf': '#,##0',    'halign': 'center'},  # G = Orders
    8:  {'bold': True,  'nf': '#,##0',    'halign': 'center'},  # H = MTD Produced
    9:  {'bold': False, 'nf': '#,##0',    'halign': 'center'},  # I = Remaining
    10: {'bold': False, 'nf': '0%',       'halign': 'center'},  # J = Compliance
    11: {'bold': True,  'nf': '#,##0',    'halign': 'center'},  # K = Dispatch
}

def apply_product_format(ws, r):
    """Apply standard product row formatting to all data columns in row r."""
    for c in DATA_COLS:
        cell = ws.cell(r, c)
        style = PRODUCT_COL_STYLES.get(c, {})
        cell.font = PRODUCT_FONT_BOLD if style.get('bold') else PRODUCT_FONT
        cell.alignment = Alignment(horizontal=style.get('halign', 'center'), vertical='center')
        cell.number_format = style.get('nf', 'General')
        cell.fill = PatternFill()  # clear any fill
        cell.border = Border()    # clear any border

# ── CLEAR ALL DATA ROWS ─────────────────────────────────────
# Clear from row 11 down to the old last used row (generous)
old_last = max(last_used_row, 70)
for r in range(11, old_last + 1):
    for c in DATA_COLS:
        ws.cell(r, c).value = None

# ── WRITE PRODUCT ROWS ──────────────────────────────────────
def write_product_row(ws, r, data):
    """Write one product row at row r with correct formulas and formatting."""
    ws.cell(r, 2).value = data['type']      # B = Type
    ws.cell(r, 3).value = data['customer']   # C = Customer
    ws.cell(r, 4).value = data['product']    # D = Product Name
    ws.cell(r, 5).value = data['dia']        # E = Dia
    ws.cell(r, 6).value = data['pid']        # F = Prod ID
    ws.cell(r, 7).value = data['orders']     # G = Orders

    # H = MTD Produced (formula — rebuild for this row position)
    if data['type'] == 'TUBE':
        ws.cell(r, 8).value = TUBE_H_TPL.format(r=r)
    else:
        ws.cell(r, 8).value = PET_H_TPL.format(r=r)

    # I = Remaining, J = Compliance (formulas)
    ws.cell(r, 9).value  = I_TPL.format(r=r)
    ws.cell(r, 10).value = J_TPL.format(r=r)

    # K = Dispatch (preserve original value/formula)
    ws.cell(r, 11).value = data['dispatch_raw']

    # Apply consistent product formatting
    apply_product_format(ws, r)


def write_total_row(ws, r, first_row, last_row, fmt_dict=None):
    """Write a TOTAL row at row r that sums from first_row to last_row."""
    # Clear all data columns first
    for c in DATA_COLS:
        ws.cell(r, c).value = None

    ws.cell(r, 4).value  = "TOTAL"                                   # D
    ws.cell(r, 7).value  = f"=SUM(G{first_row}:G{last_row})"         # G
    ws.cell(r, 8).value  = f"=SUM(H{first_row}:H{last_row})"         # H
    ws.cell(r, 9).value  = f'=SUMIF(I{first_row}:I{last_row},">"&0)' # I
    ws.cell(r, 11).value = f"=SUM(K{first_row}:K{last_row})"         # K

    # Apply TOTAL formatting
    if fmt_dict:
        for c in DATA_COLS:
            cell = ws.cell(r, c)
            f = fmt_dict.get(c)
            if f:
                cell.font      = copy(f['font'])
                cell.fill      = copy(f['fill'])
                cell.border    = copy(f['border'])
                cell.alignment = copy(f['alignment'])
                cell.number_format = f['number_format']
    else:
        # Fallback: just make bold
        bold_font = Font(bold=True)
        for c in DATA_COLS:
            ws.cell(r, c).font = bold_font


def write_blank_row(ws, r):
    """Ensure a row is blank (separator) with clean formatting."""
    for c in DATA_COLS:
        cell = ws.cell(r, c)
        cell.value = None
        cell.font = PRODUCT_FONT
        cell.fill = PatternFill()
        cell.border = Border()
        cell.alignment = Alignment()
        cell.number_format = 'General'


# ── WRITE ACTIVE TUBES ──────────────────────────────────────
for i, data in enumerate(active_tubes):
    write_product_row(ws, tube_active_start + i, data)

# ── WRITE TUBE TOTAL ────────────────────────────────────────
if N_at > 0:
    write_total_row(ws, tube_total_row, tube_active_start, tube_active_end, total_fmt)
else:
    # No active tubes — still write a TOTAL row showing zeros
    write_total_row(ws, tube_total_row, FIRST_ROW, FIRST_ROW, total_fmt)

# ── WRITE ACTIVE PETS ───────────────────────────────────────
for i, data in enumerate(active_pets):
    write_product_row(ws, pet_active_start + i, data)

# ── WRITE PET TOTAL ─────────────────────────────────────────
if N_ap > 0:
    write_total_row(ws, pet_total_row, pet_active_start, pet_active_end, total_fmt)
else:
    write_total_row(ws, pet_total_row, pet_active_start, pet_active_start, total_fmt)

# ── WRITE INACTIVE TUBES ────────────────────────────────────
for i, data in enumerate(inactive_tubes):
    write_product_row(ws, inactive_tube_start + i, data)

# ── WRITE INACTIVE PETS ─────────────────────────────────────
for i, data in enumerate(inactive_pets):
    write_product_row(ws, inactive_pet_start + i, data)

# ── CLEAR REMAINING OLD ROWS ────────────────────────────────
for r in range(last_used_row + 1, old_last + 1):
    write_blank_row(ws, r)

# ── UPDATE KPI SUMMARY FORMULAS (rows 6, 8) ─────────────────
# These use SUMIF across all product rows.  Update range to cover
# the new layout generously (11 to last_used_row).
end = last_used_row

# Row 6: Tube MTD summary
# D6 = total tube production
ws.cell(6, 4).value = f'=SUMIF($B${FIRST_ROW}:$B${end},"TUBE",$H${FIRST_ROW}:$H${end})'
# J6 = total tube dispatch
ws.cell(6, 10).value = f'=SUMIF($B${FIRST_ROW}:$B${end},"TUBE",$K${FIRST_ROW}:$K${end})'

# Row 8: PET MTD summary
ws.cell(8, 4).value = f'=SUMIF($B${FIRST_ROW}:$B${end},"PET",$H${FIRST_ROW}:$H${end})'
ws.cell(8, 10).value = f'=SUMIF($B${FIRST_ROW}:$B${end},"PET",$K${FIRST_ROW}:$K${end})'

# ── SAVE ─────────────────────────────────────────────────────
wb.save(EXCEL_PATH)

print(f"\n[OK] Dashboard sorted successfully -> {os.path.basename(EXCEL_PATH)}")
print(f"  Active:   {N_at} tubes + {N_ap} PET")
print(f"  Inactive: {N_it} tubes + {N_ip} PET")
print(f"  Open in Excel and press Ctrl+Shift+F9 to recalculate formulas.")
