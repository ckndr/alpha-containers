import openpyxl

wb = openpyxl.load_workbook(r'd:\Alpha\Tubex Records\Production report Jan-2026 till Date.xlsx', data_only=True)
ws = wb['Production Day wise']

print("--- ROWS 1 TO 5 ---")
for r in range(1, 6):
    row_vals = [ws.cell(r, c).value for c in range(1, 15)]
    print(f"Row {r}: {row_vals}")

print("\n--- FIRST 5 PF ROWS ---")
pf_count = 0
for r in range(1, ws.max_row+1):
    mach = str(ws.cell(r, 2).value or "").strip()
    if mach.startswith("PF"):
        row_vals = [ws.cell(r, c).value for c in range(1, 15)]
        print(f"Row {r}: {row_vals}")
        pf_count += 1
        if pf_count >= 5:
            break
