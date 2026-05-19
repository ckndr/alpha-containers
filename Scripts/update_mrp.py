import openpyxl
import glob
import os
import shutil
from copy import copy

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format

def is_numeric_item_id(val):
    if isinstance(val, int):
        return True
    if isinstance(val, str):
        val_clean = val.strip()
        if val_clean.isdigit():
            return True
    return False

def main():
    # 1. Find the latest AlphaContainers workbook (excluding backups)
    folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ac_files = [
        f for f in glob.glob(os.path.join(folder, "AlphaContainers*.xlsx"))
        if "backup" not in os.path.basename(f).lower()
    ]
    if not ac_files:
        print("Error: No AlphaContainers*.xlsx found!")
        return
    
    path = sorted(ac_files)[-1]
    # Name the backup with 'backup_' prefix so it does not match the 'AlphaContainers*' glob pattern
    backup_path = os.path.join(folder, "backup_" + os.path.basename(path))
    
    # Restore from backup if it already exists to guarantee clean/idempotent run
    if os.path.exists(backup_path):
        print(f"Restoring clean spreadsheet from backup: {os.path.basename(backup_path)} -> {os.path.basename(path)}")
        shutil.copyfile(backup_path, path)
    else:
        print(f"Creating backup of {os.path.basename(path)} -> {os.path.basename(backup_path)}...")
        shutil.copyfile(path, backup_path)
    
    print(f"Loading {os.path.basename(path)}...")
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb['MRP']
    
    # 2. Insert new row at row 8
    print("Inserting new row at row 8...")
    ws.insert_rows(8)
    
    # Copy styles from row 7 to the new row 8
    for col in range(1, ws.max_column + 1):
        copy_style(ws.cell(row=7, column=col), ws.cell(row=8, column=col))
        
    # Populate row 8 values
    ws.cell(row=8, column=1, value=25)  # Dia
    ws.cell(row=8, column=2, value="Samsol International Private Limited")  # Customer
    ws.cell(row=8, column=3, value="TUBES COMMON RED")  # Product Name
    ws.cell(row=8, column=4, value=6470)  # Product ID
    ws.cell(row=8, column=5, value="TBD")  # Job Order #
    ws.cell(row=8, column=6, value=23000)  # Required Qty
    
    # Produced formula
    g8_formula = (
        '=SUMPRODUCT((Production_Log!$F$3:$F$8963=D8)'
        '*(LEFT(Production_Log!$B$3:$B$8963,5)="Print")'
        '*(ISERROR(SEARCH("(Varnish)",Production_Log!$D$3:$D$8963)))'
        '*(MONTH(Production_Log!$A$3:$A$8963)=MONTH(TODAY()))'
        '*(YEAR(Production_Log!$A$3:$A$8963)=YEAR(TODAY()))'
        '*Production_Log!$H$3:$H$8963)'
    )
    ws.cell(row=8, column=7, value=g8_formula)
    
    # Remaining Balance formula
    ws.cell(row=8, column=8, value="=F8-G8")
    
    print("New Row 8 populated.")
    
    # 3. Update Total Row formulas (now at row 14)
    print("Updating Total Row (Row 14) formulas...")
    ws.cell(row=14, column=6, value="=SUM(F3:F13)")
    ws.cell(row=14, column=7, value="=SUM(G3:G13)")
    ws.cell(row=14, column=8, value='=SUMIF(H3:H13, ">0")')
    
    # 4. Update the second section (Material Requirement Plan)
    # The items start at Row 17 and go up to Row 156
    print("Updating Material Requirement Plan section formulas (Rows 17 to 156)...")
    active_rows = [3, 4, 5, 6, 7, 8, 9, 10, 11, 13]  # Skipping row 12 (V-HC Brown - on hold)
    
    for r in range(17, 157):
        item_id = ws.cell(row=r, column=1).value
        # Check if this row is an actual item row (has a numeric Item ID)
        if not is_numeric_item_id(item_id):
            print(f"  Skipping row {r} (non-item: {item_id})")
            continue
            
        # Column E: Required Qty formula
        ws.cell(row=r, column=5, value=f"=SUMPRODUCT((TableBOM[Item ID]=A{r})*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %])*SUMIF($D$3:$D$13,TableBOM[Product ID],$H$3:$H$13)/1000)")
        
        # Column H: Product Name(s) formula
        inner_parts_h = [
            f'IF((COUNTIFS(TableBOM[Product ID],$D${x},TableBOM[Item ID],$A{r})>0)*($H${x}>0),$C${x}&\", \",\"\")'
            for x in active_rows
        ]
        concat_h = " & ".join(inner_parts_h)
        formula_h = f'=IF(LEN({concat_h})>1,LEFT({concat_h},LEN({concat_h})-2),\"\")'
        ws.cell(row=r, column=8, value=formula_h)
        
        # Column I: Job Order #(s) formula
        inner_parts_i = [
            f'IF((COUNTIFS(TableBOM[Product ID],$D${x},TableBOM[Item ID],$A{r})>0)*($H${x}>0),$E${x}&\", \",\"\")'
            for x in active_rows
        ]
        concat_i = " & ".join(inner_parts_i)
        formula_i = f'=IF(LEN({concat_i})>1,LEFT({concat_i},LEN({concat_i})-2),\"\")'
        ws.cell(row=r, column=9, value=formula_i)
        
    # 5. Update CAPS ON HOLD sub-section (now at Rows 162 to 165)
    print("Updating CAPS ON HOLD sub-section formulas (Rows 162 to 165)...")
    for r in range(162, 166):
        ws.cell(row=r, column=5, value=f'=IFERROR(INDEX($C$50:$C$71,MATCH(D{r},$A$50:$A$71,0)),"-")')
        
    # 6. Save workbook
    print(f"Saving changes to {os.path.basename(path)}...")
    wb.save(path)
    print("MRP Sheet updated successfully!")

if __name__ == "__main__":
    main()
