"""
Inspect Production report Jan-2026 till Date.xlsx and Samsol_Production_and_Dispatch.xlsx
"""
import openpyxl

files = [
    r"d:\Alpha\Tubex Records\Production report Jan-2026 till Date.xlsx",
    r"d:\Alpha\Tubex Records\Samsol_Production_and_Dispatch.xlsx"
]

for fpath in files:
    print(f"\n==================================================")
    print(f" File: {fpath}")
    print(f"==================================================")
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"\n  --- Sheet: {sname} ({ws.max_row} rows, {ws.max_column} cols) ---")
            for r in range(1, min(10, ws.max_row+1)):
                vals = [ws.cell(r, c).value for c in range(1, min(12, ws.max_column+1))]
                if any(vals):
                    print(f"    R{r:2d}: {vals}")
        wb.close()
    except Exception as e:
        print(f"  ERROR: {e}")
