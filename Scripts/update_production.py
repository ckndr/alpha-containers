"""
Alpha Containers - Production Log Auto-Updater v21
──────────────────────────────────────────────────
CHANGE LOG v21 (vs v20):
  ALIAS ADD:
    ("trp bottle", "150 ml") -> TRANSPARENT BOTTLE 150ML (PID 8001)
  CUSTOMER RENAME:
    "Alpha Lab" -> "Alpha Labs PVT LTD" (PID 8001 and mappings)

CHANGE LOG v20 (vs v19):
  ALIAS ADD:
    ("alpha lab\ttrp bottle", "150 ml") -> TRANSPARENT BOTTLE 150ML (PID 8001)
    ("mabley beauty\tvince nutural", "30") -> VINCE NURTURAL (PID 5814)

CHANGE LOG v19 (vs v18):
  VARNISH POLICY — no varnish-only BOMs, no PID assigned to varnish passes.
  All varnish ALIASES entries updated: PID set to None.
  Varnish passes are logged in Production_Log by product name with "(Varnish)"
  tag for Dashboard filtering — no PID required or assigned.
  PID 9002 removed from PID_TO_CUSTOMER (no longer tracked as a product).
  Affected entries:
    ("vince nurtural (varnish)", "30")  → (name, None)  was 9002
    ("hola hair (varnish)",      "30")  → (name, None)  was 6515
    ("hola hair (varnish)",      "32")  → (name, None)  was 6515
    ("hiba h.c (varnish)",       "30")  → (name, None)  was 6312
    ("vince his (varnish)",      "25")  → (name, None)  was 6228
    ("vince his (varnish)",      "20.5")→ (name, None)  was 6077

CHANGE LOG v18 (vs v17):
  ALIAS ADD (Dia 35):
    ("vivid h.c brown", "35") -> V- HC BROWN  (PID 6530)
    ("vivid h.c black", "35") -> V-HC BLACK   (PID 6531)
    Imran writes "Vivid H.C Brown" / "Vivid H.C Black" explicitly.
    The existing generic ("vivid h.c", "35") defaulted to BROWN only and
    would miss the BLACK variant entirely. Added colour-specific entries
    so BLACK routes to 6531 (Adore) correctly.
  FG_ALIASES ADD (Dia 35):
    ("vivid h.c brown", "35", "vivid cosmetics") -> V- HC BROWN  (PID 6530)
    ("vivid h.c black", "35", "vivid cosmetics") -> V-HC BLACK   (PID 6531)
  ALIAS ADD (Dia 20.5):
    ("vince his (varnish)", "20.5") -> VINCE HIS BEARD & MUSTACHE COLOR (Varnish), PID 6077
    Imran writes "Vince His (Varnish)" for the 20.5mm varnish pass.
    Using PID 6077 (same as print pass) — Dashboard filters by "(Varnish)"
    in product name so KPIs are not double-counted.
    NOTE: A dedicated placeholder PID (like 9007) should eventually be
    added to Product_Catalog and Dashboard inactive section, matching how
    PID 9002 handles the VINCE NURTURAL (Varnish) pass at 30mm.

CHANGE LOG v17 (vs v16):
  BUGFIX — 19mm S-43 / S-45 disambiguation:
    PID 6624 was mislabeled as "S-43 DIA 19" in Product_Catalog; BOM 192 confirms
    it is "S-45 DIA 19MM". PID 6623 (BOM 193) is the separate S-43 DIA 19MM product.

    ALIASES (Dia 19) updated:
      ("samsol 43", "19")    -> S-43 DIA 19MM, PID 6623   [was 6624 — WRONG]
      ("samsol hc 43", "19") -> S-43 DIA 19MM, PID 6623   [was 6624 — WRONG]
      ADD ("samsol 45", "19")    -> S-45 DIA 19MM, PID 6624
      ADD ("samsol hc 45", "19") -> S-45 DIA 19MM, PID 6624

    FG_ALIASES (Dia 19) updated:
      ("tube 43", "19", "samsol") -> S-43 DIA 19MM, PID 6623   [was 6624]
      ("tube 45", "19", "samsol") -> S-45 DIA 19MM, PID 6624   [was 6624 w/ S-43 name]

    PID_TO_CUSTOMER:
      ADDED 6623 -> Samsol International Private Limited

CHANGE LOG v16 (vs v15):
  BUGFIX: ALIASES ("vince his","25"), ("vincehis","25"), ("vince his (varnish)","25")
  were all mapped to PID 6077 (VINCE HIS BEARD & MUSTACHE COLOR -- 20.5mm product).
  Fixed to PID 6228 (HIS ONLY HAIR COLOR CREAM 40GM -- correct 25mm product).
  FG_ALIASES already had the correct 6228 mapping; only ALIASES was wrong.

CHANGE LOG v15 (vs v14):
  NEW: FG Stock sheet integration.
  Added read_fg_stock() — reads 'FG Stock In hand' sheet from Production.xlsx,
  filters to latest date only (all older dates discarded), maps product/customer
  names to catalog names and PIDs using FG_ALIASES dict.
  Added write_fg_stock() — wipes and rewrites FG Stock sheet in Tubex.
  Title row updated with latest date on each run.
  Row colour coding: green fill = OK/Ready, orange fill = Not Ready / In Progress.
  FG_ALIASES key: (product_desc_lower, normalised_dia, customer_lower).
  Dia normalisation: strips trailing ' dia', strips whitespace, lowercases,
  collapses ' ml' -> 'ml'. Handles Imran's inconsistent spacing.
  PID 9003 (S-43 DIA 19) → PID 6624 (ERP confirmed). Updated everywhere:
    ALIASES ("samsol 43","19") and ("samsol hc 43","19") → 6624
    FG_ALIASES ("tube 43","19","samsol") → 6624
    FG_ALIASES ("tube 45","19","samsol") → 6624 (same product, Imran name varies)
    PID_TO_CUSTOMER: removed 9003, added 6624.

CHANGE LOG v14 (vs v13):
  ALIAS ADD: ("vivid h.c", "35") and ("vivid hc", "35") -> V- HC BROWN (PID 6530)
  Imran enters "Vivid H.C" for the Adore V- HC BROWN 35mm product.

CHANGE LOG v13 (vs v12):
  BUGFIX: Replaced Unicode arrow (→) with ASCII (->) in 3 print statements.
  Caused UnicodeEncodeError on Windows cp1252 consoles when output was
  redirected to log files by Run_All_Updates.bat.
  ADDED: warnings.filterwarnings to suppress openpyxl Data Validation
  UserWarning from cluttering the issues summary.

CHANGE LOG v12 (vs v11):
  STRUCTURE: ALIASES dict flattened — all customer-name section headings removed.
  One flat list sorted by dia/volume. Add new entries at the bottom.
  ALIAS ADD: ("samsol men blue", "25") -> TUBES MEN BLUE (PID 6506)
  ALIAS ADD: ("hola hair", "30") and ("hola hair (varnish)", "30") -> H.H 100GM (6515)
    Imran enters dia=30 for Hola Hair; actual product is 32mm. Both "30" and "32"
    entries retained so either Imran input is handled correctly.
  ALIAS ADD: ("samsol 43", "19") and ("samsol hc 43", "19") -> S-43 DIA 19 (PID 9003)
  ALIAS ADD: ("dowfen gel", "30") -> DOWFEN GEL 50G (PID 6595) — Seatle (Pvt) Ltd

KEY BUSINESS RULES:
  - MHK (Muhammad Hashim Khan) is a cash/no-GST placeholder account in ERP, NOT a real customer.
    Real brand customers use MHK for cash transactions. PID_TO_CUSTOMER maps each product
    to its actual brand customer. Customer resolved from PID, never from Imran's name.
  - Samsol International is a SEPARATE customer from MHK. PID disambiguates.
  - Varnish pass rule: "(Varnish)" in product name → PID=None (no PID assigned).
    Varnish passes are tracked by name only. No BOM, no PID, no ERP product.
    Regular print pass → actual ERP PID (e.g. Vince Nurtural → 5814).
  - PET BOMs are placeholder only (PIDs 8001–8013). No ERP BOMs yet.
  - Alpha Lab = Alpha Containers' Karachi subsidiary. Maps to "Alpha Labs PVT LTD".
  - Mabley's 130ml bottle is written "150ml" by Imran. Script fixes to PID 8010.
  - Samsol 43 at 20.5mm = ERP "S-43 DIA 20.5" (PID 5699). Verified Apr-2026.
  - FG Stock: only latest date rows are kept. Older dates discarded every run.
  - 19mm Samsol: PID 6623 = S-43 DIA 19MM (BOM 193); PID 6624 = S-45 DIA 19MM (BOM 192).
    These are TWO distinct products. "samsol 43"/"tube 43" -> 6623. "samsol 45"/"tube 45" -> 6624.
"""

import os
import openpyxl
import re
import glob
import warnings
warnings.filterwarnings("ignore", message=".*Data Validation.*")
warnings.filterwarnings("ignore", message=".*extension.*")
from datetime import datetime
from alpha_checks import check_freshness, check_not_locked, log_mismatches, replace_copy_export


# ─────────────────────────────────────────────────────────────────────────────
# ALIASES: (imran_product_name.lower().strip(), dia_string) → (catalog_name, PID)
# dia_string must match EXACTLY what Imran types.
# Tubes: "19","16","20.5","25","30","32","35"
# PET:   "120 ml","150 ml","200 ml","300 ml" (note space before ml)
# ─────────────────────────────────────────────────────────────────────────────
ALIASES = {
    # ── Dia 16 ──────────────────────────────────────────────────────────────
    ("eczumas ointment",           "16"): ("ECZEMUS OINTMENT 0.03% 10G",           6561),

    # ── Dia 19 ──────────────────────────────────────────────────────────────
    ("pyodine gel",                "19"): ("PYODINE GEL 20GM",                     1085),
    ("philogin gel",               "19"): ("PHLOGIN GEL 20G",                      6559),
    ("phlogin gel",                "19"): ("PHLOGIN GEL 20G",                      6559),
    ("contractubex gel",           "19"): ("CONTRATUBEX GEL 20G",                  6560),
    # v17 BUGFIX: 6623=S-43 DIA 19MM, 6624=S-45 DIA 19MM (two separate ERP products)
    ("samsol 43",                  "19"): ("S-43 DIA 19MM",                        6623),  # v17: was 6624 (WRONG)
    ("samsol hc 43",               "19"): ("S-43 DIA 19MM",                        6623),  # v17: was 6624 (WRONG)
    ("samsol 45",                  "19"): ("S-45 DIA 19MM",                        6624),  # v17: NEW entry
    ("samsol hc 45",               "19"): ("S-45 DIA 19MM",                        6624),  # v17: NEW entry

    # ── Dia 20.5 ────────────────────────────────────────────────────────────
    # VERIFIED Apr-2026: ERP (POF# 6865, 6873) confirms "S-43 DIA 20.5" = PID 5699. PID 5732 was wrong (fixed v7).
    ("samsol 43",                "20.5"): ("S-43 DIA 20.5",                        5699),
    ("samsol hc 43",             "20.5"): ("S-43 DIA 20.5",                        5699),
    ("samsol 45",                "20.5"): ("SAMSOL HAIR COLOR 45 DIA 20.5 MM",     5731),
    ("samsol hc 45",             "20.5"): ("SAMSOL HAIR COLOR 45 DIA 20.5 MM",     5731),
    ("samsol red",               "20.5"): ("S-43 DIA 20.5",                        5699),
    ("vince his",               "20.5"): ("VINCE HIS BEARD & MUSTACHE COLOR",      6077),
    ("vince his (varnish)",     "20.5"): ("VINCE HIS BEARD & MUSTACHE COLOR (Varnish)", None),  # v19: no PID for varnish passes

    # ── Dia 25 ──────────────────────────────────────────────────────────────
    ("s 43 25mm",                  "25"): ("S 43 25MM",                            3447),
    ("samsol 43",                  "25"): ("S 43 25MM",                            3447),
    ("samsol 45",                  "25"): ("S-45",                                 5389),
    ("samsol 45/43",               "25"): ("TUBES",                                3726),
    ("samsol common red",          "25"): ("TUBES COMMON RED",                     6470),
    ("tubes men blue",             "25"): ("TUBES MEN BLUE",                       6506),
    ("samsol men blue",            "25"): ("TUBES MEN BLUE",                       6506),  # Added v12
    ("samsol common purple",       "25"): ("TUBE COMMON PURPLE",                   6532),
    ("tube common purple",         "25"): ("TUBE COMMON PURPLE",                   6532),
    ("samsol red",                 "25"): ("SAMSOL RED 25MM",                      9006),
    ("vince his",                  "25"): ("HIS ONLY HAIR COLOR CREAM 40GM",      6228),  # BUGFIX v16: 25mm = PID 6228, not 6077 (6077 is 20.5mm)
    ("vincehis",                   "25"): ("HIS ONLY HAIR COLOR CREAM 40GM",      6228),  # BUGFIX v16
    ("vince his (varnish)",        "25"): ("HIS ONLY HAIR COLOR CREAM 40GM (Varnish)", None),  # v19: no PID for varnish passes

    # ── Dia 30 ──────────────────────────────────────────────────────────────
    ("hello hair",                 "30"): ("HELLO HAIR COLOR",                     6206),
    ("eazi color red",             "30"): ("EAZICOLOR PREMIUM HAIR COLOUR 60ML",   4050),
    ("hiba h.c",                   "30"): ("HIBA S HAIR COLOR 60ML",               6312),
    ("hiba h.c (varnish)",         "30"): ("HIBA S HAIR COLOR 60ML (Varnish)",     None),  # v19: no PID for varnish passes
    ("signature h.c",              "30"): ("SIGNATURE HAIR COLOR CREAM 60ML",      6416),
    ("active pro",                 "30"): ("ACTIVE PRO HAIR COLOR CREAM 60ML",     6337),
    ("hola hair",                  "30"): ("H.H 100GM",                            6515),  # Imran enters dia=30; actual product is 32mm — Added v12
    ("hola hair (varnish)",        "30"): ("H.H 100GM (Varnish)",                  None),  # v19: no PID; dia=30 alias, actual 32mm
    ("vince nurtural",             "30"): ("VINCE NURTURAL",                       5814),  # print pass — ERP BOM#95
    ("vince nurtural (varnish)",   "30"): ("VINCE NURTURAL (Varnish)",             None),  # v19: no PID for varnish passes
    ("dowfen gel",                 "30"): ("DOWFEN GEL 50G",                       6595),  # Added v12

    # ── Dia 32 ──────────────────────────────────────────────────────────────
    ("hola hair",                  "32"): ("H.H 100GM",                            6515),
    ("hola hair (varnish)",        "32"): ("H.H 100GM (Varnish)",                  None),  # v19: no PID for varnish passes
    ("mega grey",                  "32"): ("M.G",                                  5782),

    # ── Dia 35 ──────────────────────────────────────────────────────────────
    ("anvil 43",                   "35"): ("ANVIL 43",                             6020),
    ("anvil 45",                   "35"): ("ANVIL 45",                             6021),
    ("vivid h.c",                  "35"): ("V- HC BROWN",                          6530),  # Added v14
    ("vivid hc",                   "35"): ("V- HC BROWN",                          6530),  # Added v14
    ("vivid h.c brown",            "35"): ("V- HC BROWN",                          6530),  # Added v18: Imran writes colour explicitly
    ("vivid h.c black",            "35"): ("V-HC BLACK",                           6531),  # Added v18: was missing — generic entry only mapped to BROWN

    # ── PET (PF Machine — all placeholder BOMs) ─────────────────────────────
    ("transparent bottle",        "150 ml"): ("TRANSPARENT BOTTLE 150ML",          8001),
    ("transparent bottle",        "300 ml"): ("TRANSPARENT BOTTLE 300ML",          8009),
    ("yellow small bottle",       "120 ml"): ("PET BOTTLE SMALL (120ML) YELLOW",   8005),
    ("yellow large bottle",       "200 ml"): ("YELLOW LARGE BOTTLE 200ML",         8006),
    ("yellow bottle",             "200 ml"): ("YELLOW LARGE BOTTLE 200ML",         8006),
    ("white bottle",              "200 ml"): ("WHITE BOTTLE 200ML",                8007),
    ("black bottle",              "200 ml"): ("BLACK BOTTLE 200ML",                8008),
    ("samsol black bottle",       "120 ml"): ("BLACK SMALL BOTTLE 120ML",          8011),
    ("yellow bottle",             "120 ml"): ("PET BOTTLE SMALL (120ML) YELLOW",   8005),
    ("trp bottle",                "130 ml"): ("PET BOTTLE SMALL (130ML) TRANSPARENT", 8010),
    ("trp bottle",                "150 ml"): ("TRANSPARENT BOTTLE 150ML",          8001),
    ("alpha lab\ttrp bottle",      "150 ml"): ("TRANSPARENT BOTTLE 150ML",          8001),
    ("mabley beauty\tvince nutural", "30"): ("VINCE NURTURAL",                     5814),
}


# ─────────────────────────────────────────────────────────────────────────────
# FG_ALIASES: (product_desc_lower, normalised_dia, customer_lower) → (catalog_name, PID)
#
# Dia normalisation: strip whitespace, strip trailing ' dia', collapse ' ml'→'ml',
# lowercase. Examples: "25 dia"→"25", "20.5 dia"→"20.5", "150 ml"→"150ml".
# Customer normalisation: .lower().strip().
#
# Key = 3-tuple so customer disambiguates "Tube" at 30mm across multiple brands.
# Add new entries at the bottom. Entries are permanent — old dates also use this map.
# ─────────────────────────────────────────────────────────────────────────────
FG_ALIASES = {
    # ── Dia 19 ──────────────────────────────────────────────────────────────
    ("phlogin gel tube",   "19", "brooks pharma"):  ("PHLOGIN GEL 20G",                    6559),
    # v17 BUGFIX: 6623=S-43, 6624=S-45 — two distinct 19mm Samsol products
    ("tube 43",            "19", "samsol"):          ("S-43 DIA 19MM",                      6623),  # v17: was 6624
    ("tube 45",            "19", "samsol"):          ("S-45 DIA 19MM",                      6624),  # v17: now correctly maps to S-45

    # ── Dia 20.5 ────────────────────────────────────────────────────────────
    ("tube 43",          "20.5", "samsol"):          ("S-43 DIA 20.5",                      5699),
    ("tube 45",          "20.5", "samsol"):          ("SAMSOL HAIR COLOR 45 DIA 20.5 MM",   5731),
    ("vince his",        "20.5", "mabley beauty"):   ("VINCE HIS BEARD & MUSTACHE COLOR",   6077),

    # ── Dia 25 ──────────────────────────────────────────────────────────────
    ("tube 43",            "25", "samsol"):          ("S 43 25MM",                          3447),
    ("tube 45",            "25", "samsol"):          ("S-45",                               5389),
    ("common purple tube", "25", "samsol"):          ("TUBE COMMON PURPLE",                 6532),
    ("common red tube",    "25", "samsol"):          ("TUBES COMMON RED",                   6470),
    ("men blue",           "25", "samsol"):          ("TUBES MEN BLUE",                     6506),
    ("vince his",          "25", "mabley beauty"):   ("HIS ONLY HAIR COLOR CREAM 40GM",    6228),

    # ── Dia 30 ──────────────────────────────────────────────────────────────
    ("tube",               "30", "active pro"):      ("ACTIVE PRO HAIR COLOR CREAM 60ML",   6337),
    ("tube",               "30", "golden pearl"):    ("HELLO HAIR COLOR",                   6206),
    ("hello hair",         "30", "golden pearl"):    ("HELLO HAIR COLOR",                   6206),
    ("tube",               "30", "mabley beauty"):   ("VINCE NURTURAL",                     5814),  # printing pass
    ("vince nurtural",     "30", "mabley beauty"):   ("VINCE NURTURAL",                     5814),
    ("tube",               "30", "signature h.c"):   ("SIGNATURE HAIR COLOR CREAM 60ML",    6416),
    ("eazi color red tube","30", "pbs"):             ("EAZICOLOR PREMIUM HAIR COLOUR 60ML", 4050),
    ("hiba h.c tube",      "30", "pbs"):             ("HIBA S HAIR COLOR 60ML",             6312),
    ("dowfen gel",         "30", "seatle"):          ("DOWFEN GEL 50G",                     6595),

    # ── Dia 32 ──────────────────────────────────────────────────────────────
    ("tube",               "32", "hola hair"):       ("H.H 100GM",                          6515),
    ("tube",               "32", "mega grey"):       ("M.G",                                5782),

    # ── Dia 35 ──────────────────────────────────────────────────────────────
    ("anvil 43",           "35", "adore"):           ("ANVIL 43",                           6020),
    ("tube 43",            "35", "adore"):           ("ANVIL 43",                           6020),  # Imran sometimes writes "Tube 43"
    ("anvil 45",           "35", "adore"):           ("ANVIL 45",                           6021),
    ("tube 45",            "35", "adore"):           ("ANVIL 45",                           6021),  # Imran sometimes writes "Tube 45"
    ("vivid h.c",          "35", "vivid cosmetics"): ("V- HC BROWN",                        6530),
    ("vivid hc",           "35", "vivid cosmetics"): ("V- HC BROWN",                        6530),
    ("vivid h.c brown",    "35", "vivid cosmetics"): ("V- HC BROWN",                        6530),  # Added v18
    ("vivid h.c black",    "35", "vivid cosmetics"): ("V-HC BLACK",                         6531),  # Added v18

    # ── PET ─────────────────────────────────────────────────────────────────
    ("trp bottle",      "150ml", "alpha lab"):       ("TRANSPARENT BOTTLE 150ML",           8001),
    ("trp bottle",       "150ml","alpha lab"):       ("TRANSPARENT BOTTLE 150ML",           8001),
    ("trp bottle",      "150ml", "alpha labs pvt ltd"): ("TRANSPARENT BOTTLE 150ML",        8001),
    ("yellow bottle",   "120ml", "samsol"):          ("PET BOTTLE SMALL (120ML) YELLOW",    8005),
    ("yellow bottle",   "200ml", "samsol"):          ("YELLOW LARGE BOTTLE 200ML",          8006),
    ("white bottle",    "200ml", "samsol"):          ("WHITE BOTTLE 200ML",                 8007),
    ("black bottle",    "200ml", "samsol"):          ("BLACK BOTTLE 200ML",                 8008),
}


# ─────────────────────────────────────────────────────────────────────────────
# CUSTOMER_MAP: Imran shorthand → ERP name (fallback; PID_TO_CUSTOMER overrides)
# ─────────────────────────────────────────────────────────────────────────────
CUSTOMER_MAP = {
    "brooks pharma":          "Brookes Pharma Private Limited",
    "samsol international":   "Samsol International Private Limited",
    "samsol":                 "Samsol International Private Limited",
    "golden pearl":           "Golden Pearl Cosmetics (PVT) LTD",
    "golden pearl cosmetics": "Golden Pearl Cosmetics (PVT) LTD",
    "adore":                  "Muhammad Hashim Khan",
    "pbs":                    "Professional Beauty Solution (PVT) LTD.Pakistan",
    "mabley beauty":          "Mablay Beauty PVT LTD.",
    "mega grey":              "Muhammad Hashim Khan",
    "al rehman group":        "Al-Rehman Group",
    "alpha lab":              "Alpha Labs PVT LTD",
    "alpha container":        "Alpha Labs PVT LTD",
    "hashim khan":            "Muhammad Hashim Khan",
    "m hashim khan":          "Muhammad Hashim Khan",
    "active pro":             "Al-Rehman Group",
    "seatle":                 "Seatle (Private) Limited",
    "dtm":                    "Muhammad Hashim Khan",
    "bahadur":                "Bahadur",
    "swaleheen akhtar":       "Muhammad Hashim Khan",
    "mr. swaleheen akhtar":   "Muhammad Hashim Khan",
}


# ─────────────────────────────────────────────────────────────────────────────
# PID_TO_CUSTOMER: ERP-authoritative customer. ALWAYS overrides CUSTOMER_MAP.
# ─────────────────────────────────────────────────────────────────────────────
PID_TO_CUSTOMER = {
    1085: "Brookes Pharma Private Limited",
    6559: "Brookes Pharma Private Limited",
    6560: "Brookes Pharma Private Limited",
    6561: "Brookes Pharma Private Limited",
    3447: "Samsol International Private Limited",
    5699: "Samsol International Private Limited",
    5782: "Mega Grey",
    6020: "Adore",
    6021: "Adore",
    6338: "DTM",
    6470: "Samsol International Private Limited",
    6515: "Samsol International Private Limited",
    6530: "Adore",
    6531: "Adore",
    3726: "Samsol International Private Limited",
    5389: "Samsol International Private Limited",
    5731: "Samsol International Private Limited",
    5732: "Samsol International Private Limited",
    6506: "Samsol International Private Limited",
    6532: "Samsol International Private Limited",
    6556: "Samsol International Private Limited",
    6557: "Samsol International Private Limited",
    4050: "Professional Beauty Solution (PVT) LTD.Pakistan",
    6312: "Professional Beauty Solution (PVT) LTD.Pakistan",
    6206: "Golden Pearl Cosmetics (PVT) LTD",
    5814: "Mablay Beauty PVT LTD.",
    6077: "Mablay Beauty PVT LTD.",
    6228: "Mablay Beauty PVT LTD.",
    6337: "Al-Rehman Group",
    6416: "Al-Rehman Group",
    6595: "Seatle (Private) Limited",
    # 9002 removed v19 — no PID for varnish passes
    6623: "Samsol International Private Limited",  # S-43 DIA 19MM (BOM 193) — added v17
    6624: "Samsol International Private Limited",  # S-45 DIA 19MM (BOM 192) — confirmed ERP PID
    9004: "Bahadur",
    9006: "Samsol International Private Limited",
    8001: "Alpha Labs PVT LTD",
    8005: "Samsol International Private Limited",
    8006: "Samsol International Private Limited",
    8007: "Samsol International Private Limited",
    8008: "Samsol International Private Limited",
    8009: "Mablay Beauty PVT LTD.",
    8010: "Mablay Beauty PVT LTD.",
    8011: "Samsol International Private Limited",
    8012: "Samsol International Private Limited",
    8013: "Mablay Beauty PVT LTD.",
}


# ─────────────────────────────────────────────────────────────────────────────
COLUMN_CANDIDATES = {
    'Date':                 ['date', '`'],
    'Machines':             ['machines', 'machine'],
    'Customer':             ['customer'],
    'Product Name':         ['product name', 'product', 'productname'],
    'Dia(mm)/Volume':       ['dia(mm)/volume', 'dia', 'dia(mm)', 'volume', 'dia/volume'],
    'Good Production':      ['good production', 'good qty', 'good', 'goodproduction'],
    'Wastage':              ['wastage', 'reject', 'rejects', 'rejection'],
    'Total Production':     ['total production', 'total', 'totalproduction'],
    'Mechanical Downtime':  ['mechanical downtime', 'mech dt', 'mechanical'],
    'Electrical Downtime':  ['electrical downtime', 'elec dt', 'electrical'],
    'Material Shortage ':   ['material shortage', 'material shortage ', 'mat shortage'],
    'ChangeOver':           ['changeover', 'change over', 'co'],
    'Operations Downtime':  ['operations downtime', 'ops dt', 'operations'],
    'Power Shut Down':      ['power shut down', 'power shutdown', 'power'],
    'Gas Shut Down':        ['gas shut down', 'gas shutdown', 'gas'],
    'Workers Shortage':     ['workers shortage', 'worker shortage', 'workers'],
}


def build_col_map(df_columns):
    df_lower = {str(c).strip().lower(): c for c in df_columns}
    result = {}
    for key, candidates in COLUMN_CANDIDATES.items():
        for cand in candidates:
            if cand in df_lower:
                result[key] = df_lower[cand]
                break
    return result


def get_val(row, col_map, key, default=None):
    actual = col_map.get(key)
    if actual is None:
        return default
    import pandas as pd
    val = row.get(actual, default)
    try:
        if pd.isna(val):
            return default
    except Exception:
        pass
    return val




def find_files():
    folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    replace_copy_export(folder, "Production.xlsx")

    ac_files   = glob.glob(os.path.join(folder, "Tubex*.xlsx"))
    prod_files = glob.glob(os.path.join(folder, "Production*.xlsx"))
    if not ac_files:
        print("  ERROR: No Tubex*.xlsx found in: " + folder)
        return None, None
    if not prod_files:
        print("  ERROR: No Production*.xlsx found in: " + folder)
        return None, None
    ac   = sorted(ac_files)[-1]
    prod = sorted(prod_files)[-1]
    print("  Tubex File: " + os.path.basename(ac))
    print("  Production: " + os.path.basename(prod))
    return ac, prod


def detect_header_row(prod_path):
    import pandas as pd
    KNOWN = {
        'date', 'month', 'machines', 'machine', 'customer',
        'product name', 'product', 'dia', 'dia(mm)', 'volume',
        'good production', 'good', 'wastage', 'reject', 'total production',
        'mechanical downtime', 'electrical downtime', 'changeover',
        'operations downtime', 'power shut down', 'gas shut down',
        'workers shortage', 'material shortage', 'actual time', 'downtime',
    }
    raw = pd.read_excel(
        prod_path, sheet_name='Production Day wise',
        header=None, nrows=10
    )
    best_row  = 2
    best_score = 0
    for i, row in raw.iterrows():
        vals = [str(v).strip().lower() for v in row if str(v).strip() not in ('', 'nan')]
        score = sum(1 for v in vals if any(v.startswith(k) or k in v for k in KNOWN))
        if score > best_score:
            best_score = score
            best_row   = i
    print("  Header row detected at row %d (score=%d)" % (best_row + 1, best_score))
    return best_row


def parse_date(date_raw):
    import pandas as pd
    try:
        if pd.isna(date_raw):
            return None
    except Exception:
        pass
    try:
        ts = pd.Timestamp(date_raw)
        d  = ts.date()
        if not (2020 <= d.year <= 2035):
            print("  WARNING: Skipping row -- date out of range (got '%s' -> %s). "
                  "Check source file for corrupted date values." % (date_raw, d))
            return None
        return d
    except Exception:
        print("  WARNING: Skipping row -- cannot parse date value '%s'." % date_raw)
        return None


def read_production_source(prod_path):
    import pandas as pd

    header_row = detect_header_row(prod_path)
    df = pd.read_excel(
        prod_path, sheet_name='Production Day wise', skiprows=header_row
    )

    actual_cols = list(df.columns)
    print("  Columns found:")
    for c in actual_cols:
        print("    '" + str(c) + "'")

    col_map = build_col_map(actual_cols)
    missing = [k for k in ['Date', 'Machines', 'Product Name'] if k not in col_map]
    if missing:
        print("  ERROR: Missing columns: " + str(missing))
        print("  Tip: Check that Imran's file has 'Date'/'Month', 'Machines', "
              "and 'Product Name' columns on the detected header row.")
        return [], []

    rows      = []
    no_pid    = []
    bad_dates = 0

    for _, r in df.iterrows():
        date_raw    = get_val(r, col_map, 'Date')
        machine_raw = get_val(r, col_map, 'Machines')
        if date_raw is None or machine_raw is None:
            continue
        try:
            if pd.isna(machine_raw):
                continue
        except Exception:
            pass

        d = parse_date(date_raw)
        if d is None:
            bad_dates += 1
            continue

        name_raw = str(get_val(r, col_map, 'Product Name', '')).strip()
        dia_raw  = str(get_val(r, col_map, 'Dia(mm)/Volume', '')).strip()
        cust_raw = str(get_val(r, col_map, 'Customer', '')).strip()

        catalog_name, pid = ALIASES.get((name_raw.lower().strip(), dia_raw), (None, None))
        if catalog_name is None:
            catalog_name, pid = ALIASES.get((name_raw.lower().rstrip(), dia_raw), (None, None))
        if catalog_name is None:
            catalog_name = name_raw
            pid = None

        cust_norm = CUSTOMER_MAP.get(cust_raw.lower().strip(), cust_raw)

        if 'transparent' in name_raw.lower() and cust_raw.lower() == 'mabley beauty':
            if '150' in dia_raw:
                pid = 8010
                catalog_name = 'TRANSPARENT BOTTLE 130ML'
            elif '300' in dia_raw:
                pid = 8009
                catalog_name = 'TRANSPARENT BOTTLE 300ML'

        if pid is not None and pid in PID_TO_CUSTOMER:
            cust_norm = PID_TO_CUSTOMER[pid]

        try:
            dia_val = float(dia_raw)
        except ValueError:
            dia_val = dia_raw

        def sf(val, default=None):
            if val is None: return default
            try:
                if pd.isna(val): return default
            except Exception:
                pass
            try:
                return float(val)
            except Exception:
                return default

        if pid is None:
            # Varnish passes are supposed to be not counted and do not require a PID
            if not ("(varnish)" in catalog_name.lower() or "(varnish)" in name_raw.lower()):
                no_pid.append((catalog_name, dia_raw, cust_norm))

        rows.append({
            'date':                 d,
            'machine':              str(machine_raw).strip(),
            'customer':             cust_norm,
            'product_name':         catalog_name,
            'dia':                  dia_val,
            'pid':                  pid,
            'good_qty':             sf(get_val(r, col_map, 'Good Production'), 0),
            'reject_qty':           sf(get_val(r, col_map, 'Wastage'), 0),
            'total_production':     sf(get_val(r, col_map, 'Total Production'), 0),
            'mechanical_dt':        sf(get_val(r, col_map, 'Mechanical Downtime')),
            'electrical_dt':        sf(get_val(r, col_map, 'Electrical Downtime')),
            'material_shortage_dt': sf(get_val(r, col_map, 'Material Shortage ')),
            'changeover_dt':        sf(get_val(r, col_map, 'ChangeOver')),
            'operations_dt':        sf(get_val(r, col_map, 'Operations Downtime')),
            'power_shutdown_dt':    sf(get_val(r, col_map, 'Power Shut Down')),
            'gas_shutdown_dt':      sf(get_val(r, col_map, 'Gas Shut Down')),
            'workers_shortage_dt':  sf(get_val(r, col_map, 'Workers Shortage')),
            'original_name':        name_raw,
        })

    if bad_dates:
        print("  WARNING: %d row(s) skipped due to bad/corrupt date values "
              "(see warnings above). Fix in source file and re-run." % bad_dates)

    return rows, list(set(no_pid))


DATA_START_ROW = 3


def write_production_log(ac_path, source_rows):
    import openpyxl
    from copy import copy

    wb = openpyxl.load_workbook(ac_path)
    ws = wb['Production_Log']

    template_row = DATA_START_ROW
    for probe in range(DATA_START_ROW, DATA_START_ROW + 15):
        if any(ws.cell(row=probe, column=c).value is not None for c in range(1, 20)):
            template_row = probe
            break

    template_styles = {}
    for c in range(1, 20):
        cell = ws.cell(row=template_row, column=c)
        template_styles[c] = {
            'font':          copy(cell.font),
            'fill':          copy(cell.fill),
            'border':        copy(cell.border),
            'alignment':     copy(cell.alignment),
            'number_format': cell.number_format,
        }
    template_styles[1]['number_format']  = 'DD-MMM-YYYY'
    template_styles[7]['number_format']  = '#,##0'
    template_styles[8]['number_format']  = '#,##0'
    template_styles[9]['number_format']  = '#,##0'
    template_styles[10]['number_format'] = '0.0%'

    cleared = 0
    for r in range(DATA_START_ROW, ws.max_row + 1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, 20)):
            for c in range(1, 20):
                ws.cell(row=r, column=c).value = None
            cleared += 1
    print(f"  Cleared {cleared} existing rows")

    source_rows.sort(key=lambda x: (x['date'] or datetime(2099, 1, 1).date(), x['machine']))

    for i, row in enumerate(source_rows):
        r = DATA_START_ROW + i
        if row['date'] is None:
            continue

        d = row['date']
        col_a = ws.cell(row=r, column=1)
        col_a.value         = datetime(d.year, d.month, d.day)
        col_a.number_format = 'DD-MMM-YYYY'
        ws.cell(row=r, column=2).value  = row['machine']
        ws.cell(row=r, column=3).value  = row['customer']
        ws.cell(row=r, column=4).value  = row['product_name']
        ws.cell(row=r, column=5).value  = row['dia']
        ws.cell(row=r, column=6).value  = row['pid']
        ws.cell(row=r, column=7).value  = row['total_production']
        ws.cell(row=r, column=8).value  = row['good_qty']
        ws.cell(row=r, column=9).value  = row['reject_qty']
        ws.cell(row=r, column=10).value = '=IFERROR(I{0}/(H{0}+I{0}),\"\")'.format(r)
        ws.cell(row=r, column=11).value = row['mechanical_dt']
        ws.cell(row=r, column=12).value = row['electrical_dt']
        ws.cell(row=r, column=13).value = row['material_shortage_dt']
        ws.cell(row=r, column=14).value = row['changeover_dt']
        ws.cell(row=r, column=15).value = row['operations_dt']
        ws.cell(row=r, column=16).value = row['power_shutdown_dt']
        ws.cell(row=r, column=17).value = row['gas_shutdown_dt']
        ws.cell(row=r, column=18).value = row['workers_shortage_dt']

        for c in range(1, 20):
            cell = ws.cell(row=r, column=c)
            s = template_styles.get(c, {})
            if s:
                cell.font          = copy(s['font'])
                cell.fill          = copy(s['fill'])
                cell.border        = copy(s['border'])
                cell.alignment     = copy(s['alignment'])
                cell.number_format = s['number_format']

    wb.save(ac_path)
    return len(source_rows)


# ─────────────────────────────────────────────────────────────────────────────
# FG STOCK FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def _norm_dia(raw):
    s = str(raw).strip().lower()
    s = re.sub(r'\s*dia\s*$', '', s)
    s = re.sub(r'\s+ml$', 'ml', s)
    return s.strip()


def read_fg_stock(prod_path):
    import pandas as pd

    try:
        df = pd.read_excel(prod_path, sheet_name='FG Stock In hand', header=1)
    except Exception as e:
        print(f"  ERROR reading FG Stock In hand sheet: {e}")
        return [], None, []

    df.columns = [
        'Sr', 'Date', 'Customer', 'Product', 'Diameter',
        'FG_Qty', 'Prod_Remarks', 'Dispatch_Remarks'
    ] + list(df.columns[8:])

    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df = df[df['Date'].notna()].copy()
    if df.empty:
        print("  WARNING: FG Stock In hand sheet has no valid dates.")
        return [], None, []

    latest_dt = df['Date'].max()
    latest_date = latest_dt.date()
    df_latest = df[df['Date'] == latest_dt].copy()
    print(f"  FG Stock latest date: {latest_date}  ({len(df_latest)} rows)")

    rows = []
    no_pid = []

    for i, (_, r) in enumerate(df_latest.iterrows(), 1):
        raw_product  = str(r.get('Product', '')).strip()
        raw_customer = str(r.get('Customer', '')).strip()
        raw_dia      = str(r.get('Diameter', '')).strip()
        status       = str(r.get('Prod_Remarks', '')).strip()
        dispatch_rem = str(r.get('Dispatch_Remarks', '')).strip()

        try:
            fg_qty = int(float(r.get('FG_Qty', 0)))
        except (ValueError, TypeError):
            fg_qty = 0

        prod_key    = raw_product.lower().strip()
        cust_key    = raw_customer.lower().strip()
        dia_norm    = _norm_dia(raw_dia)

        result = FG_ALIASES.get((prod_key, dia_norm, cust_key))

        catalog_name = result[0] if result else raw_product
        pid          = result[1] if result else None

        if pid is not None and pid in PID_TO_CUSTOMER:
            customer_display = PID_TO_CUSTOMER[pid]
        else:
            customer_display = CUSTOMER_MAP.get(cust_key, raw_customer)

        if pid is None:
            # Varnish passes are supposed to be not counted and do not require a PID
            if not ("(varnish)" in catalog_name.lower() or "(varnish)" in raw_product.lower()):
                no_pid.append((raw_product, raw_dia, raw_customer))

        rows.append({
            'sr':               i,
            'date':             latest_date,
            'customer_display': customer_display,
            'catalog_name':     catalog_name,
            'pid':              pid,
            'dia_norm':         dia_norm,
            'fg_qty':           fg_qty,
            'status':           status,
            'dispatch_remarks': dispatch_rem,
            'raw_product':      raw_product,
            'raw_customer':     raw_customer,
        })

    return rows, latest_date, list(set(no_pid))


def write_fg_stock(ac_path, fg_rows, latest_date):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from copy import copy

    NAVY      = "FF1F3864"
    BLUE      = "FF2E4D8F"
    WHITE     = "FFFFFFFF"
    DARK      = "FF1A1A2E"
    GREEN_BG  = "FFD6EAD6"
    ORANGE_BG = "FFFDEBD0"
    FONT_NAME = "Segoe UI"

    def _font(sz=9, bold=False, color=DARK):
        return Font(name=FONT_NAME, size=sz, bold=bold, color=color)
    def _fill(color):
        return PatternFill("solid", fgColor=color) if color else PatternFill()
    def _border():
        side = Side(style='thin', color="FFB0B0B0")
        return Border(left=side, right=side, top=side, bottom=side)
    def _align(h='left', wrap=False):
        return Alignment(horizontal=h, vertical='center', wrap_text=wrap)

    wb = openpyxl.load_workbook(ac_path)

    if 'FG Stock' not in wb.sheetnames:
        print("  WARNING: 'FG Stock' sheet not found in Tubex file. Skipping FG update.")
        wb.close()
        return 0

    ws = wb['FG Stock']

    date_str = latest_date.strftime("%d-%b-%Y") if latest_date else "unknown"
    ws["A1"].value = f"FG STOCK IN HAND — Last Updated: {date_str}"

    ranges_to_remove = [str(m) for m in ws.merged_cells.ranges if m.min_row >= 4]
    for r in ranges_to_remove:
        try:
            ws.unmerge_cells(r)
        except Exception:
            pass

    max_r = ws.max_row
    for r in range(4, max_r + 1):
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.value  = None
            cell.font   = _font()
            cell.fill   = _fill(None)
            cell.border = Border()
        ws.row_dimensions[r].height = 15

    if not fg_rows:
        ws.cell(row=4, column=1).value = "No data — FG Stock In hand sheet not found or empty."
        wb.save(ac_path)
        return 0

    for i, row in enumerate(fg_rows):
        r = 4 + i
        ws.row_dimensions[r].height = 15

        status_lower = row['status'].lower()
        bg = GREEN_BG if status_lower == 'ok' else ORANGE_BG

        values = [
            row['sr'],
            row['pid'],
            row['customer_display'],
            row['catalog_name'],
            row['dia_norm'],
            row['fg_qty'],
            row['status'],
            row['dispatch_remarks'],
        ]
        aligns = ['center','center','left','left','center','center','center','left']
        bolds  = [False, False, False, False, False, True, False, False]

        for ci, (val, aln, bld) in enumerate(zip(values, aligns, bolds), 1):
            cell = ws.cell(row=r, column=ci)
            cell.value     = val
            cell.font      = _font(bold=bld)
            cell.fill      = _fill(bg)
            cell.border    = _border()
            cell.alignment = _align(aln, wrap=(ci == 8))

        ws.cell(row=r, column=6).number_format = '#,##0'

    wb.save(ac_path)
    return len(fg_rows)


# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "="*60)
    print("  Tubex -- Production Log Updater v21")
    print("="*60 + "\n")

    print("[1/4] Finding files...")
    ac_path, prod_path = find_files()
    if not ac_path:
        return

    print("\n[1a] Safety checks...")
    check_not_locked(ac_path)
    check_freshness(prod_path, max_hours=26, label="Production.xlsx")

    print("\n[2/4] Reading production data...")
    try:
        result = read_production_source(prod_path)
    except ImportError:
        print("\n  ERROR: pandas not installed. Run: pip install pandas openpyxl")
        return
    except Exception as e:
        import traceback
        print(f"\n  ERROR reading file: {e}")
        traceback.print_exc()
        return

    source_rows, no_pid = result

    if not source_rows:
        print("  No rows parsed -- check column diagnostics above.")
        return

    dates = {r['date'] for r in source_rows if r['date']}
    print(f"  Parsed {len(source_rows)} rows  |  {min(dates)} -> {max(dates)}")

    corrections = [(r['original_name'], r['product_name'], r['pid'])
                   for r in source_rows if r['original_name'] != r['product_name']]
    if corrections:
        shown = set()
        print("\n  Name corrections applied:")
        for orig, corr, pid in sorted(set(corrections)):
            if orig not in shown:
                shown.add(orig)
                print(f"    '{orig}' -> '{corr}' (PID={pid})")

    print("\n[3/4] Wiping Production_Log and rewriting from scratch...")
    written = write_production_log(ac_path, source_rows)
    print(f"  Written: {written} rows")

    print("\n[4/4] Reading FG Stock and updating FG Stock sheet...")
    try:
        fg_rows, fg_date, fg_no_pid = read_fg_stock(prod_path)
    except Exception as e:
        import traceback
        print(f"  ERROR reading FG Stock: {e}")
        traceback.print_exc()
        fg_rows, fg_date, fg_no_pid = [], None, []

    if fg_rows:
        fg_written = write_fg_stock(ac_path, fg_rows, fg_date)
        print(f"  FG Stock written: {fg_written} rows  (date: {fg_date})")
    else:
        print("  No FG Stock rows — sheet not updated.")

    print("  Saved:   " + os.path.basename(ac_path))

    # ── Post-write validation ────────────────────────────────────────────
    print("\n  Validating writes...")
    try:
        wb_check = openpyxl.load_workbook(ac_path, read_only=True, data_only=True)
        
        # Check Production_Log
        ws_check = wb_check['Production_Log']
        pl_count = sum(1 for r in range(3, ws_check.max_row + 1) 
                       if ws_check.cell(r, 1).value is not None)
        if pl_count == len(source_rows):
            print(f"  ✓ Production_Log: {pl_count} rows written (matches source)")
        else:
            print(f"  !! Production_Log: expected {len(source_rows)} rows, found {pl_count}")
        
        # Check FG Stock
        if 'FG Stock' in wb_check.sheetnames and fg_rows:
            ws_fg = wb_check['FG Stock']
            fg_count = sum(1 for r in range(4, ws_fg.max_row + 1) 
                          if ws_fg.cell(r, 1).value is not None)
            if fg_count == len(fg_rows):
                print(f"  ✓ FG Stock: {fg_count} rows written (matches source)")
            else:
                print(f"  !! FG Stock: expected {len(fg_rows)} rows, found {fg_count}")
        
        wb_check.close()
    except Exception as e:
        print(f"  !! Validation error: {e}")

    if no_pid:
        print(f"\n  WARNING -- {len(set(no_pid))} Production products with NO PID (add to ALIASES):")
        for name, dia, cust in sorted(set(no_pid)):
            print(f"    {cust}: {name}  (Dia={dia})")
        log_mismatches("production", list(set(no_pid)))

    if fg_no_pid:
        print(f"\n  WARNING -- {len(set(fg_no_pid))} FG Stock products with NO PID (add to FG_ALIASES):")
        for name, dia, cust in sorted(set(fg_no_pid)):
            print(f"    {cust}: {name}  (Dia={dia})")
        log_mismatches("fg_stock", list(set(fg_no_pid)))

    print("\n  Press Ctrl+Shift+F9 in Excel to recalculate.")
    print("="*60)


if __name__ == "__main__":
    main()
