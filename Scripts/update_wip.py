"""
Tubex - WIP (Work In Process) Updater v1
====================================================
PURPOSE:
  Parses a WhatsApp-style WIP message from Aurangzeb and writes
  Work In Process slug weights into column I of the Inventory sheet.

USAGE (two ways):
  1. Double-click "Update_WIP.bat"  -- it will ask you to paste the message
  2. python update_wip.py "#19mm 10kg #30mm 125kg #32mm 25kg"

MESSAGE FORMAT (flexible, all of these work):
  #19mm 10kg #30mm 125kg #32mm 25kg
  19mm-10kg, 30mm-125kg, 32mm-25kg
  #19 10 #30 125 #32 25
  19mm 10 30mm 125 32mm 25

HOW IT WORKS:
  - Reads slug rows from Inventory sheet (col B = "Slug")
  - Extracts diameter from slug name (e.g. "19X4.5 S/M" -> dia 19)
  - For each diameter in the WIP message, finds the matching slug row
  - If multiple slugs share a diameter, picks the one with non-zero issuance (col G)
    to identify the active slug. Falls back to first row of that diameter.
  - Clears all existing WIP values first, then writes the new ones.
  - Saves back to the SAME file. No new version is created.

WHAT IT CHANGES in Inventory sheet:
  Col I (Work In Process): slug rows only. All others untouched.
"""

import os
import re
import sys
import glob

import warnings
warnings.filterwarnings("ignore", message=".*Data Validation.*")
warnings.filterwarnings("ignore", message=".*extension.*")

from openpyxl import load_workbook

# Scripts live in Tubex/Scripts/
# Excel files live in Tubex/ (one level up)
DATA_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# -----------------------------------------------------------------------
# Column constants (1-indexed for openpyxl)
# -----------------------------------------------------------------------
INV_CAT_COL    = 2   # B: Category
INV_NAME_COL   = 3   # C: Item Name (ERP)
INV_ISSUED_COL = 7   # G: Issued to Production
INV_WIP_COL    = 9   # I: Work In Process
INV_DATA_START = 3   # first data row


# -----------------------------------------------------------------------
def find_excel():
    files = glob.glob(os.path.join(DATA_DIR, "Tubex*.xlsx"))
    if not files:
        print("  ERROR: No Tubex*.xlsx found in: " + DATA_DIR)
        return None, None
    return sorted(files)[-1], DATA_DIR


def extract_dia(name):
    """
    Pull the leading diameter from a slug name like '19X4.5 S/M' or '20.5x4.5'.
    Returns float or None.
    """
    if not name:
        return None
    m = re.match(r'^(\d+(?:\.\d+)?)[Xx]', name.strip())
    if m:
        return float(m.group(1))
    return None


def parse_wip_message(msg):
    """
    Parse WhatsApp WIP message. Returns {dia_float: kg_float}.
    Handles formats like:
      #19mm 10kg  |  #30mm 125kg  |  32mm 25  |  19 - 10  |  #19mm-10kg
    """
    # Normalise: replace - and , separators with spaces
    msg = re.sub(r'[,\-]+', ' ', msg)

    result = {}
    # Match patterns: optional # + number + optional mm + optional space + number + optional kg
    pattern = re.compile(
        r'#?\s*(\d+(?:\.\d+)?)\s*(?:mm)?\s*(\d+(?:\.\d+)?)\s*(?:kg)?',
        re.IGNORECASE
    )
    for m in pattern.finditer(msg):
        dia = float(m.group(1))
        kg  = float(m.group(2))
        result[dia] = result.get(dia, 0) + kg

    return result


def build_slug_map(ws):
    """
    Scan Inventory sheet. For each Slug row, map dia -> list of (row, issued_qty).
    Returns {dia: [(row_num, issued_qty, name), ...]}
    """
    slug_map = {}
    for r in range(INV_DATA_START, ws.max_row + 1):
        if str(ws.cell(r, INV_CAT_COL).value or '').strip().upper() != 'SLUG':
            continue
        name   = ws.cell(r, INV_NAME_COL).value
        issued = ws.cell(r, INV_ISSUED_COL).value or 0
        dia    = extract_dia(str(name)) if name else None
        if dia is None:
            continue
        if dia not in slug_map:
            slug_map[dia] = []
        slug_map[dia].append((r, issued, str(name)))
    return slug_map


def pick_row(candidates):
    """
    Given [(row, issued, name)...] for one diameter,
    prefer the row with non-zero issuance (= active slug this month).
    Fall back to the first row.
    """
    active = [(r, issued, name) for r, issued, name in candidates if issued and issued > 0]
    if active:
        return active[0][0], active[0][2]
    return candidates[0][0], candidates[0][2]


# -----------------------------------------------------------------------
def main():
    SEP = "=" * 55

    print("")
    print(SEP)
    print("  Tubex - WIP Updater v1")
    print(SEP)
    print("")

    # Get WIP message
    if len(sys.argv) > 1:
        msg = " ".join(sys.argv[1:])
    else:
        print("  Paste the WIP message from Aurangzeb and press Enter:")
        print("  Example:  #19mm 10kg #30mm 125kg #32mm 25kg")
        print("")
        msg = input("  Message: ").strip()

    if not msg:
        print("  No message entered. Nothing updated.")
        return

    print("")
    print("  Message received: " + msg)

    # Parse message
    wip_data = parse_wip_message(msg)
    if not wip_data:
        print("  ERROR: Could not parse any diameter/weight pairs from message.")
        print("  Expected format: #19mm 10kg #30mm 125kg")
        return

    print("")
    print("  Parsed WIP values:")
    for dia, kg in sorted(wip_data.items()):
        print("    %gmm -> %g KGS" % (dia, kg))

    # Find Excel file
    print("")
    print("[1/2] Finding Excel file...")
    excel_path, folder = find_excel()
    if not excel_path:
        return
    print("  File: " + os.path.basename(excel_path))

    # Load and update
    print("")
    print("[2/2] Updating Inventory sheet...")
    wb = load_workbook(excel_path)
    ws = wb['Inventory']

    slug_map = build_slug_map(ws)

    # Clear all existing WIP values in slug rows
    cleared = 0
    for dia, candidates in slug_map.items():
        for r, issued, name in candidates:
            if ws.cell(r, INV_WIP_COL).value is not None:
                ws.cell(r, INV_WIP_COL).value = None
                cleared += 1
    if cleared:
        print("  Cleared %d existing WIP value(s)" % cleared)

    # Write new WIP values
    written     = []
    not_matched = []

    for dia, kg in sorted(wip_data.items()):
        if dia not in slug_map:
            not_matched.append((dia, kg))
            continue
        candidates = slug_map[dia]
        row, name = pick_row(candidates)
        ws.cell(row, INV_WIP_COL).value = kg
        written.append((dia, kg, row, name))

    wb.save(excel_path)

    # Report
    print("")
    if written:
        print("  WIP values written:")
        for dia, kg, row, name in written:
            print("    %gmm (%s) row %d -> %g KGS" % (dia, name, row, kg))

    if not_matched:
        print("")
        print("  !! Diameters NOT found in Inventory slug rows:")
        for dia, kg in not_matched:
            print("    %gmm %g KGS -- no matching slug row" % (dia, kg))
        known_dias = sorted(slug_map.keys())
        print("  Known slug diameters: %s" % ", ".join("%g" % d for d in known_dias))

    print("")
    print("  Saved: " + os.path.basename(excel_path))
    print(SEP)


if __name__ == "__main__":
    main()
