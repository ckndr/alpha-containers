import openpyxl

wb = openpyxl.load_workbook(r'd:\Alpha\Tubex Records\Tubex_July26.xlsx', data_only=True)
ws = wb['Production_Log']
dates = {}
tube_sum = 0
for r in range(3, ws.max_row+1):
    d = ws.cell(r, 1).value
    g = ws.cell(r, 8).value
    dia = str(ws.cell(r, 5).value or '')
    if d and g is not None:
        m = d.strftime('%Y-%m') if hasattr(d, 'strftime') else str(d)[:7]
        dates[m] = dates.get(m, 0) + int(g)
        if 'ml' not in dia.lower():
            tube_sum += int(g)
print('Dates in Tubex_July26.xlsx Production_Log:', dates)
print('Total Tubes:', tube_sum)
