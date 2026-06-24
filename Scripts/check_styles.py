import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.load_workbook(r'd:\Alpha\Aerosol_BOM.xlsx')

def copy_style(src_cell, dest_cell):
    if src_cell.has_style:
        dest_cell.font = Font(
            name=src_cell.font.name,
            size=src_cell.font.size,
            bold=src_cell.font.bold,
            italic=src_cell.font.italic,
            charset=src_cell.font.charset,
            color=src_cell.font.color,
            underline=src_cell.font.underline,
            strike=src_cell.font.strike
        )
        dest_cell.fill = PatternFill(
            fill_type=src_cell.fill.fill_type,
            start_color=src_cell.fill.start_color,
            end_color=src_cell.fill.end_color
        )
        dest_cell.border = Border(
            left=src_cell.border.left,
            right=src_cell.border.right,
            top=src_cell.border.top,
            bottom=src_cell.border.bottom
        )
        dest_cell.alignment = Alignment(
            horizontal=src_cell.alignment.horizontal,
            vertical=src_cell.alignment.vertical,
            text_rotation=src_cell.alignment.text_rotation,
            wrap_text=src_cell.alignment.wrap_text,
            shrink_to_fit=src_cell.alignment.shrink_to_fit,
            indent=src_cell.alignment.indent
        )
        dest_cell.number_format = src_cell.number_format

# Print existing row 9 style in BOM sheet to match
ws_bom = wb['BOM']
print("BOM Row 9:")
for col in range(1, 12):
    cell = ws_bom.cell(row=9, column=col)
    print(f"Col {col} value: {cell.value}, Number Format: {cell.number_format}")

# Print existing row 10 style in Req. Calculator sheet to match
ws_calc = wb['Req. Calculator']
print("\nReq. Calculator Row 10:")
for col in range(1, 9):
    cell = ws_calc.cell(row=10, column=col)
    print(f"Col {col} value: {cell.value}, Number Format: {cell.number_format}")
