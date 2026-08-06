import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

wb = openpyxl.load_workbook('Tubex_Aug26.xlsx')
ws = wb['MRP']

# ── helpers ──────────────────────────────────────────────────────────────
def clone_border(b):
    def cs(s):
        return Side(border_style=s.border_style, color=copy(s.color)) if s else Side()
    return Border(left=cs(b.left), right=cs(b.right),
                  top=cs(b.top), bottom=cs(b.bottom))

def apply_header_style(cell, ref_cell):
    cell.font      = Font(bold=True, color="FFFFFF",
                          size=ref_cell.font.size or 10,
                          name=ref_cell.font.name or "Calibri")
    cell.fill      = PatternFill("solid", fgColor="2E4D8F")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = clone_border(ref_cell.border)

def apply_data_style(cell, ref_cell):
    cell.font      = Font(bold=ref_cell.font.bold,
                          size=ref_cell.font.size or 10,
                          name=ref_cell.font.name or "Calibri")
    cell.fill      = PatternFill("solid", fgColor="F7F9FC")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = clone_border(ref_cell.border)
    cell.number_format = "#,##0"

# ─────────────────────────────────────────────────────────────────────────
# JULY TUBES MRP
# One product order: Product ID in $D$3, Remaining Balance in $H$3
# Formula logic:
#   Rate  = SUMPRODUCT( (BOM Item ID = A{r}) * (BOM Product ID = $D$3)
#                       * ($H$3 > 0) * BOM[Per 1000 Units] )
#   If Rate = 0  → "-"   (item not in this product's BOM, or no remaining qty)
#   Else         → ROUND( F{r} / (Rate / 1000), 0 )
#
# Because there is exactly one product, the "weighted average" collapses to
# simply the rate for that specific product — which is the most accurate value.
# ─────────────────────────────────────────────────────────────────────────
JULY_HDR  = 6
JULY_DATA = list(range(7, 89))

j6 = ws.cell(row=JULY_HDR, column=10, value="Pieces Can Produce")
apply_header_style(j6, ws["I6"])

for r in JULY_DATA:
    # SUMPRODUCT filters: item matches AND product = D3 AND remaining > 0
    rate_expr = (
        'SUMPRODUCT((TableBOM[Item ID]=A{r})*(TableBOM[Product ID]=$D$3)'
        '*($H$3>0)*TableBOM[Per 1000 Units])'
    ).format(r=r)

    formula = (
        '=IFERROR(IF({rate}=0,"-",'
        'ROUND(F{r}/({rate}/1000),0)),"-")'
    ).format(rate=rate_expr, r=r)

    cell = ws.cell(row=r, column=10, value=formula)
    apply_data_style(cell, ws['I' + str(r)])

# ─────────────────────────────────────────────────────────────────────────
# AUGUST PET MRP
# Four product orders: Product IDs in $D$93:$D$96, Remaining in $H$93:$H$96
# Formula logic (weighted average over products with remaining > 0):
#
#   For each BOM row that uses Item A{r}:
#     rem_qty   = SUMIF($D$93:$D$96, BOM[Product ID], $H$93:$H$96)
#     Include   = (rem_qty > 0)
#
#   Num = SUMPRODUCT( (BOM Item = A{r}) * Include * rem_qty * BOM[Per 1000 Units] )
#   Den = SUMPRODUCT( (BOM Item = A{r}) * Include * rem_qty )
#
#   Weighted avg rate = Num / Den
#   Pieces = ROUND( F{r} * Den / Num * 1000, 0 )
#          = ROUND( F{r} * 1000 * Den / Num, 0 )
#
# If Num = 0 (no remaining orders use this item) → "-"
# ─────────────────────────────────────────────────────────────────────────
PET_HDR  = 99
PET_DATA = list(range(100, 107))

j99 = ws.cell(row=PET_HDR, column=10, value="Pieces Can Produce")
apply_header_style(j99, ws["I99"])

for r in PET_DATA:
    # rem() = remaining balance lookup for each BOM product ID in the PET orders
    rem  = 'IFERROR(SUMIF($D$93:$D$96,TableBOM[Product ID],$H$93:$H$96),0)'
    # item filter
    item = '(TableBOM[Item ID]=A{r})'.format(r=r)
    # positive-remaining filter
    pos  = '({rem}>0)'.format(rem=rem)

    num = 'SUMPRODUCT({item}*{pos}*{rem}*TableBOM[Per 1000 Units])'.format(
              item=item, pos=pos, rem=rem)
    den = 'SUMPRODUCT({item}*{pos}*{rem})'.format(
              item=item, pos=pos, rem=rem)

    formula = (
        '=IFERROR(IF({num}=0,"-",'
        'ROUND(F{r}*1000*{den}/{num},0)),"-")'
    ).format(num=num, den=den, r=r)

    cell = ws.cell(row=r, column=10, value=formula)
    apply_data_style(cell, ws['I' + str(r)])

# ── Column width ──────────────────────────────────────────────────────────
ws.column_dimensions["J"].width = 18

# ── Save ──────────────────────────────────────────────────────────────────
wb.save('Tubex_Aug26.xlsx')
print("Done! Pieces Can Produce (weighted by remaining orders) updated.")
print("  July: J{} header + J{}:J{} data".format(JULY_HDR, JULY_DATA[0], JULY_DATA[-1]))
print("  Aug PET: J{} header + J{}:J{} data".format(PET_HDR, PET_DATA[0], PET_DATA[-1]))
print()
# Print the generated formulas for review
print("=== Sample July formula (J7) ===")
print(ws['J7'].value)
print()
print("=== Sample Aug PET formula (J100) ===")
print(ws['J100'].value)
