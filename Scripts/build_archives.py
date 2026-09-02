"""
build_archives.py
=================
Creates two unified archive files from all Tubex monthly files & legacy records:

  1.  Dashboard_Archive.xlsx   (Tubex Records/)
        - One tab per archived month  (value-snapshot of Tubex_Dashboard)
        - "Annual Summary" tab        (KPIs across all months + charts)

  2.  Production_Archive.xlsx  (Tubex Records/)
        - One tab per archived month  (Production_Log data from Nov 2025 - Jul 2026+)
        - "All Months" tab            (all rows stacked with Month column + AutoFilter)
        - "Customer Breakdown" tab    (Summary of production per customer, TUBE vs PET, per month)

Usage:
    python Scripts/build_archives.py

Add new months to MONTH_FILES and re-run to rebuild from scratch.
"""

import os, sys, time, datetime, zipfile, re, shutil
import win32com.client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

from parse_legacy_xls import get_legacy_production_records
from parse_legacy_dispatch import parse_dispatch_xls
from alpha_checks import get_active_tubex_file

# ============================================================
#  CONFIGURATION
# ============================================================

MONTH_FILES = [
    ("July 2026",      r"d:\Alpha\Tubex Records\Tubex_July26.xlsx",   7),
    ("August 2026",    r"d:\Alpha\Tubex Records\Tubex_Aug26.xlsx",    8),
]

# Dynamically find the active month file in d:\Alpha using standard resolution (Rule R1-22)
latest = get_active_tubex_file(r"d:\Alpha")
if latest:
    month_name = datetime.datetime.now().strftime("%B %Y")
    month_num = datetime.datetime.now().month
    if "Aug" in os.path.basename(latest):
        month_name, month_num = "August 2026", 8
    elif "Sep" in os.path.basename(latest):
        month_name, month_num = "September 2026", 9
    
    # Check if already added
    if not any(f[1] == latest or f[0] == month_name for f in MONTH_FILES):
        try:
            wb_test = openpyxl.load_workbook(latest, data_only=True)
            if "Production_Log" in wb_test.sheetnames:
                has_data = any(wb_test["Production_Log"].cell(r, 1).value is not None for r in range(3, 10))
                if has_data:
                    MONTH_FILES.append((month_name, latest, month_num))
            wb_test.close()
        except Exception:
            pass

DASHBOARD_SHEET    = "Tubex_Dashboard"
PRODUCTION_SHEET   = "Production_Log"

OUTPUT_DIR         = r"d:\Alpha\Tubex Records"
DASHBOARD_ARCHIVE  = os.path.join(OUTPUT_DIR, "Dashboard_Archive.xlsx")
PRODUCTION_ARCHIVE = os.path.join(OUTPUT_DIR, "Production_Archive.xlsx")
TEMP_DIR           = os.path.join(OUTPUT_DIR, "_tmp_archive")

KPI_CELLS = {
    "TUBE_MTD":      (6,  4),
    "TUBE_REJECT":   (6,  7),
    "TUBE_DISPATCH": (6, 10),
    "PET_MTD":       (8,  4),
    "PET_REJECT":    (8,  7),
    "PET_DISPATCH":  (8, 10),
}

PROD_HEADERS = [
    "Date", "Machine", "Customer", "Product Name",
    "Dia / Volume", "Product ID", "Target Quantity",
    "Good Quantity Produced", "Reject/Scrap Quantity", "Waste%"
]

C_NAVY  = "1A2B4A"
C_MID   = "2E4A7A"
C_GOLD  = "D4AF37"
C_WHITE = "FFFFFF"
C_LIGHT = "EDF0F7"
C_ALT   = "F8F9FC"

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color=C_GOLD)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(ws, row, col, value, bg=C_MID, fg=C_WHITE, bold=True, size=10, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, size=size, color=fg, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border    = thin_border()
    return c

EXCEL_XLSX = 51

def build_dashboard_archive(available_months):
    print("\n[1/5] Building Dashboard_Archive via Excel COM...")
    xl = None
    kpi_data = {}

    try:
        xl = win32com.client.DispatchEx("Excel.Application")
        xl.Visible          = False
        xl.DisplayAlerts    = False
        xl.AskToUpdateLinks = False

        if os.path.exists(DASHBOARD_ARCHIVE):
            os.remove(DASHBOARD_ARCHIVE)
        os.makedirs(TEMP_DIR, exist_ok=True)

        archive_wb = None

        for idx, (label, src_path, month_num) in enumerate(available_months):
            tab     = label[:31]
            src_abs = os.path.abspath(src_path)
            print(f"\n     {label}")

            src_wb = xl.Workbooks.Open(src_abs, UpdateLinks=0, ReadOnly=False)
            xl.Calculate()
            time.sleep(1.0)

            kpis = {}
            try:
                ds = src_wb.Sheets(DASHBOARD_SHEET)
                for name, (r, c) in KPI_CELLS.items():
                    kpis[name] = ds.Cells(r, c).Value
                tube = kpis.get("TUBE_MTD")
                pet  = kpis.get("PET_MTD")
                if tube is not None and pet is not None:
                    print(f"       KPIs => TUBE: {int(tube):,}  PET: {int(pet):,}")
                else:
                    print(f"       KPIs => {kpis}")
            except Exception as e:
                print(f"       [WARN] KPI read: {e}")
            kpi_data[label] = {"kpis": kpis, "month_num": month_num}

            try:
                src_wb.Sheets(DASHBOARD_SHEET).Copy()
                temp_wb = xl.ActiveWorkbook
                temp_ws = temp_wb.Sheets(1)

                try:
                    temp_ws.UsedRange.Value = temp_ws.UsedRange.Value
                except Exception as ve:
                    print(f"       [WARN] value-freeze: {ve}")

                xl.CutCopyMode = False

                if archive_wb is None:
                    temp_ws.Name = tab
                    archive_wb = temp_wb
                else:
                    temp_ws.Copy(None, archive_wb.Sheets(archive_wb.Sheets.Count))
                    archive_wb.Sheets(archive_wb.Sheets.Count).Name = tab
                    try:
                        temp_wb.Close(SaveChanges=False)
                    except Exception:
                        pass

                print(f"       [OK] Dashboard  -> '{tab}'")
            except Exception as e:
                import traceback
                print(f"       [ERROR] Dashboard copy: {e}")
                traceback.print_exc()

            src_wb.Close(SaveChanges=False)

        if archive_wb is not None:
            dest = os.path.abspath(DASHBOARD_ARCHIVE)
            archive_wb.SaveAs(dest, FileFormat=EXCEL_XLSX)
            print(f"\n       [OK] Dashboard_Archive.xlsx saved  ({os.path.getsize(dest)//1024} KB)")

    finally:
        if xl is not None:
            try:
                xl.Quit()
            except Exception:
                pass
        del xl

    # Fill KPI data for legacy months from parse_legacy_xls & parse_legacy_dispatch
    legacy_recs = get_legacy_production_records()
    tube_disp = parse_dispatch_xls("dispatch nov to jul.xls", "TUBE")
    pet_disp  = parse_dispatch_xls("dispatch pet nov to jul.xls", "PET")

    disp_months = {}
    for r in (tube_disp + pet_disp):
        m = r["month"]
        t = r["type"]
        if m not in disp_months:
            disp_months[m] = {"TUBE": 0, "PET": 0}
        disp_months[m][t] += r["disp_qty"]

    legacy_months = {}
    for r in legacy_recs:
        m = r["month"]
        if m not in legacy_months:
            legacy_months[m] = {"TUBE": 0, "PET": 0, "REJ_TUBE": 0, "REJ_PET": 0}
        if r["type"] == "PET":
            legacy_months[m]["PET"] += r["good"]
            legacy_months[m]["REJ_PET"] += r["reject"]
        else:
            legacy_months[m]["TUBE"] += r["good"]
            legacy_months[m]["REJ_TUBE"] += r["reject"]

    month_num_map = {
        "January 2026": 1, "February 2026": 2, "March 2026": 3, "April 2026": 4,
        "May 2026": 5, "June 2026": 6, "July 2026": 7, "August 2026": 8,
        "November 2025": 11, "December 2025": 12
    }

    for mname, mvals in legacy_months.items():
        if mname not in kpi_data and mname in month_num_map:
            tot_t = mvals["TUBE"]
            tot_p = mvals["PET"]
            rej_t = (mvals["REJ_TUBE"] / tot_t) if tot_t > 0 else 0
            rej_p = (mvals["REJ_PET"] / tot_p) if tot_p > 0 else 0
            disp_t = disp_months.get(mname, {}).get("TUBE", tot_t)
            disp_p = disp_months.get(mname, {}).get("PET", tot_p)
            kpi_data[mname] = {
                "month_num": month_num_map[mname],
                "kpis": {
                    "TUBE_MTD": tot_t,
                    "TUBE_REJECT": rej_t,
                    "TUBE_DISPATCH": disp_t,
                    "PET_MTD": tot_p,
                    "PET_REJECT": rej_p,
                    "PET_DISPATCH": disp_p
                }
            }

    return kpi_data


def build_production_archive(available_months):
    print("\n[2/5] Building Production_Archive via openpyxl & legacy parser...")

    if os.path.exists(PRODUCTION_ARCHIVE):
        os.remove(PRODUCTION_ARCHIVE)

    archive_wb = openpyxl.Workbook()
    archive_wb.remove(archive_wb.active)

    # 1. Load legacy records from parse_legacy_xls
    legacy_recs = get_legacy_production_records()
    by_month = {}
    for r in legacy_recs:
        m = r["month"]
        if m not in by_month:
            by_month[m] = []
        by_month[m].append(r)

    month_order = ["November 2025", "December 2025", "January 2026", "February 2026", "March 2026", "April 2026", "May 2026", "June 2026"]

    for mname in month_order:
        if mname in by_month:
            tab = mname[:31]
            dst_ws = archive_wb.create_sheet(tab)
            # Add header
            dst_ws.cell(row=1, column=1, value=f"PRODUCTION LOG -- {mname}")
            for ci, h in enumerate(PROD_HEADERS, 1):
                dst_ws.cell(row=2, column=ci, value=h)

            row_idx = 3
            for rec in by_month[mname]:
                waste_pct = (rec["reject"] / (rec["good"] + rec["reject"])) if (rec["good"] + rec["reject"]) > 0 else 0
                row_vals = [rec["date"], rec["machine"], rec["customer"], rec["product"], rec["dia"], rec["pid"], rec["target"], rec["good"], rec["reject"], waste_pct]
                for ci, val in enumerate(row_vals, 1):
                    dst_ws.cell(row=row_idx, column=ci, value=val)
                row_idx += 1
            dst_ws.freeze_panes = "A3"
            print(f"       [OK] Legacy Production -> '{tab}' ({row_idx-3} rows)")

    # 2. Add XLSX available months
    for mi, (label, src_path, month_num) in enumerate(available_months):
        tab = label[:31]
        if tab in archive_wb.sheetnames:
            continue
        print(f"\n     {label}")
        try:
            src_wb = openpyxl.load_workbook(src_path, data_only=True)
            if PRODUCTION_SHEET not in src_wb.sheetnames:
                src_wb.close()
                continue

            src_ws = src_wb[PRODUCTION_SHEET]
            dst_ws = archive_wb.create_sheet(tab)

            row_count = 0
            for row in src_ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if cell.column == 3 and cell.row > 2 and val:
                        from customer_normalization import normalize_customer_name, correct_customer_by_product
                        val = normalize_customer_name(val, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
                        prod_name = src_ws.cell(row=cell.row, column=4).value
                        val = correct_customer_by_product(val, prod_name)
                    new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=val)
                    if cell.number_format:
                        new_cell.number_format = cell.number_format
                row_count += 1

            dst_ws.freeze_panes = "A3"
            src_wb.close()
            print(f"       [OK] Production -> '{tab}' ({row_count} rows copied)")
        except Exception as e:
            print(f"       [ERROR] {e}")

    archive_wb.save(PRODUCTION_ARCHIVE)
    archive_wb.close()
    sz = os.path.getsize(PRODUCTION_ARCHIVE) // 1024
    print(f"\n       [OK] Production_Archive.xlsx saved ({sz} KB)")


ALL_MONTHS = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]

def add_dashboard_summary(kpi_data, year=2026):
    print("\n[3/5] Adding Annual Summary -> Dashboard_Archive...")
    wb = openpyxl.load_workbook(DASHBOARD_ARCHIVE)
    ws = wb.create_sheet("Annual Summary", 0)

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = f"TUBEX -- Annual Production Summary {year}"
    c.font      = Font(bold=True, size=20, color=C_GOLD, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 46

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value     = f"Generated {datetime.datetime.now().strftime('%d %B %Y %H:%M')}"
    c.font      = Font(italic=True, size=9, color="AAAAAA", name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

    HDR_ROW = 4
    headers = [
        ("Month",          14),
        ("TUBE Produced",  16),
        ("TUBE Dispatch",  16),
        ("TUBE Reject %",  14),
        ("PET Produced",   15),
        ("PET Dispatch",   15),
        ("PET Reject %",   13),
        ("Total Produced", 16),
        ("Status",         12),
    ]
    for ci, (h, w) in enumerate(headers, 1):
        hdr_cell(ws, HDR_ROW, ci, h, bg=C_MID, size=10, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[HDR_ROW].height = 32

    lookup = {}
    for label, entry in kpi_data.items():
        mnum = entry["month_num"]
        lookup[mnum] = entry["kpis"]

    for mi, abbr in enumerate(ALL_MONTHS, 1):
        row      = HDR_ROW + mi
        bg_color = C_LIGHT if mi % 2 == 0 else C_ALT

        if mi in lookup:
            k    = lookup[mi]
            def _num(v):
                if v is None or v == "" or v == "-": return 0.0
                try: return float(v)
                except Exception: return 0.0
            t_m  = _num(k.get("TUBE_MTD"))
            t_d  = _num(k.get("TUBE_DISPATCH"))
            t_r  = _num(k.get("TUBE_REJECT"))
            p_m  = _num(k.get("PET_MTD"))
            p_d  = _num(k.get("PET_DISPATCH"))
            p_r  = _num(k.get("PET_REJECT"))
            tot  = t_m + p_m
            status = "Archived"
        else:
            t_m = t_d = t_r = p_m = p_d = p_r = tot = None
            status = "-- Pending"

        row_vals = [f"{abbr} {year}", t_m, t_d, t_r, p_m, p_d, p_r, tot, status]
        row_fmts = [None, "#,##0", "#,##0", "0.00%", "#,##0", "#,##0", "0.00%", "#,##0", None]

        for ci, (val, fmt) in enumerate(zip(row_vals, row_fmts), 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill      = PatternFill("solid", fgColor=bg_color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border()
            c.font      = Font(size=10, name="Calibri", bold=(ci == 1),
                               color="888888" if val is None else "000000",
                               italic=(val is None))
            if fmt and val is not None:
                c.number_format = fmt
        ws.row_dimensions[row].height = 22

    TOT_ROW = HDR_ROW + 13
    ws.row_dimensions[TOT_ROW].height = 26
    hdr_cell(ws, TOT_ROW, 1, "TOTAL / AVG", bg=C_NAVY, fg=C_GOLD, size=10)
    agg_cols = {
        2: ("#,##0",  lambda v: sum(v)),
        3: ("#,##0",  lambda v: sum(v)),
        4: ("0.00%",  lambda v: sum(v) / len(v)),
        5: ("#,##0",  lambda v: sum(v)),
        6: ("#,##0",  lambda v: sum(v)),
        7: ("0.00%",  lambda v: sum(v) / len(v)),
        8: ("#,##0",  lambda v: sum(v)),
    }
    for ci, (fmt, fn) in agg_cols.items():
        vals = [ws.cell(row=HDR_ROW + mi, column=ci).value
                for mi in range(1, 13)
                if isinstance(ws.cell(row=HDR_ROW + mi, column=ci).value, (int, float))]
        result = fn(vals) if vals else None
        c = ws.cell(row=TOT_ROW, column=ci, value=result)
        c.font      = Font(bold=True, size=10, color=C_WHITE, name="Calibri")
        c.fill      = PatternFill("solid", fgColor=C_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thick_border()
        if fmt and result is not None:
            c.number_format = fmt
    ws.cell(row=TOT_ROW, column=9, value="").fill = PatternFill("solid", fgColor=C_NAVY)

    wb.save(DASHBOARD_ARCHIVE)
    wb.close()
    print("       [OK] Annual Summary added")


def add_production_summary(available_months):
    print("\n[4/5] Adding 'All Months' tab -> Production_Archive...")
    wb = openpyxl.load_workbook(PRODUCTION_ARCHIVE)
    ws = wb.create_sheet("All Months", 0)

    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value     = "TUBEX -- Production Log (All Archived Months)"
    c.font      = Font(bold=True, size=16, color=C_GOLD, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    all_headers = ["Month"] + PROD_HEADERS
    col_widths  = [15, 13, 14, 26, 26, 12, 11, 16, 18, 18, 8, 8]
    HDR_ROW = 3
    for ci, (h, w) in enumerate(zip(all_headers, col_widths), 1):
        hdr_cell(ws, HDR_ROW, ci, h, bg=C_MID, size=10, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[HDR_ROW].height = 28

    data_row = HDR_ROW + 1
    month_sheets = [s for s in wb.sheetnames if s != "All Months"]
    for mi, tab_name in enumerate(month_sheets):
        src_ws   = wb[tab_name]
        bg_color = C_LIGHT if mi % 2 == 0 else C_ALT

        for src_row in src_ws.iter_rows(min_row=3, values_only=True):
            if not any(v is not None for v in src_row):
                continue
            row_data = [tab_name] + list(src_row[:len(PROD_HEADERS)])
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=data_row, column=ci, value=val)
                c.fill      = PatternFill("solid", fgColor=bg_color)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = thin_border()
                c.font      = Font(bold=(ci == 1), size=9, name="Calibri")
                if ci == 2 and isinstance(val, datetime.datetime):
                    c.number_format = "DD-MMM-YY"
            ws.row_dimensions[data_row].height = 16
            data_row += 1

    ws.freeze_panes             = "B4"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = (
        f"A{HDR_ROW}:{get_column_letter(len(all_headers))}{max(data_row - 1, HDR_ROW)}"
    )
    wb.save(PRODUCTION_ARCHIVE)
    wb.close()
    row_count = data_row - HDR_ROW - 1
    print(f"       [OK] All Months added ({row_count} data rows)")


def extract_mtd_dispatch(file_path, month_label, month_num):
    import openpyxl
    recs = []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "Tubex_Dashboard" not in wb.sheetnames:
        return recs
    ws = wb["Tubex_Dashboard"]
    for r in range(11, 60):
        customer = ws.cell(row=r, column=3).value
        product = ws.cell(row=r, column=4).value
        dispatch = ws.cell(row=r, column=11).value
        if dispatch and product and str(product).strip() != "TOTAL":
            try:
                disp_val = int(dispatch)
                if disp_val > 0:
                    prod_str = str(product).strip()
                    ptype = "PET" if "PET" in prod_str.upper() or "BOTTLE" in prod_str.upper() or "MUSTARD" in prod_str.upper() or prod_str.upper().startswith("BT-") else "TUBE"
                    
                    from customer_normalization import normalize_customer_name, correct_customer_by_product
                    import os
                    alpha_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                    norm_customer = normalize_customer_name(customer, alpha_dir)
                    norm_customer = correct_customer_by_product(norm_customer, prod_str)
                    
                    recs.append({
                        "month": month_label,
                        "date": datetime.datetime(2026, month_num, 28), # Approximate end of month for summary
                        "type": ptype,
                        "customer": norm_customer,
                        "product": prod_str,
                        "dia": "",
                        "ord_qty": 0,
                        "disp_qty": disp_val,
                        "pof": "",
                        "client_po": "",
                        "sv_no": ""
                    })
            except:
                pass
    wb.close()
    return recs

def add_dispatch_log(available_months):
    print("\n[5/6] Adding 'Dispatch_Log' tab -> Production_Archive...")
    tube_disp = parse_dispatch_xls("dispatch nov to jul.xls", "TUBE")
    pet_disp  = parse_dispatch_xls("dispatch pet nov to jul.xls", "PET")
    all_disp  = tube_disp + pet_disp
    
    # Extract MTD dispatch from active files
    for label, path, mnum in available_months:
        active_disp = extract_mtd_dispatch(path, label, mnum)
        if active_disp:
            all_disp.extend(active_disp)

    wb = openpyxl.load_workbook(PRODUCTION_ARCHIVE)
    if "Dispatch_Log" in wb.sheetnames:
        del wb["Dispatch_Log"]

    ws = wb.create_sheet("Dispatch_Log", 2)
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value     = "TUBEX -- Historical Dispatch Log (Nov 2025 - Present)"
    c.font      = Font(bold=True, size=16, color=C_GOLD, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    disp_headers = [
        ("Month", 15), ("Date", 12), ("Type", 8), ("Customer", 32),
        ("Product Name", 32), ("Dia / Volume", 12), ("Ordered Qty", 14),
        ("Dispatched Qty", 14), ("POF #", 10), ("Client PO #", 14), ("SV #", 10)
    ]
    HDR_ROW = 3
    for ci, (h, w) in enumerate(disp_headers, 1):
        hdr_cell(ws, HDR_ROW, ci, h, bg=C_MID, size=10, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[HDR_ROW].height = 28

    data_row = HDR_ROW + 1
    for mi, r in enumerate(all_disp):
        bg_color = C_LIGHT if mi % 2 == 0 else C_ALT
        row_vals = [
            r["month"], r["date"], r["type"], r["customer"], r["product"],
            r["dia"], r["ord_qty"], r["disp_qty"], r["pof"], r["client_po"], r["sv_no"]
        ]
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=data_row, column=ci, value=val)
            c.fill      = PatternFill("solid", fgColor=bg_color)
            c.alignment = Alignment(horizontal="center" if ci in (1,2,3,6,9,10,11) else ("right" if ci in (7,8) else "left"), vertical="center")
            c.border    = thin_border()
            c.font      = Font(bold=(ci == 1), size=9, name="Calibri")
            if ci in (7, 8) and isinstance(val, (int, float)):
                c.number_format = "#,##0"
        ws.row_dimensions[data_row].height = 16
        data_row += 1

    ws.freeze_panes             = "B4"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f"A{HDR_ROW}:K{max(data_row - 1, HDR_ROW)}"
    wb.save(PRODUCTION_ARCHIVE)
    wb.close()
    print(f"       [OK] Dispatch_Log added ({len(all_disp)} records)")


def add_customer_breakdown(available_months):
    print("\n[6/6] Adding 'Customer Breakdown' tab -> Production_Archive...")
    from generate_customer_report import extract_all_customer_records
    recs = extract_all_customer_records()

    # Aggregate by customer and month
    cust_summary = {}
    for r in recs:
        cname = r["customer"]
        mname = r["month"]
        if cname not in cust_summary:
            cust_summary[cname] = {}
        if mname not in cust_summary[cname]:
            cust_summary[cname][mname] = {
                "tube_prod": 0, "tube_disp": 0,
                "pet_prod": 0,  "pet_disp": 0,
                "reject": 0
            }
        if r["type"] == "TUBE":
            cust_summary[cname][mname]["tube_prod"] += r["produced"]
            cust_summary[cname][mname]["tube_disp"] += r["dispatched"]
        else:
            cust_summary[cname][mname]["pet_prod"]  += r["produced"]
            cust_summary[cname][mname]["pet_disp"]  += r["dispatched"]
        cust_summary[cname][mname]["reject"] += r["reject"]

    wb = openpyxl.load_workbook(PRODUCTION_ARCHIVE)
    if "Customer Breakdown" in wb.sheetnames:
        del wb["Customer Breakdown"]
    ws = wb.create_sheet("Customer Breakdown", 1)

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = "TUBEX -- Customer Monthly Production & Dispatch Summary"
    c.font      = Font(bold=True, size=16, color=C_GOLD, name="Calibri")
    c.fill      = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    headers = [
        ("Customer", 32),
        ("Month", 15),
        ("TUBE Prod", 14),
        ("TUBE Disp", 14),
        ("PET Prod", 14),
        ("PET Disp", 14),
        ("Total Prod", 15),
        ("Total Disp", 15),
        ("Reject Qty", 13),
        ("Reject %", 11)
    ]

    HDR_ROW = 3
    for ci, (h, w) in enumerate(headers, 1):
        hdr_cell(ws, HDR_ROW, ci, h, bg=C_MID, size=10, wrap=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[HDR_ROW].height = 28

    data_row = HDR_ROW + 1
    cust_idx = 0
    monthOrder = ["November 2025", "December 2025", "January 2026", "February 2026", "March 2026", "April 2026", "May 2026", "June 2026", "July 2026", "August 2026"]

    for cust in sorted(cust_summary.keys()):
        bg_color = C_LIGHT if cust_idx % 2 == 0 else C_ALT
        sorted_months = sorted(cust_summary[cust].keys(), key=lambda m: monthOrder.index(m) if m in monthOrder else 99)
        for month in sorted_months:
            d = cust_summary[cust][month]
            tot_p = d["tube_prod"] + d["pet_prod"]
            tot_d = d["tube_disp"] + d["pet_disp"]
            rej_q = d["reject"]
            rej_p = (rej_q / tot_p) if tot_p > 0 else 0.0

            row_vals = [
                cust, month, d["tube_prod"], d["tube_disp"],
                d["pet_prod"], d["pet_disp"], tot_p, tot_d, rej_q, rej_p
            ]
            row_fmts = [None, None, "#,##0", "#,##0", "#,##0", "#,##0", "#,##0", "#,##0", "#,##0", "0.00%"]

            for ci, (val, fmt) in enumerate(zip(row_vals, row_fmts), 1):
                c = ws.cell(row=data_row, column=ci, value=val)
                c.fill      = PatternFill("solid", fgColor=bg_color)
                c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
                c.border    = thin_border()
                c.font      = Font(bold=(ci == 1 or ci == 7), size=9, name="Calibri")
                if fmt:
                    c.number_format = fmt
            ws.row_dimensions[data_row].height = 20
            data_row += 1
        cust_idx += 1

    ws.freeze_panes             = "B4"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = (
        f"A{HDR_ROW}:J{max(data_row - 1, HDR_ROW)}"
    )
    wb.save(PRODUCTION_ARCHIVE)
    wb.close()
    print(f"       [OK] Customer Breakdown added ({data_row - HDR_ROW - 1} summary rows)")


def cleanup_temp():
    if os.path.exists(TEMP_DIR):
        shutil.rmtree(TEMP_DIR, ignore_errors=True)

if __name__ == "__main__":
    print("=" * 62)
    print("  Tubex Archive Builder -- Dashboard + Production")
    print("=" * 62)

    available = [(l, p, m) for l, p, m in MONTH_FILES if os.path.exists(p)]

    try:
        kpi_data = build_dashboard_archive(available)
        build_production_archive(available)
        add_dashboard_summary(kpi_data)
        add_production_summary(available)
        add_dispatch_log(available)
        add_customer_breakdown(available)
    finally:
        cleanup_temp()

    print("\n" + "=" * 62)
    print("  [DONE] Archives built successfully!")
    print(f"  >> {DASHBOARD_ARCHIVE}")
    print(f"  >> {PRODUCTION_ARCHIVE}")
    print("=" * 62)
