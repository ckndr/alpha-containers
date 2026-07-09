import pandas as pd

file_path = r'd:\Alpha\Tubex_July26.xlsx'
df_inv = pd.read_excel(file_path, sheet_name='Inventory', header=None)

inv_full = {}
item_id_col = None
store_bal_col = None
wip_col = None

for _, row in df_inv.iterrows():
    for i, val in enumerate(row):
        if str(val).strip() == 'Item ID':
            item_id_col = i
        elif str(val).strip() == 'Store Balance':
            store_bal_col = i
        elif str(val).strip() == 'Work In Process':
            wip_col = i
    if item_id_col is not None and store_bal_col is not None:
        try:
            item_id = str(int(row[item_id_col]))
            store_bal = row[store_bal_col]
            wip = row[wip_col] if wip_col is not None else 0
            if pd.isna(wip): wip = 0
            if pd.isna(store_bal): store_bal = 0
            inv_full[item_id] = {'store': float(store_bal), 'wip': float(wip), 'total': float(store_bal) + float(wip)}
        except:
            pass

target_items = ['6', '406', '3595', '68', '578', '3598', '6935', '4155', '185', '186', '194']
print("Item ID | Store Balance |   WIP | Store + WIP")
print("-" * 55)
for item_id in target_items:
    if item_id in inv_full:
        d = inv_full[item_id]
        print(f"{item_id:>7} | {d['store']:>13.2f} | {d['wip']:>5.1f} | {d['total']:>11.2f}")
    else:
        print(f"{item_id:>7} | NOT FOUND")

# Cross check with MRP 2 sheet values
import openpyxl
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['MRP 2']

print("\n\nMRP 2 Sheet Material Rows (from actual sheet):")
print(f"{'Item ID':>7} | {'Category':>15} | {'Item Name':>35} | {'UOM':>5} | {'Required':>12} | {'Store Bal':>10} | {'Net Buy':>12} | Used In")
print("-" * 140)
# Material rows start at row 10 (after gap)
for row in ws.iter_rows(min_row=10, max_row=ws.max_row, values_only=True):
    if row[0] is not None and row[1] is not None and row[0] != 'Item ID':
        item_id = str(row[0])
        cat = str(row[1])
        name = str(row[2])
        uom = str(row[3])
        req = row[4] if row[4] else 0
        store = row[5] if row[5] else 0
        net = row[6] if row[6] else 0
        used = str(row[7]) if row[7] else ''
        print(f"{item_id:>7} | {cat:>15} | {name:>35} | {uom:>5} | {req:>12.2f} | {store:>10.2f} | {net:>12.2f} | {used}")
        
        # Compare with inventory
        if item_id in inv_full:
            inv_store = inv_full[item_id]['store']
            inv_wip = inv_full[item_id]['wip']
            inv_total = inv_full[item_id]['total']
            if abs(float(store) - inv_store) > 0.01:
                print(f"  ** MISMATCH: Sheet says {store}, Inventory Store={inv_store}, WIP={inv_wip}, Total={inv_total}")
            elif inv_wip > 0:
                print(f"  ** NOTE: WIP exists ({inv_wip}). Sheet uses Store only ({inv_store}), not Store+WIP ({inv_total})")
