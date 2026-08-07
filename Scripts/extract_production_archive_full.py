import openpyxl
import os
import sys
from datetime import datetime, date

EXCEL_PATH = r"D:\Alpha\Tubex Records\Production_Archive.xlsx"
OUTPUT_MD_PATH = r"C:\Users\HP\.gemini\antigravity\brain\bf493533-bad9-4325-ae4c-2e333aee3cbe\production_archive_full_audit.md"
WORKSPACE_MD_PATH = r"D:\Alpha\Tubex Records\Production_Archive_Extracted.md"

def format_cell_value(val):
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
        if 0 < abs(val) < 1 and round(val, 6) != round(val, 2):
            return f"{val * 100:.2f}%" if val < 0.2 else f"{val:.4f}"
        return f"{val:.2f}".rstrip('0').rstrip('.')
    return str(val).strip()

def extract_production_archive():
    if not os.path.exists(EXCEL_PATH):
        print(f"Error: {EXCEL_PATH} does not exist.")
        sys.exit(1)
        
    print(f"Loading workbook: {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet_names = wb.sheetnames
    
    print("\n" + "="*80)
    print("WORKBOOK SUMMARY")
    print("="*80)
    print(f"File Path: {EXCEL_PATH}")
    print(f"Total Sheet Count: {len(sheet_names)}")
    print("Sheet Names List:")
    for idx, name in enumerate(sheet_names, 1):
        print(f"  {idx:2d}. {name}")
    print("="*80 + "\n")
    
    md_lines = []
    md_lines.append("# Production Archive Complete Data Audit Report")
    md_lines.append(f"**Source File:** `{EXCEL_PATH}`  ")
    md_lines.append(f"**Total Sheets:** {len(sheet_names)}  ")
    md_lines.append(f"**Extraction Timestamp:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  \n")
    
    md_lines.append("## 1. Sheet Overview\n")
    md_lines.append("| # | Sheet Name | Dimensions | Max Row | Max Col | Merged Ranges | Non-Empty Rows | Header Row # | Data Rows Count |")
    md_lines.append("|---|------------|------------|---------|---------|---------------|----------------|--------------|-----------------|")

    sheets_data = {}

    for i, sname in enumerate(sheet_names, 1):
        ws = wb[sname]
        
        # Unroll merged cells
        merged_cells_map = {}
        merged_ranges_str = [str(r) for r in ws.merged_cells.ranges]
        for rng in ws.merged_cells.ranges:
            top_left_val = ws.cell(row=rng.min_row, column=rng.min_col).value
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    merged_cells_map[(r, c)] = top_left_val
                    
        # Extract row values
        non_empty_rows = []
        for r in range(1, ws.max_row + 1):
            row_vals = []
            for c in range(1, ws.max_column + 1):
                if (r, c) in merged_cells_map:
                    val = merged_cells_map[(r, c)]
                else:
                    val = ws.cell(row=r, column=c).value
                row_vals.append(val)
                
            # Check if empty row
            if any(v is not None and str(v).strip() != "" for v in row_vals):
                non_empty_rows.append((r, row_vals))

        # Detect Header Row vs Title vs Data
        header_row_idx = None
        
        for idx, (r_num, r_vals) in enumerate(non_empty_rows[:5]):
            distinct_str_vals = set(format_cell_value(v) for v in r_vals if v is not None and str(v).strip() != "")
            # Skip title rows where all non-empty values are identical (merged title banner)
            if len(distinct_str_vals) <= 1:
                continue
                
            str_vals = [format_cell_value(v) for v in r_vals if v is not None and str(v).strip() != ""]
            joined = " ".join(str_vals).lower()
            
            # Check for header indicator keywords
            if any(kw in joined for kw in ["date", "machine", "customer", "product name", "product id", "dia / volume", "tube prod"]):
                header_row_idx = idx
                break

        if header_row_idx is None:
            # Fallback
            for idx, (r_num, r_vals) in enumerate(non_empty_rows[:5]):
                distinct_str_vals = set(format_cell_value(v) for v in r_vals if v is not None and str(v).strip() != "")
                if len(distinct_str_vals) > 2:
                    header_row_idx = idx
                    break
            if header_row_idx is None:
                header_row_idx = 0

        header_r_num, header_r_vals = non_empty_rows[header_row_idx]
        data_rows = non_empty_rows[header_row_idx + 1:]

        md_lines.append(f"| {i} | {sname} | `{ws.dimensions}` | {ws.max_row} | {ws.max_column} | {len(merged_ranges_str)} | {len(non_empty_rows)} | Row #{header_r_num} | {len(data_rows)} |")
        
        sheets_data[sname] = {
            "sheet_name": sname,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_ranges": merged_ranges_str,
            "non_empty_rows": non_empty_rows,
            "header_row_idx": header_row_idx,
            "header_r_num": header_r_num,
            "header_r_vals": header_r_vals,
            "data_rows": data_rows
        }

    md_lines.append("\n---\n")
    md_lines.append("## 2. Detailed Data Per Sheet\n")

    for i, sname in enumerate(sheet_names, 1):
        sinfo = sheets_data[sname]
        data_rows = sinfo["data_rows"]
        merged_ranges = sinfo["merged_ranges"]
        header_r_vals = sinfo["header_r_vals"]
        header_r_num = sinfo["header_r_num"]
        all_rows = sinfo["non_empty_rows"]
        
        print(f"Processing Sheet {i:2d}/{len(sheet_names)}: '{sname}' ({len(data_rows)} data rows)...")
        
        md_lines.append(f"### {i}. Sheet: `{sname}`")
        md_lines.append(f"- **Dimensions:** `{sinfo['dimensions']}` (Max Row: {sinfo['max_row']}, Max Column: {sinfo['max_column']})")
        md_lines.append(f"- **Merged Cell Ranges ({len(merged_ranges)}):** {', '.join([f'`{r}`' for r in merged_ranges]) if merged_ranges else 'None'}")
        md_lines.append(f"- **Total Non-Empty Rows:** {len(all_rows)}")
        md_lines.append(f"- **Header Row:** Row #{header_r_num}")
        md_lines.append(f"- **Data Rows Count:** {len(data_rows)}\n")
        
        # Calculate max non-empty column index
        max_col_used = 0
        for _, r_vals in all_rows:
            for c_idx in range(len(r_vals) - 1, -1, -1):
                if r_vals[c_idx] is not None and str(r_vals[c_idx]).strip() != "":
                    if c_idx + 1 > max_col_used:
                        max_col_used = c_idx + 1
                    break
        max_col_used = max(max_col_used, 1)

        # Build column header names
        headers_formatted = []
        for c in range(max_col_used):
            val = header_r_vals[c] if c < len(header_r_vals) else None
            h_text = format_cell_value(val)
            headers_formatted.append(h_text if h_text != "" else f"Col_{c+1}")

        md_lines.append("#### Headers & Data Table\n")
        
        table_headers = ["Row #"] + headers_formatted
        table_headers = [h.replace("|", "\\|") for h in table_headers]
        
        md_lines.append("| " + " | ".join(table_headers) + " |")
        md_lines.append("| " + " | ".join(["---"] * len(table_headers)) + " |")
        
        # Print Title rows (rows before header) if any
        for r_idx in range(sinfo["header_row_idx"]):
            r_num, r_vals = all_rows[r_idx]
            # Use raw row vals without merged cell unrolling duplication for title display
            raw_title = format_cell_value(r_vals[0])
            row_cells = [f"**Title (Row #{r_num})**"] + [raw_title if c == 0 else "" for c in range(max_col_used)]
            md_lines.append("| " + " | ".join(row_cells) + " |")

        # Header Row representation
        hdr_cells = [f"**Header (Row #{header_r_num})**"] + [h.replace("|", "\\|") for h in headers_formatted]
        md_lines.append("| " + " | ".join(hdr_cells) + " |")

        # Data Rows
        for r_num, r_vals in data_rows:
            row_cells = [str(r_num)]
            for c in range(max_col_used):
                val = r_vals[c] if c < len(r_vals) else None
                formatted_val = format_cell_value(val).replace("|", "\\|").replace("\n", " ")
                row_cells.append(formatted_val)
            md_lines.append("| " + " | ".join(row_cells) + " |")
            
        md_lines.append("\n" + "="*40 + "\n")

    full_md_content = "\n".join(md_lines)
    
    # Write artifact
    os.makedirs(os.path.dirname(OUTPUT_MD_PATH), exist_ok=True)
    with open(OUTPUT_MD_PATH, "w", encoding="utf-8") as f:
        f.write(full_md_content)
    print(f"\nSaved audit markdown report to: {OUTPUT_MD_PATH}")

    # Write workspace copy
    with open(WORKSPACE_MD_PATH, "w", encoding="utf-8") as f:
        f.write(full_md_content)
    print(f"Saved workspace markdown report to: {WORKSPACE_MD_PATH}")

if __name__ == "__main__":
    extract_production_archive()
