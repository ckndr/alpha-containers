import openpyxl

wb = openpyxl.load_workbook(r'd:\Alpha\Aerosol\Aerosol BOM.xlsx')
ws = wb['Req. Calculator']

# Let's inspect Row 11, Row 12, Row 13
print("Before rewriting:")
for r in [11, 12, 13]:
    print(f"Row {r}: {[cell.value for cell in ws[r]]}")

# Let's write the values for Row 11 explicitly
ws.cell(row=11, column=2, value='=BOM!F10')
ws.cell(row=11, column=3, value='=BOM!I10')
ws.cell(row=11, column=4, value='=BOM!G10')
ws.cell(row=11, column=5, value='=BOM!H10')
ws.cell(row=11, column=6, value='=BOM!J10')

print("After rewriting:")
for r in [11, 12, 13]:
    print(f"Row {r}: {[cell.value for cell in ws[r]]}")

wb.save(r'd:\Alpha\Aerosol\Aerosol BOM.xlsx')
