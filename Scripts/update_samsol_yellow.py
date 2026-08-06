import openpyxl
wb = openpyxl.load_workbook('Tubex_Aug26.xlsx')

# BOM: Change Item 4070 to 4002 in rows 254 and 271
ws_bom = wb['BOM']

for rn in [254, 271]:
    ws_bom.cell(row=rn, column=7, value=4002)                            # G: Item ID
    ws_bom.cell(row=rn, column=8, value='MASTER BATCH YELLOW CMN# 3087') # H: Item Name
    print('BOM Row {}: Item ID changed 4070 -> 4002, name updated'.format(rn))

# MRP July: Change row 77 (Item 4070 -> 4002)
ws_mrp = wb['MRP']

ws_mrp.cell(row=77, column=1, value=4002)                            # A: Item ID
ws_mrp.cell(row=77, column=3, value='MASTER BATCH YELLOW CMN# 3087') # C: Item Name
print('MRP Row 77 (July): Item ID changed 4070 -> 4002, name updated')

# MRP Aug PET: Change row 101 (Item 4070 -> 4002)
ws_mrp.cell(row=101, column=1, value=4002)                            # A: Item ID
ws_mrp.cell(row=101, column=3, value='MASTER BATCH YELLOW CMN# 3087') # C: Item Name
print('MRP Row 101 (Aug PET): Item ID changed 4070 -> 4002, name updated')

# Save
wb.save('Tubex_Aug26.xlsx')
print()
print('Done! Saved successfully.')
print()

# Verify
wb2 = openpyxl.load_workbook('Tubex_Aug26.xlsx')
print('=== Verification ===')
print('BOM 254 G:', wb2['BOM'].cell(254,7).value, '| H:', wb2['BOM'].cell(254,8).value)
print('BOM 271 G:', wb2['BOM'].cell(271,7).value, '| H:', wb2['BOM'].cell(271,8).value)
print('MRP  77 A:', wb2['MRP'].cell(77,1).value,  '| C:', wb2['MRP'].cell(77,3).value)
print('MRP 101 A:', wb2['MRP'].cell(101,1).value, '| C:', wb2['MRP'].cell(101,3).value)
