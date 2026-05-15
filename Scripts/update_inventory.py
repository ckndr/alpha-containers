"""
Alpha Containers - Inventory Auto-Updater v3
=============================================
CHANGE LOG v3 (vs v2):
  ADDED: warnings.filterwarnings to suppress openpyxl Data Validation
  UserWarning from cluttering the issues summary.

CHANGE LOG v2:
  - Reads from inventory.xls (ERP Excel export) instead of PDF.
    File must be named "inventory.xls" in the same folder.
  - Overwrites the same AlphaContainers file in place (no new version).

USAGE: Double-click "Update Inventory.bat" in the Scripts folder.

FOLDER STRUCTURE:
    AlphaContainers/
        AlphaContainers_v9_xx.xlsx    (any version - latest is used)
        inventory.xls                 (fresh ERP export - Item Wise Consolidated)
        Scripts/
            Update_Inventory.bat      (double-click this)
            update_inventory.py       (this script)

ERP XLS FORMAT (Item Wise Consolidated Report):
  Col 0: Item ID (numeric rows) or category header (text rows)
  Col 2: Item Name
  Col 6: Opening qty
  Col 7: In (received)
  Col 8: Out (issued)
  Col 9: Balance
  Col 10: Unit
  Row 4 (index): contains date range string "From : DD-MM-YYYY To : DD-MM-YYYY"

WHAT THIS SCRIPT UPDATES in Inventory sheet:
  Col E (5): Opening
  Col F (6): Received
  Col G (7): Issued
  Matched by Item ID (col A). Items not in Inventory sheet are listed
  as warnings but NOT added. Only existing rows are updated.
"""

import os
import re
import glob

import warnings
warnings.filterwarnings("ignore", message=".*Data Validation.*")
warnings.filterwarnings("ignore", message=".*extension.*")
import pandas as pd
from openpyxl import load_workbook

# Scripts live in AlphaContainers/Scripts/
# Excel and ERP files live in AlphaContainers/ (one level up)
DATA_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def find_files():
    excels = glob.glob(os.path.join(DATA_DIR, "AlphaContainers*.xlsx"))
    xls    = os.path.join(DATA_DIR, "inventory.xls")

    if not excels:
        print("  ERROR: No AlphaContainers*.xlsx found in folder:")
        print("    " + DATA_DIR)
        return None, None
    if not os.path.exists(xls):
        print("  ERROR: inventory.xls not found in folder:")
        print("    " + DATA_DIR)
        print("  Export from ERP and save as 'inventory.xls' there.")
        return None, None

    excel_file = sorted(excels)[-1]
    print("  Folder:  " + DATA_DIR)
    print("  Excel:   " + os.path.basename(excel_file))
    print("  Source:  inventory.xls")
    return excel_file, xls


def parse_inventory_xls(xls_path):
    """
    Parse ERP inventory.xls (Item Wise Consolidated Report).
    Returns ({item_id: {name, opening, inward, out, balance, unit}}, date_range_str).
    Only rows where col 0 is a valid integer ID are treated as data rows.
    """
    df = pd.read_excel(xls_path, sheet_name=0, engine='xlrd', header=None)

    items = {}
    date_range = ""

    for _, row in df.iterrows():
        col0 = row[0]

        # Detect date range from the header row (col 0 contains the full string)
        if isinstance(col0, str) and 'From :' in col0 and 'To :' in col0:
            m = re.search(r'From\s*:\s*(\S+)\s+To\s*:\s*(\S+)', col0)
            if m:
                date_range = m.group(1) + " to " + m.group(2)
            continue

        # Data row: col 0 must be a valid integer item ID
        try:
            item_id = int(float(col0))
        except (ValueError, TypeError):
            continue

        name    = str(row[2]).strip() if pd.notna(row[2]) else ""
        opening = float(row[6]) if pd.notna(row[6]) else 0.0
        inward  = float(row[7]) if pd.notna(row[7]) else 0.0
        out     = float(row[8]) if pd.notna(row[8]) else 0.0
        balance = float(row[9]) if pd.notna(row[9]) else 0.0
        unit    = str(row[10]).strip() if pd.notna(row[10]) else ""

        items[item_id] = {
            'name':    name,
            'opening': opening,
            'inward':  inward,
            'out':     out,
            'balance': balance,
            'unit':    unit,
        }

    return items, date_range


def _n(v):
    if v is None:
        return 0.0
    try:
        if isinstance(v, str) and v.startswith('='):
            return '(formula)'
        return float(v)
    except Exception:
        return 0.0


def update_excel(excel_path, xls_items, date_range):
    wb = load_workbook(excel_path)
    ws = wb['Inventory']

    # Update date range in cell A1 if present
    if date_range:
        cell = ws.cell(row=1, column=1)
        if cell.value:
            cell.value = re.sub(r'\(.*?\)', '(' + date_range + ')', str(cell.value))

    # Build map of item ID -> row number from Inventory sheet (col A)
    excel_ids = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            try:
                excel_ids[int(float(str(val)))] = row
            except (ValueError, TypeError):
                pass

    updated      = []
    not_in_excel = []

    for item_id, data in sorted(xls_items.items()):
        if item_id in excel_ids:
            row = excel_ids[item_id]
            old = [ws.cell(row=row, column=c).value for c in [5, 6, 7]]
            new = [data['opening'], data['inward'], data['out']]

            ws.cell(row=row, column=5).value = new[0]
            ws.cell(row=row, column=6).value = new[1]
            ws.cell(row=row, column=7).value = new[2]

            if any(_n(old[i]) != new[i] for i in range(3)):
                updated.append((
                    item_id, data['name'],
                    "Open: %s->%s" % (_n(old[0]), new[0]),
                    "In: %s->%s"   % (_n(old[1]), new[1]),
                    "Out: %s->%s"  % (_n(old[2]), new[2]),
                ))
        else:
            not_in_excel.append((item_id, data['name'], data['balance'], data['unit']))

    # Overwrite same file
    wb.save(excel_path)
    return updated, not_in_excel


def main():
    SEP = "=" * 55

    print("")
    print(SEP)
    print("  Alpha Containers - Inventory Updater v3")
    print(SEP)

    print("")
    print("[1/3] Finding files...")
    excel_path, xls_path = find_files()
    if not excel_path:
        return

    print("")
    print("[2/3] Reading ERP inventory.xls...")
    xls_items, date_range = parse_inventory_xls(xls_path)
    print("  Found %d items" % len(xls_items))
    if date_range:
        print("  Period: " + date_range)

    print("")
    print("[3/3] Updating Excel...")
    updated, not_in_excel = update_excel(excel_path, xls_items, date_range)
    print("  Saved: " + os.path.basename(excel_path))

    print("")
    print(SEP)
    if updated:
        print("  %d items UPDATED:" % len(updated))
        print("-" * 55)
        for item_id, name, e, f, g in updated:
            print("  ID %5d: %s" % (item_id, name[:30]))
            for c in [e, f, g]:
                parts = c.split('->')
                if len(parts) == 2:
                    old_val = parts[0].split(': ')[-1].strip()
                    new_val = parts[1].strip()
                    if old_val != new_val:
                        print("           " + c)
    else:
        print("  No changes - inventory already matches ERP.")

    if not_in_excel:
        print("")
        print("  %d ERP items NOT in Inventory sheet (not added):" % len(not_in_excel))
        for item_id, name, bal, unit in not_in_excel[:8]:
            print("    ID %5d: %-38s  Bal=%s %s" % (item_id, name[:38], bal, unit))
        if len(not_in_excel) > 8:
            print("    ...and %d more" % (len(not_in_excel) - 8))

    print("")
    print("  Open Excel and press Ctrl+Shift+F9 to recalculate.")
    print(SEP)


if __name__ == "__main__":
    main()
