"""
update_html.py
Tubex — Dashboard HTML Auto-Updater
------------------------------------------------
Reads Tubex_vX_XX.xlsx → calculates live KPIs, MTD production,
downtime, and order compliance → injects into Tubex.html.

Also rebuilds PRODUCTS array and BOM dict directly from Excel on every run.
New or re-activated products are automatically picked up — no manual HTML edits needed.

Run manually or add to Run All Updates.bat:
    python update_html.py

Author: Sikander / Claude
Version: 2.1
Changes from 2.0:
  - read_orders_dynamic: also captures rows where dispatch > 0 even when ordered = 0
    (fixes PHLOGIN GEL and any future completed-but-dispatched products being missed)
  - DASH_DATA kpi now includes tubeMTDDispatch / petMTDDispatch totals
    requires dispatch.xls date column mapping to compute; add when file structure known)
"""

import os, re, json, glob
from datetime import datetime, date

# ── PATH SETUP ──────────────────────────────────────────────
# Scripts live in Tubex/Scripts/ — Excel and HTML are one level up
DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

excel_pattern = os.path.join(DIR, 'Tubex*.xlsx')
excel_files   = sorted(glob.glob(excel_pattern))
if not excel_files:
    raise FileNotFoundError(f"No Tubex*.xlsx found in {DIR}")
EXCEL_PATH = excel_files[-1]
HTML_PATH  = os.path.join(DIR, 'Tubex.html')

print(f"Reading:  {os.path.basename(EXCEL_PATH)}")
print(f"Updating: {os.path.basename(HTML_PATH)}")

def recalculate_formulas_via_com(file_path):
    import platform
    if platform.system() != 'Windows':
        return False
    try:
        import win32com.client
        abs_path = os.path.abspath(file_path)
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_com = excel.Workbooks.Open(abs_path)
        wb_com.Save()
        wb_com.Close(SaveChanges=True)
        excel.Quit()
        print(f"  Recalculated formulas in Excel via COM: {os.path.basename(file_path)}")
        return True
    except Exception as e:
        print(f"  Warning: Could not recalculate formulas via Excel COM: {e}")
        return False

# Evaluate safe math formulas like "=400000-98056"
def evaluate_math_formula(formula_str):
    if not formula_str:
        return 0
    val_str = str(formula_str).strip()
    if not val_str.startswith('='):
        try:
            return int(float(val_str))
        except ValueError:
            return 0
    expr = val_str[1:].strip()
    clean_expr = re.sub(r'[^0-9\+\-\*\/\.\(\)\s]', '', expr)
    if not clean_expr.strip():
        return 0
    try:
        return int(eval(clean_expr, {"__builtins__": None}, {}))
    except Exception:
        return 0

wb_form = None # Lazy load only when needed

def get_mrp_formula_value(row, col):
    global wb_form
    if wb_form is None:
        try:
            wb_form = openpyxl.load_workbook(EXCEL_PATH, data_only=False, read_only=True)
        except Exception:
            return None
    try:
        ws_mrp_form = wb_form['MRP']
        return ws_mrp_form.cell(row=row, column=col).value
    except Exception:
        return None

# ── LOAD EXCEL ───────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    raise ImportError("openpyxl not installed. Run: pip install openpyxl --break-system-packages")

# Recalculate formulas via COM before reading if possible
recalculate_formulas_via_com(EXCEL_PATH)

wb     = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws_pl  = wb['Production_Log']
ws_db  = wb['Tubex_Dashboard']
ws_bom = wb['BOM']
ws_cat = wb['Product_Catalog']

# ── MONTH FILTER ─────────────────────────────────────────────
now        = datetime.now()

# Dynamically resolve reporting month from Production_Log
dates_in_log = []
for r in range(3, ws_pl.max_row + 1):
    val = ws_pl.cell(row=r, column=1).value
    if val and isinstance(val, (date, datetime)):
        dates_in_log.append(val)

if dates_in_log:
    latest_log_date = max(dates_in_log)
    cur_month  = latest_log_date.month
    cur_year   = latest_log_date.year
    month_name = latest_log_date.strftime('%B %Y')
else:
    cur_month  = now.month
    cur_year   = now.year
    month_name = now.strftime('%B %Y')

# ── READ PRODUCTION LOG ──────────────────────────────────────
# Cols (1-based): A=Date B=Machine C=Customer D=Product E=Dia F=PID
#                 G=Target H=Good I=Reject J=Waste%
#                 K=Mech L=Elec M=MatShort N=Changeover O=Operations
#                 P=PowerShutdown Q=GasShutdown R=WorkersShort

mtd_by_pid  = {}
yest_tube   = 0
yest_pet    = 0
latest_date = None

downtime_cols = {
    'Mechanical':        11,
    'Electrical':        12,
    'Material Shortage': 13,
    'Changeover':        14,
    'Operations':        15,
    'Power Shutdown':    16,
    'Gas Shutdown':      17,
    'Workers Shortage':  18,
}
dt_totals = {k: 0.0 for k in downtime_cols}

for row in ws_pl.iter_rows(min_row=3, values_only=True):
    row_date  = row[0]
    machine   = row[1]
    prod_name = row[3]
    pid       = row[5]
    good_qty  = row[7]

    if not row_date or not machine: continue
    if not isinstance(row_date, (date, datetime)): continue

    row_dt = row_date if isinstance(row_date, datetime) else datetime(row_date.year, row_date.month, row_date.day)
    if row_dt.month != cur_month or row_dt.year != cur_year: continue

    if latest_date is None or row_dt.date() > latest_date:
        latest_date = row_dt.date()

    mach_up = str(machine).upper()
    is_press_print = mach_up.startswith('PRESS') or mach_up.startswith('PRINT') or mach_up.startswith('PLINE')
    if is_press_print:
        for cat, col_idx in downtime_cols.items():
            val = row[col_idx - 1]
            if val and isinstance(val, (int, float)):
                dt_totals[cat] += float(val)

    if not pid or not good_qty: continue
    good_qty = int(good_qty)

    mach_up  = str(machine).upper()
    is_print = mach_up.startswith('PRINT') or mach_up.startswith('PLINE')
    is_pet   = mach_up.startswith('PF') or mach_up.startswith('PET')
    is_varn  = '(VARNISH)' in str(prod_name).upper()

    if (is_print and not is_varn) or is_pet:
        pid_int = int(pid)
        mtd_by_pid[pid_int] = mtd_by_pid.get(pid_int, 0) + good_qty

# Yesterday's production
if latest_date:
    for row in ws_pl.iter_rows(min_row=3, values_only=True):
        row_date  = row[0]
        machine   = row[1]
        prod_name = row[3]
        good_qty  = row[7]
        if not row_date or not machine or not good_qty: continue
        if not isinstance(row_date, (date, datetime)): continue
        row_dt = row_date if isinstance(row_date, datetime) else datetime(row_date.year, row_date.month, row_date.day)
        if row_dt.date() != latest_date: continue
        mach_up  = str(machine).upper()
        is_print = mach_up.startswith('PRINT') or mach_up.startswith('PLINE')
        is_pet   = mach_up.startswith('PF') or mach_up.startswith('PET')
        is_varn  = '(VARNISH)' in str(prod_name).upper()
        if is_print and not is_varn:
            yest_tube += int(good_qty)
        elif is_pet:
            yest_pet += int(good_qty)

tube_mtd = sum(v for k, v in mtd_by_pid.items() if k < 8000)
pet_mtd  = sum(v for k, v in mtd_by_pid.items() if k >= 8000)

print(f"\nKPIs for {month_name}:")
print(f"  Tube MTD: {tube_mtd:,}  |  Yesterday: {yest_tube:,}")
print(f"  PET MTD:  {pet_mtd:,}   |  Yesterday: {yest_pet:,}")

# ── READ ACTIVE ORDERS (DYNAMIC SCAN) ────────────────────────
# Dashboard layout (1-based columns):
#   A=blank  B=Type(TUBE/PET)  C=Customer  D=Product  E=Dia
#   F=ProdID  G=Orders  H=MTDProd  I=Remaining  J=Compliance  K=Dispatch
#
# A row is ACTIVE if col B matches the type AND at least ONE of:
#   • col G (ordered qty)   > 0  — has an open or in-progress order
#   • MTD produced (Log)    > 0  — was produced this month even if order closed
#   • col K (dispatch)      > 0  — dispatched this month (e.g. PHLOGIN ordered=0, dispatch=35,280)
# A row is inactive ONLY when all three are zero simultaneously.
# Scanning rows 11–100 catches any future rows added when more orders come in.

# Read MRP sheet to resolve Tube order formula lookup values (since openpyxl loads return None for formula cells)
mrp_orders_map = {}
if 'MRP' in wb.sheetnames:
    ws_mrp_temp = wb['MRP']
    for r_mrp in range(3, 150):
        pid_val = ws_mrp_temp.cell(row=r_mrp, column=4).value
        ord_val = ws_mrp_temp.cell(row=r_mrp, column=6).value
        if pid_val is not None:
            try:
                pid_int = int(pid_val)
                ord_val_int = int(ord_val) if ord_val else 0
                if ord_val_int == 0:
                    formula_str = get_mrp_formula_value(r_mrp, 6)
                    if formula_str:
                        ord_val_int = evaluate_math_formula(formula_str)
                mrp_orders_map[pid_int] = ord_val_int
            except (TypeError, ValueError):
                pass

def read_orders_dynamic(ws, order_type):
    orders = []
    is_pet = order_type.upper() == 'PET'
    for r in range(11, 200):
        type_val = ws.cell(row=r, column=2).value   # col B
        pid      = ws.cell(row=r, column=6).value   # col F
        ordered  = ws.cell(row=r, column=7).value   # col G
        dispatch = ws.cell(row=r, column=11).value  # col K

        if not type_val:
            continue
        if str(type_val).strip().upper() != order_type.upper():
            continue
        if not pid:
            continue
        try:
            pid_int = int(pid)
        except (TypeError, ValueError):
            continue
        try:
            ordered_int = int(ordered) if ordered else 0
        except (TypeError, ValueError):
            ordered_int = 0

        # Fallback to MRP sheet for TUBE orders if it evaluates to 0 or is None/formula
        if not is_pet and ordered_int == 0:
            ordered_int = mrp_orders_map.get(pid_int, 0)

        try:
            dispatch_int = int(dispatch) if dispatch else 0
        except (TypeError, ValueError):
            dispatch_int = 0

        # Resolve MTD produced here so we can use it in the active check
        produced = mtd_by_pid.get(pid_int, 0)

        # Skip truly inactive rows — must have NO order, NO production, AND NO dispatch
        if ordered_int == 0 and produced == 0 and dispatch_int == 0:
            continue

        product  = ws.cell(row=r, column=4).value
        customer = ws.cell(row=r, column=3).value
        dia      = ws.cell(row=r, column=5).value

        dia_raw = str(dia).replace(' ml', '').replace('ml', '') \
                          .replace(' mm', '').replace('mm', '').strip() if dia else ''
        dia_str = (dia_raw + ('ml' if is_pet else 'mm')) if dia_raw else '—'

        orders.append({
            'pid':      pid_int,
            'product':  str(product) if product else '',
            'customer': str(customer) if customer else '—',
            'dia':      dia_str,
            'ordered':  ordered_int,
            'produced': produced,
            'dispatch': dispatch_int,
        })
    return orders

tube_orders = read_orders_dynamic(ws_db, 'TUBE')
pet_orders  = read_orders_dynamic(ws_db, 'PET')

# ── DISPATCH TOTALS ───────────────────────────────────────────
# Monthly: sum dispatch column across all active order rows
tube_mtd_dispatch = sum(o['dispatch'] for o in tube_orders)
pet_mtd_dispatch  = sum(o['dispatch'] for o in pet_orders)

print(f"\nTube orders found: {len(tube_orders)}")
for o in tube_orders:
    print(f"  PID {o['pid']} — {o['product'][:35]} | ordered={o['ordered']:,} | produced={o['produced']:,} | dispatch={o['dispatch']:,}")
print(f"\nPET orders found: {len(pet_orders)}")
for o in pet_orders:
    print(f"  PID {o['pid']} — {o['product'][:35]} | ordered={o['ordered']:,} | produced={o['produced']:,} | dispatch={o['dispatch']:,}")
print(f"\nDispatch MTD — Tubes: {tube_mtd_dispatch:,}  |  PET: {pet_mtd_dispatch:,}")

# ── BUILD PRODUCTS FROM PRODUCT_CATALOG ──────────────────────
# Shorten long customer names for display in the app
CUST_SHORT = {
    'Brookes Pharma Private Limited':              'Brookes Pharma',
    'Samsol International Private Limited':        'Samsol International',
    'Mablay Beauty PVT LTD.':                      'Mablay Beauty',
    'Golden Pearl Cosmetics (PVT) LTD':            'Golden Pearl Cosmetics',
    'Al-Rehman Group':                             'Al-Rehman Group',
    'Professional Beauty Solution (PVT) LTD.Pakistan': 'Professional Beauty Solution',
    'Professional Beauty Solution (PVT) LTD. Pakistan':'Professional Beauty Solution',
    'Seatle (Private) Limited':                    'Seatle (Private) Limited',
    'Mega Grey':   'Mega Grey',
    'Hola Hair':   'Hola Hair',
    'Adore':       'Adore',
    'Bahadur':     'Bahadur',
    'DTM':         'DTM',
    'Alpha Labs PVT LTD':  'Alpha Labs',
    'Alpha Labs PVT LTD.': 'Alpha Labs',
}

products_list = []
for row in ws_cat.iter_rows(min_row=3, values_only=True):
    pid      = row[0]   # col A
    customer = row[2]   # col C
    name     = row[3]   # col D
    dia      = row[4]   # col E
    if not pid or not name:
        continue
    try:
        pid_int = int(pid)
    except (TypeError, ValueError):
        continue

    dia_str = str(dia).strip() if dia else '?'
    is_pet  = 'ml' in dia_str.lower()
    p_type  = 'pet' if is_pet else 'tube'

    if is_pet:
        dia_out = dia_str.replace(' ', '').lower()   # "120 ml" → "120ml"
    else:
        try:
            d = float(dia_str.replace('mm', '').strip())
            dia_out = str(int(d)) if d == int(d) else str(d)
        except (ValueError, TypeError):
            dia_out = dia_str

    cust_short = CUST_SHORT.get(str(customer), str(customer)) if customer else '—'

    products_list.append({
        'pid':      pid_int,
        'type':     p_type,
        'dia':      dia_out,
        'customer': cust_short,
        'name':     str(name),
    })

print(f"\nProducts loaded from catalog: {len(products_list)}")

# ── BUILD BOM FROM BOM SHEET ─────────────────────────────────
# Categories excluded from the Material Calculator (packing jali / metal frames)
SKIP_CATS = {'JALI', 'IRON AND METAL'}

# UOM normalisation — match what MRP/Inventory sheets use
UOM_NORM = {'KGS': 'kg', 'LTRS': 'L', 'NO': 'pcs', 'PCS': 'pcs'}

bom_dict = {}   # pid_int → [mat, ...]
for row in ws_bom.iter_rows(min_row=3, values_only=True):
    pid        = row[0]    # col A = Product ID
    mat_cat    = row[5]    # col F = Material Category
    item_id    = row[6]    # col G = Item ID
    item_name  = row[7]    # col H = Item Name
    uom        = row[8]    # col I = UOM
    per_1000   = row[9]    # col J = Per 1000 Units
    mat_group  = row[10]   # col K = Material Group
    scrap      = row[11]   # col L = Scrap %
    change_note= row[12]   # col M = Change Note

    if not pid or not mat_cat:
        continue
    try:
        pid_int = int(pid)
    except (TypeError, ValueError):
        continue

    cat_upper = str(mat_cat).strip().upper()
    if cat_upper in SKIP_CATS:
        continue
    if item_id is None:
        continue        # no item assigned yet
    if per_1000 is None:
        continue        # no rate at all — row is completely empty

    # est flag: red-font items are estimated/placeholder
    est = False
    try:
        if int(item_id) >= 9000:
            est = True
    except (TypeError, ValueError):
        pass
    if mat_group and str(mat_group).strip().lower() == 'estimated':
        est = True
    if item_name and '(est)' in str(item_name).lower():
        est = True
    if change_note and ('PLACEHOLDER' in str(change_note).upper() or
                        'EST.' in str(change_note).upper()):
        est = True
    # Zero-rate line = not yet confirmed, treat as estimated
    if per_1000 == 0:
        est = True

    uom_raw = str(uom).strip() if uom else 'kg'
    uom_out = UOM_NORM.get(uom_raw.upper(), uom_raw)

    mat = {
        'cat':  cat_upper,
        'id':   int(item_id),
        'name': str(item_name).strip() if item_name else '',
        'uom':  uom_out,
        'r':    round(float(per_1000), 7),
        's':    round(float(scrap), 4) if scrap else 0.1,
        'est':  est,
    }

    bom_dict.setdefault(pid_int, []).append(mat)

print(f"BOM entries loaded: {sum(len(v) for v in bom_dict.values())} lines across {len(bom_dict)} products")

# ── JS HELPERS ───────────────────────────────────────────────
def js_esc(s):
    return str(s).replace('\\', '\\\\').replace("'", "\\'").replace('\n', '').replace('\r', '')

def build_products_js(pl):
    lines = [
        f"  {{pid:{p['pid']},type:'{p['type']}',dia:'{js_esc(p['dia'])}'"
        f",customer:'{js_esc(p['customer'])}',name:'{js_esc(p['name'])}'}}"
        for p in pl
    ]
    return "const PRODUCTS = [\n" + ",\n".join(lines) + "\n];"

def build_bom_js(bd):
    pid_lines = []
    for pid in sorted(bd.keys()):
        mats = bd[pid]
        mat_parts = []
        for m in mats:
            est_js = 'true' if m['est'] else 'false'
            # Compact number formatting: drop trailing zeros
            r_str = f"{m['r']:.7f}".rstrip('0').rstrip('.')
            s_str = f"{m['s']:.4f}".rstrip('0').rstrip('.')
            mat_parts.append(
                f"{{cat:'{m['cat']}',id:{m['id']},name:'{js_esc(m['name'])}'"
                f",uom:'{m['uom']}',r:{r_str},s:{s_str},est:{est_js}}}"
            )
        pid_lines.append(f"  {pid}:[{','.join(mat_parts)}]")
    return "const BOM = {\n" + ",\n".join(pid_lines) + "\n};"

# ── BUILD DASH_DATA ──────────────────────────────────────────
DOWNTIME_ICONS = {
    'Mechanical':        '🔧',
    'Electrical':        '⚡',
    'Workers Shortage':  '👷',
    'Power Shutdown':    '🔌',
    'Material Shortage': '📦',
    'Gas Shutdown':      '🔥',
    'Changeover':        '🔄',
    'Operations':        '⚙️',
}

dash_data = {
    'month':       month_name,
    'lastUpdated': f"Updated {now.strftime('%d-%b-%Y %H:%M')} from {os.path.basename(EXCEL_PATH)}",
    'kpi': {
        'tubeYest':         yest_tube,
        'tubeMTD':          tube_mtd,
        'petYest':          yest_pet,
        'petMTD':           pet_mtd,
        'tubeMTDDispatch':  tube_mtd_dispatch,
        'petMTDDispatch':   pet_mtd_dispatch,
    },
    'downtime': [
        {'cat': cat, 'icon': DOWNTIME_ICONS.get(cat, '⏱'), 'hrs': round(dt_totals[cat] / 60, 2)}
        for cat in downtime_cols
    ],
    'tubeOrders': tube_orders,
    'petOrders':  pet_orders,
}

# ── READ INVENTORY ───────────────────────────────────────────
inventory_data = []
ws_inv = wb['Inventory']
for r in range(2, ws_inv.max_row + 1):
    item_id = ws_inv.cell(row=r, column=1).value
    if item_id is None:
        continue
    try:
        item_id = int(item_id)
    except (TypeError, ValueError):
        continue
    
    cat      = str(ws_inv.cell(row=r, column=2).value or '').strip()
    name     = str(ws_inv.cell(row=r, column=3).value or '').strip()
    uom      = str(ws_inv.cell(row=r, column=4).value or '').strip()
    opening  = float(ws_inv.cell(row=r, column=5).value or 0)
    received = float(ws_inv.cell(row=r, column=6).value or 0)
    issued   = float(ws_inv.cell(row=r, column=7).value or 0)
    balance  = ws_inv.cell(row=r, column=8).value
    wip      = ws_inv.cell(row=r, column=9).value
    
    # balance and wip may be formulas — data_only=True gives calculated values.
    # However, if the Excel file was modified and saved by openpyxl, cached values for
    # formulas are lost and return None. Calculate balance as: opening + received - issued.
    if balance is None or not isinstance(balance, (int, float)):
        balance = opening + received - issued
    else:
        balance = float(balance)
        
    wip = float(wip) if wip and isinstance(wip, (int, float)) else 0.0
    
    if not name:
        continue
    
    inventory_data.append({
        'id': item_id,
        'cat': cat,
        'name': name,
        'uom': uom,
        'opening': round(opening, 2),
        'received': round(received, 2),
        'issued': round(issued, 2),
        'balance': round(balance, 2),
        'wip': round(wip, 2),
    })

inv_title = str(ws_inv.cell(row=1, column=1).value or 'Inventory')

# ── READ PRODUCTION LOG FOR HTML ─────────────────────────────
prodlog_data = []
for row in ws_pl.iter_rows(min_row=3, values_only=True):
    row_date = row[0]
    machine  = row[1]
    if not row_date or not machine:
        continue
    if not isinstance(row_date, (date, datetime)):
        continue
    row_dt = row_date if isinstance(row_date, datetime) else datetime(row_date.year, row_date.month, row_date.day)
    if row_dt.month != cur_month or row_dt.year != cur_year:
        continue
    
    good    = int(row[7]) if row[7] and isinstance(row[7], (int, float)) else 0
    reject  = int(row[8]) if row[8] and isinstance(row[8], (int, float)) else 0
    total   = int(row[6]) if row[6] and isinstance(row[6], (int, float)) else 0
    
    prodlog_data.append({
        'date': row_dt.strftime('%d-%b'),
        'machine': str(machine).strip(),
        'customer': str(row[2] or '').strip(),
        'product': str(row[3] or '').strip(),
        'dia': str(row[4] or '').strip(),
        'pid': int(row[5]) if row[5] else None,
        'total': total,
        'good': good,
        'reject': reject,
    })

# ── READ FG STOCK ────────────────────────────────────────────
fg_data = []
ws_fg = wb['FG Stock']
fg_title = str(ws_fg.cell(row=1, column=1).value or 'FG Stock')

for r in range(4, ws_fg.max_row + 1):
    pid_val = ws_fg.cell(row=r, column=2).value
    product = ws_fg.cell(row=r, column=4).value
    if not product:
        continue
    
    fg_qty = ws_fg.cell(row=r, column=6).value
    fg_qty = int(fg_qty) if fg_qty and isinstance(fg_qty, (int, float)) else 0
    
    fg_data.append({
        'sr': ws_fg.cell(row=r, column=1).value,
        'pid': int(pid_val) if pid_val else None,
        'customer': str(ws_fg.cell(row=r, column=3).value or '').strip(),
        'product': str(product).strip(),
        'dia': str(ws_fg.cell(row=r, column=5).value or '').strip(),
        'qty': fg_qty,
        'status': str(ws_fg.cell(row=r, column=7).value or '').strip(),
        'remarks': str(ws_fg.cell(row=r, column=8).value or '').strip(),
    })

# ── READ MRP DATA ────────────────────────────────────────────
ws_mrp = wb['MRP']
mrp_title = str(ws_mrp.cell(row=1, column=1).value or 'Material Requirement Plan')

# Dynamically find starting rows for all sections in the MRP sheet
sec_rows = {
    'TUBE_ORDERS': None,
    'TUBE_MATERIALS': None,
    'PET_ORDERS': None,
    'PET_MATERIALS': None,
    'INK': None
}

for r in range(1, ws_mrp.max_row + 1):
    val = ws_mrp.cell(row=r, column=1).value
    val_str = str(val).strip() if val is not None else ""
    
    if "tube required orders" in val_str.lower():
        sec_rows['TUBE_ORDERS'] = r
    elif "tubes material requirement plan" in val_str.lower():
        sec_rows['TUBE_MATERIALS'] = r
    elif "pet required" in val_str.lower():
        sec_rows['PET_ORDERS'] = r
    elif "pet material requirement plan" in val_str.lower():
        sec_rows['PET_MATERIALS'] = r
    elif "ink" in val_str.lower() and ("avg monthly" in val_str.lower() or "usage" in val_str.lower()):
        sec_rows['INK'] = r

# Fallback to defaults if any section is not found
tube_orders_start = sec_rows['TUBE_ORDERS'] if sec_rows['TUBE_ORDERS'] is not None else 1
tube_mats_start = sec_rows['TUBE_MATERIALS'] if sec_rows['TUBE_MATERIALS'] is not None else 8
pet_orders_start = sec_rows['PET_ORDERS'] if sec_rows['PET_ORDERS'] is not None else 94
pet_mats_start = sec_rows['PET_MATERIALS'] if sec_rows['PET_MATERIALS'] is not None else 103
ink_start = sec_rows['INK'] if sec_rows['INK'] is not None else 113

def read_orders_from_range(start_row, end_row):
    orders = []
    for r in range(start_row, end_row):
        pid = ws_mrp.cell(row=r, column=4).value
        if not pid:
            continue
        try:
            pid = int(pid)
        except (TypeError, ValueError):
            continue
        
        required = ws_mrp.cell(row=r, column=6).value
        required_val = float(required) if required and isinstance(required, (int, float)) else 0.0
        if required_val == 0.0:
            formula_str = get_mrp_formula_value(r, 6)
            if formula_str:
                required_val = float(evaluate_math_formula(formula_str))
                
        produced = ws_mrp.cell(row=r, column=7).value
        produced_val = float(produced) if produced and isinstance(produced, (int, float)) else 0.0
        if produced_val == 0.0:
            produced_val = float(mtd_by_pid.get(pid, 0))
            
        remaining = ws_mrp.cell(row=r, column=8).value
        remaining_val = float(remaining) if remaining and isinstance(remaining, (int, float)) else 0.0
        if remaining_val == 0.0:
            remaining_val = max(0.0, required_val - produced_val)
            
        orders.append({
            'dia': ws_mrp.cell(row=r, column=1).value,
            'customer': str(ws_mrp.cell(row=r, column=2).value or '').strip(),
            'product': str(ws_mrp.cell(row=r, column=3).value or '').strip(),
            'pid': pid,
            'jobOrder': str(ws_mrp.cell(row=r, column=5).value or '').strip(),
            'required': required_val,
            'produced': produced_val,
            'remaining': remaining_val,
            'remarks': str(ws_mrp.cell(row=r, column=9).value or '').strip(),
        })
    return orders

mrp_orders = read_orders_from_range(tube_orders_start + 2, tube_mats_start)
mrp_pet_orders = read_orders_from_range(pet_orders_start + 2, pet_mats_start)

# Read material requirements (rows 18-101 for tubes)
mrp_materials = []
for r in range(tube_mats_start + 2, pet_orders_start):
    item_id = ws_mrp.cell(row=r, column=1).value
    if item_id is None:
        continue
    try:
        item_id_int = int(item_id)
    except (TypeError, ValueError):
        continue
    
    cat      = str(ws_mrp.cell(row=r, column=2).value or '').strip()
    name     = str(ws_mrp.cell(row=r, column=3).value or '').strip()
    uom      = str(ws_mrp.cell(row=r, column=4).value or '').strip()
    req_qty  = ws_mrp.cell(row=r, column=5).value
    stock    = ws_mrp.cell(row=r, column=6).value
    surplus  = ws_mrp.cell(row=r, column=7).value
    products = str(ws_mrp.cell(row=r, column=8).value or '').strip()
    status   = str(ws_mrp.cell(row=r, column=9).value or '').strip()
    
    req_qty = round(float(req_qty), 2) if req_qty and isinstance(req_qty, (int, float)) else 0
    stock   = round(float(stock), 2) if stock and isinstance(stock, (int, float)) else 0
    surplus = round(float(surplus), 2) if surplus and isinstance(surplus, (int, float)) else 0
    
    if not name:
        continue
    
    mrp_materials.append({
        'id': item_id_int,
        'cat': cat,
        'name': name,
        'uom': uom,
        'required': req_qty,
        'stock': stock,
        'surplus': surplus,
        'products': products,
        'status': status,
        'section': 'tube',
    })

# Read PET material requirements
for r in range(pet_mats_start + 2, ink_start):
    item_id = ws_mrp.cell(row=r, column=1).value
    if item_id is None:
        continue
    try:
        item_id_int = int(item_id)
    except (TypeError, ValueError):
        continue
    
    cat      = str(ws_mrp.cell(row=r, column=2).value or '').strip()
    name     = str(ws_mrp.cell(row=r, column=3).value or '').strip()
    uom      = str(ws_mrp.cell(row=r, column=4).value or '').strip()
    req_qty  = ws_mrp.cell(row=r, column=5).value
    stock    = ws_mrp.cell(row=r, column=6).value
    surplus  = ws_mrp.cell(row=r, column=7).value
    products = str(ws_mrp.cell(row=r, column=8).value or '').strip()
    status   = str(ws_mrp.cell(row=r, column=9).value or '').strip()
    
    req_qty = round(float(req_qty), 2) if req_qty and isinstance(req_qty, (int, float)) else 0
    stock   = round(float(stock), 2) if stock and isinstance(stock, (int, float)) else 0
    surplus = round(float(surplus), 2) if surplus and isinstance(surplus, (int, float)) else 0
    
    if not name:
        continue
    
    mrp_materials.append({
        'id': item_id_int,
        'cat': cat,
        'name': name,
        'uom': uom,
        'required': req_qty,
        'stock': stock,
        'surplus': surplus,
        'products': products,
        'status': status,
        'section': 'pet',
    })

# Read INK table
mrp_inks = []
for r in range(ink_start + 2, ws_mrp.max_row + 1):
    item_id = ws_mrp.cell(row=r, column=1).value
    if item_id is None:
        continue
    try:
        item_id_int = int(item_id)
    except (TypeError, ValueError):
        continue
    
    name    = str(ws_mrp.cell(row=r, column=3).value or '').strip()
    uom     = str(ws_mrp.cell(row=r, column=4).value or '').strip()
    avg_use = ws_mrp.cell(row=r, column=5).value  # Avg Monthly Usage
    days    = ws_mrp.cell(row=r, column=6).value   # Days of Stock
    status  = str(ws_mrp.cell(row=r, column=7).value or '').strip()
    stock   = ws_mrp.cell(row=r, column=8).value   # Current Stock
    
    avg_use = round(float(avg_use), 2) if avg_use and isinstance(avg_use, (int, float)) else 0
    days = round(float(days), 1) if days and isinstance(days, (int, float)) else 0
    stock   = round(float(stock), 2) if stock and isinstance(stock, (int, float)) else 0
    
    if not name:
        continue
    
    mrp_inks.append({
        'id': item_id_int,
        'name': name,
        'uom': uom,
        'avgUse': avg_use,
        'daysLeft': days,
        'status': status,
        'stock': stock,
    })

print(f"\nInventory items loaded: {len(inventory_data)}")
print(f"Production Log rows (current month): {len(prodlog_data)}")
print(f"FG Stock rows: {len(fg_data)}")
print(f"MRP orders: {len(mrp_orders)}  |  Materials: {len(mrp_materials)}  |  Inks: {len(mrp_inks)}")

# ── INJECT INTO HTML ─────────────────────────────────────────
if not os.path.exists(HTML_PATH):
    raise FileNotFoundError(f"HTML not found: {HTML_PATH}")

with open(HTML_PATH, 'r', encoding='utf-8') as f:
    html = f.read()

# ── SELF-PATCH: add CATALOG markers on first run ──────────────
# This only runs once — subsequent runs find the markers already in place.
if '/* CATALOG_START */' not in html:
    print("  First run: adding CATALOG markers to HTML...")
    # Insert CATALOG_START right after DATA_END (on a new line)
    html = html.replace(
        '/* DATA_END */',
        '/* DATA_END */\n/* CATALOG_START */'
    )
    # Insert CATALOG_END just before const CAT_ICON
    if '\nconst CAT_ICON' in html:
        html = html.replace('\nconst CAT_ICON', '\n/* CATALOG_END */\nconst CAT_ICON', 1)
    else:
        raise RuntimeError("Could not locate 'const CAT_ICON' to place CATALOG_END marker.")

if '/* INVENTORY_START */' not in html:
    print("  First run: adding new data markers to HTML...")
    # Insert after CATALOG_END
    html = html.replace(
        '/* CATALOG_END */',
        '/* CATALOG_END */\n'
        '/* INVENTORY_START */\n/* INVENTORY_END */\n'
        '/* PRODLOG_START */\n/* PRODLOG_END */\n'
        '/* FGSTOCK_START */\n/* FGSTOCK_END */\n'
        '/* MRP_START */\n/* MRP_END */'
    )

# ── INJECT DASH_DATA ─────────────────────────────────────────
marker_start = '/* DATA_START */'
marker_end   = '/* DATA_END */'
pos_start    = html.find(marker_start)
pos_end      = html.find(marker_end)
if pos_start == -1 or pos_end == -1:
    raise RuntimeError("DATA_START / DATA_END markers not found.")

new_data_block = (
    f"{marker_start}\n"
    f"const DASH_DATA = {json.dumps(dash_data, indent=2, ensure_ascii=False)};\n"
    f"{marker_end}"
)
html = html[:pos_start] + new_data_block + html[pos_end + len(marker_end):]

# Update timestamp comment if present
ts_old = re.search(r'Last generated: .+', html)
if ts_old:
    html = html.replace(ts_old.group(), f"Last generated: {now.strftime('%d-%b-%Y %H:%M')}")

# ── INJECT PRODUCTS + BOM ────────────────────────────────────
cat_start = '/* CATALOG_START */'
cat_end   = '/* CATALOG_END */'
pos_cs    = html.find(cat_start)
pos_ce    = html.find(cat_end)
if pos_cs == -1 or pos_ce == -1:
    raise RuntimeError("CATALOG_START / CATALOG_END markers not found after patching.")

products_js = build_products_js(products_list)
bom_js      = build_bom_js(bom_dict)

new_catalog_block = (
    f"{cat_start}\n"
    f"{products_js}\n\n"
    f"{bom_js}\n"
    f"{cat_end}"
)
html = html[:pos_cs] + new_catalog_block + html[pos_ce + len(cat_end):]

# ── INJECT NEW DATA CONSTANTS ────────────────────────────────
def inject_block(html, start_marker, end_marker, js_content):
    ps = html.find(start_marker)
    pe = html.find(end_marker)
    if ps == -1 or pe == -1:
        raise RuntimeError(f"{start_marker} / {end_marker} markers not found.")
    return html[:ps] + f"{start_marker}\n{js_content}\n{end_marker}" + html[pe + len(end_marker):]

html = inject_block(html, '/* INVENTORY_START */', '/* INVENTORY_END */',
    f"const INVENTORY_DATA = {json.dumps({'title': inv_title, 'items': inventory_data}, ensure_ascii=False)};")
html = inject_block(html, '/* PRODLOG_START */', '/* PRODLOG_END */',
    f"const PRODUCTION_LOG_DATA = {json.dumps({'month': month_name, 'rows': prodlog_data}, ensure_ascii=False)};")
html = inject_block(html, '/* FGSTOCK_START */', '/* FGSTOCK_END */',
    f"const FG_STOCK_DATA = {json.dumps({'title': fg_title, 'rows': fg_data}, ensure_ascii=False)};")
html = inject_block(html, '/* MRP_START */', '/* MRP_END */',
    f"const MRP_DATA = {json.dumps({'title': mrp_title, 'orders': mrp_orders, 'pet_orders': mrp_pet_orders, 'materials': mrp_materials, 'inks': mrp_inks}, ensure_ascii=False)};")

# ── WRITE HTML ───────────────────────────────────────────────
with open(HTML_PATH, 'w', encoding='utf-8') as f:
    f.write(html)

# ── UPDATE SERVICE WORKER CACHE VERSION ──────────────────────
SW_PATH = os.path.join(DIR, 'sw.js')
if os.path.exists(SW_PATH):
    with open(SW_PATH, 'r', encoding='utf-8') as f:
        sw = f.read()
    new_cache = f"tubex-{now.strftime('%Y%m%d%H%M')}"
    sw = re.sub(r"const CACHE_NAME = '[^']+';",
                f"const CACHE_NAME = '{new_cache}';", sw)
    with open(SW_PATH, 'w', encoding='utf-8') as f:
        f.write(sw)
    print(f"  SW cache version -> {new_cache}")

if wb_form:
    try:
        wb_form.close()
    except Exception:
        pass

print(f"\n[OK] HTML updated successfully -> {os.path.basename(HTML_PATH)}")
print(f"  Dashboard orders: {len(tube_orders)} tube + {len(pet_orders)} PET")
print(f"  Products in catalog: {len(products_list)}  |  BOM products: {len(bom_dict)}")
print(f"  Open in Chrome on PC or Android to view dashboard.")
