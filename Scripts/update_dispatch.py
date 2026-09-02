"""
Tubex - Dispatch Updater v10
----------------------------------------------------------------------
CHANGE LOG v10:
  BUGFIX — 19mm S-43 / S-45 disambiguation:
    PID 6624 was mislabeled as "S-43 DIA 19" in Product_Catalog; BOM 192 confirms
    it is "S-45 DIA 19MM". PID 6623 (BOM 193) is the S-43 DIA 19MM product.
    Both are now in Product_Catalog with correct names, so dispatch auto-resolves
    by exact name match — no NAME_FIXES entries required.
    NAME_FIXES entries added as explicit safety anchors regardless:
      "S-43 DIA 19MM" -> "S-43 DIA 19MM"  (PID 6623 — auto-match safety)
      "S-45 DIA 19MM" -> "S-45 DIA 19MM"  (PID 6624 — auto-match safety)
    These are identity mappings; they make the intent explicit and ensure
    that even if catalog names change slightly the script flags the mismatch.

CHANGE LOG v9:
  - Added dynamic date filtering: The script now automatically detects and
    ignores any dispatch records matching the current system date (today's data).
    If it's May 8th, all May 8th data is excluded.

CHANGE LOG v8:
  - NAME_FIXES: Added 3 new ERP name mappings confirmed from dispatch
    mismatch log:
      "PET BOTTLE (150ML)TRANSPARENT BODY MIST"              -> "TRANSPARENT BOTTLE 150ML"          (PID 8001)
      "PET BOTTLE SMALL (120 ML) YELLOW"                     -> "PET BOTTLE SMALL (120ML) YELLOW"   (PID 8005)
      "PET BOTTLE SMALL (130 ML) (TRANSPARENT) (WITHOUT CAP)"-> "PET BOTTLE SMALL (130ML) TRANSPARENT" (PID 8010)

CHANGE LOG v7:
  - NAME_FIXES updated after catalog rename (v10.3):
      REMOVED "BT-200 ML YELLOW" -> old "YELLOW LARGE BOTTLE 200ML"
              (ERP name now matches catalog exactly -- auto-resolves)
      REMOVED "PET BOTTLE LARGE 200ML WHITE" -> old "WHITE BOTTLE 200ML"
              (ERP name now matches catalog exactly -- auto-resolves)
      ADDED   "BT-120 ML YELLOW" -> "PET BOTTLE SMALL (120ML) YELLOW"
              (ERP uses BT- prefix; catalog uses descriptive name)
  - PET_TOTAL_ROW updated: 30 -> 32 (Dashboard restructured in v10.3,
    PET total row shifted from row 30 to row 32)
  - Stale K30 formula patch updated to K32.

CHANGE LOG v5:
  - Reads dispatch.xls (TUBEX-ALUM) and dispatch_pet.xls (TUBEX-PET)
  from the same folder, then writes MTD dispatch quantities into
  column K of the Tubex_Dashboard sheet in the latest
  Tubex*.xlsx file.

HOW IT WORKS:
  - Both dispatch files are ERP "Dispatch Report (Date Wise)" exports.
  - They are always full MTD. Every run replaces all K values.
  - ERP product names matched to Dashboard rows via Product_Catalog lookup.
  - Saves back to the SAME file. No new version is created.

NAME_FIXES:
  Only add entries here when the ERP dispatch name differs from the
  Product_Catalog name. Most tube products match automatically.
  Key   = product name exactly as it appears in dispatch .xls (stripped)
  Value = product name exactly as it appears in Product_Catalog col D
"""

import os
import glob
from datetime import datetime, date

import warnings
warnings.filterwarnings("ignore", message=".*Data Validation.*")
warnings.filterwarnings("ignore", message=".*extension.*")
import pandas as pd
from openpyxl import load_workbook
from alpha_checks import check_freshness, check_not_locked, log_mismatches, replace_copy_export


# -----------------------------------------------------------------------
# NAME_FIXES
# Only needed when ERP dispatch name != Product_Catalog name.
# If ERP name matches catalog name exactly (case-insensitive), no entry
# is needed here -- it is matched automatically.
#
# TO ADD A NEW ENTRY: paste the ERP name exactly as it appears in the
# dispatch .xls on the left, and the Product_Catalog col D name on right.
# -----------------------------------------------------------------------
NAME_FIXES = {
    # 19mm Samsol — two distinct products added v10.22:
    #   6623 = S-43 DIA 19MM (BOM 193)
    #   6624 = S-45 DIA 19MM (BOM 192)
    # Both names now exist in Product_Catalog and auto-resolve.
    # Entries below are safety anchors to make intent explicit.
    "S-43 DIA 19MM":  "S-43 DIA 19MM",   # PID 6623 — identity mapping, safety anchor
    "S-45 DIA 19MM":  "S-45 DIA 19MM",   # PID 6624 — identity mapping, safety anchor
    "HELLO HAIR":     "HELLO HAIR COLOR", # PID 6206
    "H.H 100GM":      "H.H 100GM",        # PID 6515

    # PET products -- ERP uses a different naming convention than catalog
    "BT-120 ML YELLOW":                                       "PET BOTTLE SMALL (120ML) YELLOW",   # PID 8005 — ERP name
    "BT-120ML YELLOW":                                        "PET BOTTLE SMALL (120ML) YELLOW",   # PID 8005 — ERP name without space
    "PET BOTTLE SMALL (120 ML) YELLOW":                       "PET BOTTLE SMALL (120ML) YELLOW",   # PID 8005 — ERP space: "120 ML" vs "120ML"
    "YELLOW SMALL BOTTLE 120ML":                              "PET BOTTLE SMALL (120ML) YELLOW",   # PID 8005 — old catalog name synonym
    "BT-200 ML YELLOW":                                       "PET BOTTLE LARGE (200 ML) YELLOW",  # PID 8006 — ERP name synonym
    "BT-200ML YELLOW":                                        "PET BOTTLE LARGE (200 ML) YELLOW",  # PID 8006 — ERP name synonym without space
    "PET BOTTLE LARGE 200ML YELLOW":                          "PET BOTTLE LARGE (200 ML) YELLOW",  # PID 8006 — variant name
    "PET BOTTLE LARGE (200 ML) YELLOW":                       "PET BOTTLE LARGE (200 ML) YELLOW",  # PID 8006 — variant name
    "YELLOW LARGE BOTTLE 200ML":                              "PET BOTTLE LARGE (200 ML) YELLOW",  # PID 8006 — old catalog name synonym
    "YELLOW BOTTLE 200ML":                                    "PET BOTTLE LARGE (200 ML) YELLOW",  # PID 8006 — variant name
    "PET BOTTLE SMALL (120ML) COMPACT BLACK":                 "BLACK SMALL BOTTLE 120ML",
    "BLACK SMALL BOTTLE 120ML":                               "BLACK SMALL BOTTLE 120ML",
    "WHITE SMALL BOTTLE 120ML":                               "WHITE SMALL BOTTLE 120ML",
    "TRANSPARENT BOTTLE 150ML":                               "TRANSPARENT BOTTLE 150ML",
    "PET BOTTLE (150ML)TRANSPARENT BODY MIST":                "TRANSPARENT BOTTLE 150ML",          # PID 8001 fallback
    "PET BOTTLE SMALL (130 ML) (TRANSPARENT) (WITHOUT CAP)":  "PET BOTTLE SMALL (130ML) TRANSPARENT",  # PID 8010 — ERP adds "(WITHOUT CAP)"
    "PET BOTTLE SMALL (130ML) TRANSPARENT":                   "PET BOTTLE SMALL (130ML) TRANSPARENT",
    "PET BOTTLE 130 ML WHITE":                                "PET BOTTLE 130ML WHITE",              # PID 8015 — ERP space: "130 ML" vs "130ML"
    "PET BOTTLE 130ML WHITE":                                 "PET BOTTLE 130ML WHITE",
}


# -----------------------------------------------------------------------
# Sheet / column constants
# -----------------------------------------------------------------------
DASHBOARD_SHEET     = "Tubex_Dashboard"
CATALOG_SHEET       = "Product_Catalog"

DASHBOARD_PID_COL   = 6    # col F -- PID in Dashboard
DASHBOARD_DISP_COL  = 11   # col K -- Dispatch qty in Dashboard
DASHBOARD_ROW_MIN   = 11
DASHBOARD_ROW_MAX   = 200

CATALOG_PID_COL     = 1    # col A -- Product ID
CATALOG_NAME_COL    = 4    # col D -- Product Name
CATALOG_DATA_START  = 3    # first data row (rows 1-2 are headers)




# -----------------------------------------------------------------------
def load_catalog(wb):
    """
    Read Product_Catalog sheet from already-open workbook.
    Returns {product_name_upper: pid} for all rows with a valid PID and name.
    """
    ws = wb[CATALOG_SHEET]
    catalog = {}
    for row in range(CATALOG_DATA_START, ws.max_row + 1):
        pid_val  = ws.cell(row, CATALOG_PID_COL).value
        name_val = ws.cell(row, CATALOG_NAME_COL).value
        if pid_val is None or name_val is None:
            continue
        try:
            pid = int(float(str(pid_val)))
        except (ValueError, TypeError):
            continue
        name = str(name_val).strip()
        if name:
            catalog[name.upper()] = pid
    return catalog


def resolve_pid(product_entry, catalog):
    """
    Given an ERP dispatch product entry (either string or (product_name, party_name)),
    return (catalog_name, pid) or (erp_name, None).
    """
    party_name = ""
    if isinstance(product_entry, tuple):
        erp_name, party_name = product_entry
    else:
        erp_name = str(product_entry)

    party_upper = party_name.upper()
    erp_upper = erp_name.upper()

    # 1. Party-specific disambiguation (e.g. Horizon vs Alpha for 150ml body mist)
    if "HORIZON" in party_upper and ("150" in erp_upper or "BODY MIST" in erp_upper):
        return "PET BOTTLE (150ML)TRANSPARENT BODY MIST", 8017
    if "ALPHA" in party_upper and ("150" in erp_upper or "BODY MIST" in erp_upper):
        return "TRANSPARENT BOTTLE 150ML", 8001

    # 2. Check NAME_FIXES
    catalog_name = NAME_FIXES.get(erp_name, erp_name)
    pid = catalog.get(catalog_name.upper())
    if pid is None:
        pid = catalog.get(erp_upper)
    return catalog_name, pid


def parse_dispatch_file(path, target_month_abbr=None):
    """
    Parse one ERP dispatch .xls (Date Wise format).
    Ignores any dispatch rows matching the current day (Rule R1-06 / AUDIT_NOTES.md).
    Returns {(product_name, party_name): total_dispatched_qty}.
    """
    import xlrd
    df = pd.read_excel(path, sheet_name=0, engine='xlrd', header=None)

    # Check file month header against active workbook month
    if target_month_abbr:
        file_month = None
        for idx in range(min(15, len(df))):
            row_vals = [str(x).strip() for x in df.iloc[idx].values if pd.notna(x)]
            for v in row_vals:
                if v.lower().startswith("month :") or v.lower().startswith("month:"):
                    file_month = v.upper()
                    break
            if file_month:
                break
        if file_month and target_month_abbr.upper() not in file_month:
            print(f"  [WARN] {os.path.basename(path)} header is '{file_month}', but active file is '{target_month_abbr}'. Stale previous-month dispatch ignored.")
            return {}

    SKIP_PREFIXES = ('dispatch report', 'month :', 'no.')
    SKIP_EXACT    = {'end of file'}

    result = {}
    current_product = None
    ignored_today = 0

    today = datetime.now()
    today_date = today.date()
    today_strs = [
        today.strftime("%d/%m/%Y").lower(),
        today.strftime("%d-%m-%Y").lower(),
        today.strftime("%d-%b-%Y").lower(),
        today.strftime("%Y-%m-%d").lower(),
        f"{today.day}/{today.month}/{today.year}",
        f"{today.day}-{today.month}-{today.year}",
        today.strftime("%d.%m.%Y").lower(),
        today.strftime("%d-%b-%y").lower(),
        today.strftime("%d/%m/%y").lower()
    ]

    # Dynamically detect dispatch quantity and party columns (Rule R1-07)
    col_disp_idx = 7
    col_party_idx = 12
    for idx, r in df.iterrows():
        row_str = [str(x).lower().strip() for x in r.values if pd.notna(x)]
        if any('disp' in s and 'qty' in s for s in row_str):
            for i, s in enumerate(r.values):
                if pd.notna(s) and 'disp' in str(s).lower() and 'qty' in str(s).lower():
                    col_disp_idx = i
                elif pd.notna(s) and 'party' in str(s).lower():
                    col_party_idx = i
            break

    for _, row in df.iterrows():
        col0 = row[0]
        col_disp = row[col_disp_idx] if col_disp_idx < len(row) else (row[7] if len(row) > 7 else None)

        if isinstance(col0, str) and str(col0).strip():
            c0       = col0.strip()
            c0_lower = c0.lower()
            if c0_lower in SKIP_EXACT:
                continue
            if any(c0_lower.startswith(p) for p in SKIP_PREFIXES):
                continue
            if 'grand total' in c0_lower:
                continue
            if len(row) > 1 and pd.isna(row[1]):
                current_product = c0
                continue

        try:
            int(float(col0))
            if pd.isna(col_disp):
                continue

            # Robust date check including xlrd serial floats (Rule R1-06)
            skip_row = False
            for val in row:
                if pd.isna(val) or val is None:
                    continue
                if hasattr(val, 'date') and callable(getattr(val, 'date')):
                    if val.date() == today_date:
                        skip_row = True
                        break
                elif isinstance(val, (datetime, date)):
                    d = val.date() if isinstance(val, datetime) else val
                    if d == today_date:
                        skip_row = True
                        break
                elif isinstance(val, (int, float)):
                    if 40000 <= val <= 55000:
                        try:
                            d = xlrd.xldate_as_datetime(val, 0).date()
                            if d == today_date:
                                skip_row = True
                                break
                        except Exception:
                            pass
                elif isinstance(val, str):
                    v_str = val.strip().lower()
                    if v_str in today_strs:
                        skip_row = True
                        break
                    for ts in today_strs:
                        if v_str.startswith(ts + ' ') or v_str.startswith(ts + 't'):
                            skip_row = True
                            break
                    try:
                        ts = pd.to_datetime(val, dayfirst=True, errors='coerce')
                        if pd.notna(ts) and ts.date() == today_date:
                            skip_row = True
                            break
                    except Exception:
                        pass

            if skip_row:
                ignored_today += 1
                continue

            party_name = str(row[col_party_idx]).strip() if col_party_idx < len(row) and pd.notna(row[col_party_idx]) else ""
            disp_qty = float(col_disp)
            if current_product is not None:
                key = (current_product, party_name) if party_name else current_product
                result[key] = result.get(key, 0) + disp_qty
        except (ValueError, TypeError):
            pass

    if ignored_today > 0:
        print(f"  -> Ignored {ignored_today} dispatch row(s) from today ({today_date}) in {os.path.basename(path)}")

    return result


def find_files(folder):
    replace_copy_export(folder, "dispatch.xls")
    replace_copy_export(folder, "dispatch_pet.xls")

    ac_files = glob.glob(os.path.join(folder, "Tubex*.xlsx"))
    if not ac_files:
        print("  ERROR: No Tubex*.xlsx found in: " + folder)
        return None, None, None
    ac = sorted(ac_files)[-1]

    dispatch_tube = os.path.join(folder, "dispatch.xls")
    dispatch_pet  = os.path.join(folder, "dispatch_pet.xls")

    for p, label in [(dispatch_tube, "dispatch.xls"), (dispatch_pet, "dispatch_pet.xls")]:
        if not os.path.exists(p):
            print("  ERROR: " + label + " not found in: " + folder)
            return None, None, None

    return ac, dispatch_tube, dispatch_pet


def build_pid_row_map(ws):
    """Scan Dashboard col F for PIDs. Returns {pid: row_number}."""
    pid_map = {}
    for r in range(DASHBOARD_ROW_MIN, DASHBOARD_ROW_MAX + 1):
        pid_val = ws.cell(r, DASHBOARD_PID_COL).value
        if pid_val is not None:
            try:
                pid_map[int(pid_val)] = r
            except (ValueError, TypeError):
                pass
    return pid_map


def update_dispatch(ac_path, dispatch_by_pid):
    wb = load_workbook(ac_path)
    ws = wb[DASHBOARD_SHEET]

    pid_row_map = build_pid_row_map(ws)

    for r in range(DASHBOARD_ROW_MIN, DASHBOARD_ROW_MAX + 1):
        cell = ws.cell(r, DASHBOARD_DISP_COL)
        if cell.value is not None and not isinstance(cell.value, str):
            cell.value = None

    updated = 0
    skipped = []
    for pid, qty in dispatch_by_pid.items():
        if pid not in pid_row_map:
            skipped.append((pid, qty))
            continue
        ws.cell(pid_row_map[pid], DASHBOARD_DISP_COL).value = int(qty)
        updated += 1

    wb.save(ac_path)
    return updated, skipped


def main():
    SEP = "=" * 60

    print("")
    print(SEP)
    print("  Tubex - Dispatch Updater v10")
    print(SEP)
    print("")

    folder = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    print("[1/4] Finding files...")
    ac_path, tube_path, pet_path = find_files(folder)
    if not ac_path:
        return
    print("  Tubex File:    " + os.path.basename(ac_path))
    print("  Dispatch Tube: " + os.path.basename(tube_path))
    print("  Dispatch PET:  " + os.path.basename(pet_path))

    print("")
    print("[1a] Safety checks...")
    check_not_locked(ac_path)
    check_freshness(tube_path, max_hours=26, label="dispatch.xls")
    check_freshness(pet_path, max_hours=26, label="dispatch_pet.xls")

    print("")
    print("[1b] Loading Product_Catalog from Excel...")
    wb_temp = load_workbook(ac_path, read_only=True, data_only=True)
    catalog = load_catalog(wb_temp)
    wb_temp.close()
    print("  %d products loaded from catalog." % len(catalog))

    print("")
    print("[2/4] Parsing dispatch files...")
    base = os.path.basename(ac_path)
    target_month = None
    for m in ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]:
        if m in base.upper():
            target_month = m
            break

    tube_dispatch = parse_dispatch_file(tube_path, target_month_abbr=target_month)
    pet_dispatch  = parse_dispatch_file(pet_path, target_month_abbr=target_month)

    print("")
    print("  TUBE dispatches (TUBEX-ALUM):")
    tube_total = 0
    for product in sorted(tube_dispatch, key=lambda x: str(x)):
        qty           = tube_dispatch[product]
        cat_name, pid = resolve_pid(product, catalog)
        disp_name     = f"{product[0]} ({product[1]})" if isinstance(product, tuple) else str(product)
        flag = ("-> matched: %s (PID %s)" % (cat_name, pid)) if pid else "-> !! NOT IN CATALOG"
        print("    %-45s %9d  %s" % (disp_name[:45], int(qty), flag))
        tube_total += qty
    print("    %-45s %9d" % ("GRAND TOTAL", int(tube_total)))

    print("")
    print("  PET dispatches (TUBEX-PET):")
    pet_total = 0
    for product in sorted(pet_dispatch, key=lambda x: str(x)):
        qty           = pet_dispatch[product]
        cat_name, pid = resolve_pid(product, catalog)
        disp_name     = f"{product[0]} ({product[1]})" if isinstance(product, tuple) else str(product)
        flag = ("-> matched: %s (PID %s)" % (cat_name, pid)) if pid else "-> !! NOT IN CATALOG"
        print("    %-45s %9d  %s" % (disp_name[:45], int(qty), flag))
        pet_total += qty
    print("    %-45s %9d" % ("GRAND TOTAL", int(pet_total)))

    print("")
    print("[3/4] Resolving PIDs and checking for mismatches...")
    dispatch_by_pid   = {}
    unmapped_products = []

    all_dispatches = {}
    all_dispatches.update(tube_dispatch)
    all_dispatches.update(pet_dispatch)

    for product, qty in all_dispatches.items():
        cat_name, pid = resolve_pid(product, catalog)
        if pid is None:
            unmapped_products.append((product, cat_name, qty))
        else:
            dispatch_by_pid[pid] = dispatch_by_pid.get(pid, 0) + qty

    if unmapped_products:
        print("")
        print("  !! PRODUCTS NOT MATCHED (not written to Dashboard):")
        for product, cat_name, qty in unmapped_products:
            if product in NAME_FIXES:
                print("    ERP: '%s' -> NAME_FIXES target '%s' not found in catalog. Check spelling." % (product, cat_name))
            else:
                print("    ERP: '%s'  qty=%d" % (product, int(qty)))
                print("    Fix options:")
                print("      A) If ERP name matches catalog col D exactly -> will auto-resolve next run.")
                print("      B) If ERP uses a different name -> add to NAME_FIXES in script.")
        log_mismatches("dispatch", unmapped_products)

    print("")
    print("[4/4] Updating Dashboard column K...")
    updated_count, skipped_pids = update_dispatch(ac_path, dispatch_by_pid)

    print("")
    print("  PIDs written to K:")
    for pid in sorted(dispatch_by_pid):
        print("    PID %-8d qty=%9d" % (pid, int(dispatch_by_pid[pid])))

    if skipped_pids:
        print("")
        print("  !! PIDs resolved from catalog but NOT found in Dashboard col F:")
        for pid, qty in skipped_pids:
            print("    PID %d  qty=%d  (check Dashboard col F has this PID)" % (pid, int(qty)))

    print("")
    print("  Updated %d product row(s)" % updated_count)
    print("  Saved:  " + os.path.basename(ac_path))

    # ── Post-write validation ────────────────────────────────────────────
    print("  Validating writes...")
    try:
        wb_check = load_workbook(ac_path, read_only=True, data_only=True)
        ws_check = wb_check[DASHBOARD_SHEET]
        dispatch_cells = sum(1 for r in range(DASHBOARD_ROW_MIN, DASHBOARD_ROW_MAX + 1) 
                            if ws_check.cell(r, DASHBOARD_DISP_COL).value is not None
                            and isinstance(ws_check.cell(r, DASHBOARD_DISP_COL).value, (int, float)))
        if dispatch_cells == updated_count:
            msg = f"  ✓ Dashboard col K: {dispatch_cells} dispatch values verified"
        else:
            msg = f"  !! Dashboard col K: expected {updated_count} values, found {dispatch_cells}"
        try:
            print(msg)
        except UnicodeEncodeError:
            print(msg.replace('✓', 'OK'))
        wb_check.close()
    except Exception as e:
        print(f"  !! Validation error: {e}")
    print("")
    print("  Press Ctrl+Shift+F9 in Excel to recalculate formulas.")
    print(SEP)


if __name__ == "__main__":
    main()
