import openpyxl
wb = openpyxl.load_workbook('Production.xlsx', data_only=True)

# Check Summary sheet for PET details
ws = wb['Summary 12-08-2026']
print('=== Summary 12-08-2026 ===')
for r in range(1, ws.max_row+1):
    row_data = []
    for c in range(1, min(15, ws.max_column+1)):
        v = ws.cell(r, c).value
        if v is not None:
            row_data.append(str(v)[:50])
    if row_data:
        print("R{}: {}".format(r, " | ".join(row_data)))

# Check Production Day wise for PET columns
print('\n=== Production Day wise (first 15 rows) ===')
ws2 = wb['Production Day wise']
for r in range(1, min(16, ws2.max_row+1)):
    row_data = []
    for c in range(1, min(26, ws2.max_column+1)):
        v = ws2.cell(r, c).value
        if v is not None:
            row_data.append("C{}={}".format(c, str(v)[:40]))
    if row_data:
        print("R{}: {}".format(r, " | ".join(row_data)))

# Check Dashboard
print('\n=== Dashboard (first 30 rows) ===')
ws3 = wb['Dashbord']
for r in range(1, min(31, ws3.max_row+1)):
    row_data = []
    for c in range(1, min(15, ws3.max_column+1)):
        v = ws3.cell(r, c).value
        if v is not None:
            row_data.append("C{}={}".format(c, str(v)[:40]))
    if row_data:
        print("R{}: {}".format(r, " | ".join(row_data)))

# Downtime Summary
print('\n=== Downtime Summary (all rows) ===')
ws4 = wb['Downtime Summary']
for r in range(1, ws4.max_row+1):
    row_data = []
    for c in range(1, min(11, ws4.max_column+1)):
        v = ws4.cell(r, c).value
        if v is not None:
            row_data.append("C{}={}".format(c, str(v)[:60]))
    if row_data:
        print("R{}: {}".format(r, " | ".join(row_data)))

# Pet Bottles dispatch
print('\n=== Pet Bottles dispatch (first 20 rows) ===')
ws5 = wb['Pet Bottles dispatch']
for r in range(1, min(21, ws5.max_row+1)):
    row_data = []
    for c in range(1, min(8, ws5.max_column+1)):
        v = ws5.cell(r, c).value
        if v is not None:
            row_data.append("C{}={}".format(c, str(v)[:50]))
    if row_data:
        print("R{}: {}".format(r, " | ".join(row_data)))

# Sheet3 - Product catalog
print('\n=== Sheet3 - Product Catalog (first 30 rows) ===')
ws6 = wb['Sheet3']
for r in range(1, min(31, ws6.max_row+1)):
    row_data = []
    for c in range(1, min(17, ws6.max_column+1)):
        v = ws6.cell(r, c).value
        if v is not None:
            row_data.append("C{}={}".format(c, str(v)[:50]))
    if row_data:
        print("R{}: {}".format(r, " | ".join(row_data)))
