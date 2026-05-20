import openpyxl

wb = openpyxl.load_workbook(r'd:\Alpha\Tubex_v10_27.xlsx', data_only=False)
ws = wb['MRP']

print("=== MRP Sheet Full Dump ===")
for row in range(1, ws.max_row + 1):
    cells = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        v = cell.value
        if v is not None:
            col_letter = openpyxl.utils.get_column_letter(col)
            cells.append("{}{}={}".format(col_letter, row, repr(v)))
    if cells:
        print("Row {}: {}".format(row, " | ".join(cells)))
    else:
        print("Row {}: (empty)".format(row))
