import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

wb = openpyxl.load_workbook('Tubex_Aug26.xlsx')
ws = wb['MRP']

def clone_border(b):
    def cs(s):
        return Side(border_style=s.border_style, color=copy(s.color)) if s else Side()
    return Border(left=cs(b.left), right=cs(b.right),
                  top=cs(b.top), bottom=cs(b.bottom))

def apply_header_style(cell, ref_cell):
    cell.font      = Font(bold=True, color="FFFFFF",
                          size=ref_cell.font.size or 10,
                          name=ref_cell.font.name or "Calibri")
    cell.fill      = PatternFill("solid", fgColor="2E4D8F")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = clone_border(ref_cell.border)

def apply_data_style(cell, ref_cell, num_fmt=None):
    cell.font      = Font(bold=ref_cell.font.bold,
                          size=ref_cell.font.size or 10,
                          name=ref_cell.font.name or "Calibri")
    cell.fill      = PatternFill("solid", fgColor="F7F9FC")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = clone_border(ref_cell.border)
    if num_fmt:
        cell.number_format = num_fmt


# =====================================================================
# PART 1: Restore J column = "Product Name(s)"
# This is the formula that was originally in column H before we started,
# showing which products with remaining balance use each material item.
# =====================================================================

# --- JULY section (1 product: D3/H3/C3) ---
# Header
j6 = ws.cell(row=6, column=10, value="Product Name(s)")
apply_header_style(j6, ws["I6"])

# Data rows 7-88
# Original formula (was in H, user moved to J):
# Checks if item A{r} is in product D3's BOM AND H3 (remaining) > 0
# If yes, shows product name C3
for r in range(7, 89):
    # The inner expression that builds the comma-separated name
    inner = (
        'IF((COUNTIFS(TableBOM[Product ID],$D$3,TableBOM[Item ID],$A{r})>0)'
        '*($H$3>0),$C$3&", ","")'
    ).format(r=r)

    formula = (
        '=IF(LEN({inner})>1,'
        'LEFT({inner},LEN({inner})-2),"")'
    ).format(inner=inner)

    cell = ws.cell(row=r, column=10, value=formula)
    apply_data_style(cell, ws['I{}'.format(r)])

# --- AUGUST PET section (4 products: rows 93-96) ---
# Header
j99 = ws.cell(row=99, column=10, value="Product Name(s)")
apply_header_style(j99, ws["I99"])

# Data rows 100-106
# Concatenates product names from all 4 orders where item is in BOM and remaining > 0
for r in range(100, 107):
    # Build the concatenated string of all matching product names
    parts = []
    for pr in [93, 94, 95, 96]:
        part = (
            'IF((COUNTIFS(TableBOM[Product ID],$D${pr},TableBOM[Item ID],$A{r})>0)'
            '*($H${pr}>0),$C${pr}&", ","")'
        ).format(pr=pr, r=r)
        parts.append(part)

    concat = ' & '.join(parts)

    formula = (
        '=IF(LEN({concat})>1,'
        'LEFT({concat},LEN({concat})-2),"")'
    ).format(concat=concat)

    cell = ws.cell(row=r, column=10, value=formula)
    apply_data_style(cell, ws['I{}'.format(r)])


# =====================================================================
# PART 2: Update H column formula = Pieces Can Produce
# Change from AVERAGEIF (all BOM items) to remaining-order-weighted
# =====================================================================

# --- JULY section (1 product) ---
# Since there's only 1 product (D3), use its exact BOM rate, not an average
# Rate = SUMPRODUCT matching Item ID AND Product ID, gated by remaining > 0
for r in range(7, 89):
    rate = (
        'SUMPRODUCT((TableBOM[Item ID]=A{r})'
        '*(TableBOM[Product ID]=$D$3)'
        '*($H$3>0)'
        '*TableBOM[Per 1000 Units])'
    ).format(r=r)

    formula = (
        '=IFERROR(IF({rate}=0,"-",'
        'ROUND(F{r}/({rate}/1000),0)),"-")'
    ).format(rate=rate, r=r)

    cell = ws.cell(row=r, column=8, value=formula)
    # Keep existing data style
    apply_data_style(cell, ws['I{}'.format(r)], num_fmt="#,##0")

# --- AUGUST PET section (4 products, weighted average) ---
# Weighted avg: sum(remaining_qty * rate) / sum(remaining_qty)
# Only for products with remaining > 0
for r in range(100, 107):
    rem  = 'IFERROR(SUMIF($D$93:$D$96,TableBOM[Product ID],$H$93:$H$96),0)'
    item = '(TableBOM[Item ID]=A{r})'.format(r=r)
    pos  = '({rem}>0)'.format(rem=rem)

    num = 'SUMPRODUCT({item}*{pos}*{rem}*TableBOM[Per 1000 Units])'.format(
              item=item, pos=pos, rem=rem)
    den = 'SUMPRODUCT({item}*{pos}*{rem})'.format(
              item=item, pos=pos, rem=rem)

    formula = (
        '=IFERROR(IF({num}=0,"-",'
        'ROUND(F{r}*1000*{den}/{num},0)),"-")'
    ).format(num=num, den=den, r=r)

    cell = ws.cell(row=r, column=8, value=formula)
    apply_data_style(cell, ws['I{}'.format(r)], num_fmt="#,##0")


# =====================================================================
# Save and verify
# =====================================================================
wb.save('Tubex_Aug26.xlsx')
print('Done! Both columns fixed.')
print()
print('Column J restored: Product Name(s)')
print('  July: J6 header + J7:J88')
print('  Aug PET: J99 header + J100:J106')
print()
print('Column H updated: Pieces Can Produce (remaining-order-weighted)')
print('  July: H7:H88')
print('  Aug PET: H100:H106')
print()

# Verify
wb2 = openpyxl.load_workbook('Tubex_Aug26.xlsx')
ws2 = wb2['MRP']
print('=== Verification ===')
print('J6:', ws2['J6'].value)
print('J7:', ws2['J7'].value[:80])
print('J99:', ws2['J99'].value)
print('J100:', ws2['J100'].value[:80])
print()
print('H6:', ws2['H6'].value)
print('H7:', ws2['H7'].value[:80])
print('H99:', ws2['H99'].value)
print('H100:', ws2['H100'].value[:80])
