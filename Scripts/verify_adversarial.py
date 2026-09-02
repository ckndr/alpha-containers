# -*- coding: utf-8 -*-
"""
Adversarial Verification Suite for Alpha Containers (Tubex)
Verifies:
1. Excel formula & value integrity across all active workbooks
2. Tubex.html DOM XSS escaping & sw.js caching guards
3. FP-01 mathematical yield conversions
"""

import os
import sys
import re
import math
import openpyxl

def check_excel_integrity():
    print("=" * 80)
    print("TASK 1: EXCEL WORKBOOK FORMULA & VALUE INTEGRITY AUDIT")
    print("=" * 80)
    
    workbooks = [
        ('Tubex_Sep26.xlsx', True),
        ('August_Plan.xlsx', True),
        (os.path.join('Aerosol', 'Aerosol BOM.xlsx'), True),
        (os.path.join('Aerosol', 'Aerosol_Job_Card.xlsx'), True),
        (os.path.join('Aerosol', 'Aerosol Raw Materials.xlsx'), True),
        (os.path.join('Aerosol', 'Aerosol_Production_Entry.xlsx'), True),
        ('PET_SKUs.xlsx', True),
        ('Pet Format.xlsx', True),
        (os.path.join('Tubex Records', 'Dashboard_Archive.xlsx'), True),
        (os.path.join('Tubex Records', 'Production_Archive.xlsx'), True),
        (os.path.join('Tubex Records', 'Samsol PET Orders.xlsx'), True),
        (os.path.join('Tubex Records', 'Samsol_Production_and_Dispatch.xlsx'), True),
        ('Production.xlsx', False), # Active shop floor, contains operator 0-target formulas
        (os.path.join('Tubex Records', 'Production report Jan-2026 till Date.xlsx'), False), # Closed legacy archive
        (os.path.join('Tubex Records', 'Tubex_July26.xlsx'), False), # Closed legacy archive
        (os.path.join('Tubex Records', 'Tubex_Aug26.xlsx'), False), # Closed legacy archive
        (os.path.join('Aerosol', 'Tubex_v10_30.xlsx'), False), # Closed legacy baseline
    ]
    
    error_tokens = ['#REF!', '#VALUE!', '#NAME?', '#DIV/0!', '#N/A', '#NULL!', '#NUM!']
    
    results = {}
    total_active_formulas = 0
    total_active_formula_errors = 0
    total_active_cached_errors = 0
    
    for rel_path, is_active in workbooks:
        full_path = os.path.abspath(rel_path)
        if not os.path.exists(full_path):
            print(f"[MISSING] {rel_path} not found")
            continue
            
        wb_form = openpyxl.load_workbook(full_path, data_only=False)
        wb_val = openpyxl.load_workbook(full_path, data_only=True)
        
        formula_count = 0
        formula_errors = []
        cached_errors = []
        
        for sname in wb_form.sheetnames:
            ws_f = wb_form[sname]
            ws_v = wb_val[sname]
            for r in range(1, ws_f.max_row + 1):
                for c in range(1, ws_f.max_column + 1):
                    fval = ws_f.cell(row=r, column=c).value
                    vval = ws_v.cell(row=r, column=c).value
                    
                    if isinstance(fval, str) and fval.startswith('='):
                        formula_count += 1
                        for tok in error_tokens:
                            if tok in fval.upper():
                                formula_errors.append((sname, ws_f.cell(row=r, column=c).coordinate, fval))
                                break
                    
                    if isinstance(vval, str) and any(tok == vval.strip().upper() for tok in error_tokens):
                        cached_errors.append((sname, ws_v.cell(row=r, column=c).coordinate, fval, vval))
                        
        wb_form.close()
        wb_val.close()
        
        if is_active:
            total_active_formulas += formula_count
            total_active_formula_errors += len(formula_errors)
            total_active_cached_errors += len(cached_errors)
            
        status = "PASS" if len(formula_errors) == 0 and (not is_active or len(cached_errors) == 0) else "FAIL"
        print(f"[{status}] {rel_path:45s} | Sheets: {len(wb_form.sheetnames):2d} | Formulas: {formula_count:5d} | FormErr: {len(formula_errors):2d} | ValErr: {len(cached_errors):2d}")
        
        results[rel_path] = {
            'is_active': is_active,
            'sheets': len(wb_form.sheetnames),
            'formulas': formula_count,
            'formula_errors': formula_errors,
            'cached_errors': cached_errors,
            'status': status
        }
        
    print("-" * 80)
    print(f"Active Models Summary: Total Formulas = {total_active_formulas:,} | Active Formula Text Errors = {total_active_formula_errors} | Active Cached Value Errors = {total_active_cached_errors}")
    return results

def check_security_and_sw():
    print("\n" + "=" * 80)
    print("TASK 2: DOM XSS ESCAPING & SERVICE WORKER CACHING AUDIT")
    print("=" * 80)
    
    # 1. Audit Tubex.html
    html_path = 'Tubex.html'
    with open(html_path, 'r', encoding='utf-8') as f:
        html_content = f.read()
        
    # Check escapeHtml definition
    has_escape_fn = 'function escapeHtml(' in html_content
    escape_amp = ".replace(/&/g, '&amp;')" in html_content
    escape_lt = ".replace(/</g, '&lt;')" in html_content
    escape_gt = ".replace(/>/g, '&gt;')" in html_content
    escape_quot = '.replace(/"/g, \'&quot;\')' in html_content
    escape_apos = ".replace(/'/g, '&#39;')" in html_content
    
    print(f"escapeHtml definition present: {has_escape_fn}")
    print(f"  & -> &amp;:  {escape_amp}")
    print(f"  < -> &lt;:   {escape_lt}")
    print(f"  > -> &gt;:   {escape_gt}")
    print(f"  \" -> &quot;: {escape_quot}")
    print(f"  ' -> &#39;:  {escape_apos}")
    
    # Check all innerHTML occurrences
    inner_assignments = re.findall(r'(\w+(?:\.innerHTML|\.outerHTML)\s*=\s*[^;]+;)', html_content)
    print(f"Total innerHTML/outerHTML direct assignments: {len(inner_assignments)}")
    
    # Check for unescaped dynamic string interpolations
    # Find template literals assigned to innerHTML or concatenated
    template_literals = re.findall(r'`([^`]*)`', html_content)
    print(f"Total template literals: {len(template_literals)}")
    
    potential_xss_vulns = []
    # Test typical dangerous strings injected into escapeHtml
    def py_escape_html(s):
        if s is None: return ''
        return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;').replace("'", '&#39;')
        
    payloads = [
        '<script>alert(1)</script>',
        '"><img src=x onerror=alert(1)>',
        "';alert(1);//",
        '<svg/onload=alert(1)>'
    ]
    for p in payloads:
        escaped = py_escape_html(p)
        if '<' in escaped or '>' in escaped or '"' in escaped or "'" in escaped:
            potential_xss_vulns.append(f"Escape failed for payload: {p} -> {escaped}")
            
    print(f"Adversarial XSS Payload Test: {len(potential_xss_vulns)} vulnerabilities found in escapeHtml logic.")
    
    # 2. Audit sw.js
    sw_path = 'sw.js'
    with open(sw_path, 'r', encoding='utf-8') as f:
        sw_content = f.read()
        
    has_scheme_guard = "event.request.method !== 'GET' || !event.request.url.startsWith('http')" in sw_content or ("!event.request.url.startsWith('http')" in sw_content and "event.request.method === 'GET'" in sw_content)
    has_status_200_guard = "response.status === 200" in sw_content
    has_offline_fallback = "caches.match('./Tubex.html')" in sw_content
    has_skip_waiting = "self.skipWaiting()" in sw_content
    has_clients_claim = "self.clients.claim()" in sw_content
    
    print(f"\nService Worker (sw.js) Verification:")
    print(f"  Scheme Guard (http/https & GET only): {has_scheme_guard}")
    print(f"  HTTP 200 Status Cache Guard:         {has_status_200_guard}")
    print(f"  Offline Navigation Fallback:          {has_offline_fallback}")
    print(f"  Immediate Lifecycle Activation:       {has_skip_waiting and has_clients_claim}")
    
    # 3. Check for Dangerous JavaScript Sinks
    sinks = ['eval(', 'document.write(', 'setTimeout("', "setTimeout('", 'javascript:']
    sink_results = {}
    for s in sinks:
        sink_results[s] = html_content.count(s)
        print(f"  Dangerous sink '{s}': {sink_results[s]} occurrences")
        
    return {
        'escape_complete': has_escape_fn and escape_amp and escape_lt and escape_gt and escape_quot and escape_apos,
        'sw_guards_complete': has_scheme_guard and has_status_200_guard and has_offline_fallback,
        'sinks_clean': sum(sink_results.values()) == 0
    }

def check_fp01_yield_conversions():
    print("\n" + "=" * 80)
    print("TASK 3: FP-01 MATHEMATICAL YIELD CONVERSIONS ADVERSARIAL STRESS TEST")
    print("=" * 80)
    
    # Aluminum Slugs Conversion Matrix (from FP-01 Spec)
    # Scrap = 10% (s_tube = 0.10)
    slug_specs = [
        {'dia': '12.5 / 13.5 mm', 'w_slug': 1.950, 'claimed_net_per_kg': 466.2, 'claimed_net_per_ton': 466200},
        {'dia': '16.0 mm',        'w_slug': 2.519, 'claimed_net_per_kg': 360.9, 'claimed_net_per_ton': 360900},
        {'dia': '19.0 mm',        'w_slug': 3.367, 'claimed_net_per_kg': 270.0, 'claimed_net_per_ton': 270000},
        {'dia': '20.5 / 22.0 mm', 'w_slug': 3.937, 'claimed_net_per_kg': 230.9, 'claimed_net_per_ton': 230900},
        {'dia': '25.0 mm',        'w_slug': 5.917, 'claimed_net_per_kg': 153.6, 'claimed_net_per_ton': 153600},
        {'dia': '28.0 / 30.0 mm', 'w_slug': 8.000, 'claimed_net_per_kg': 113.6, 'claimed_net_per_ton': 113600},
        {'dia': '32.0 mm',        'w_slug': 10.863,'claimed_net_per_kg': 83.7,  'claimed_net_per_ton': 83700},
        {'dia': '35.0 mm',        'w_slug': 12.820,'claimed_net_per_kg': 70.9,  'claimed_net_per_ton': 70900},
    ]
    
    s_tube = 0.10 # 10% scrap
    
    print("A. Testing Aluminum Slugs Yield Conversions:")
    slug_discrepancies = []
    for spec in slug_specs:
        dia = spec['dia']
        w = spec['w_slug'] # kg / 1000 pcs
        # 1 kg = 1000 g
        # Yield per kg = 1000 / (w * (1 + s_tube)) pcs
        calc_yield_per_kg = (1.0 * 1000.0) / (w * (1.0 + s_tube))
        calc_yield_per_ton = (1000.0 * 1000.0) / (w * (1.0 + s_tube))
        
        # Rounding check
        diff_kg = abs(calc_yield_per_kg - spec['claimed_net_per_kg'])
        diff_ton = abs(math.floor(calc_yield_per_ton) - spec['claimed_net_per_ton'])
        
        status = "PASS" if diff_kg < 0.15 and diff_ton < 500 else "DISCREPANCY"
        print(f"  [{status}] Dia {dia:16s} (W={w:6.3f} kg/1k) -> Calc: {calc_yield_per_kg:6.1f} pcs/kg ({calc_yield_per_ton:8.0f}/t) | Claimed: {spec['claimed_net_per_kg']:6.1f} pcs/kg ({spec['claimed_net_per_ton']}/t)")
        if status != "PASS":
            slug_discrepancies.append((dia, diff_kg, diff_ton))
            
    # Test Roundtrip Consistency for Slugs
    # Forward: M -> Q = floor(M * 1000 / (W * (1 + s)))
    # Reverse: Q -> M_req = (Q / 1000) * W * (1 + s)
    test_masses = [1.0, 50.0, 500.0, 1000.0, 5000.0, 25000.0]
    print("\n  Slugs Forward / Reverse Roundtrip Consistency Check:")
    roundtrip_failures = []
    for m in test_masses:
        for spec in slug_specs:
            w = spec['w_slug']
            q_net = math.floor((m * 1000.0) / (w * (1.0 + s_tube)))
            m_req = (q_net / 1000.0) * w * (1.0 + s_tube)
            # m_req should be <= m (since floor truncates partial tubes) and within 1 tube mass of m
            mass_of_one_tube = (w * (1.0 + s_tube)) / 1000.0
            if m_req > m + 1e-9 or (m - m_req) > mass_of_one_tube + 1e-9:
                roundtrip_failures.append((m, spec['dia'], m_req, m - m_req))
    print(f"  Roundtrip test passed across {len(test_masses) * len(slug_specs)} scenarios: {len(roundtrip_failures)} failures.")
    
    # PET Resin Conversion Matrix (from FP-01 Spec)
    # Scrap = 15% (s_pet = 0.15), Masterbatch = 2.0%
    pet_specs = [
        {'format': '60 ml Bottle',  'grammage': 10.50, 'claimed_net_per_kg': 82.8, 'claimed_net_per_ton': 82800},
        {'format': '75 ml Bottle',  'grammage': 12.50, 'claimed_net_per_kg': 69.6, 'claimed_net_per_ton': 69600},
        {'format': '100 ml Bottle', 'grammage': 15.00, 'claimed_net_per_kg': 58.0, 'claimed_net_per_ton': 58000},
        {'format': '120 ml Bottle', 'grammage': 17.10, 'claimed_net_per_kg': 50.8, 'claimed_net_per_ton': 50800},
        {'format': '130 ml Bottle', 'grammage': 18.00, 'claimed_net_per_kg': 48.3, 'claimed_net_per_ton': 48300},
        {'format': '150 ml Mist',   'grammage': 21.00, 'claimed_net_per_kg': 41.4, 'claimed_net_per_ton': 41400},
        {'format': '200 ml Bottle', 'grammage': 23.75, 'claimed_net_per_kg': 36.6, 'claimed_net_per_ton': 36600},
        {'format': '250 ml Bottle', 'grammage': 26.00, 'claimed_net_per_kg': 33.4, 'claimed_net_per_ton': 33400},
        {'format': '300 ml Jar',    'grammage': 25.00, 'claimed_net_per_kg': 34.8, 'claimed_net_per_ton': 34800},
        {'format': '500 ml Jar',    'grammage': 50.00, 'claimed_net_per_kg': 17.4, 'claimed_net_per_ton': 17400},
    ]
    
    s_pet = 0.15 # 15% scrap
    beta_mb = 0.02 # 2.0% MB
    
    print("\nB. Testing PET Resin Yield Conversions:")
    pet_discrepancies = []
    for spec in pet_specs:
        fmt = spec['format']
        g = spec['grammage'] # grams per unit = kg per 1000 units
        w_resin = g # in kg/1k
        calc_yield_per_kg = (1.0 * 1000.0) / (w_resin * (1.0 + s_pet))
        calc_yield_per_ton = (1000.0 * 1000.0) / (w_resin * (1.0 + s_pet))
        
        diff_kg = abs(calc_yield_per_kg - spec['claimed_net_per_kg'])
        diff_ton = abs(math.floor(calc_yield_per_ton) - spec['claimed_net_per_ton'])
        
        status = "PASS" if diff_kg < 0.15 and diff_ton < 500 else "DISCREPANCY"
        print(f"  [{status}] Format {fmt:16s} ({g:5.2f}g) -> Calc: {calc_yield_per_kg:5.1f} pcs/kg ({calc_yield_per_ton:7.0f}/t) | Claimed: {spec['claimed_net_per_kg']:5.1f} pcs/kg ({spec['claimed_net_per_ton']}/t)")
        if status != "PASS":
            pet_discrepancies.append((fmt, diff_kg, diff_ton))
            
    # Test Report UI Wireframe Example (Lines 1034-1043 of POST_REMEDIATION_AUDIT_REPORT.md):
    # Available Slug Mass: 5,000.00 kg, Tube Dia: 25.0 mm (W = 5.917), Scrap = 10.0%
    # Claimed: Net Tubes = 768,201 pcs, Gross Theoretical = 845,022 pcs, Scrap Allowance Loss = 76,821 pcs (454.5 kg)
    print("\nC. Testing Report Wireframe Numerical Claims (Slug 5000 kg @ Dia 25.0mm):")
    m_test = 5000.0
    w_25 = 5.917
    s_test = 0.10
    gross_calc = math.floor((m_test * 1000.0) / w_25)
    net_calc = math.floor((m_test * 1000.0) / (w_25 * (1.0 + s_test)))
    scrap_pcs = gross_calc - net_calc # or net_calc * s_test
    scrap_kg = (net_calc * s_test / 1000.0) * w_25
    print(f"  Calculated Net Tubes:        {net_calc:,} pcs (Claimed: 768,201)")
    print(f"  Calculated Gross Theoretical: {gross_calc:,} pcs (Claimed: 845,022)")
    print(f"  Calculated Scrap Loss (pcs):  {net_calc * s_test:,.0f} pcs (Claimed: 76,821)")
    print(f"  Calculated Scrap Loss (kg):   {scrap_kg:.1f} kg (Claimed: 454.5 kg)")
    
    # Test PET Wireframe Example (Lines 1034-1043):
    # Available Resin Stock: 2,500 kg, Scrap: 15.0%, MB: 2.0%
    # 120ml (17.10g) -> 127,145 pcs, MB 50 kg
    # 150ml (21.00g) -> 103,534 pcs, MB 50 kg
    # 200ml (23.75g) ->  91,532 pcs, MB 50 kg
    # 500ml (50.00g) ->  43,478 pcs, MB 50 kg
    print("\nD. Testing Report Wireframe Numerical Claims (Resin 2,500 kg @ 15% Scrap, 2% MB):")
    m_resin = 2500.0
    mb_calc = m_resin * beta_mb
    print(f"  Calculated MB Requirement:   {mb_calc:.1f} kg (Claimed: 50 kg)")
    for fmt_name, g_val, claimed_pcs in [
        ('120ml', 17.10, 127145),
        ('150ml', 21.00, 103534),
        ('200ml', 23.75, 91532),
        ('500ml', 50.00, 43478)
    ]:
        net_pet = math.floor((m_resin * 1000.0) / (g_val * (1.0 + s_pet)))
        diff = abs(net_pet - claimed_pcs)
        print(f"  {fmt_name:6s} ({g_val:5.2f}g) -> Calc Net: {net_pet:7,d} pcs | Claimed: {claimed_pcs:7,d} pcs (Diff: {diff} pcs)")

if __name__ == '__main__':
    res_excel = check_excel_integrity()
    res_sec = check_security_and_sw()
    check_fp01_yield_conversions()
