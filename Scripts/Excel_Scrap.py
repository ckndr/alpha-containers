import openpyxl
import json
import os
import re
from pathlib import Path
from datetime import datetime

# ==============================================================================
# CONFIGURATION — Edit this section to control what gets exported
# ==============================================================================

# Scripts live in AlphaContainers/Scripts/
# Excel files live in AlphaContainers/ (one level up)
SCRIPT_DIR = Path(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR   = SCRIPT_DIR.parent   # AlphaContainers root

# Output JSON saved in root folder (easy to find and upload to Claude)
OUTPUT_JSON = DATA_DIR / "excel_for_ai.json"

# ---------------------------------------------------------------------------
# SHEET EXPORT MODES
# Three modes available per sheet:
#   "full"    — All rows, all columns, formulas + calculated values + flag colors
#   "summary" — Column headers + column stats + last N non-empty rows (saves tokens)
#   "skip"    — Sheet not exported at all (maximum token saving)
#
# To exclude a sheet: set its mode to "skip"
# To add a new sheet: add a line here, or leave it out (defaults to "full")
# ---------------------------------------------------------------------------

SHEET_MODES = {
    "Tubex_Dashboard": "full",
    "MRP":             "skip",
    "BOM":             "skip",
    "Inventory":       "skip",
    "Product_Catalog": "full",
    "Production_Log":  "full",   # Large sheet — summary + last rows only
    "BOM Issues":      "skip",
    "Notes & Guide":   "skip",
}

# For "summary" mode: how many tail non-empty rows to include
SUMMARY_TAIL_ROWS = 20

# Default mode for any sheet NOT listed in SHEET_MODES above
DEFAULT_MODE = "full"

# ---------------------------------------------------------------------------
# FONT COLOR FILTER
# Only these colors are exported — everything else is silently dropped.
# Standard dark text (FF1A1A2E) and white headers (FFFFFFFF) are excluded
# because they are styling noise with no data meaning.
# Add new colors here if you introduce new semantic highlighting in future.
# ---------------------------------------------------------------------------

MEANINGFUL_COLORS = {
    "FFFF0000": "RED_PLACEHOLDER",   # Red = unconfirmed / placeholder
    "FFC62828": "RED_PLACEHOLDER",   # Dark red variant
    "FF00B050": "GREEN_CONFIRMED",   # Green = confirmed / verified
    "FF2E7D32": "GREEN_CONFIRMED",   # Dark green variant
    "FFBDD7EE": "BLUE_HIGHLIGHT",    # Light blue
    "FFB4E0E0": "TEAL_HIGHLIGHT",    # Teal
    "FF595959": "GREY_NOTE",         # Grey = secondary/note
    "FF1F3864": "DARK_HEADER",       # Dark section header
}

# ==============================================================================
# END OF CONFIGURATION
# ==============================================================================


def get_cell_color_flag(cell):
    """Return (hex_color, flag_name) if meaningful, else (None, None)."""
    try:
        color = cell.font.color
        if color and color.type == "rgb":
            rgb = color.rgb.upper()
            if rgb in MEANINGFUL_COLORS:
                return rgb, MEANINGFUL_COLORS[rgb]
    except Exception:
        pass
    return None, None


def export_cell(cell_f, cell_v):
    """
    Build dict for a single cell. Returns None if cell is empty and colorless.
    cell_f = formula workbook cell, cell_v = value workbook cell.
    """
    formula = cell_f.value
    value   = cell_v.value
    color, flag = get_cell_color_flag(cell_f)

    if formula is None and color is None:
        return None

    result = {}

    if isinstance(formula, str) and formula.startswith("="):
        result["formula"] = formula
        if value is not None:
            result["value"] = value
        if isinstance(value, str) and value in ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"):
            result["ERROR"] = True
    elif formula is not None:
        result["value"] = formula

    if flag == "RED_PLACEHOLDER":
        result["RED_PLACEHOLDER"] = True
    elif flag == "GREEN_CONFIRMED":
        result["CONFIRMED"] = True
    elif flag:
        result["color_flag"] = flag

    return result if result else None


def get_active_columns(ws):
    """Return list of column indices with at least one non-empty cell."""
    active = []
    for col in range(1, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            if ws.cell(row, col).value not in (None, ""):
                active.append(col)
                break
    return active


def export_sheet_full(ws_f, ws_v):
    """Export all rows and columns. Returns list of {row, cells} dicts."""
    active_cols = get_active_columns(ws_f)
    rows_out = []
    for row_idx in range(1, ws_f.max_row + 1):
        row_dict = {}
        for col_idx in active_cols:
            exp = export_cell(ws_f.cell(row_idx, col_idx), ws_v.cell(row_idx, col_idx))
            if exp is not None:
                row_dict[ws_f.cell(row_idx, col_idx).coordinate] = exp
        if row_dict:
            rows_out.append({"row": row_idx, "cells": row_dict})
    return rows_out


def export_sheet_summary(ws_f, ws_v, tail_rows):
    """
    Summary mode: title + headers + column stats + last N non-empty rows.
    Scans backwards to find non-empty rows — skips trailing blanks correctly.
    """
    active_cols = get_active_columns(ws_f)
    out = {}

    # Detect header row (first row with >1 non-empty cell)
    header_row_idx = 1
    for r in range(1, min(5, ws_f.max_row + 1)):
        non_empty = sum(1 for c in active_cols if ws_f.cell(r, c).value not in (None, ""))
        if non_empty > 1:
            header_row_idx = r
            break

    # Title row (if header is not row 1)
    if header_row_idx > 1:
        title_val = ws_f.cell(1, active_cols[0]).value
        out["title"] = str(title_val) if title_val else None

    # Header row
    header = {}
    for col_idx in active_cols:
        exp = export_cell(ws_f.cell(header_row_idx, col_idx), ws_v.cell(header_row_idx, col_idx))
        if exp:
            header[ws_f.cell(header_row_idx, col_idx).coordinate] = exp
    out["header_row"] = {"row": header_row_idx, "cells": header}

    # Column stats
    stats = {}
    for col_idx in active_cols:
        col_letter = ws_f.cell(1, col_idx).column_letter
        non_empty = sum(
            1 for r in range(header_row_idx + 1, ws_f.max_row + 1)
            if ws_f.cell(r, col_idx).value not in (None, "")
        )
        stats[col_letter] = {
            "header": str(ws_f.cell(header_row_idx, col_idx).value or ""),
            "non_empty_data_rows": non_empty,
            "total_data_rows": ws_f.max_row - header_row_idx
        }
    out["column_stats"] = stats

    # Last N NON-EMPTY rows (scan backwards, skip blanks)
    collected = []
    for row_idx in range(ws_f.max_row, header_row_idx, -1):
        if any(ws_f.cell(row_idx, c).value not in (None, "") for c in active_cols):
            collected.append(row_idx)
        if len(collected) >= tail_rows:
            break

    tail_out = []
    for row_idx in sorted(collected):
        row_dict = {}
        for col_idx in active_cols:
            exp = export_cell(ws_f.cell(row_idx, col_idx), ws_v.cell(row_idx, col_idx))
            if exp is not None:
                row_dict[ws_f.cell(row_idx, col_idx).coordinate] = exp
        if row_dict:
            tail_out.append({"row": row_idx, "cells": row_dict})

    out[f"last_{tail_rows}_nonempty_rows"] = tail_out
    return out


def get_table_ranges(wb):
    """
    Read Excel Table definitions (TableBOM, TableProduction, etc.)
    These are NOT in wb.defined_names — they live in each worksheet's XML.
    """
    tables = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for tbl in ws.tables.values():
            tables[tbl.name] = {
                "sheet": sheet_name,
                "range": tbl.ref,
                "display_name": tbl.displayName
            }
    return tables


def get_named_ranges(wb):
    """Extract workbook-level named ranges (separate from Excel Tables)."""
    ranges = {}
    for name, defn in wb.defined_names.items():
        try:
            dests = list(defn.destinations)
            ranges[name] = [f"{sheet}!{ref}" for sheet, ref in dests] if dests else str(defn.attr_text)
        except Exception:
            ranges[name] = str(defn)
    return ranges


def extract_version(filename):
    match = re.search(r'v(\d+)[_.](\d+)', filename, re.IGNORECASE)
    return f"v{match.group(1)}.{match.group(2)}" if match else "unknown"


# ==============================================================================
# MAIN
# ==============================================================================

print("=" * 60)
print(f"SCANNING: {DATA_DIR}")
print("=" * 60)

excel_files = list(DATA_DIR.glob("AlphaContainers*.xlsx"))
if not excel_files:
    print(f"ERROR: No AlphaContainers*.xlsx files found in {DATA_DIR}.")
    exit()

latest_file = max(excel_files, key=lambda x: x.stat().st_mtime)
version = extract_version(latest_file.name)

print(f"FILE:    {latest_file.name}  (version: {version})")
print("Loading formula workbook (data_only=False)...")
wb_f = openpyxl.load_workbook(str(latest_file), data_only=False)
print("Loading value workbook  (data_only=True)...")
wb_v = openpyxl.load_workbook(str(latest_file), data_only=True)

output = {
    "meta": {
        "file": latest_file.name,
        "version": version,
        "exported_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "export_script": "Excel_Scrap.py",
        "color_legend": {
            "RED_PLACEHOLDER": "Red font = unconfirmed/estimated. Do not trust value without verification.",
            "CONFIRMED": "Green font = verified/confirmed data.",
            "color_flag": "Other semantic highlight — see MEANINGFUL_COLORS in script config.",
        },
        "cell_format": "{value, formula?, RED_PLACEHOLDER?, CONFIRMED?, color_flag?, ERROR?}"
    },
    "excel_tables": get_table_ranges(wb_f),
    "named_ranges": get_named_ranges(wb_f),
    "sheets": {}
}

# Warn about sheets in workbook not in SHEET_MODES (uses DEFAULT_MODE)
unlisted  = [s for s in wb_f.sheetnames if s not in SHEET_MODES]
missing   = [s for s in SHEET_MODES     if s not in wb_f.sheetnames]
if unlisted:
    print(f"\n⚠  Not in SHEET_MODES → using DEFAULT_MODE='{DEFAULT_MODE}': {unlisted}")
if missing:
    print(f"\n⚠  In SHEET_MODES but not in workbook (skipped): {missing}")
print()

for sheet_name in wb_f.sheetnames:
    mode = SHEET_MODES.get(sheet_name, DEFAULT_MODE)

    if mode == "skip":
        print(f"  [SKIP]    {sheet_name}")
        continue

    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]
    print(f"  [{mode.upper():<7}] {sheet_name}  ({ws_f.max_row}r × {ws_f.max_column}c)")

    base = {"mode": mode, "dimensions": f"{ws_f.max_row} rows × {ws_f.max_column} cols"}

    if mode == "full":
        output["sheets"][sheet_name] = {**base, "rows": export_sheet_full(ws_f, ws_v)}
    elif mode == "summary":
        output["sheets"][sheet_name] = {**base, **export_sheet_summary(ws_f, ws_v, SUMMARY_TAIL_ROWS)}

with open(OUTPUT_JSON, "w", encoding="utf-8") as jf:
    json.dump(output, jf, indent=2, default=str)

size_kb = OUTPUT_JSON.stat().st_size / 1024
print()
print("=" * 60)
print(f"DONE  →  {OUTPUT_JSON.name}   ({size_kb:.1f} KB)")
print(f"         Saved to: {OUTPUT_JSON}")
print("=" * 60)
