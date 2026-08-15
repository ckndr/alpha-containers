"""
Rebuild 'Improved PET Format' as a proper WIDE register-style printed form.
- Landscape A4 / register paper
- 8 columns for a spacious layout
- Labels + wide input boxes side by side
- Big clear fonts for operators
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.load_workbook('Pet formate.xlsx')

if "Improved PET Format" in wb.sheetnames:
    del wb["Improved PET Format"]

ws = wb.create_sheet("Improved PET Format")

# ── STYLES ──
thin = Border(left=Side("thin"), right=Side("thin"),
              top=Side("thin"), bottom=Side("thin"))
no_border = Border()

blue_dark = PatternFill("solid", fgColor="2F5496")
blue_band = PatternFill("solid", fgColor="4472C4")
section_bg = PatternFill("solid", fgColor="D6E4F0")
green_bg = PatternFill("solid", fgColor="E2EFDA")
orange_bg = PatternFill("solid", fgColor="FCE4D6")
red_bg = PatternFill("solid", fgColor="F4CCCC")
yellow_bg = PatternFill("solid", fgColor="FFF2CC")
white_bg = PatternFill("solid", fgColor="FFFFFF")
check_bg = PatternFill("solid", fgColor="FBE5D6")
dt_header_bg = PatternFill("solid", fgColor="E6B8AF")

title_f = Font("Calibri", 20, bold=True, color="FFFFFF")
doc_f = Font("Calibri", 11, color="FFFFFF")
section_f = Font("Calibri", 15, bold=True, color="2F5496")
label_f = Font("Calibri", 14)
label_fb = Font("Calibri", 14, bold=True)
hint_f = Font("Calibri", 10, italic=True, color="999999")
check_f = Font("Calibri", 12, bold=True, color="C00000")

ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
la = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 8 columns: A-H
# Widths designed for landscape register paper
col_widths = {"A": 22, "B": 18, "C": 6, "D": 22, "E": 18, "F": 6, "G": 22, "H": 18}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

LAST_COL = 8  # H

def apply_border_range(row_start, row_end, col_start, col_end):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).border = thin

def full_merge(row, text, font, fill, height=35):
    """Full width merged row across all 8 columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST_COL)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font
    c.fill = fill
    c.alignment = ca
    for col in range(1, LAST_COL + 1):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).border = thin
    ws.row_dimensions[row].height = height

def field_pair(row, label1, label2=None, label3=None, height=32):
    """
    Up to 3 label+value pairs in one row across 8 columns:
      A(label) B(value) | C(spacer) | D(label) E(value) | F(spacer) | G(label) H(value)
    Actually let's use: A=label, B=value, C=spacer, D=label, E=value, F=spacer, G=label, H=value
    But spacer cols are narrow dividers. Let's just use pairs:
      Col A-B = pair 1,  Col D-E = pair 2,  Col G-H = pair 3
      Col C and F are thin spacers
    """
    # Pair 1
    c1 = ws.cell(row=row, column=1, value=label1)
    c1.font = label_fb
    c1.alignment = la
    c1.border = thin
    ws.cell(row=row, column=2).border = thin
    ws.cell(row=row, column=2).alignment = ca
    ws.cell(row=row, column=2).font = label_f

    # Spacer C
    ws.cell(row=row, column=3).border = thin
    ws.cell(row=row, column=3).fill = section_bg

    # Pair 2
    if label2:
        c2 = ws.cell(row=row, column=4, value=label2)
        c2.font = label_fb
        c2.alignment = la
        c2.border = thin
        ws.cell(row=row, column=5).border = thin
        ws.cell(row=row, column=5).alignment = ca
        ws.cell(row=row, column=5).font = label_f
    else:
        ws.cell(row=row, column=4).border = thin
        ws.cell(row=row, column=5).border = thin

    # Spacer F
    ws.cell(row=row, column=6).border = thin
    ws.cell(row=row, column=6).fill = section_bg

    # Pair 3
    if label3:
        c3 = ws.cell(row=row, column=7, value=label3)
        c3.font = label_fb
        c3.alignment = la
        c3.border = thin
        ws.cell(row=row, column=8).border = thin
        ws.cell(row=row, column=8).alignment = ca
        ws.cell(row=row, column=8).font = label_f
    else:
        ws.cell(row=row, column=7).border = thin
        ws.cell(row=row, column=8).border = thin

    ws.row_dimensions[row].height = height

def label_value_row(row, label, height=32, bold=False, value_fill=None, hint=None):
    """Single label spanning A-D, value spanning E-H — full width."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value=label)
    c.font = label_fb if bold else label_f
    c.alignment = la
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=LAST_COL)
    v = ws.cell(row=row, column=5)
    v.alignment = ca
    v.font = label_f
    if value_fill:
        v.fill = value_fill
    if hint:
        v.value = hint
        v.font = hint_f
    for col in range(1, LAST_COL + 1):
        ws.cell(row=row, column=col).border = thin
    ws.row_dimensions[row].height = height

r = 0

def R():
    global r
    r += 1
    return r

# ═══════════════════════════════════════════════════════════
# ROW 1: TITLE
# ═══════════════════════════════════════════════════════════
full_merge(R(), "PET  SHIFT  SUMMARY", title_f, blue_dark, 42)

# ROW 2: Doc No
full_merge(R(), "Doc No: PET-SF-001   |   Rev. 01   |   Alpha Containers", doc_f, blue_band, 24)

# ═══════════════════════════════════════════════════════════
# SHIFT INFORMATION
# ═══════════════════════════════════════════════════════════
full_merge(R(), "SHIFT INFORMATION", section_f, section_bg, 30)

field_pair(R(), "Date:", "Shift (Day/Night):", "Machine:")
field_pair(R(), "Shift Incharge:", "Shift Start Time:", "Shift End Time:")

# Workers row — full width
row = R()
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST_COL)
ws.cell(row=row, column=1,
        value="Workers:   1. ________________    2. ________________    3. ________________    4. ________________")
ws.cell(row=row, column=1).font = label_f
ws.cell(row=row, column=1).alignment = la
for c in range(1, LAST_COL + 1):
    ws.cell(row=row, column=c).border = thin
ws.row_dimensions[row].height = 35

# ═══════════════════════════════════════════════════════════
# PRODUCT DETAILS
# ═══════════════════════════════════════════════════════════
full_merge(R(), "PRODUCT DETAILS", section_f, section_bg, 30)

field_pair(R(), "Customer Name:", "Job Card No.:", "Product Name:")
field_pair(R(), "Volume (ml):", "Color:", "Preform Weight (g):")

# Bottle weight alone
row = R()
ws.cell(row=row, column=1, value="Bottle Weight (g):").font = label_fb
ws.cell(row=row, column=1).alignment = la
ws.cell(row=row, column=1).border = thin
ws.cell(row=row, column=2).border = thin
ws.cell(row=row, column=2).alignment = ca
# Fill remaining cols
for c in range(3, LAST_COL + 1):
    ws.cell(row=row, column=c).border = thin
ws.row_dimensions[row].height = 32

# ═══════════════════════════════════════════════════════════
# PRODUCTION  +  WASTE  side by side (the core data)
# ═══════════════════════════════════════════════════════════
row = R()
# Left half header: PRODUCTION (cols A-D)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1, value="PRODUCTION").font = section_f
ws.cell(row=row, column=1).fill = green_bg
ws.cell(row=row, column=1).alignment = ca
for c in range(1, 5):
    ws.cell(row=row, column=c).fill = green_bg
    ws.cell(row=row, column=c).border = thin
# Right half header: WASTE (cols E-H)
ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=LAST_COL)
ws.cell(row=row, column=5, value="WASTE").font = section_f
ws.cell(row=row, column=5).fill = orange_bg
ws.cell(row=row, column=5).alignment = ca
for c in range(5, LAST_COL + 1):
    ws.cell(row=row, column=c).fill = orange_bg
    ws.cell(row=row, column=c).border = thin
ws.row_dimensions[row].height = 30

# Sub-headers
row = R()
for c, txt in [(1, "Item"), (2, "Qty"), (5, "Item"), (6, "Qty")]:
    # merge label cols and value cols
    pass
# Actually let's do: A-C = label, D = value  |  E-G = label, H = value
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws.cell(row=row, column=1).font = label_fb
ws.cell(row=row, column=1).alignment = ca
ws.cell(row=row, column=4).font = label_fb
ws.cell(row=row, column=4).alignment = ca
ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
ws.cell(row=row, column=5).font = label_fb
ws.cell(row=row, column=5).alignment = ca
ws.cell(row=row, column=8).font = label_fb
ws.cell(row=row, column=8).alignment = ca
# Don't need sub-headers, go straight to data rows

# Combined production + waste rows
prod_waste = [
    ("Total Production (pcs)", True,    "Waste Bottles (pcs)", False),
    ("Good Production (pcs)", True,     "Preform Waste (pcs)", False),
    ("Material Consumed (kg)", True,    "Purging Waste (pcs)", False),
    ("", False,                         "Total Waste (pcs)", True),
]

for lbl_l, bold_l, lbl_r, bold_r in prod_waste:
    # Use the sub-header row for first item since we didn't write sub-headers
    # Actually let's overwrite row we created above
    pass

# Let me redo - remove the empty sub-header row
# Actually it's already at row r, let me just overwrite
# Back up: let's just delete what we did for sub-headers and do data rows

# Overwrite row r (the sub-header row) with first data pair
current = r
for i, (lbl_l, bold_l, lbl_r, bold_r) in enumerate(prod_waste):
    row_num = current + i
    if i > 0:
        R()  # increment r
        row_num = r

    # Left side: production
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
    ws.cell(row=row_num, column=1, value=lbl_l).font = label_fb if bold_l else label_f
    ws.cell(row=row_num, column=1).alignment = la
    ws.cell(row=row_num, column=4).alignment = ca
    ws.cell(row=row_num, column=4).font = label_f
    for c in range(1, 5):
        ws.cell(row=row_num, column=c).border = thin

    # Right side: waste
    ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=7)
    ws.cell(row=row_num, column=5, value=lbl_r).font = label_fb if bold_r else label_f
    ws.cell(row=row_num, column=5).alignment = la
    ws.cell(row=row_num, column=8).alignment = ca
    ws.cell(row=row_num, column=8).font = label_f
    if bold_r:
        ws.cell(row=row_num, column=8).fill = yellow_bg
    for c in range(5, LAST_COL + 1):
        ws.cell(row=row_num, column=c).border = thin

    ws.row_dimensions[row_num].height = 34

# Total Waste hint
ws.cell(row=r, column=8).value = "= add above"
ws.cell(row=r, column=8).font = hint_f

# CHECK row
row = R()
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST_COL)
ws.cell(row=row, column=1,
        value="CHECK:   Total Production   =   Good Production   +   Total Waste")
ws.cell(row=row, column=1).font = check_f
ws.cell(row=row, column=1).alignment = ca
ws.cell(row=row, column=1).fill = check_bg
for c in range(1, LAST_COL + 1):
    ws.cell(row=row, column=c).fill = check_bg
    ws.cell(row=row, column=c).border = thin
ws.row_dimensions[row].height = 26

# ═══════════════════════════════════════════════════════════
# REWORK & OT  +  STOCK  side by side
# ═══════════════════════════════════════════════════════════
row = R()
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1, value="REWORK & OVERTIME").font = section_f
ws.cell(row=row, column=1).fill = section_bg
ws.cell(row=row, column=1).alignment = ca
for c in range(1, 5):
    ws.cell(row=row, column=c).fill = section_bg
    ws.cell(row=row, column=c).border = thin

ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=LAST_COL)
ws.cell(row=row, column=5, value="STOCK").font = section_f
ws.cell(row=row, column=5).fill = green_bg
ws.cell(row=row, column=5).alignment = ca
for c in range(5, LAST_COL + 1):
    ws.cell(row=row, column=c).fill = green_bg
    ws.cell(row=row, column=c).border = thin
ws.row_dimensions[row].height = 28

ot_stock = [
    ("Rework (pcs)", "Resin in WIP (kg)"),
    ("OT Production (pcs)", "Raw Material Stock (kg)"),
    ("OT Waste (pcs)", ""),
]
for lbl_l, lbl_r in ot_stock:
    row = R()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row=row, column=1, value=lbl_l).font = label_f
    ws.cell(row=row, column=1).alignment = la
    ws.cell(row=row, column=4).alignment = ca
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = thin

    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
    ws.cell(row=row, column=5, value=lbl_r).font = label_fb if lbl_r else label_f
    ws.cell(row=row, column=5).alignment = la
    ws.cell(row=row, column=8).alignment = ca
    for c in range(5, LAST_COL + 1):
        ws.cell(row=row, column=c).border = thin
    ws.row_dimensions[row].height = 32

# ═══════════════════════════════════════════════════════════
# DOWNTIME — full width table
# ═══════════════════════════════════════════════════════════
full_merge(R(), "DOWNTIME", section_f, red_bg, 28)

# Downtime laid out in 2 side-by-side columns (5 reasons each side)
# Left: cols A-B (reason) C-D (minutes)  |  Right: cols E-F (reason) G-H (minutes)
row = R()
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws.cell(row=row, column=1, value="Reason").font = label_fb
ws.cell(row=row, column=1).alignment = ca
ws.cell(row=row, column=1).fill = dt_header_bg
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
ws.cell(row=row, column=3, value="Minutes").font = label_fb
ws.cell(row=row, column=3).alignment = ca
ws.cell(row=row, column=3).fill = dt_header_bg

ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
ws.cell(row=row, column=5, value="Reason").font = label_fb
ws.cell(row=row, column=5).alignment = ca
ws.cell(row=row, column=5).fill = dt_header_bg
ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=LAST_COL)
ws.cell(row=row, column=7, value="Minutes").font = label_fb
ws.cell(row=row, column=7).alignment = ca
ws.cell(row=row, column=7).fill = dt_header_bg

for c in range(1, LAST_COL + 1):
    ws.cell(row=row, column=c).fill = dt_header_bg
    ws.cell(row=row, column=c).border = thin
ws.row_dimensions[row].height = 26

# 10 downtime reasons in 2 columns of 5
dt_left = [
    "Mechanical Problem",
    "Electrical Problem",
    "Operations Problem",
    "Power Shutdown",
    "Workers Not Available",
]
dt_right = [
    "Material Not Available",
    "Compressor Problem",
    "Mould Change",
    "No Order",
    "Other:",
]

for i in range(5):
    row = R()
    # Left
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value=dt_left[i]).font = label_f
    ws.cell(row=row, column=1).alignment = la
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3).alignment = ca
    # Right
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    ws.cell(row=row, column=5, value=dt_right[i]).font = label_f
    ws.cell(row=row, column=5).alignment = la
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=LAST_COL)
    ws.cell(row=row, column=7).alignment = ca

    for c in range(1, LAST_COL + 1):
        ws.cell(row=row, column=c).border = thin
    ws.row_dimensions[row].height = 30

# Total downtime row
row = R()
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws.cell(row=row, column=1, value="TOTAL DOWNTIME").font = label_fb
ws.cell(row=row, column=1).alignment = la
ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=LAST_COL)
ws.cell(row=row, column=3).alignment = ca
ws.cell(row=row, column=3).fill = yellow_bg
for c in range(1, LAST_COL + 1):
    ws.cell(row=row, column=c).border = thin
ws.row_dimensions[row].height = 30

# ═══════════════════════════════════════════════════════════
# SIGNATURES + NOTES — side by side at bottom
# ═══════════════════════════════════════════════════════════
full_merge(R(), "SIGNATURES", section_f, section_bg, 26)

row = R()
# 3 signature blocks side by side
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
ws.cell(row=row, column=1, value="Shift Incharge:").font = label_fb
ws.cell(row=row, column=1).alignment = la
ws.cell(row=row, column=3).border = thin

ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
ws.cell(row=row, column=4, value="Production Manager:").font = label_fb
ws.cell(row=row, column=4).alignment = la

ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=LAST_COL)
ws.cell(row=row, column=7, value="Notes:").font = label_fb
ws.cell(row=row, column=7).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

for c in range(1, LAST_COL + 1):
    ws.cell(row=row, column=c).border = thin
ws.row_dimensions[row].height = 45

# ── Print settings: Landscape, fit to 1 page ──
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.page_setup.orientation = "landscape"
ws.page_setup.paperSize = 9  # A4
ws.print_area = "A1:H{}".format(r)

# Set margins tight for maximum print area
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.3
ws.page_margins.bottom = 0.3
ws.page_margins.header = 0.2
ws.page_margins.footer = 0.2

wb.save('Pet formate.xlsx')
print("Done! Wide landscape register-style form. Rows:", r)
