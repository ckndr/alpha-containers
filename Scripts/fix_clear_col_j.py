import openpyxl
wb = openpyxl.load_workbook('Tubex_Aug26.xlsx')
ws = wb['MRP']

# Clear column J completely - user had swapped H/J manually and
# my script incorrectly wrote new formulas into J
# J should be left alone (was empty / user's own content)

cleared = 0
for row_num in range(1, ws.max_row + 1):
    cell = ws.cell(row=row_num, column=10)
    if cell.value is not None:
        cell.value = None
        cleared += 1

wb.save('Tubex_Aug26.xlsx')
print('Cleared {} cells in column J'.format(cleared))
print('Column J is now empty.')
