"""
add_order.py
Tubex — Order Manager, MRP & Dashboard Auto-Updater
──────────────────────────────────────────────────────────
Features:
  1. Add / Increase Orders:
     - If product already has an active order in MRP, adds quantity to the existing row
       (e.g. 200k + 100k -> 300k) and appends the JOF # without creating duplicate rows.
     - If product is new to MRP, inserts a new order row in the correct section (Tube or PET),
       preserving 100% exact formatting, borders, fonts, and row heights.
  2. Remove / Delete Orders:
     - Removes an order from MRP by Product ID or Product Name, adjusts boundaries,
       and recalculates all material and dashboard totals.
  3. Bill of Materials (BOM) Verification:
     - Checks TableBOM for recipes and alerts if components are missing.
  4. Dynamic Formula Recalculations:
     - Automatically updates all MRP Material Requirement formulas across the active ranges.
  5. Dashboard Sorting:
     - Dynamically synchronizes Tubex_Dashboard and positions active products at the top.
  6. Live Excel Preview & Terminal Inspection:
     - Allows opening a full live preview copy in Microsoft Excel to inspect before deciding.
     - Allows viewing a formatted terminal summary table.
  7. Automated Backup & Keep / Revert Option:
     - Creates a safety backup before editing.
     - Prompts you to KEEP or REVERT changes after your review!

Usage:
  Interactive Menu Wizard:
    python Scripts/add_order.py
    (or double-click Scripts/Add_Order.bat)

  CLI Mode — Add / Accumulate Order:
    python Scripts/add_order.py -p 6206 -q 100000 -j 6905
    python Scripts/add_order.py -p "VINCE NURTURAL" -q 25000 -j 6901

  CLI Mode — Remove Order:
    python Scripts/add_order.py --remove 3447
    python Scripts/add_order.py --delete "S 43 25MM"

  CLI Mode — Auto-Keep:
    python Scripts/add_order.py -p 6206 -q 50000 -y

Author: Sikander / Antigravity
Version: 2.1
"""

import os
import sys
import glob
import re
import shutil
import difflib
import argparse
from datetime import datetime
from copy import copy

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ── PATH SETUP ──────────────────────────────────────────────
DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LOGS_DIR = os.path.join(DIR, 'Logs')
os.makedirs(LOGS_DIR, exist_ok=True)

# Find active Tubex workbook
excel_pattern = os.path.join(DIR, 'Tubex*.xlsx')
excel_files   = sorted(glob.glob(excel_pattern))
if not excel_files:
    print(f"[ERROR] No Tubex*.xlsx workbook found in {DIR}")
    sys.exit(1)
EXCEL_PATH = excel_files[-1]

# ── SAFETY LOCK CHECK ───────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from alpha_checks import check_not_locked
    check_not_locked(EXCEL_PATH)
except ImportError:
    pass


def backup_workbook(filepath):
    """Creates a timestamped backup in Logs directory."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = os.path.basename(filepath).replace('.xlsx', '')
    backup_name = f"backup_{ts}_before_add_order_{base}.xlsx"
    backup_path = os.path.join(LOGS_DIR, backup_name)
    shutil.copy2(filepath, backup_path)
    print(f"[BACKUP] Created: {backup_name}")
    return backup_path


PET_SHORT_NAMES = {
    8001: "Alpha 150ml TRP",
    8005: "Samsol Yellow 120ml",
    8006: "Samsol Yellow 200ml",
    8007: "Samsol White 200ml",
    8008: "Samsol Black 200ml",
    8009: "Mablay 300ml TRP",
    8010: "Mablay 130ml TRP",
    8011: "Samsol Black 120ml",
    8012: "Samsol White 120ml",
    8013: "Mablay 500ml Jar TRP",
    8014: "Samsol Mustard Oil 200ml TRP",
    8015: "Mablay 130ml White",
    8016: "Samsol Coconut Oil 200ml White",
    8017: "Horizon 150ml TRP",
}


def load_product_catalog(wb):
    """Reads Product_Catalog to map PIDs, Names, Customers, Dia, Category."""
    catalog = {}
    ws_cat = wb['Product_Catalog']
    for r in range(3, ws_cat.max_row + 1):
        pid = ws_cat.cell(r, 1).value
        if pid is None:
            continue
        try:
            pid_int = int(pid)
        except (ValueError, TypeError):
            continue

        bom_id   = ws_cat.cell(r, 2).value
        customer = str(ws_cat.cell(r, 3).value or '').strip()
        pname    = str(ws_cat.cell(r, 4).value or '').strip()
        dia      = ws_cat.cell(r, 5).value

        p_type = 'PET' if (8000 <= pid_int <= 8999 or 'PET' in pname.upper() or 'BOTTLE' in pname.upper() or (isinstance(dia, str) and 'ML' in dia.upper())) else 'TUBE'

        catalog[pid_int] = {
            'pid': pid_int,
            'bom_id': bom_id,
            'customer': customer,
            'product_name': pname,
            'dia': dia,
            'type': p_type
        }
    return catalog


def resolve_product(catalog, search_term):
    """Resolves search term (PID or string) to a product entry."""
    search_str = str(search_term).strip()

    # 1. Exact numeric PID match
    try:
        pid_int = int(search_str)
        if pid_int in catalog:
            return catalog[pid_int]
    except ValueError:
        pass

    # 2. Exact name match (case-insensitive)
    for prod in catalog.values():
        if prod['product_name'].upper() == search_str.upper() or prod.get('raw_product_name', '').upper() == search_str.upper():
            return prod

    # 3. Substring match
    matches = [
        prod for prod in catalog.values()
        if search_str.upper() in prod['product_name'].upper()
        or search_str.upper() in prod.get('raw_product_name', '').upper()
        or search_str.upper() in prod['customer'].upper()
    ]
    if len(matches) == 1:
        return matches[0]
    elif len(matches) > 1:
        print(f"\n[SEARCH] Multiple matching products found for '{search_str}':")
        for idx, m in enumerate(matches, 1):
            print(f"  [{idx}] PID: {m['pid']} | {m['product_name']} ({m['customer']}) - {m['type']} {m['dia']}")
        try:
            choice = input(f"Select product [1-{len(matches)}] (or Enter to cancel): ").strip()
            if choice.isdigit() and 1 <= int(choice) <= len(matches):
                return matches[int(choice) - 1]
        except (EOFError, KeyboardInterrupt):
            pass
        return None

    # 4. Fuzzy match
    names = [p['product_name'] for p in catalog.values()] + [p.get('raw_product_name', '') for p in catalog.values() if p.get('raw_product_name')]
    close = difflib.get_close_matches(search_str, names, n=3, cutoff=0.4)
    if close:
        matched_prods = [p for p in catalog.values() if p['product_name'] in close or p.get('raw_product_name') in close]
        # Remove duplicates
        seen_pids = set()
        unique_matched = []
        for p in matched_prods:
            if p['pid'] not in seen_pids:
                seen_pids.add(p['pid'])
                unique_matched.append(p)
        print(f"\n[SEARCH] Close matches found for '{search_str}':")
        for idx, m in enumerate(unique_matched, 1):
            print(f"  [{idx}] PID: {m['pid']} | {m['product_name']} ({m['customer']}) - {m['type']} {m['dia']}")
        try:
            choice = input(f"Select product [1-{len(unique_matched)}] (or Enter to cancel): ").strip()
            if choice.isdigit() and 1 <= int(choice) <= len(unique_matched):
                return unique_matched[int(choice) - 1]
        except (EOFError, KeyboardInterrupt):
            pass

    return None


def check_bom(wb, pid):
    """Checks if TableBOM has recipe items for this PID."""
    ws_bom = wb['BOM']
    bom_items = []
    for r in range(3, ws_bom.max_row + 1):
        row_pid = ws_bom.cell(r, 1).value
        try:
            if row_pid is not None and int(row_pid) == int(pid):
                bom_items.append({
                    'category': ws_bom.cell(r, 6).value,
                    'item_id': ws_bom.cell(r, 7).value,
                    'item_name': ws_bom.cell(r, 8).value,
                    'uom': ws_bom.cell(r, 9).value,
                    'rate': ws_bom.cell(r, 10).value,
                    'scrap': ws_bom.cell(r, 12).value
                })
        except (ValueError, TypeError):
            pass
    return bom_items


def get_mrp_boundaries(ws_mrp):
    """Scans and returns dynamic row boundaries for Tubes and PET sections in MRP."""
    tube_start = 3
    tube_total_row = None
    for r in range(3, ws_mrp.max_row + 1):
        if str(ws_mrp.cell(r, 5).value or '').strip().upper() == 'TOTAL:':
            tube_total_row = r
            break
    if not tube_total_row:
        raise ValueError("Could not locate Tube TOTAL row in MRP sheet.")

    tube_end = tube_total_row - 1
    tube_items_start = tube_total_row + 3
    tube_items_end = None
    for r in range(tube_items_start, ws_mrp.max_row + 1):
        if ws_mrp.cell(r, 1).value is None:
            tube_items_end = r - 1
            break

    pet_header_row = None
    for r in range(tube_items_end + 1, ws_mrp.max_row + 1):
        if 'PET REQUIRED' in str(ws_mrp.cell(r, 1).value or '').upper():
            pet_header_row = r
            break

    pet_start = pet_header_row + 2
    pet_total_row = None
    for r in range(pet_start, ws_mrp.max_row + 1):
        if str(ws_mrp.cell(r, 5).value or '').strip().upper() == 'TOTAL:':
            pet_total_row = r
            break

    pet_end = pet_total_row - 1
    pet_items_start = pet_total_row + 3
    pet_items_end = None
    for r in range(pet_items_start, ws_mrp.max_row + 2):
        if ws_mrp.cell(r, 1).value is None:
            pet_items_end = r - 1
            break

    return {
        'tube_start': tube_start,
        'tube_end': tube_end,
        'tube_total_row': tube_total_row,
        'tube_items_start': tube_items_start,
        'tube_items_end': tube_items_end,
        'pet_start': pet_start,
        'pet_end': pet_end,
        'pet_total_row': pet_total_row,
        'pet_items_start': pet_items_start,
        'pet_items_end': pet_items_end
    }


def update_tube_material_formulas(ws_mrp, tube_end, tube_items_start, tube_items_end):
    """Updates all formula columns in the Tube Material Plan."""
    for r in range(tube_items_start, tube_items_end + 1):
        f_req = f'=SUMPRODUCT((TableBOM[Item ID]=A{r})*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %])*SUMIF($D$3:$D${tube_end},TableBOM[Product ID],$H$3:$H${tube_end})/1000)'
        ws_mrp.cell(r, 5, f_req)

        f_stk = f'=IFERROR(INDEX(TableInventory[Store Balance],MATCH(A{r},TableInventory[Item ID],0)),0)+IFERROR(INDEX(TableInventory[WIP],MATCH(A{r},TableInventory[Item ID],0)),0)'
        ws_mrp.cell(r, 6, f_stk)

        f_surp = f'=F{r}-E{r}'
        ws_mrp.cell(r, 7, f_surp)

        pcp_conds = '+'.join([f'(TableBOM[Product ID]=$D${pr})*($H${pr}>0)' for pr in range(3, tube_end + 1)])
        f_pcp = f'=IFERROR(IF(SUMPRODUCT((TableBOM[Item ID]=A{r})*({pcp_conds}))=0,"-",ROUND(F{r}/((SUMPRODUCT((TableBOM[Item ID]=A{r})*TableBOM[Per 1000 Units]*({pcp_conds}))/SUMPRODUCT((TableBOM[Item ID]=A{r})*({pcp_conds})))/1000),0)),"-")'
        ws_mrp.cell(r, 8, f_pcp)

        f_stat = f'=IF(E{r}=0,"Not needed",IF(G{r}<0,"SHORTAGE",IF(G{r}<F{r}*0.1,"LOW","OK")))'
        ws_mrp.cell(r, 9, f_stat)

        pname_conds = ' & '.join([f'IF((COUNTIFS(TableBOM[Product ID],$D${pr},TableBOM[Item ID],$A{r})>0)*($H${pr}>0),$C${pr}&", ","")' for pr in range(3, tube_end + 1)])
        f_pnames = f'=IF(LEN({pname_conds})>1,LEFT({pname_conds},LEN({pname_conds})-2),"")'
        ws_mrp.cell(r, 10, f_pnames)


def update_pet_material_formulas(ws_mrp, pet_start, pet_end, pet_items_start, pet_items_end):
    """Updates all formula columns in the PET Material Plan."""
    for r in range(pet_items_start, pet_items_end + 1):
        f_req = f'=SUMPRODUCT((TableBOM[Item ID]=A{r})*TableBOM[Per 1000 Units]*(1+TableBOM[Scrap %])*SUMIF($D${pet_start}:$D${pet_end},TableBOM[Product ID],$H${pet_start}:$H${pet_end})/1000)'
        ws_mrp.cell(r, 5, f_req)

        f_stk = f'=IFERROR(INDEX(TableInventory[Store Balance],MATCH(A{r},TableInventory[Item ID],0)),0)+IFERROR(INDEX(TableInventory[WIP],MATCH(A{r},TableInventory[Item ID],0)),0)'
        ws_mrp.cell(r, 6, f_stk)

        f_surp = f'=F{r}-E{r}'
        ws_mrp.cell(r, 7, f_surp)

        f_pcp = f'=IFERROR(IF(SUMPRODUCT((TableBOM[Item ID]=A{r})*((SUMIF($D${pet_start}:$D${pet_end},TableBOM[Product ID],$H${pet_start}:$H${pet_end}))>0)*(SUMIF($D${pet_start}:$D${pet_end},TableBOM[Product ID],$H${pet_start}:$H${pet_end}))*TableBOM[Per 1000 Units])=0,"-",ROUND(F{r}*1000*SUMPRODUCT((TableBOM[Item ID]=A{r})*((SUMIF($D${pet_start}:$D${pet_end},TableBOM[Product ID],$H${pet_start}:$H${pet_end}))>0)*(SUMIF($D${pet_start}:$D${pet_end},TableBOM[Product ID],$H${pet_start}:$H${pet_end})))/SUMPRODUCT((TableBOM[Item ID]=A{r})*((SUMIF($D${pet_start}:$D${pet_end},TableBOM[Product ID],$H${pet_start}:$H${pet_end}))>0)*(SUMIF($D${pet_start}:$D${pet_end},TableBOM[Product ID],$H${pet_start}:$H${pet_end}))*TableBOM[Per 1000 Units]),0)),"-")'
        ws_mrp.cell(r, 8, f_pcp)

        f_stat = f'=IF(E{r}=0,"Not needed",IF(G{r}<0,"SHORTAGE",IF(G{r}<F{r}*0.1,"LOW","OK")))'
        ws_mrp.cell(r, 9, f_stat)

        pet_order_names = []
        for pr in range(pet_start, pet_end + 1):
            pid_val = ws_mrp.cell(pr, 4).value
            try:
                pid_int = int(pid_val)
                sname = PET_SHORT_NAMES.get(pid_int, str(ws_mrp.cell(pr, 3).value or ''))
            except (ValueError, TypeError):
                sname = str(ws_mrp.cell(pr, 3).value or '')
            pet_order_names.append((pr, sname))

        pname_conds = ' & '.join([f'IF((COUNTIFS(TableBOM[Product ID],$D${pr},TableBOM[Item ID],$A{r})>0)*($H${pr}>0),"{sname}, ","")' for pr, sname in pet_order_names])
        f_pnames = f'=IF(LEN({pname_conds})>1,LEFT({pname_conds},LEN({pname_conds})-2),"")'
        ws_mrp.cell(r, 10, f_pnames)


def add_or_update_order(filepath, product, order_qty, jof_num=None, customer_override=None, remarks=None, auto_keep=False):
    """Adds a new order row or accumulates quantity into an existing row."""
    backup_file = backup_workbook(filepath)
    wb = load_workbook(filepath, data_only=False)
    ws_mrp = wb['MRP']

    p_type   = product['type']
    pid      = product['pid']
    pname    = product['product_name']
    customer = customer_override or product['customer']
    dia      = product['dia']

    # Check BOM
    bom_items = check_bom(wb, pid)
    if not bom_items:
        print(f"   [WARN] No BOM recipe found for Product ID {pid} in TableBOM.")
        print(f"          Material requirements will show 0 until BOM is entered.")
    else:
        print(f"   [OK] BOM Verified: {len(bom_items)} recipe components found in TableBOM.")

    b = get_mrp_boundaries(ws_mrp)

    # Check if PID already exists in MRP
    existing_row = None
    if p_type == 'TUBE':
        for r in range(b['tube_start'], b['tube_end'] + 1):
            if ws_mrp.cell(r, 4).value is not None and int(ws_mrp.cell(r, 4).value) == pid:
                existing_row = r
                break
    else:
        for r in range(b['pet_start'], b['pet_end'] + 1):
            if ws_mrp.cell(r, 4).value is not None and int(ws_mrp.cell(r, 4).value) == pid:
                existing_row = r
                break

    # ── CASE A: PRODUCT ALREADY HAS AN ORDER IN MRP (ACCUMULATE QTY) ──
    if existing_row:
        old_qty_val = ws_mrp.cell(existing_row, 6).value
        old_jof_val = str(ws_mrp.cell(existing_row, 5).value or '').strip()
        old_rem_val = str(ws_mrp.cell(existing_row, 9).value or '').strip()

        # Update Quantity (Column F)
        if isinstance(old_qty_val, str) and old_qty_val.startswith('='):
            new_qty_formula = f"{old_qty_val}+{order_qty}"
        else:
            new_qty_formula = f"={old_qty_val}+{order_qty}"
        ws_mrp.cell(existing_row, 6, new_qty_formula)

        # Update JOF # (Column E)
        if jof_num and str(jof_num) not in old_jof_val:
            new_jof_val = f"{old_jof_val} & {jof_num}" if old_jof_val else str(jof_num)
            ws_mrp.cell(existing_row, 5, new_jof_val)
        else:
            new_jof_val = old_jof_val

        # Update Remarks (Column I)
        if remarks and remarks not in old_rem_val:
            new_rem_val = f"{old_rem_val}; {remarks}" if old_rem_val else remarks
            ws_mrp.cell(existing_row, 9, new_rem_val)

        print(f"\n[MERGED] Product ID {pid} already exists at MRP Row {existing_row}.")
        print(f"         Updated Required Qty: {old_qty_val} + {order_qty:,} -> {new_qty_formula}")
        print(f"         Updated JOF #: {new_jof_val}")

    # ── CASE B: NEW PRODUCT (INSERT NEW ORDER ROW) ────────────────────
    else:
        if p_type == 'TUBE':
            insert_row = b['tube_total_row']
            ws_mrp.insert_rows(insert_row)

            ws_mrp.cell(insert_row, 1, dia)
            ws_mrp.cell(insert_row, 2, customer)
            ws_mrp.cell(insert_row, 3, pname)
            ws_mrp.cell(insert_row, 4, pid)
            ws_mrp.cell(insert_row, 5, jof_num or '')
            ws_mrp.cell(insert_row, 6, order_qty)
            ws_mrp.cell(insert_row, 7, f'=INDEX(Tubex_Dashboard!$H$11:$H$100,MATCH(MRP!D{insert_row},Tubex_Dashboard!$F$11:$F$100,0))')
            ws_mrp.cell(insert_row, 8, f'=F{insert_row}-G{insert_row}')
            ws_mrp.cell(insert_row, 9, remarks or '')

            # Copy styles from Row 3
            for c in range(1, 10):
                src = ws_mrp.cell(3, c)
                dst = ws_mrp.cell(insert_row, c)
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format

            if 3 in ws_mrp.row_dimensions:
                ws_mrp.row_dimensions[insert_row].height = ws_mrp.row_dimensions[3].height

            # Boundaries shift by +1
            new_tube_end = b['tube_end'] + 1
            new_tube_total = b['tube_total_row'] + 1
            new_tube_items_start = new_tube_total + 3
            new_tube_items_end = b['tube_items_end'] + 1

            ws_mrp.cell(new_tube_total, 6, f'=SUM(F3:F{new_tube_end})')
            ws_mrp.cell(new_tube_total, 7, f'=SUM(G3:G{new_tube_end})')
            ws_mrp.cell(new_tube_total, 8, f'=SUMIF(H3:H{new_tube_end}, ">0")')

            update_tube_material_formulas(ws_mrp, new_tube_end, new_tube_items_start, new_tube_items_end)

            new_pet_start = b['pet_start'] + 1
            new_pet_end = b['pet_end'] + 1
            new_pet_total = b['pet_total_row'] + 1
            new_pet_items_start = new_pet_total + 3
            new_pet_items_end = b['pet_items_end'] + 1

            for r in range(new_pet_start, new_pet_end + 1):
                ws_mrp.cell(r, 7, f'=INDEX(Tubex_Dashboard!$H$11:$H$100,MATCH(MRP!D{r},Tubex_Dashboard!$F$11:$F$100,0))')
                ws_mrp.cell(r, 8, f'=F{r}-G{r}')

            ws_mrp.cell(new_pet_total, 6, f'=SUM(F{new_pet_start}:F{new_pet_end})')
            ws_mrp.cell(new_pet_total, 7, f'=SUM(G{new_pet_start}:G{new_pet_end})')
            ws_mrp.cell(new_pet_total, 8, f'=SUMIF(H{new_pet_start}:H{new_pet_end}, ">0")')

            update_pet_material_formulas(ws_mrp, new_pet_start, new_pet_end, new_pet_items_start, new_pet_items_end)

        else:
            insert_row = b['pet_total_row']
            ws_mrp.insert_rows(insert_row)

            ws_mrp.cell(insert_row, 1, dia)
            ws_mrp.cell(insert_row, 2, customer)
            ws_mrp.cell(insert_row, 3, pname)
            ws_mrp.cell(insert_row, 4, pid)
            ws_mrp.cell(insert_row, 5, jof_num or '')
            ws_mrp.cell(insert_row, 6, order_qty)
            ws_mrp.cell(insert_row, 7, f'=INDEX(Tubex_Dashboard!$H$11:$H$100,MATCH(MRP!D{insert_row},Tubex_Dashboard!$F$11:$F$100,0))')
            ws_mrp.cell(insert_row, 8, f'=F{insert_row}-G{insert_row}')
            ws_mrp.cell(insert_row, 9, remarks or '')

            for c in range(1, 10):
                src = ws_mrp.cell(b['pet_start'], c)
                dst = ws_mrp.cell(insert_row, c)
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format

            if b['pet_start'] in ws_mrp.row_dimensions:
                ws_mrp.row_dimensions[insert_row].height = ws_mrp.row_dimensions[b['pet_start']].height

            new_pet_end = b['pet_end'] + 1
            new_pet_total = b['pet_total_row'] + 1
            new_pet_items_start = new_pet_total + 3
            new_pet_items_end = b['pet_items_end'] + 1

            ws_mrp.cell(new_pet_total, 6, f'=SUM(F{b["pet_start"]}:F{new_pet_end})')
            ws_mrp.cell(new_pet_total, 7, f'=SUM(G{b["pet_start"]}:G{new_pet_end})')
            ws_mrp.cell(new_pet_total, 8, f'=SUMIF(H{b["pet_start"]}:H{new_pet_end}, ">0")')

            update_pet_material_formulas(ws_mrp, b['pet_start'], new_pet_end, new_pet_items_start, new_pet_items_end)

        print(f"\n[INSERTED] New order row inserted at MRP Row {insert_row}.")

    # Save workbook
    wb.save(filepath)
    print(f"[OK] Changes written to {os.path.basename(filepath)}.")

    # Re-sort Dashboard
    sort_script = os.path.join(os.path.dirname(__file__), 'sort_dashboard.py')
    if os.path.exists(sort_script):
        os.system(f'python "{sort_script}"')

    # Confirmation
    print(f"\n" + "=" * 60)
    print(f"   ORDER UPDATE SUMMARY")
    print(f"=" * 60)
    print(f"Product:  [{p_type}] {pid} - {pname}")
    print(f"Customer: {customer}")
    print(f"Quantity: {order_qty:,}")
    print(f"JOF #:    {jof_num or 'N/A'}")
    print(f"=" * 60)

    return confirm_keep_or_revert(filepath, backup_file, auto_keep)


def remove_order_from_workbook(filepath, search_term, auto_keep=False):
    """Removes an order from MRP and updates formulas and dashboard."""
    backup_file = backup_workbook(filepath)
    wb = load_workbook(filepath, data_only=False)
    ws_mrp = wb['MRP']

    catalog = load_product_catalog(wb)
    prod = resolve_product(catalog, search_term)
    if not prod:
        print(f"[ERROR] Could not resolve product for '{search_term}'.")
        return False

    pid = prod['pid']
    p_type = prod['type']
    pname = prod['product_name']

    b = get_mrp_boundaries(ws_mrp)

    # Locate order row in MRP
    target_row = None
    if p_type == 'TUBE':
        for r in range(b['tube_start'], b['tube_end'] + 1):
            if ws_mrp.cell(r, 4).value is not None and int(ws_mrp.cell(r, 4).value) == pid:
                target_row = r
                break
    else:
        for r in range(b['pet_start'], b['pet_end'] + 1):
            if ws_mrp.cell(r, 4).value is not None and int(ws_mrp.cell(r, 4).value) == pid:
                target_row = r
                break

    if not target_row:
        print(f"[WARN] Product ID {pid} ({pname}) does not have an active order in MRP to remove.")
        return False

    print(f"\n[REMOVING] Found active order at MRP Row {target_row}: PID {pid} - {pname}")
    ws_mrp.delete_rows(target_row)

    # Re-scan boundaries after deletion
    b_new = get_mrp_boundaries(ws_mrp)

    if p_type == 'TUBE':
        ws_mrp.cell(b_new['tube_total_row'], 6, f'=SUM(F3:F{b_new["tube_end"]})')
        ws_mrp.cell(b_new['tube_total_row'], 7, f'=SUM(G3:G{b_new["tube_end"]})')
        ws_mrp.cell(b_new['tube_total_row'], 8, f'=SUMIF(H3:H{b_new["tube_end"]}, ">0")')
        update_tube_material_formulas(ws_mrp, b_new['tube_end'], b_new['tube_items_start'], b_new['tube_items_end'])

        # Shift PET row formulas
        for r in range(b_new['pet_start'], b_new['pet_end'] + 1):
            ws_mrp.cell(r, 7, f'=INDEX(Tubex_Dashboard!$H$11:$H$100,MATCH(MRP!D{r},Tubex_Dashboard!$F$11:$F$100,0))')
            ws_mrp.cell(r, 8, f'=F{r}-G{r}')

        ws_mrp.cell(b_new['pet_total_row'], 6, f'=SUM(F{b_new["pet_start"]}:F{b_new["pet_end"]})')
        ws_mrp.cell(b_new['pet_total_row'], 7, f'=SUM(G{b_new["pet_start"]}:G{b_new["pet_end"]})')
        ws_mrp.cell(b_new['pet_total_row'], 8, f'=SUMIF(H{b_new["pet_start"]}:H{b_new["pet_end"]}, ">0")')
        update_pet_material_formulas(ws_mrp, b_new['pet_start'], b_new['pet_end'], b_new['pet_items_start'], b_new['pet_items_end'])
    else:
        ws_mrp.cell(b_new['pet_total_row'], 6, f'=SUM(F{b_new["pet_start"]}:F{b_new["pet_end"]})')
        ws_mrp.cell(b_new['pet_total_row'], 7, f'=SUM(G{b_new["pet_start"]}:G{b_new["pet_end"]})')
        ws_mrp.cell(b_new['pet_total_row'], 8, f'=SUMIF(H{b_new["pet_start"]}:H{b_new["pet_end"]}, ">0")')
        update_pet_material_formulas(ws_mrp, b_new['pet_start'], b_new['pet_end'], b_new['pet_items_start'], b_new['pet_items_end'])

    wb.save(filepath)
    print(f"[OK] Removed row from MRP and updated formulas.")

    sort_script = os.path.join(os.path.dirname(__file__), 'sort_dashboard.py')
    if os.path.exists(sort_script):
        os.system(f'python "{sort_script}"')

    print(f"\n" + "=" * 60)
    print(f"   ORDER REMOVAL SUMMARY")
    print(f"=" * 60)
    print(f"Removed Product: [{p_type}] {pid} - {pname}")
    print(f"=" * 60)

    return confirm_keep_or_revert(filepath, backup_file, auto_keep)


def print_terminal_summary(filepath):
    """Prints a detailed terminal summary of MRP and Dashboard active tables."""
    wb = load_workbook(filepath, data_only=False)
    ws_dash = wb['Tubex_Dashboard']
    ws_mrp = wb['MRP']
    b = get_mrp_boundaries(ws_mrp)

    print("\n" + "=" * 78)
    print("   [TERMINAL PREVIEW] ACTIVE TUBE ORDERS (MRP)")
    print("=" * 78)
    print(f"  {'Row':<5} | {'PID':<5} | {'Product Name':<32} | {'Dia':<4} | {'Quantity':<12} | {'JOF':<10}")
    print("  " + "-" * 74)
    for r in range(b['tube_start'], b['tube_end'] + 1):
        dia = ws_mrp.cell(r, 1).value
        name = str(ws_mrp.cell(r, 3).value or '')[:32]
        pid = ws_mrp.cell(r, 4).value
        jof = ws_mrp.cell(r, 5).value or '-'
        qty = str(ws_mrp.cell(r, 6).value)
        print(f"  {r:<5} | {str(pid):<5} | {name:<32} | {str(dia):<4} | {qty:<12} | {str(jof):<10}")
    tot_f = str(ws_mrp.cell(b['tube_total_row'], 6).value)
    print("  " + "-" * 74)
    print(f"  {'TOTAL':<46} | Formula: {tot_f}")

    print("\n" + "=" * 78)
    print("   [TERMINAL PREVIEW] ACTIVE PET ORDERS (MRP)")
    print("=" * 78)
    print(f"  {'Row':<5} | {'PID':<5} | {'Product Name':<32} | {'Dia':<8} | {'Quantity':<12} | {'JOF':<10}")
    print("  " + "-" * 74)
    for r in range(b['pet_start'], b['pet_end'] + 1):
        dia = ws_mrp.cell(r, 1).value
        name = str(ws_mrp.cell(r, 3).value or '')[:32]
        pid = ws_mrp.cell(r, 4).value
        jof = ws_mrp.cell(r, 5).value or '-'
        qty = str(ws_mrp.cell(r, 6).value)
        print(f"  {r:<5} | {str(pid):<5} | {name:<32} | {str(dia):<8} | {qty:<12} | {str(jof):<10}")
    tot_pet = str(ws_mrp.cell(b['pet_total_row'], 6).value)
    print("  " + "-" * 74)
    print(f"  {'TOTAL':<50} | Formula: {tot_pet}")

    print("\n" + "=" * 78)
    print("   [TERMINAL PREVIEW] DASHBOARD ACTIVE SUMMARY TABLE")
    print("=" * 78)
    print(f"  {'Row':<5} | {'Type':<5} | {'PID':<5} | {'Product Name':<32} | {'Orders Lookup':<20}")
    print("  " + "-" * 74)
    for r in range(11, ws_dash.max_row + 1):
        p_type = ws_dash.cell(r, 2).value
        if not p_type and ws_dash.cell(r, 4).value == "TOTAL":
            tot_g = ws_dash.cell(r, 7).value
            print(f"  {r:<5} | {'---':<5} | {'---':<5} | {'TOTAL':<32} | {str(tot_g):<20}")
            continue
        if p_type not in ('TUBE', 'PET'):
            continue
        # Check if active row (before inactive section starts)
        pid = ws_dash.cell(r, 6).value
        name = str(ws_dash.cell(r, 4).value or '')[:32]
        orders = str(ws_dash.cell(r, 7).value or '')
        print(f"  {r:<5} | {str(p_type):<5} | {str(pid):<5} | {name:<32} | {orders:<20}")
        if r >= 25:
            # Reached lower inactive rows
            break
    print("=" * 78)
    wb.close()


def confirm_keep_or_revert(filepath, backup_file, auto_keep):
    """Prompts user to Preview in Excel, View Terminal Summary, Keep changes, or Revert."""
    if auto_keep:
        print(f"\n[SAVED] Auto-keep enabled. All changes saved.")
        return True

    preview_file = None

    while True:
        print("\n" + "=" * 60)
        print("   WHAT WOULD YOU LIKE TO DO?")
        print("=" * 60)
        print("  [P] Preview in Excel (Opens full spreadsheet for review)")
        print("  [V] View Terminal Table Summary")
        print("  [K] Keep changes (Finalize & Save) [Default]")
        print("  [R] Revert to backup (Discard all changes)")
        print("=" * 60)

        try:
            choice = input("\nEnter choice [P/V/K/R] (default: K): ").strip().upper()
        except (EOFError, KeyboardInterrupt):
            choice = 'K'

        if not choice or choice == 'K':
            if preview_file and os.path.exists(preview_file):
                try:
                    os.remove(preview_file)
                except Exception:
                    pass
            print(f"\n[SAVED] All changes kept successfully! Report is ready.")
            return True

        elif choice == 'P':
            ts = datetime.now().strftime("%H%M%S")
            preview_name = f"PREVIEW_Tubex_{ts}.xlsx"
            preview_file = os.path.join(LOGS_DIR, preview_name)
            shutil.copy2(filepath, preview_file)
            print(f"\n[PREVIEW] Opening temporary preview in Excel: {preview_name}")
            print(f"          Take your time to inspect the sheets.")
            print(f"          When done, return here to Keep [K] or Revert [R].")
            try:
                os.startfile(preview_file)
            except Exception as e:
                print(f"  (Could not auto-launch Excel: {e}. File saved at: {preview_file})")

        elif choice == 'V':
            print_terminal_summary(filepath)

        elif choice == 'R':
            shutil.copy2(backup_file, filepath)
            sort_script = os.path.join(os.path.dirname(__file__), 'sort_dashboard.py')
            if os.path.exists(sort_script):
                os.system(f'python "{sort_script}"')
            if preview_file and os.path.exists(preview_file):
                try:
                    os.remove(preview_file)
                except Exception:
                    pass
            print(f"\n[REVERTED] All changes were discarded. Workbook restored to backup state.")
            return False

        else:
            print(f"Invalid option '{choice}'. Please select P, V, K, or R.")


def list_active_orders():
    """Prints all active orders currently in MRP."""
    wb = load_workbook(EXCEL_PATH, data_only=False)
    ws_mrp = wb['MRP']
    b = get_mrp_boundaries(ws_mrp)

    print("\n" + "=" * 70)
    print("   CURRENT ACTIVE TUBE ORDERS (MRP)")
    print("=" * 70)
    for r in range(b['tube_start'], b['tube_end'] + 1):
        dia = ws_mrp.cell(r, 1).value
        cust = ws_mrp.cell(r, 2).value
        name = ws_mrp.cell(r, 3).value
        pid = ws_mrp.cell(r, 4).value
        jof = ws_mrp.cell(r, 5).value or ''
        qty = ws_mrp.cell(r, 6).value
        print(f"  Row {r}: PID {pid} | {name} ({cust}) Dia: {dia} | Qty: {qty} | JOF: {jof}")

    print("\n" + "=" * 70)
    print("   CURRENT ACTIVE PET ORDERS (MRP)")
    print("=" * 70)
    for r in range(b['pet_start'], b['pet_end'] + 1):
        dia = ws_mrp.cell(r, 1).value
        cust = ws_mrp.cell(r, 2).value
        name = ws_mrp.cell(r, 3).value
        pid = ws_mrp.cell(r, 4).value
        jof = ws_mrp.cell(r, 5).value or ''
        qty = ws_mrp.cell(r, 6).value
        print(f"  Row {r}: PID {pid} | {name} ({cust}) Dia: {dia} | Qty: {qty} | JOF: {jof}")
    print("=" * 70)
    wb.close()


def main():
    parser = argparse.ArgumentParser(description="Add, Accumulate, or Remove Tube/PET Job Orders in Tubex")
    parser.add_argument('-p', '--product', '--pid', dest='product', help="Product ID (e.g. 5814) or Product Name (e.g. 'VINCE NURTURAL')")
    parser.add_argument('-q', '--qty', dest='qty', type=int, help="Order Quantity (e.g. 25000)")
    parser.add_argument('-j', '--jof', dest='jof', help="Job Order # / JOF No. (e.g. 6901, 349)")
    parser.add_argument('-c', '--customer', dest='customer', help="Customer Name (optional override)")
    parser.add_argument('-r', '--remarks', dest='remarks', help="Order Remarks (e.g. 'Cap Shortage')")
    parser.add_argument('-y', '--yes', '--keep', dest='auto_keep', action='store_true', help="Automatically keep changes without prompting")
    parser.add_argument('--remove', '--delete', dest='remove_target', help="Remove an order by Product ID or Product Name")
    parser.add_argument('--list', action='store_true', help="List all active orders currently in MRP")

    args = parser.parse_args()

    if args.list:
        list_active_orders()
        return

    if args.remove_target:
        remove_order_from_workbook(EXCEL_PATH, args.remove_target, auto_keep=args.auto_keep)
        return

    # Load catalog
    wb_temp = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    catalog = load_product_catalog(wb_temp)
    wb_temp.close()

    product_input = args.product
    qty_input = args.qty
    jof_input = args.jof
    cust_input = args.customer
    remarks_input = args.remarks

    # Interactive wizard mode if no CLI options supplied
    if not product_input:
        print("=" * 60)
        print("   ALPHA CONTAINERS -- ORDER MANAGEMENT WIZARD")
        print("=" * 60)
        print("  [1] Add or Increase Order Quantity")
        print("  [2] Remove / Delete an Active Order")
        print("  [3] List Current Active Orders")
        print("  [Q] Quit")

        try:
            mode = input("\nSelect option [1/2/3/Q] (default: 1): ").strip().upper()
        except (EOFError, KeyboardInterrupt):
            return

        if mode == '2':
            list_active_orders()
            try:
                target = input("\nEnter Product ID or Name to remove: ").strip()
                if target:
                    remove_order_from_workbook(EXCEL_PATH, target, auto_keep=args.auto_keep)
            except (EOFError, KeyboardInterrupt):
                pass
            return
        elif mode == '3':
            list_active_orders()
            return
        elif mode == 'Q':
            print("Cancelled.")
            return

        try:
            product_input = input("\nEnter Product ID or Product Name: ").strip()
        except (EOFError, KeyboardInterrupt):
            return

    if not product_input:
        print("[ERROR] No product provided.")
        return

    prod = resolve_product(catalog, product_input)
    if not prod:
        print(f"[ERROR] Could not find product matching '{product_input}' in Product_Catalog.")
        return

    print(f"\nSelected: [{prod['type']}] PID {prod['pid']} - {prod['product_name']} ({prod['customer']}) Dia: {prod['dia']}")

    if qty_input is None:
        try:
            qty_str = input(f"Enter Order Quantity to Add for {prod['product_name']}: ").strip().replace(',', '')
            qty_input = int(qty_str)
        except (ValueError, EOFError, KeyboardInterrupt):
            print(f"[ERROR] Invalid quantity.")
            return

    if jof_input is None and not args.product:
        try:
            jof_input = input("Enter JOF # / Job Order Number (press Enter to skip): ").strip()
        except (EOFError, KeyboardInterrupt):
            jof_input = ''

    if remarks_input is None and not args.product:
        try:
            remarks_input = input("Enter Remarks (press Enter to skip): ").strip()
        except (EOFError, KeyboardInterrupt):
            remarks_input = ''

    add_or_update_order(
        filepath=EXCEL_PATH,
        product=prod,
        order_qty=qty_input,
        jof_num=jof_input or None,
        customer_override=cust_input or None,
        remarks=remarks_input or None,
        auto_keep=args.auto_keep
    )


if __name__ == '__main__':
    main()
