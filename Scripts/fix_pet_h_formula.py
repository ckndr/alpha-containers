import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

wb = openpyxl.load_workbook('Tubex_Aug26.xlsx')
ws = wb['MRP']

def clone_border(b):
    def cs(s):
        return Side(border_style=s.border_style, color=copy(s.color)) if s else Side()
    return Border(left=cs(b.left), right=cs(b.right),
                  top=cs(b.top), bottom=cs(b.bottom))

def apply_data_style(cell, ref_cell, num_fmt=None):
    cell.font      = Font(bold=ref_cell.font.bold,
                          size=ref_cell.font.size or 10,
                          name=ref_cell.font.name or "Calibri")
    cell.fill      = PatternFill("solid", fgColor="F7F9FC")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = clone_border(ref_cell.border)
    if num_fmt:
        cell.number_format = num_fmt

# =====================================================================
# Fix H100:H106 - August PET Pieces Can Produce
#
# Problem: openpyxl/Excel converted TableBOM[Product ID] to
# TableBOM[[#This Row],[Product ID]] which doesn't work outside the BOM table.
#
# Solution: Use the actual cell range for the BOM Product ID column
# instead of structured table references inside SUMIF.
# BOM table = A2:M353, so Product ID (col A) = BOM!$A$3:$A$353
# Item ID (col G) = BOM!$G$3:$G$353
# Per 1000 Units (col J) = BOM!$J$3:$J$353
# =====================================================================

# BOM ranges (avoiding structured refs that get mangled)
BOM_PROD_ID   = 'BOM!$A$3:$A$353'
BOM_ITEM_ID   = 'BOM!$G$3:$G$353'
BOM_PER1000   = 'BOM!$J$3:$J$353'

for r in range(100, 107):
    # rem = remaining qty for each BOM product in the PET orders
    rem  = 'IFERROR(SUMIF($D$93:$D$96,{prod},$H$93:$H$96),0)'.format(prod=BOM_PROD_ID)
    item = '({bom_item}=A{r})'.format(bom_item=BOM_ITEM_ID, r=r)
    pos  = '({rem}>0)'.format(rem=rem)

    num = 'SUMPRODUCT({item}*{pos}*{rem}*{per1000})'.format(
              item=item, pos=pos, rem=rem, per1000=BOM_PER1000)
    den = 'SUMPRODUCT({item}*{pos}*{rem})'.format(
              item=item, pos=pos, rem=rem)

    formula = (
        '=IFERROR(IF({num}=0,"-",'
        'ROUND(F{r}*1000*{den}/{num},0)),"-")'
    ).format(num=num, den=den, r=r)

    cell = ws.cell(row=r, column=8, value=formula)
    apply_data_style(cell, ws['I{}'.format(r)], num_fmt="#,##0")

wb.save('Tubex_Aug26.xlsx')
print('Fixed H100:H106 - using direct cell ranges instead of structured table refs.')
print()

# Verify
wb2 = openpyxl.load_workbook('Tubex_Aug26.xlsx')
ws2 = wb2['MRP']
print('=== H100 formula ===')
print(ws2['H100'].value)
print()
print('=== H101 formula ===')
print(ws2['H101'].value)
