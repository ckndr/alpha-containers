import os
import openpyxl

# Fallback mapping if reading from catalog fails or to force standard names
MANUAL_MAPPING = {
    "SAMSOL": "Samsol International Private Limited",
    "ALPHA": "Alpha Labs PVT LTD",
    "ALPHA LAB": "Alpha Labs PVT LTD",
    "ALPHA CONTAINER": "Alpha Labs PVT LTD",
    "MABLEY": "Mablay Beauty PVT LTD.",
    "MABLAY": "Mablay Beauty PVT LTD.",
    "BROOKES": "Brookes Pharma Private Limited",
    "GOLDEN": "Golden Pearl Cosmetics (PVT) LTD",
    "SEATLE": "Seatle (Private) Limited",
    "PROFESSIONAL": "Professional Beauty Solution (Pvt) Ltd"
}

_master_customers = []

def load_master_customers(alpha_dir):
    global _master_customers
    if _master_customers:
        return _master_customers
        
    # Search for latest Tubex_*.xlsx
    import glob
    active_files = sorted(glob.glob(os.path.join(alpha_dir, "Tubex_*.xlsx")), key=os.path.getmtime)
    if active_files:
        latest_file = active_files[-1]
        try:
            wb = openpyxl.load_workbook(latest_file, data_only=True)
            if "Product_Catalog" in wb.sheetnames:
                ws = wb["Product_Catalog"]
                custs = set()
                for r in range(4, ws.max_row + 1):
                    val = ws.cell(r, 3).value
                    if val:
                        custs.add(str(val).strip())
                _master_customers = sorted(list(custs))
            wb.close()
        except Exception:
            pass
            
    if not _master_customers:
        # Fallback to unique values from MANUAL_MAPPING
        _master_customers = sorted(list(set(MANUAL_MAPPING.values())))
        
    return _master_customers

def normalize_customer_name(raw_name, alpha_dir):
    """
    Takes a raw customer name and normalizes it to the master catalog name.
    """
    if not raw_name or raw_name == "None":
        return "Other Customers"
        
    raw_str = str(raw_name).strip()
    raw_upper = raw_str.upper()
    
    # 1. Exact match in MANUAL_MAPPING
    if raw_upper in MANUAL_MAPPING:
        return MANUAL_MAPPING[raw_upper]
        
    # 2. Check if a key from MANUAL_MAPPING is in the raw string (e.g. "Samsol International" contains "SAMSOL")
    for key, official_name in MANUAL_MAPPING.items():
        if key in raw_upper:
            return official_name
            
    # 3. Check against master customers dynamically
    master_list = load_master_customers(alpha_dir)
    
    # Exact match check
    for mc in master_list:
        if raw_upper == mc.upper():
            return mc
            
    # Partial match check (e.g., "Al-Rehman" in "Al-Rehman Group")
    for mc in master_list:
        # Prevent very short matches from overriding
        if len(raw_upper) > 3 and (mc.upper() in raw_upper or raw_upper in mc.upper()):
            return mc
            
    return raw_str

def correct_customer_by_product(customer_name, product_name):
    """
    Overrides customer name based on product name rules.
    This should be called AFTER normalize_customer_name if possible, 
    or just returns the corrected customer name.
    """
    if not product_name:
        return customer_name
        
    p_upper = str(product_name).upper()
    
    # Rule 1: "if its yellow and is either 120 or 200 then its Samsol"
    if "YELLOW" in p_upper and ("120" in p_upper or "200" in p_upper):
        return MANUAL_MAPPING["SAMSOL"]
        
    # Rule 2: "PET BOTTLE (150ML)TRANSPARENT BODY MIST" belongs to Alpha Lab
    if "PET BOTTLE (150ML)TRANSPARENT BODY MIST" in p_upper:
        return MANUAL_MAPPING["ALPHA"]
        
    return customer_name

if __name__ == "__main__":
    alpha_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    test_cases = [
        "Samsol",
        "Samsol International",
        "Alpha Container",
        "Alpha Lab",
        "Mabley Beauty",
        "Professional Beauty Solution (PVT) LTD.Pakistan",
        "Adore"
    ]
    for tc in test_cases:
        print(f"'{tc}' -> '{normalize_customer_name(tc, alpha_dir)}'")
