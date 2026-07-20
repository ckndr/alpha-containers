"""
Tubex — Daily Update Master Script
───────────────────────────────────
Complete replacement for Run_All_Updates.bat. One script handles:
  1. Pre-run backup of Excel workbook
  2. Detect fresh ERP exports (inventory, dispatch, dispatch_pet)
  3. Find Imran's production report from Downloads
  4. Apply WIP message from Mehmood
  5. Run full pipeline (production → inventory → dispatch → sort → HTML)
  6. Cross-check results against Imran's source data
  7. Take dashboard screenshot for WhatsApp
  8. Git push to GitHub

Usage:
  python daily.py              (interactive — walks you through everything)
  python daily.py --skip-wip   (skip WIP prompt)
  python daily.py --skip-prod  (skip Production file search)
  python daily.py --skip-git   (skip Git push)

Author: Sikander
"""

import os
import sys
import glob
import shutil
import time
import subprocess
import logging
from datetime import datetime

# ── PATH SETUP ──────────────────────────────────────────────────────────────
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
ALPHA_DIR   = os.path.dirname(SCRIPTS_DIR)
LOGS_DIR    = os.path.join(ALPHA_DIR, "Logs")

# ── CONFIGURABLE ────────────────────────────────────────────────────────────
DOWNLOADS_DIR = r"C:\Users\HP\Downloads"

# Production file pattern: "Production report June-2026.xlsx" etc.
PROD_FILE_PATTERN = "Production report "   # files starting with this
PROD_TARGET_NAME  = "Production.xlsx"      # what we rename/copy it to

# How many backup files to keep
MAX_BACKUPS = 3

# ── COLORS (Windows terminal) ──────────────────────────────────────────────
if sys.platform == 'win32':
    os.system('')  # enables ANSI escape codes

GREEN  = '\033[92m'
YELLOW = '\033[93m'
RED    = '\033[91m'
CYAN   = '\033[96m'
BOLD   = '\033[1m'
DIM    = '\033[2m'
RESET  = '\033[0m'

TOTAL_STEPS = 9

def ok(msg):
    try:
        print(f"    {GREEN}✓{RESET} {msg}")
    except UnicodeEncodeError:
        safe_msg = msg.replace('✓', '[OK]').replace('─', '-').replace('⚠', '[WARN]').replace('✗', '[FAIL]')
        print(f"    [OK] {safe_msg}")

def warn(msg):
    try:
        print(f"    {YELLOW}⚠{RESET} {msg}")
    except UnicodeEncodeError:
        safe_msg = msg.replace('✓', '[OK]').replace('─', '-').replace('⚠', '[WARN]').replace('✗', '[FAIL]')
        print(f"    [WARN] {safe_msg}")

def fail(msg):
    try:
        print(f"    {RED}✗{RESET} {msg}")
    except UnicodeEncodeError:
        safe_msg = msg.replace('✓', '[OK]').replace('─', '-').replace('⚠', '[WARN]').replace('✗', '[FAIL]')
        print(f"    [FAIL] {safe_msg}")

def timed_input(prompt, timeout=2.0):
    print(prompt, end='', flush=True)
    if sys.platform != 'win32':
        try:
            import select
            rlist, _, _ = select.select([sys.stdin], [], [], timeout)
            if rlist:
                return sys.stdin.readline().strip()
            else:
                print()
                return None
        except Exception:
            return input().strip()

    import msvcrt
    start_time = time.time()
    input_str = ""
    has_started = False

    while True:
        if not has_started and (time.time() - start_time >= timeout):
            print()
            return None

        if msvcrt.kbhit():
            has_started = True
            ch = msvcrt.getwch()
            if ch == '\r' or ch == '\n':
                print()
                return input_str.strip()
            elif ch == '\b':  # Backspace
                if len(input_str) > 0:
                    input_str = input_str[:-1]
                    sys.stdout.write('\b \b')
                    sys.stdout.flush()
            elif ord(ch) >= 32:  # Printable
                input_str += ch
                sys.stdout.write(ch)
                sys.stdout.flush()
        time.sleep(0.05)

def header(step, title):
    print(f"\n  {CYAN}{BOLD}[{step}/{TOTAL_STEPS}]{RESET} {BOLD}{title}{RESET}")


def banner():
    msg = f"""
  {CYAN}╔══════════════════════════════════════════════════╗
  ║  {BOLD}Tubex — Daily Update{RESET}{CYAN}                              ║
  ║  {DIM}{datetime.now().strftime('%A, %d-%b-%Y %H:%M')}{RESET}{CYAN}                     ║
  ╚══════════════════════════════════════════════════╝{RESET}
"""
    try:
        print(msg)
    except UnicodeEncodeError:
        ascii_msg = f"""
  +--------------------------------------------------+
     Tubex - Daily Update
     {datetime.now().strftime('%A, %d-%b-%Y %H:%M')}
  +--------------------------------------------------+
"""
        print(ascii_msg)


# ── LOGGING SETUP ──────────────────────────────────────────────────────────
def setup_logging():
    """Mirror all print output to a timestamped log file."""
    os.makedirs(LOGS_DIR, exist_ok=True)
    log_name = f"update_{datetime.now().strftime('%Y%m%d_%H%M')}.log"
    log_path = os.path.join(LOGS_DIR, log_name)

    # Custom stream that writes to both console and file
    class TeeStream:
        def __init__(self, original, log_file):
            self.original = original
            self.log_file = log_file
        def write(self, data):
            try:
                self.original.write(data)
            except UnicodeEncodeError:
                # Fallback: replace common box-drawing / checkmark symbols for console output
                clean_data = data.replace('✓', '[OK]').replace('⚠', '[WARN]').replace('✗', '[FAIL]')
                clean_data = clean_data.replace('─', '-').replace('╔', '+').replace('═', '-').replace('╗', '+')
                clean_data = clean_data.replace('║', '|').replace('╚', '+').replace('╝', '+')
                try:
                    self.original.write(clean_data)
                except UnicodeEncodeError:
                    enc = getattr(self.original, 'encoding', 'ascii') or 'ascii'
                    self.original.write(clean_data.encode(enc, errors='replace').decode(enc))
            self.log_file.write(data)
            self.log_file.flush()
        def flush(self):
            self.original.flush()
            self.log_file.flush()

    log_file = open(log_path, 'w', encoding='utf-8')
    sys.stdout = TeeStream(sys.__stdout__, log_file)
    sys.stderr = TeeStream(sys.__stderr__, log_file)
    return log_path


# ═══════════════════════════════════════════════════════════════════════════
# STEP 1: PRE-RUN BACKUP
# ═══════════════════════════════════════════════════════════════════════════
def step_backup():
    header(1, "Pre-run backup...")
    os.makedirs(LOGS_DIR, exist_ok=True)

    date_str = datetime.now().strftime('%Y%m%d')
    excel_files = sorted(glob.glob(os.path.join(ALPHA_DIR, "Tubex*.xlsx")))

    if not excel_files:
        warn("No Tubex*.xlsx found — nothing to back up")
        return

    for src in excel_files:
        name = os.path.basename(src)
        dst = os.path.join(LOGS_DIR, f"backup_{date_str}_{name}")
        shutil.copy2(src, dst)
        ok(f"Backed up: {name}")

    # Clean old backups — keep only the last MAX_BACKUPS
    backups = sorted(glob.glob(os.path.join(LOGS_DIR, "backup_*.xlsx")),
                     key=os.path.getmtime, reverse=True)
    for old in backups[MAX_BACKUPS:]:
        os.remove(old)
        print(f"    {DIM}Cleaned old backup: {os.path.basename(old)}{RESET}")


# ═══════════════════════════════════════════════════════════════════════════
# STEP 2: CHECK ERP EXPORTS
# ═══════════════════════════════════════════════════════════════════════════
def step_check_erp():
    header(2, "Checking ERP exports...")
    warnings_list = []

    erp_files = {
        'inventory.xls':    'Inventory',
        'dispatch.xls':     'Dispatch (Tube)',
        'dispatch_pet.xls': 'Dispatch (PET)',
    }

    all_ok = True
    for filename, label in erp_files.items():
        target = os.path.join(ALPHA_DIR, filename)

        # Check for "- copy" variant first (Sikander's ERP download workflow)
        import re
        stem, ext = os.path.splitext(filename)
        pattern = re.compile(rf"^{re.escape(stem)} - copy(?: \(\d+\))?{re.escape(ext)}$", re.IGNORECASE)
        copy_matches = [
            os.path.join(ALPHA_DIR, f) for f in os.listdir(ALPHA_DIR)
            if pattern.match(f)
        ]

        if copy_matches:
            latest_copy = max(copy_matches, key=os.path.getmtime)
            age_min = (time.time() - os.path.getmtime(latest_copy)) / 60
            ok(f"{label}: fresh copy found ({os.path.basename(latest_copy)}, {int(age_min)} min ago)")
            continue

        if os.path.exists(target):
            age_h = (time.time() - os.path.getmtime(target)) / 3600
            if age_h < 26:
                ok(f"{label}: {filename} ({age_h:.1f}h old)")
            else:
                warn(f"{label}: {filename} is {age_h:.0f}h old — stale?")
                warnings_list.append(f"{label}: file is stale ({age_h:.0f} hours old)")
                all_ok = False
        else:
            fail(f"{label}: {filename} NOT FOUND")
            warnings_list.append(f"{label}: file ({filename}) not found")
            all_ok = False

    if not all_ok:
        print(f"\n    {DIM}Copy fresh ERP exports to {ALPHA_DIR}")
        print(f"    as 'filename - copy.xls' — pipeline auto-replaces.{RESET}")

    return warnings_list


# ═══════════════════════════════════════════════════════════════════════════
# STEP 3: FIND PRODUCTION FILE
# ═══════════════════════════════════════════════════════════════════════════
def step_find_production(skip=False):
    header(3, "Finding Production report...")

    if skip:
        warn("Skipped (--skip-prod)")
        return True

    target = os.path.join(ALPHA_DIR, PROD_TARGET_NAME)

    # Search Downloads for files matching "Production report *.xlsx"
    candidates = []
    if os.path.isdir(DOWNLOADS_DIR):
        for f in os.listdir(DOWNLOADS_DIR):
            if (f.lower().startswith(PROD_FILE_PATTERN.lower())
                    and f.lower().endswith('.xlsx')
                    and '~$' not in f):
                full = os.path.join(DOWNLOADS_DIR, f)
                age_h = (time.time() - os.path.getmtime(full)) / 3600
                candidates.append((full, f, age_h))

    # Sort by modification time — most recent first
    candidates.sort(key=lambda x: x[2])

    if candidates:
        best_path, best_name, best_age = candidates[0]

        # Format age string
        if best_age < 1:
            age_str = f"{int(best_age * 60)} min ago"
        elif best_age < 24:
            age_str = f"{best_age:.1f}h ago"
        else:
            age_str = f"{best_age / 24:.1f} days ago"

        print(f"    Found: {CYAN}{best_name}{RESET}  ({age_str})")

        if len(candidates) > 1:
            print(f"    {DIM}({len(candidates)} total files found — using most recent){RESET}")

        # Check if we already have a newer file in Alpha
        if os.path.exists(target):
            target_mtime = os.path.getmtime(target)
            source_mtime = os.path.getmtime(best_path)
            if target_mtime >= source_mtime:
                ok(f"Existing {PROD_TARGET_NAME} is already up-to-date")
                resp = input(f"    {DIM}Replace anyway? [y/N]: {RESET}").strip().lower()
                if resp not in ('y', 'yes'):
                    return True

        # Copy (not move — keep original in Downloads as backup)
        shutil.copy2(best_path, target)
        ok(f"Copied → {PROD_TARGET_NAME}")
        return True

    # No candidates found in Downloads
    if os.path.exists(target):
        age_h = (time.time() - os.path.getmtime(target)) / 3600
        if age_h < 26:
            ok(f"Using existing {PROD_TARGET_NAME} ({age_h:.1f}h old)")
            return True
        else:
            warn(f"Existing {PROD_TARGET_NAME} is {age_h:.0f}h old — no fresh file found in Downloads")
    else:
        fail(f"No 'Production report *.xlsx' found in {DOWNLOADS_DIR}")
        fail(f"No existing {PROD_TARGET_NAME} in {ALPHA_DIR}")

    # Manual fallback
    manual = input(f"    {DIM}Drag+drop file here, or Enter to skip: {RESET}").strip().strip('"')
    if manual and os.path.exists(manual):
        shutil.copy2(manual, target)
        ok(f"Copied → {PROD_TARGET_NAME}")
        return True

    warn(f"Continuing without fresh Production data")
    return True


# ═══════════════════════════════════════════════════════════════════════════
# STEP 4: WIP UPDATE
# ═══════════════════════════════════════════════════════════════════════════
def step_wip(skip=False):
    header(4, "WIP Update (Mehmood's message)...")

    if skip:
        warn("Skipped (--skip-wip)")
        return True

    msg = timed_input(f"    Paste WIP message (2s timeout, or Enter/timeout to skip):\n    {CYAN}>{RESET} ", timeout=2.0)

    if not msg:
        warn("No WIP message — skipped")
        return True

    # Import WIP functions from existing update_wip.py
    sys.path.insert(0, SCRIPTS_DIR)
    try:
        from update_wip import parse_wip_message, build_slug_map, find_excel, pick_row
        from update_wip import INV_WIP_COL
        from openpyxl import load_workbook
    except ImportError as e:
        fail(f"Could not import WIP module: {e}")
        return False

    wip_data = parse_wip_message(msg)
    if not wip_data:
        fail("Could not parse. Expected: #19mm 10kg #30mm 125kg")
        return False

    excel_path, _ = find_excel()
    if not excel_path:
        fail("No Tubex*.xlsx found")
        return False

    wb = load_workbook(excel_path)
    ws = wb['Inventory']
    slug_map = build_slug_map(ws)

    # Clear existing WIP
    for dia, rows in slug_map.items():
        for r, issued, name in rows:
            if ws.cell(r, INV_WIP_COL).value is not None:
                ws.cell(r, INV_WIP_COL).value = None

    # Write new
    written = []
    for dia, kg in sorted(wip_data.items()):
        if dia in slug_map:
            row, name = pick_row(slug_map[dia])
            ws.cell(row, INV_WIP_COL).value = kg
            written.append(f"{dia}mm→{kg}kg")

    wb.save(excel_path)
    if written:
        ok(f"WIP updated: {', '.join(written)}")
    else:
        warn("No matching slug rows found")

    return True


# ═══════════════════════════════════════════════════════════════════════════
# STEP 5: RUN PIPELINE
# ═══════════════════════════════════════════════════════════════════════════
def step_pipeline():
    header(5, "Running update pipeline...")

    # Check Excel not open
    excel_files = sorted(glob.glob(os.path.join(ALPHA_DIR, "Tubex*.xlsx")))
    if excel_files:
        try:
            with open(excel_files[-1], 'r+b'):
                pass
        except PermissionError:
            fail(f"{os.path.basename(excel_files[-1])} is OPEN in Excel!")
            fail("Close Excel and run again.")
            return False

    # Clear old mismatches log
    mismatch_log = os.path.join(LOGS_DIR, "mismatches.log")
    if os.path.exists(mismatch_log):
        os.remove(mismatch_log)

    # Pipeline order is CRITICAL:
    # 1. Production first (populates Production_Log that sort_dashboard reads)
    # 2. Inventory (independent)
    # 3. Dispatch (can add dispatch to inactive products)
    # 4. Sort Dashboard AFTER 1-3 (rearranges based on fresh production+dispatch)
    # 5. HTML last (reads everything)
    scripts = [
        ("update_production.py",  "Production Log + FG Stock"),
        ("update_inventory.py",   "Inventory"),
        ("update_dispatch.py",    "Dispatch"),
        ("sort_dashboard.py",     "Sort Dashboard"),
        ("update_html.py",        "HTML Dashboard"),
    ]

    failures = []
    for script_name, label in scripts:
        path = os.path.join(SCRIPTS_DIR, script_name)
        if not os.path.exists(path):
            fail(f"{script_name} NOT FOUND")
            failures.append(label)
            continue

        print(f"\n    {DIM}── {label} ──{RESET}")
        result = subprocess.run(
            [sys.executable, path],
            cwd=SCRIPTS_DIR
        )

        if result.returncode != 0:
            fail(f"{label} FAILED (exit code {result.returncode})")
            failures.append(label)
        else:
            ok(label)

    # Show mismatches if any
    if os.path.exists(mismatch_log):
        try:
            print(f"\n    {YELLOW}{'─'*50}")
        except UnicodeEncodeError:
            print(f"\n    {YELLOW}{'-'*50}")
        print(f"    {BOLD}MISMATCHES DETECTED:{RESET}")
        with open(mismatch_log, 'r') as f:
            for line in f:
                print(f"    {line.rstrip()}")
        try:
            print(f"    {YELLOW}{'─'*50}{RESET}")
        except UnicodeEncodeError:
            print(f"    {YELLOW}{'-'*50}{RESET}")

    if failures:
        fail(f"Pipeline had {len(failures)} failure(s): {', '.join(failures)}")
        return False

    ok(f"{BOLD}Pipeline completed successfully{RESET}")
    return True


# ═══════════════════════════════════════════════════════════════════════════
# STEP 6: CROSS-CHECK
# ═══════════════════════════════════════════════════════════════════════════
def step_crosscheck():
    header(6, "Cross-checking with Imran's data...")

    errors = []
    prod_path = os.path.join(ALPHA_DIR, PROD_TARGET_NAME)
    if not os.path.exists(prod_path):
        warn("Production.xlsx not found — skipping machine-level and summary checks")
        errors.append("Production.xlsx not found - skipped cross-checks")
        return errors

    try:
        import pandas as pd
        from openpyxl import load_workbook
    except ImportError:
        warn("pandas or openpyxl not installed — skipping cross-check")
        errors.append("pandas or openpyxl not installed - skipped cross-checks")
        return errors

    # --- Part A: Machine-level production totals comparison ---
    try:
        # Detect header row dynamically by scanning the first 10 rows
        raw = pd.read_excel(prod_path, sheet_name='Production Day wise', header=None, nrows=10)
        best_row = 0
        best_score = 0
        KNOWN_KEYWORDS = {'date', 'machines', 'machine', 'customer', 'product name', 'good production'}
        for idx, r_vals in raw.iterrows():
            vals_clean = [str(v).strip().lower() for v in r_vals if str(v).strip() not in ('', 'nan')]
            score = sum(1 for v in vals_clean if any(k in v for k in KNOWN_KEYWORDS))
            if score > best_score:
                best_score = score
                best_row = idx
        
        # Read Imran's raw data using the detected header row
        df_imran = pd.read_excel(prod_path, sheet_name='Production Day wise', header=best_row)

        # Find the good-production and machine columns (flexible matching)
        good_col = None
        for col in df_imran.columns:
            c = str(col).lower().strip()
            if 'good' in c and ('qty' in c or 'prod' in c or 'quantity' in c):
                good_col = col
                break
        if good_col is None:
            for col in df_imran.columns:
                if 'good' in str(col).lower():
                    good_col = col
                    break

        machine_col = None
        for col in df_imran.columns:
            if 'machine' in str(col).lower():
                machine_col = col
                break

        if not good_col or not machine_col:
            warn("Could not identify Machine/Good columns in Imran's file")
            errors.append("Could not identify Machine/Good columns in Production.xlsx")
        else:
            df_imran[good_col] = pd.to_numeric(df_imran[good_col], errors='coerce').fillna(0)
            imran_totals = df_imran.groupby(machine_col)[good_col].sum()

            # Read our Production_Log
            excel_files = sorted(glob.glob(os.path.join(ALPHA_DIR, "Tubex*.xlsx")))
            if not excel_files:
                warn("No Tubex*.xlsx — skipping machine totals check")
                errors.append("No Tubex*.xlsx found for machine totals check")
            else:
                df_dash = pd.read_excel(excel_files[-1], sheet_name='Production_Log', header=1)

                good_col_d = None
                for col in df_dash.columns:
                    if 'good' in str(col).lower():
                        good_col_d = col
                        break
                machine_col_d = None
                for col in df_dash.columns:
                    if 'machine' in str(col).lower():
                        machine_col_d = col
                        break

                if not good_col_d or not machine_col_d:
                    warn("Could not identify columns in Production_Log")
                    errors.append("Could not identify columns in Production_Log")
                else:
                    df_dash[good_col_d] = pd.to_numeric(df_dash[good_col_d], errors='coerce').fillna(0)
                    dash_totals = df_dash.groupby(machine_col_d)[good_col_d].sum()

                    # Compare machine by machine
                    all_machines = sorted(set(imran_totals.index) | set(dash_totals.index))
                    mismatches = []
                    print()
                    for m in all_machines:
                        iv = int(imran_totals.get(m, 0))
                        dv = int(dash_totals.get(m, 0))
                        if iv == dv:
                            ok(f"{str(m):14s} Imran={iv:>8,}  Dashboard={dv:>8,}")
                        else:
                            diff = dv - iv
                            fail(f"{str(m):14s} Imran={iv:>8,}  Dashboard={dv:>8,}  ({diff:+,})")
                            mismatches.append(m)
                            errors.append(f"Machine total mismatch for {m}: Imran={iv:,}, Dashboard={dv:,} (diff={diff:+,})")

                    it = int(imran_totals.sum())
                    dt = int(dash_totals.sum())
                    try:
                        print(f"    {'─'*52}")
                    except UnicodeEncodeError:
                        print(f"    {'-'*52}")

                    if it == dt:
                        ok(f"{'TOTAL':14s} Imran={it:>8,}  Dashboard={dt:>8,}")
                    else:
                        fail(f"{'TOTAL':14s} Imran={it:>8,}  Dashboard={dt:>8,}  ({dt-it:+,})")
                        errors.append(f"Grand Total production mismatch: Imran={it:,}, Dashboard={dt:,} (diff={dt-it:+,})")

                    if mismatches:
                        warn(f"Mismatches in: {', '.join(str(m) for m in mismatches)}")
                    else:
                        ok("All machines match!")
    except Exception as e:
        warn(f"Machine totals cross-check error: {e}")
        errors.append(f"Machine totals cross-check error: {e}")

    # --- Part B: Summary Sheet comparison ---
    try:
        wb_imran = load_workbook(prod_path, data_only=True)
        summary_sheet_name = None
        for name in wb_imran.sheetnames:
            if name.lower().startswith("summary"):
                summary_sheet_name = name
                break
        
        if not summary_sheet_name:
            warn("No Summary sheet found in Production.xlsx")
            errors.append("No Summary sheet found in Production.xlsx")
        else:
            ws_sum = wb_imran[summary_sheet_name]
            excel_files = sorted(glob.glob(os.path.join(ALPHA_DIR, "Tubex*.xlsx")))
            if not excel_files:
                warn("No Tubex*.xlsx found — skipping summary check")
            else:
                wb_dash = load_workbook(excel_files[-1], data_only=True)
                ws_dash = wb_dash['Tubex_Dashboard']
                
                def _to_int(v):
                    if v is None: return 0
                    try: return int(float(str(v).replace(',', '').strip()))
                    except Exception: return 0

                # Define the checks: (Label, Imran Cell, Dash Cell)
                checks = [
                    ("Printing Production (Today)", "B14", "B6"),
                    ("Printing Production (MTD)",   "B15", "D6"),
                    ("PET Production (Today)",      "B3",  "B8"),
                    ("PET Production (MTD)",        "B4",  "D8"),
                    ("Tube Dispatch (MTD)",         "B22", "J6"),
                    ("PET Dispatch (MTD)",          "B11", "J8"),
                ]

                print(f"\n    {DIM}── Summary Sheet ({summary_sheet_name}) vs Dashboard ──{RESET}")
                summary_mismatches = 0
                for label, imran_cell, dash_cell in checks:
                    imran_val = _to_int(ws_sum[imran_cell].value)
                    dash_val = _to_int(ws_dash[dash_cell].value)
                    
                    if imran_val == dash_val:
                        ok(f"{label:28s} Imran ({imran_cell})={imran_val:>8,}  Dashboard ({dash_cell})={dash_val:>8,}")
                    else:
                        diff = dash_val - imran_val
                        fail(f"{label:28s} Imran ({imran_cell})={imran_val:>8,}  Dashboard ({dash_cell})={dash_val:>8,}  ({diff:+,})")
                        errors.append(f"Summary mismatch - {label}: Imran={imran_val:,}, Dashboard={dash_val:,} (diff={diff:+,})")
                        summary_mismatches += 1
                
                if summary_mismatches == 0:
                    ok("All Summary sheet KPIs match Dashboard!")
                else:
                    warn(f"{summary_mismatches} Summary sheet KPI mismatch(es) found!")
                
                wb_dash.close()
        wb_imran.close()
    except Exception as e:
        warn(f"Summary sheet cross-check error: {e}")
        errors.append(f"Summary sheet cross-check error: {e}")

    # --- Part C: Pending Tube Orders comparison ---
    try:
        print(f"\n    {DIM}── Pending Tube Orders: MRP vs PENDING ORDER file ──{RESET}")
        pending_files = []
        if os.path.exists(DOWNLOADS_DIR):
            for f in os.listdir(DOWNLOADS_DIR):
                if f.upper().startswith("PENDING ORDER ") and (f.endswith(".xlsx") or f.endswith(".xls")):
                    path = os.path.join(DOWNLOADS_DIR, f)
                    pending_files.append((path, os.path.getmtime(path)))
        
        if not pending_files:
            warn("No PENDING ORDER file found in Downloads — skipping comparison")
            errors.append("No PENDING ORDER file found in Downloads")
        else:
            pending_files.sort(key=lambda x: x[1], reverse=True)
            most_recent_pending = pending_files[0][0]
            pending_basename = os.path.basename(most_recent_pending)
            
            # Find the most recent Tubex file
            excel_files = sorted(glob.glob(os.path.join(ALPHA_DIR, "Tubex*.xlsx")))
            if not excel_files:
                warn("No Tubex*.xlsx found — skipping Pending Order check")
                errors.append("No Tubex*.xlsx found for Pending Order check")
            else:
                # 1. Read Tubex MRP sheet
                df_mrp = pd.read_excel(excel_files[-1], sheet_name='MRP', header=None)
                mrp_total = None
                for idx, row in df_mrp.iterrows():
                    val_4 = str(row[4]).strip() if pd.notna(row[4]) else ""
                    if val_4.upper() == 'TOTAL:':
                        mrp_total = row[7]
                        break
                
                if mrp_total is None:
                    warn(f"Could not find tube total 'TOTAL:' in MRP sheet of {os.path.basename(excel_files[-1])}")
                    errors.append(f"Could not find tube total in MRP sheet of {os.path.basename(excel_files[-1])}")
                else:
                    try:
                        mrp_total = int(float(str(mrp_total).replace(',', '').strip()))
                    except Exception as e:
                        warn(f"Error parsing MRP tube total: {e}")
                        errors.append(f"Error parsing MRP tube total: {e}")
                        mrp_total = None
                
                if mrp_total is not None:
                    # 2. Read Pending Order sheet
                    xls_pending = pd.ExcelFile(most_recent_pending)
                    pending_sheet_names = xls_pending.sheet_names
                    
                    # Today's date representation in DD-MM-YYYY
                    today_str = datetime.now().strftime("%d-%m-%Y")
                    
                    target_sheet = None
                    if today_str in pending_sheet_names:
                        target_sheet = today_str
                    else:
                        # Find most recent date sheet
                        date_sheets = []
                        for s in pending_sheet_names:
                            try:
                                dt = datetime.strptime(s.strip(), "%d-%m-%Y")
                                date_sheets.append((s, dt))
                            except ValueError:
                                continue
                        if date_sheets:
                            date_sheets.sort(key=lambda x: x[1], reverse=True)
                            target_sheet = date_sheets[0][0]
                    
                    if not target_sheet:
                        warn(f"No date sheet (DD-MM-YYYY) found in {pending_basename} — skipping comparison")
                        errors.append(f"No date sheet (DD-MM-YYYY) found in {pending_basename}")
                    else:
                        df_pending = pd.read_excel(most_recent_pending, sheet_name=target_sheet, header=None)
                        pending_total = None
                        for idx, row in df_pending.iterrows():
                            val_0 = str(row[0]).strip().upper() if pd.notna(row[0]) else ""
                            if val_0 == 'GRAND TOTAL':
                                pending_total = row[7]
                                break
                        
                        if pending_total is None:
                            warn(f"Could not find 'GRAND TOTAL' in {pending_basename} sheet {target_sheet}")
                            errors.append(f"Could not find 'GRAND TOTAL' in {pending_basename} sheet {target_sheet}")
                        else:
                            try:
                                pending_total = int(float(str(pending_total).replace(',', '').strip()))
                                
                                # Compare values
                                if mrp_total == pending_total:
                                    ok(f"Pending Tube Orders Match: {pending_basename} ({target_sheet}) = {pending_total:,}  MRP Sheet = {mrp_total:,}")
                                else:
                                    diff = pending_total - mrp_total
                                    fail(f"Pending Tube Orders Mismatch: {pending_basename} ({target_sheet})={pending_total:,}  MRP Sheet={mrp_total:,} (diff={diff:+,})")
                                    errors.append(f"Pending Tube Orders mismatch: {pending_basename} ({target_sheet})={pending_total:,}, MRP Sheet={mrp_total:,} (diff={diff:+,})")
                            except Exception as e:
                                warn(f"Error parsing Grand Total in {pending_basename} sheet {target_sheet}: {e}")
                                errors.append(f"Error parsing Grand Total in {pending_basename} sheet {target_sheet}: {e}")
    except Exception as e:
        warn(f"Pending Tube Orders cross-check error: {e}")
        errors.append(f"Pending Tube Orders cross-check error: {e}")

    return errors


# ═══════════════════════════════════════════════════════════════════════════
# STEP 7: SCREENSHOT
# ═══════════════════════════════════════════════════════════════════════════
def step_screenshot():
    header(7, "Dashboard screenshot...")

    html_path = os.path.join(ALPHA_DIR, "Tubex.html")
    if not os.path.exists(html_path):
        warn("Tubex.html not found")
        return

    date_str = datetime.now().strftime('%Y%m%d')
    ss_path = os.path.join(LOGS_DIR, f"dashboard_{date_str}.png")

    # Try Playwright
    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(viewport={'width': 430, 'height': 932})
            page.goto(f"file:///{html_path.replace(os.sep, '/')}")
            page.wait_for_timeout(2000)
            try:
                page.click('text=Dashboard', timeout=3000)
                page.wait_for_timeout(1000)
            except Exception:
                pass
            page.screenshot(path=ss_path, full_page=False)
            browser.close()

        ok(f"Saved: Logs/dashboard_{date_str}.png")
        if sys.platform == 'win32':
            os.startfile(ss_path)
            print(f"    {DIM}Image opened — share to WhatsApp{RESET}")
        return
    except ImportError:
        pass
    except Exception as e:
        warn(f"Playwright error: {e}")

    # Fallback — just open in browser
    warn("Playwright not installed — opening in browser instead")
    print(f"    {DIM}To enable: pip install playwright && playwright install chromium{RESET}")
    if sys.platform == 'win32':
        os.startfile(html_path)


# ═══════════════════════════════════════════════════════════════════════════
# STEP 8: ONEDRIVE BACKUP
# ═══════════════════════════════════════════════════════════════════════════
def step_onedrive_backup(skip=False):
    header(8, "Copying to OneDrive...")

    if skip:
        warn("Skipped (--skip-onedrive)")
        return

    onedrive_dir = r"C:\Users\HP\OneDrive\Alpha"
    
    try:
        cmd = ["robocopy", ALPHA_DIR, onedrive_dir, "/MIR", "/XD", ".git", "Logs", "__pycache__", "/R:1", "/W:1"]
        result = subprocess.run(cmd, capture_output=True)
        if result.returncode < 8:
            ok("Copied to OneDrive ✓")
        else:
            fail(f"OneDrive copy failed (exit code {result.returncode})")
    except Exception as e:
        fail(f"Error copying to OneDrive: {e}")


# ═══════════════════════════════════════════════════════════════════════════
# STEP 9: GIT PUSH
# ═══════════════════════════════════════════════════════════════════════════
def step_git_push(skip=False):
    header(9, "Pushing to GitHub...")

    if skip:
        warn("Skipped (--skip-git)")
        return

    # Check git is available
    try:
        subprocess.run(["git", "--version"], capture_output=True, check=True)
    except (FileNotFoundError, subprocess.CalledProcessError):
        warn("Git not installed or not in PATH — skipping push")
        return

    # Check it's a git repo
    result = subprocess.run(
        ["git", "-C", ALPHA_DIR, "rev-parse", "--git-dir"],
        capture_output=True
    )
    if result.returncode != 0:
        warn("Not a git repo — skipping push")
        return

    # Stage all changes
    subprocess.run(["git", "-C", ALPHA_DIR, "add", "-A"], capture_output=True)

    # Check if there's anything to commit
    result = subprocess.run(
        ["git", "-C", ALPHA_DIR, "diff", "--quiet", "--cached"],
        capture_output=True
    )
    if result.returncode == 0:
        ok("No changes to push")
        return

    # Commit and push
    msg = f"Daily update {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    subprocess.run(
        ["git", "-C", ALPHA_DIR, "commit", "-m", msg],
        capture_output=True
    )
    result = subprocess.run(
        ["git", "-C", ALPHA_DIR, "push", "origin", "main"],
        capture_output=True
    )

    if result.returncode == 0:
        ok("Pushed to GitHub ✓")
    else:
        fail("Git push failed — check internet/credentials")
        stderr = result.stderr.decode(errors='replace').strip()
        if stderr:
            print(f"    {DIM}{stderr}{RESET}")


def read_mismatches_log(log_path):
    import json
    import re
    from datetime import datetime
    
    inventory_warnings = []
    mapping_warnings = []
    
    state_file = os.path.join(LOGS_DIR, "previous_missing_items.json")
    state_data = {}
    if os.path.exists(state_file):
        try:
            with open(state_file, 'r', encoding='utf-8') as f:
                state_data = json.load(f)
        except Exception:
            pass

    today_str = datetime.now().strftime('%Y-%m-%d')
    last_run_date = state_data.get("last_run_date", "")
    
    if last_run_date != today_str:
        prev_missing = set(state_data.get("missing_today", []))
        state_data["missing_yesterday"] = list(prev_missing)
        state_data["last_run_date"] = today_str
    else:
        prev_missing = set(state_data.get("missing_yesterday", []))

    current_missing = set()

    if os.path.exists(log_path):
        with open(log_path, 'r', encoding='utf-8') as f:
            for line in f:
                l = line.strip()
                if not l or l.startswith('---') or l.startswith('==='):
                    continue
                if 'missing from inventory.xls' in l:
                    clean = l.replace('WARNING:', '').strip()
                    lower_clean = clean.lower()
                    
                    # 1. Ignore INKs completely in daily summary
                    if re.search(r'\binks?\b', lower_clean):
                        continue
                        
                    # Extract ID
                    m = re.search(r'Item ID\s+(\d+)', clean)
                    item_id = m.group(1) if m else None
                    if item_id:
                        current_missing.add(item_id)
                        
                    # 2. Hide if already missing yesterday, except exceptions
                    is_exception = re.search(r'\b(pet resin|master batch|slugs?)\b', lower_clean)
                    if item_id and item_id in prev_missing and not is_exception:
                        continue
                        
                    inventory_warnings.append(clean)
                else:
                    mapping_warnings.append(l)

    state_data["missing_today"] = list(current_missing)

    try:
        with open(state_file, 'w', encoding='utf-8') as f:
            json.dump(state_data, f)
    except Exception:
        pass

    return inventory_warnings, mapping_warnings


# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    banner()

    skip_wip  = '--skip-wip'  in sys.argv
    skip_prod = '--skip-prod' in sys.argv
    skip_git  = '--skip-git'  in sys.argv

    # Start logging
    log_path = setup_logging()

    start = time.time()
    
    all_errors = []

    step_backup()              # 1. Backup Excel
    
    erp_warnings = step_check_erp()  # 2. Check ERP exports
    all_errors.extend(erp_warnings)
    
    step_find_production(      # 3. Find Production file
        skip=skip_prod)
    step_wip(skip=skip_wip)    # 4. WIP update
    
    success = step_pipeline()  # 5. Run all 5 scripts
    if not success:
        all_errors.append("Pipeline execution had failures (one or more scripts failed)")

    # Read mismatches.log (populated during step 5)
    mismatch_log = os.path.join(LOGS_DIR, "mismatches.log")
    inv_warns, map_warns = read_mismatches_log(mismatch_log)

    crosscheck_errors = []
    if success:
        crosscheck_errors = step_crosscheck()  # 6. Cross-check
        
    step_screenshot()          # 7. Screenshot
    step_onedrive_backup()     # 8. OneDrive backup
    step_git_push(             # 9. Git push
        skip=skip_git)

    elapsed = time.time() - start

    # ── Unified Error Summary ──
    has_issues = bool(all_errors or inv_warns or map_warns or crosscheck_errors)
    error_summary_path = os.path.join(LOGS_DIR, "error_summary.txt")
    
    with open(error_summary_path, 'w', encoding='utf-8') as f_sum:
        def print_both(msg=""):
            print(msg)
            # Remove ANSI colors from file output
            clean_msg = msg
            for color in [GREEN, YELLOW, RED, CYAN, BOLD, DIM, RESET]:
                clean_msg = clean_msg.replace(color, "")
            f_sum.write(clean_msg + "\n")

        if has_issues:
            print_both()
            print_both(f"  {RED}{BOLD}╔══════════════════════════════════════════════════════════╗{RESET}")
            print_both(f"  {RED}{BOLD}║  ⚠ DAILY WORKFLOW ERROR SUMMARY                          ║{RESET}")
            print_both(f"  {RED}{BOLD}╚══════════════════════════════════════════════════════════╝{RESET}")
            print_both()
            
            if all_errors:
                print_both(f"  {YELLOW}{BOLD}[SYSTEM / FILE CHECK ISSUES]{RESET}")
                for err in all_errors:
                    print_both(f"    • {err}")
                print_both()
                
            if inv_warns:
                print_both(f"  {RED}{BOLD}[INVENTORY: ITEMS MISSING FROM ERP]{RESET}")
                print_both(f"  {DIM}  (These rows are highlighted in RED in Excel and zeroed out){RESET}")
                for err in inv_warns:
                    print_both(f"    • {err}")
                print_both()
                
            if crosscheck_errors:
                print_both(f"  {RED}{BOLD}[CROSS-CHECK MISMATCHES]{RESET}")
                for err in crosscheck_errors:
                    print_both(f"    • {err}")
                print_both()
                
            if map_warns:
                print_both(f"  {YELLOW}{BOLD}[MAPPING MISMATCHES]{RESET}")
                print_both(f"  {DIM}  (Unmapped products found during production/dispatch/FG processing){RESET}")
                for err in map_warns:
                    print_both(f"    • {err}")
                print_both()
        else:
            print_both()
            print_both(f"  {GREEN}{BOLD}✓ ALL CHECKS PASSED: No errors, missing items, or mismatches detected!{RESET}")

    # ── Final Summary ──
    print(f"\n  {'='*52}")
    if not has_issues:
        print(f"  {GREEN}{BOLD}  ALL DONE{RESET} in {elapsed:.0f} seconds")
    else:
        print(f"  {YELLOW}{BOLD}  COMPLETED WITH ISSUES{RESET} — see summary above")
    print(f"  {DIM}  Log: {os.path.basename(log_path)}{RESET}")
    print(f"  {DIM}  Error Summary: {os.path.basename(error_summary_path)}{RESET}")
    print(f"  {'='*52}\n")

    input("  Press Enter to close...")


if __name__ == '__main__':
    main()
