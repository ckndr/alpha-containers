import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl

wb = openpyxl.load_workbook(r'd:\Alpha\Aerosol_BOM.xlsx')
print('Sheets:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    print()
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        vals = [cell.value for cell in row]
        if any(v is not None for v in vals):
            print(vals)
