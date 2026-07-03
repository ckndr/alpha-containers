import openpyxl

wb = openpyxl.load_workbook('Tubex_v10_30.xlsx')
sheet = wb['MRP']

def get_formula(row_idx):
    # Products are in rows 3 to 8. We dynamically create the IF checks for these rows.
    parts = []
    for i in range(3, 9):
        part = f'IF((COUNTIFS(TableBOM[Product ID],$D${i},TableBOM[Item ID],$A{row_idx})>0)*($H${i}>0),$C${i}&", ","")'
        parts.append(part)
    
    concatenated = ' & '.join(parts)
    
    formula = f'=IF(LEN({concatenated})>1,LEFT({concatenated},LEN({concatenated})-2),"")'
    return formula

count = 0
for row_idx in range(12, sheet.max_row + 1):
    cell = sheet.cell(row=row_idx, column=8) # Column H
    if isinstance(cell.value, str) and cell.value.startswith('=IF(LEN('):
        cell.value = get_formula(row_idx)
        count += 1
        
wb.save('Tubex_v10_30.xlsx')
print(f"Formulas updated successfully for {count} rows.")
